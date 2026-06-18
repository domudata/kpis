# -*- coding: utf-8 -*-
import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import numpy as np
import io, locale, random, time, os, hashlib, json, base64
from datetime import datetime
import plotly.express as px
import plotly.graph_objects as go
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ============================================================
st.set_page_config(layout="wide", page_title="Dashboard KPI", initial_sidebar_state="expanded")
# ============================================================

QK = ["TAUX_REALISATION_CORRECTIF/PT","OT préparation <1 mois","OT préparation >3 mois",
      "OT préparation 1mois< <3mois","OT planification <1 mois","OT planification >3 mois",
      "OT planification 1mois< <3mois","OT exécution <1 mois","OT exécution >3 mois",
      "OT exécution 1mois< <3mois",
      "Performance Graissage","Performance Inspection","Performance Appels Systématiques"]
PK = ["appel avis approuvé","OT LANC ESTIME","Backlog préparation caractérisé",
      "Backlog planification caractérisé","OT CONFIME","OT_COR_EGAL",
      "OT Fiabilité","Total Avis de Panne"]
ALL_KPI = QK + PK

CIBLE = {"TAUX_REALISATION_CORRECTIF/PT":85,"OT préparation <1 mois":80,"OT préparation >3 mois":5,
         "OT préparation 1mois< <3mois":15,"OT planification <1 mois":80,"OT planification >3 mois":5,
         "OT planification 1mois< <3mois":15,"OT exécution <1 mois":80,"OT exécution >3 mois":5,
         "OT exécution 1mois< <3mois":15,"appel avis approuvé":95,"OT LANC ESTIME":100,
         "Backlog préparation caractérisé":100,"Backlog planification caractérisé":100,
         "OT CONFIME":100,"OT_COR_EGAL":100,
         "Performance Graissage":95,"Performance Inspection":95,"Performance Appels Systématiques":95,
         "OT Fiabilité":100,"Total Avis de Panne":100}

ACT_MAP = {"TAUX_REALISATION_CORRECTIF/PT":"Ameliorer le taux de realisation des OT.",
           "OT préparation <1 mois":"Reduire l'age de preparation des OT (< 1 mois).",
           "OT préparation >3 mois":"Traiter les OT avec preparation > 3 mois.",
           "OT planification <1 mois":"Reduire l'age de planification des OT (< 1 mois).",
           "OT planification >3 mois":"Traiter les OT avec planification > 3 mois.",
           "OT exécution <1 mois":"Reduire l'age d'execution des OT (< 1 mois).",
           "OT exécution >3 mois":"Traiter les OT avec execution > 3 mois.",
           "OT LANC ESTIME":"Estimer les couts des OT lances.",
           "Backlog préparation caractérisé":"Caracteriser le backlog de preparation.",
           "Backlog planification caractérisé":"Caracteriser le backlog de planification.",
           "OT CONFIME":"Confirmer les OT termines.",
           "OT_COR_EGAL":"Rapprocher les couts reels et budgetes.",
           "appel avis approuvé":"Creer un OT pour les avis sans ordre.",
           "OT préparation 1mois< <3mois":"Reduire les OT entre 1 et 3 mois.",
           "OT planification 1mois< <3mois":"Reduire les OT entre 1 et 3 mois.",
           "OT exécution 1mois< <3mois":"Reduire les OT entre 1 et 3 mois.",
           "Performance Graissage":"Ameliorer le taux de realisation des OT de graissage (Type 350).",
           "Performance Inspection":"Ameliorer le taux de realisation des OT d'inspection (Types 290,300,310).",
           "Performance Appels Systématiques":"Ameliorer le taux de realisation des appels systematiques (Type 360).",
           "OT Fiabilité":"Maintenir la fiabilite des OT a 100%.",
           "Total Avis de Panne":"Maintenir le suivi des avis de panne a 100%."}

KPI_RESP_MAP = {
    "TAUX_REALISATION_CORRECTIF/PT": "Chef d'atelier",
    "OT préparation <1 mois": "Préparateur BM",
    "OT préparation 1mois< <3mois": "Préparateur BM",
    "OT préparation >3 mois": "Préparateur BM",
    "OT planification <1 mois": "Planificateur BM",
    "OT planification 1mois< <3mois": "Planificateur BM",
    "OT planification >3 mois": "Planificateur BM",
    "OT exécution <1 mois": "Chef d'atelier",
    "OT exécution 1mois< <3mois": "Chef d'atelier",
    "OT exécution >3 mois": "Chef d'atelier",
    "appel avis approuvé": "Chef d'atelier",
    "OT LANC ESTIME": "Fiabilité",
    "Backlog préparation caractérisé": "Préparateur BM",
    "Backlog planification caractérisé": "Planificateur BM",
    "OT CONFIME": "Agent de saisie",
    "OT_COR_EGAL": "Agent de saisie",
    "Performance Graissage": "Chef d'atelier",
    "Performance Inspection": "Chef d'atelier",
    "Performance Appels Systématiques": "Chef d'atelier",
    "OT Fiabilité": "Fiabilité",
    "Total Avis de Panne": "Fiabilité"
}

LOWER_BETTER = ["OT préparation >3 mois","OT planification >3 mois","OT exécution >3 mois",
                "OT préparation 1mois< <3mois","OT planification 1mois< <3mois","OT exécution 1mois< <3mois"]

MP_KW = ["CRPR ATPD","CRPR ATMR","CRPR ATER","CRPR ATRS","CRPR ATMO","ATPD","ATMR","ATER","ATRS","ATMO"]
MPLAN_KW = ["ATPL ATEI","ATPL ATAL","ATPL ATER","ATPL AGAR","ATPL ATHS","ATEI","ATAL","ATAS","AGAR","ATHS"]

CONSIGNES_HSE = [
    "Port obligatoire des EPI avant toute intervention.","Port obligatoire du casque de securite.",
    "Port obligatoire des lunettes de protection.","Port obligatoire des gants adaptes au travail.",
    "Utiliser les protections auditives dans les zones bruyantes.","Verifier l'absence de tension avant toute intervention electrique.",
    "Respecter la procedure de consignation et deconsignation.","Ne jamais intervenir sur un equipement en marche.",
    "Baliser et securiser la zone de travail.","Maintenir le poste de travail propre et ordonne.",
    "Verifier l'etat des outils avant utilisation.","Utiliser uniquement du materiel homologue.",
    "Respecter les permis de travail en vigueur.","Identifier les risques avant de commencer une tache.",
    "Signaler immediatement toute situation dangereuse.","Signaler tout incident ou presque accident.",
    "Ne jamais neutraliser un dispositif de securite.","Verifier les detecteurs de gaz avant utilisation.",
    "Verifier la bonne ventilation des zones de travail.","Respecter les regles des espaces confines.",
    "Controler l'atmosphere avant d'entrer dans un espace confine.","Utiliser les points d'ancrage pour les travaux en hauteur.",
    "Verifier l'etat des echafaudages avant utilisation.","Securiser les outils lors des travaux en hauteur.",
    "Ne pas travailler seul lors d'operations a risque.","Controler les elingues avant chaque levage.",
    "Respecter les limites de charge des equipements.","Verifier l'etat des appareils de levage.",
    "Maintenir les voies de circulation degagees.","Respecter la signalisation de securite.",
    "Verifier les extincteurs a proximite du chantier.","Connaitre les issues de secours les plus proches.",
    "Respecter les procedures d'arret d'urgence.","Verifier les flexibles et raccords avant mise en service.",
    "Controler les fuites avant demarrage d'un equipement.","Respecter les distances de securite.",
    "Ne jamais contourner une procedure HSE.","Porter les EPI adaptes au risque identifie.",
    "Prevenir son responsable avant toute intervention particuliere.","Analyser les risques avant chaque demarrage de chantier.",
    "Verifier la stabilite des equipements.","Utiliser les bons outils pour la bonne tache.",
    "Respecter les consignes specifiques du chantier.","Ne jamais prendre de raccourci au detriment de la securite.",
    "Arreter immediatement les travaux en cas de danger.","Proteger l'environnement lors des interventions.",
    "Collecter et trier correctement les dechets.","Eviter toute pollution accidentelle.",
    "Respecter les consignes de stockage des produits dangereux.","Lire les fiches de securite avant manipulation.",
    "Verifier les equipements avant chaque prise de poste.","S'assurer de la disponibilite des moyens de secours.",
    "Communiquer clairement avec l'equipe avant intervention.","Respecter les regles de circulation des engins.",
    "Garder une vigilance permanente sur son environnement.","Prendre le temps d'effectuer le travail en securite.",
    "La securite est l'affaire de tous.","Chaque incident peut etre evite par la prevention.",
    "Aucun travail n'est plus urgent que la securite.","Zero accident commence par un comportement sur."]

def get_logo_base64():
    for path in ["logo.png", "./logo.png", "../logo.png"]:
        if os.path.exists(path):
            try:
                with open(path, "rb") as f:
                    return base64.b64encode(f.read()).decode()
            except Exception:
                pass
    return None

# ============================================================
# ✅ CORRECTION 2 : Date toujours lue depuis date.txt, plus de date en dur
# ============================================================
def get_date_from_file():
    if os.path.exists("date.txt"):
        try:
            with open("date.txt","r",encoding="utf-8") as f:
                date_lue = f.read().strip()
                if date_lue:
                    return date_lue
        except Exception: pass
    return datetime.now().strftime("%d/%m/%Y")

def contient_mot(t,lm):
    t=str(t); return any(m in t for l in lm for m in l.split())
def cat_age(a):
    if pd.isna(a): return "Inconnu"
    if a<=1: return "<1 mois"
    elif a>=3: return ">3 mois"
    return "1 mois < <3 mois"
def excr(df):
    if "Poste travail princ." in df.columns:
        return df[~df["Poste travail princ."].astype(str).str.contains("cresseur",case=False,na=False)].copy()
    return df

# ============================================================
# CACHED HEAVY DATA PREPARATION
# ============================================================
@st.cache_data(show_spinner=False)
def prepare_data(ot_bytes, av_bytes, date_str):
    raw_ot = pd.read_excel(io.BytesIO(ot_bytes))
    raw_av = pd.read_excel(io.BytesIO(av_bytes))
    raw_ot = excr(raw_ot)
    raw_av = excr(raw_av)
    
    for c in ["Créé le","Date de début planifiée","Date de clôture","Début réel","Fin réelle"]:
        if c in raw_ot.columns: raw_ot[c]=pd.to_datetime(raw_ot[c],errors="coerce")
    for c in ["Créé le","Début souhaité","Date de la clôture"]:
        if c in raw_av.columns: raw_av[c]=pd.to_datetime(raw_av[c],errors="coerce")
        
    now_ts = pd.to_datetime(date_str, format="%d/%m/%Y", errors='coerce')
    if pd.isna(now_ts): now_ts = pd.Timestamp.now()
    
    df = raw_ot.copy()
    
    df["Backlog preparation"]=np.where(df["Statut utilisateur"].apply(lambda x:contient_mot(x,MP_KW)),"CARACTERISE","NON CARACTERISE")
    df["Backlog planification"]=np.where(df["Statut utilisateur"].apply(lambda x:contient_mot(x,MPLAN_KW)),"CARACTERISE","NON CARACTERISE")
    df["Type Carac Prep"]=df["Statut utilisateur"].apply(lambda x: next((kw for kw in MP_KW if kw in str(x)), "NON CARACTERISE"))
    df["Type Carac Plan"]=df["Statut utilisateur"].apply(lambda x: next((kw for kw in MPLAN_KW if kw in str(x)), "NON CARACTERISE"))
    
    for dc,am,ac in [('Créé le',"amp","ap"),('Date de début planifiée',"amlp","alp"),('Date de début planifiée',"amex","aex")]:
        if dc in df.columns:
            df[am]=((now_ts.year-df[dc].dt.year)*12+(now_ts.month-df[dc].dt.month)).round(2)
            df[ac]=df[am].apply(cat_age)
        else: df[am]=np.nan; df[ac]="Inconnu"
        
    # ============================================================
    # ✅ CORRECTION 1 : OT CONFIME avec CLOT ou TCLO (au lieu de CLO)
    # ============================================================
    df["OT CONFIME"]=np.where((df["Statut système"].str.contains("CLOT",na=False)|df["Statut système"].str.contains("TCLO",na=False))&df["Statut système"].str.contains("CONF",na=False),"OUI","NON")
    
    df["Contient SOPL"]=df["Statut utilisateur"].str.contains("SOPL",na=False).map({True:1,False:0})
    df["OT LANC ESTIME"]=np.where(df["Total coûts budgétés"].fillna(0)==0,"NON","OUI")
    df["OT_COR_EGAL"]=np.where((df["Total coûts budgétés"].fillna(0)-df["Total coûts réels"].fillna(0))==0,"OUI","NON")
    df["_tw_num"]=pd.to_numeric(df.get("Type de travail",pd.Series(dtype=float)),errors="coerce")
    
    if "Statut système" in df.columns: df["Statut OT"]=df["Statut système"].fillna("").astype(str).str.strip().str.split().str[0]
    
    avf = raw_av[(raw_av["Ordre"].isna())|(raw_av["Ordre"].astype(str).str.strip()=="")].copy()
    
    apm = sorted(df[df["Poste travail princ."].astype(str).str.startswith(("SF1","SF2"),na=False)]["Poste travail princ."].dropna().unique().tolist())
    
    return df, avf, apm, now_ts

def save_kpis_to_excel(prows,pcols,qrows,qcols,ano_p_r,ano_p_c,ano_q_r,ano_q_c,sheet_name):
    kpis_dir="kpis"; os.makedirs(kpis_dir,exist_ok=True)
    filepath=os.path.join(kpis_dir,"indicateurs_kpis.xlsx")
    sn=str(sheet_name).replace("/","-").replace("\\","-").replace("*","").replace("?","").replace("[","").replace("]","")[:31]
    hf=Font(bold=True,color="FFFFFF",size=10); hfl=PatternFill(start_color="1E3A5F",end_color="1E3A5F",fill_type="solid")
    tf=Font(bold=True,size=12,color="1E3A5F")
    tb=Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
    try: wb=load_workbook(filepath)
    except Exception: wb=Workbook()
    if "Sheet" in wb.sheetnames: del wb["Sheet"]
    if sn in wb.sheetnames: del wb[sn]
    ws=wb.create_sheet(sn); rn=1
    def ws_sec(title,cols,rows,sr):
        ws.cell(row=sr,column=1,value=title).font=tf; sr+=1
        for j,c in enumerate(cols,1):
            cl=ws.cell(row=sr,column=j,value=c); cl.font=hf; cl.fill=hfl; cl.alignment=Alignment(horizontal='center'); cl.border=tb
        sr+=1
        for r in rows:
            for j,c in enumerate(cols,1):
                cl=ws.cell(row=sr,column=j,value=r.get(c,"")); cl.border=tb; cl.alignment=Alignment(horizontal='center')
            sr+=1
        return sr+1
    rn=ws_sec("INDICATEURS DE PERFORMANCE",pcols,prows,rn)
    if ano_p_c and ano_p_r: rn=ws_sec("ANOMALIES PERFORMANCE",ano_p_c,ano_p_r,rn)
    rn=ws_sec("INDICATEURS DE QUALITE",qcols,qrows,rn)
    if ano_q_c and ano_q_r: rn=ws_sec("ANOMALIES QUALITE",ano_q_c,ano_q_r,rn)
    try: wb.save(filepath)
    except Exception: pass

def load_historical_kpis(filepath):
    if not os.path.exists(filepath): return pd.DataFrame()
    try: wb=load_workbook(filepath,read_only=True,data_only=True)
    except Exception: return pd.DataFrame()
    records=[]; section=None; headers=None
    for sheet_name in wb.sheetnames:
        try:
            ws=wb[sheet_name]; rows_data=list(ws.iter_rows(values_only=True))
            for row in rows_data:
                cell0=str(row[0]).strip() if row[0] else ""
                if "INDICATEURS DE PERFORMANCE" in cell0.upper(): section="perf"; headers=None; continue
                elif "INDICATEURS DE QUALITE" in cell0.upper(): section="qual"; headers=None; continue
                elif "ANOMALIES" in cell0.upper(): section=None; continue
                if section and headers is None and cell0:
                    headers=[str(c).strip() if c else "" for c in row]; continue
                if section and headers and cell0 and cell0 not in ("Cible","Total general",""):
                    entry={"Date":sheet_name}
                    for j,h in enumerate(headers):
                        if j<len(row): entry[h]=row[j]
                    entry["_section"]=section; records.append(entry)
        except Exception: continue
    wb.close()
    if not records: return pd.DataFrame()
    df=pd.DataFrame(records)
    df["Date_parsed"]=pd.to_datetime(df["Date"].str.replace("-","/"),format="%d/%m/%Y",errors="coerce")
    return df.sort_values("Date_parsed").reset_index(drop=True)

def calculate_variations(hist_df):
    if hist_df.empty or "Date" not in hist_df.columns: return pd.DataFrame()
    dates=sorted(hist_df["Date"].unique())
    if len(dates)<2: return pd.DataFrame()
    perf_df=hist_df[hist_df["_section"]=="perf"].copy()
    qual_df=hist_df[hist_df["_section"]=="qual"].copy()
    variations=[]
    for i in range(1,len(dates)):
        prev_date,curr_date=dates[i-1],dates[i]
        prev_perf=perf_df[perf_df["Date"]==prev_date].set_index("Poste de travail") if "Poste de travail" in perf_df.columns else pd.DataFrame()
        curr_perf=perf_df[perf_df["Date"]==curr_date].set_index("Poste de travail") if "Poste de travail" in perf_df.columns else pd.DataFrame()
        prev_qual=qual_df[qual_df["Date"]==prev_date].set_index("Poste de travail") if "Poste de travail" in qual_df.columns else pd.DataFrame()
        curr_qual=qual_df[qual_df["Date"]==curr_date].set_index("Poste de travail") if "Poste de travail" in qual_df.columns else pd.DataFrame()
        for sec_name,prev_d,curr_d,kpi_list in [("Performance",prev_perf,curr_perf,QK+["Score Performance"]),("Qualite",prev_qual,curr_qual,PK+["Score Qualite"])]:
            for poste in set(prev_d.index)&set(curr_d.index):
                for kpi in kpi_list:
                    if kpi not in prev_d.columns or kpi not in curr_d.columns: continue
                    try: pv=float(prev_d.loc[poste,kpi])
                    except Exception: continue
                    try: cv=float(curr_d.loc[poste,kpi])
                    except Exception: continue
                    diff=cv-pv; pct=(diff/pv*100) if pv!=0 else (100 if cv!=0 else 0)
                    if abs(diff)<=0.5: trend="stabilite"
                    elif diff>0.5: trend="hausse"
                    else: trend="baisse"
                    sens = "Stable"
                    if trend != "stabilite":
                        if (trend == "hausse" and kpi not in LOWER_BETTER) or (trend == "baisse" and kpi in LOWER_BETTER):
                            sens = "Amelioration"
                        else:
                            sens = "Degradation"
                    variations.append({"Date precedente":prev_date,"Date actuelle":curr_date,"Poste":poste,
                        "Type":sec_name,"KPI":kpi,"Valeur precedente":round(pv,2),"Valeur actuelle":round(cv,2),
                        "Ecart":round(diff,2),"Ecart %":round(pct,2),"Tendance":trend, "Sens":sens})
    return pd.DataFrame(variations)

def generate_journal(var_df):
    if var_df.empty: return pd.DataFrame()
    j=var_df.copy(); j["Significatif"]=j["Ecart %"].abs()>=5
    j=j[j["Significatif"]].copy()
    return j.sort_values(["Date actuelle","Ecart %"],ascending=[True,False])

def calculate_rankings(var_df):
    if var_df.empty: return pd.DataFrame(),pd.DataFrame()
    scores={}
    for poste in var_df["Poste"].unique():
        pv=var_df[var_df["Poste"]==poste].copy()
        scores[poste]=sum((-r["Ecart %"] if r["KPI"] in LOWER_BETTER else r["Ecart %"]) for _,r in pv.iterrows())
    ranked=sorted(scores.items(),key=lambda x:x[1],reverse=True)
    return pd.DataFrame(ranked[:5],columns=["Poste","Score variation"]),pd.DataFrame(ranked[-5:][::-1],columns=["Poste","Score variation"])

def inject_custom_css():
    st.markdown("""<style>
    section[data-testid="stSidebar"]{width:250px!important}
    .main .block-container{max-width:100%!important;width:100%!important;padding-left:0.5rem!important;padding-right:0.5rem!important}
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700;800&display=swap');
    :root{--p:#1e3a5f;--pl:#2c5282;--b:#e2e8f0;--r:10px}
    *{box-sizing:border-box;margin:0;padding:0}
    .stApp{background:#edf2f7;font-family:'Inter',sans-serif}
    .main .block-container{padding-top:.8rem;padding-bottom:.8rem}
    .stTabs,.stTabs>div,.stTabs [data-baseweb="tab-list"]{width:100%!important;max-width:100%!important}
    .mh{background:linear-gradient(135deg,var(--p),var(--pl));padding:10px 20px;border-radius:var(--r);margin-bottom:6px;box-shadow:0 6px 20px rgba(0,0,0,.1);overflow:hidden;display:flex;align-items:center;gap:12px}
    .mh h1{color:#fff;font-size:40px;font-weight:800;margin:0;display:inline;flex:1}
    .mh .logo{height:47px;width:auto;max-width:140px;object-fit:contain;flex-shrink:0;border-radius:4px}
    .mh .db{background:rgba(255,255,255,.15);padding:3px 12px;border-radius:14px;color:#fff;font-size:21px;font-weight:500;border:1px solid rgba(255,255,255,.2);white-space:nowrap;flex-shrink:0}
    .cr{display:grid;grid-template-columns:repeat(4,1fr);gap:6px;margin-bottom:6px}
    .cc{background:#fff;border-radius:var(--r);padding:10px 12px;box-shadow:0 2px 8px rgba(0,0,0,.04);border:1px solid var(--b);text-align:center}
    .cc .cv{font-size:26px;font-weight:900;line-height:1}
    .cc .cl{font-size:11px;color:#718096;font-weight:700;text-transform:uppercase;letter-spacing:.5px;margin-top:2px}
    .cc.c1{border-top:3px solid #3182ce}.cc.c1 .cv{color:#2b6cb0}
    .cc.c2{border-top:3px solid #38a169}.cc.c2 .cv{color:#276749}
    .cc.c3{border-top:3px solid #805ad5}.cc.c3 .cv{color:#6b46c1}
    .cc.c4{border-top:3px solid #e53e3e}.cc.c4 .cv{color:#c53030}
    .cc.c5{border-top:3px solid #2b6cb0}.cc.c5 .cv{color:#2b6cb0}
    .cc.c6{border-top:3px solid #3182ce}.cc.c6 .cv{color:#3182ce}
    .cc.c7{border-top:3px solid #d69e2e}.cc.c7 .cv{color:#975a16}
    .cc.c8{border-top:3px solid #ecc94b}.cc.c8 .cv{color:#b7791f}
    .stl{font-size:15px;font-weight:700;color:var(--p);margin:6px 0 2px 0;padding-left:10px;border-left:3px solid var(--pl)}
    .stl.q{border-left-color:#3182ce}.stl.p{border-left-color:#38a169}.stl.a{border-left-color:#e53e3e}.stl.c{border-left-color:#805ad5}.stl.s{border-left-color:#d69e2e}
    .tw{width:100%;border-collapse:collapse;font-family:'Inter',sans-serif;font-size:12px;display:block;overflow-x:auto;-webkit-overflow-scrolling:touch;margin:0}
    .tw thead th{background:var(--p);color:#fff;font-weight:700;font-size:11px;text-transform:uppercase;letter-spacing:.3px;padding:5px 6px;border:none;white-space:nowrap;position:sticky;top:0;z-index:10}
    .tw.qt thead th{background:linear-gradient(135deg,#2b6cb0,#3182ce)}
    .tw.pt thead th{background:linear-gradient(135deg,#276749,#38a169)}
    .tw.at thead th{background:linear-gradient(135deg,#c53030,#e53e3e)}
    .tw.st thead th{background:linear-gradient(135deg,#975a16,#d69e2e)}
    .tw.omt thead th{background:linear-gradient(135deg,#6b46c1,#805ad5)}
    .tw.tht thead th{background:linear-gradient(135deg,#9b2c2c,#e53e3e)}
    .tw thead th:first-child{z-index:11;left:0}
    .tw tbody td:first-child{position:sticky;left:0;background:#fff;z-index:5;border-right:1px solid #edf2f7;color:#1a202c !important}
    .tw tbody tr:nth-child(even) td:first-child{background:#f7fafc}
    .tw tbody tr:hover td:first-child{background:#ebf8ff}
    .tw.cb td:first-child{background:#2b6cb0!important;color:#fff!important}
    .tw.tr td:first-child{background:#e2e8f0!important;color:#1a202c!important}
    .tw tbody td{padding:4px 6px;border-bottom:1px solid #edf2f7;white-space:nowrap;color:#1a202c !important}
    .tw tbody tr:nth-child(even) td{background:#f7fafc}
    .tw tbody tr:hover td{background:#ebf8ff!important}
    .cb td{background:#2b6cb0!important;color:#fff!important;font-weight:700!important;font-size:12px!important}
    .tr td{background:#e2e8f0!important;color:#1a202c !important;font-weight:800!important;font-size:12px!important}
    
    .stTabs [data-baseweb="tab-list"]{gap:6px;background:#e2e8f0;padding:6px;border-radius:8px;margin-bottom:8px}
    .stTabs [data-baseweb="tab"]{
        border-radius:6px;
        padding:12px 22px;
        font-weight:700;
        font-size:20px;
        line-height:1.5;
        min-height:48px;
    }
    .stTabs [data-baseweb="tab"] span,
    .stTabs [data-baseweb="tab"] > div{
        font-size:22px !important;
    }
    .stTabs [aria-selected="true"]{
        background:#fff!important;
        color:var(--p)!important;
        box-shadow:0 3px 8px rgba(0,0,0,.1);
        font-size:21px;
    }
    .stTabs [data-baseweb="tab"] svg{width:22px;height:22px}
    
    .ca{background:#fff;border-radius:var(--r);padding:10px;margin-top:4px;border:1px solid var(--b);box-shadow:0 1px 4px rgba(0,0,0,.02)}
    .ca .ct{font-size:14px;font-weight:700;margin-bottom:6px;padding-bottom:4px;border-bottom:1px solid var(--b)}
    .car{display:flex;align-items:center;margin-bottom:6px;font-size:12px}
    .car:last-child{margin-bottom:0}
    .car .cal{width:260px;font-weight:600;color:var(--p);text-align:right;padding-right:8px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .car .cab{flex:1;height:26px;background:#edf2f7;border-radius:4px;overflow:visible;position:relative}
    .car .caf{height:100%;border-radius:4px;transition:width .3s}
    
    .car .target-mark{
        position:absolute;
        top:-4px;
        bottom:-4px;
        width:3px;
        background:#e53e3e;
        z-index:20;
        transform:translateX(-50%);
        box-shadow:0 0 6px rgba(229,62,62,.9),0 0 2px rgba(0,0,0,.4);
        border-radius:2px;
    }
    .car .target-mark::before{
        content:"";
        position:absolute;
        top:-5px;
        left:50%;
        transform:translateX(-50%);
        width:0;height:0;
        border-left:5px solid transparent;
        border-right:5px solid transparent;
        border-top:6px solid #e53e3e;
    }
    .car .cav-out{font-size:12px;font-weight:800;color:#1a202c;min-width:55px;text-align:right;padding-left:6px}
    .car .cav-tgt{font-size:10px;font-weight:700;color:#1a202c;min-width:42px;text-align:right;padding-left:4px;opacity:.7}
    .gbr{display:flex;align-items:center;padding:3px 0;font-size:12px;border-bottom:1px solid #f7fafc}
    .gbr:last-child{border:none}
    .gbr-l{width:160px;font-weight:600;color:#1a202c;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;font-size:11px}
    .gbr-g{display:flex;align-items:center;gap:4px;flex:1;position:relative}
    .gbr-target{position:absolute;left:90%;top:-4px;bottom:-4px;width:3px;background:#e53e3e;z-index:10;box-shadow:0 0 6px rgba(229,62,62,.8);border-radius:2px}
    .gbr-target-label{position:absolute;left:90%;top:-20px;transform:translateX(-50%);font-size:9px;font-weight:800;color:#fff;background:#e53e3e;padding:1px 5px;border-radius:3px;white-space:nowrap;z-index:11;box-shadow:0 1px 3px rgba(0,0,0,.2)}
    .gbr-w{flex:1;height:22px;background:#edf2f7;border-radius:3px;overflow:hidden}
    .gbr-f{height:100%;border-radius:3px}
    .gb-p{background:linear-gradient(90deg,#2b6cb0,#4299e1)}.gb-q{background:linear-gradient(90deg,#276749,#48bb78)}
    .gbr-v{font-size:11px;font-weight:800;min-width:48px;text-align:right;color:#1a202c}
    .gbr-legend{display:flex;gap:14px;margin-bottom:10px;font-size:12px;font-weight:700;align-items:center}
    .gbr-legend span{display:flex;align-items:center;gap:5px}
    .gbr-legend i{display:inline-block;width:14px;height:14px;border-radius:2px}
    .gbr-legend .target-icon{display:inline-block;width:3px;height:14px;background:#e53e3e;border-radius:1px;box-shadow:0 0 3px rgba(229,62,62,.6)}
    .cg{display:grid;grid-template-columns:1fr 1fr;gap:6px}
    .cg>div{background:#fff;border-radius:var(--r);padding:8px 10px;border:1px solid var(--b)}
    .cg .ct{font-size:13px;font-weight:700;margin-bottom:4px;padding-bottom:3px;border-bottom:1px solid var(--b)}
    .cgr{display:flex;align-items:center;padding:3px 0;font-size:12px;border-bottom:1px solid #f7fafc}
    .cgr:last-child{border:none}
    .cgr .rk{width:18px;font-weight:800;text-align:center}
    .cgr .pn{flex:1;font-weight:600;color:#1a202c;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .cgr .ps{font-weight:800;min-width:55px;text-align:right}
    .dgrid{display:grid;grid-template-columns:1fr 1fr;gap:6px}
    .stButton>button[kind="primary"]{background:linear-gradient(135deg,var(--p),var(--pl));border:none;border-radius:6px;padding:8px 14px;font-weight:700;font-size:15px;width:100%}
    ::-webkit-scrollbar{width:5px;height:5px}::-webkit-scrollbar-track{background:#f1f1f1}::-webkit-scrollbar-thumb{background:#cbd5e0;border-radius:3px}
    
    div[data-testid="stSidebar"]{
        background:linear-gradient(180deg,#1e40af 0%,#1e3a8a 50%,#1e3a5f 100%)!important;
    }
    div[data-testid="stSidebar"]*{color:rgba(255,255,255,.9)!important}
    div[data-testid="stSidebar"] .stSelectbox label,div[data-testid="stSidebar"] .stMultiSelect label,div[data-testid="stSidebar"] .stDateInput label,div[data-testid="stSidebar"] .stCheckbox label,div[data-testid="stSidebar"] .stTextInput label{color:rgba(255,255,255,.9)!important;font-weight:600;font-size:13px;text-transform:uppercase;letter-spacing:.5px}
    div[data-testid="stSidebar"] div[data-testid="stWidget"]{background:rgba(255,255,255,.1);border-radius:6px;padding:5px 10px;margin-bottom:5px;border:1px solid rgba(255,255,255,.15)}
    div[data-testid="stSidebar"] .stSelectbox>div>div,div[data-testid="stSidebar"] .stMultiSelect>div>div,div[data-testid="stSidebar"] .stDateInput>div>div,div[data-testid="stSidebar"] .stTextInput>div>div{background:rgba(255,255,255,.95)!important;border-radius:5px}
    
    .es{text-align:center;padding:14px;color:#718096;font-size:14px}
    .synth-tbl{width:100%;border-collapse:collapse;font-family:'Inter',sans-serif;font-size:12px}
    .synth-tbl thead th{background:var(--p);color:#fff;font-weight:700;font-size:11px;padding:5px 8px;border:none;white-space:nowrap;position:sticky;top:0}
    .synth-tbl tbody td{padding:4px 8px;border-bottom:1px solid #edf2f7;text-align:center;color:#1a202c !important}
    .synth-tbl tbody tr:nth-child(even) td{background:#f7fafc}
    .synth-tbl tbody tr:hover td{background:#ebf8ff!important}
    .synth-tbl .poste-cell{text-align:left;font-weight:700;white-space:nowrap;min-width:140px;color:#1a202c !important}
    div[data-testid="stHorizontalBlock"]{align-items:center!important}
    
    [data-testid="stHeaderActionElements"]{display:none !important;}
    [data-testid="stActionButtonContainer"]{display:none !important;}
    
    .footer {
        text-align: center;
        margin-top: 30px;
        padding: 15px;
        color: #718096;
        font-size: 13px;
        border-top: 1px solid #e2e8f0;
        font-weight: 600;
    }

    div[data-testid="stDataEditor"] table, 
    div[data-testid="stDataEditor"] th, 
    div[data-testid="stDataEditor"] td {
        font-size: 18px !important;
        line-height: 1.4 !important;
        white-space: normal !important;
        word-wrap: break-word !important;
    }
    div[data-testid="stDataEditor"] [data-testid="stMarkdownContainer"] {
        font-size: 18px !important;
    }
    div[data-testid="stDataEditor"] [data-testid="stTable"] {
        overflow-x: hidden !important;
        width: 100% !important;
    }

    @media(max-width:768px){
        .cr{grid-template-columns:repeat(2,1fr)}
        .mh{padding:8px 10px;gap:8px}
        .mh h1{font-size:18px}
        .mh .logo{height:35px;max-width:70px}
        .mh .db{font-size:11px;padding:2px 8px}
        .cg,.dgrid{grid-template-columns:1fr}
        .car{flex-wrap:wrap;gap:2px}
        .car .cal{width:100%;text-align:left;padding-right:0;margin-bottom:2px}
        .car .cab{flex:1 1 70%}
        .car .cav-out,.car .cav-tgt{flex:1 1 15%;min-width:40px}
        .gbr{flex-direction:column;align-items:flex-start;gap:4px}
        .gbr-l{width:100%;margin-bottom:2px}
        .gbr-g{width:100%;flex-wrap:wrap}
        .gbr-w{flex:1 1 45%}
        .gbr-v{flex:1 1 10%;min-width:40px}
        .tw{font-size:10px}
        .tw thead th,.tw tbody td{padding:3px 4px}
        .stl{font-size:13px}
        .stTabs [data-baseweb="tab"]{padding:8px 12px;font-size:15px}
        .stTabs [data-baseweb="tab"] span{font-size:16px !important}
        div[data-testid="stDataEditor"] table, div[data-testid="stDataEditor"] th, div[data-testid="stDataEditor"] td { font-size: 14px !important; }
    }
    
    @media print {
        section[data-testid="stSidebar"], 
        header[data-testid="stHeader"], 
        div[data-testid="stToolbar"], 
        div[data-testid="stHeaderActionElements"],
        footer, 
        .stDeployButton, 
        #MainMenu { display: none !important; }
        
        .main .block-container { 
            padding-top: 0 !important; 
            padding-left: 0 !important; 
            padding-right: 0 !important; 
            max-width: 100% !important; 
        }
        
        .stButton, .stDownloadButton { display: none !important; }
        
        * {
            -webkit-print-color-adjust: exact !important;
            print-color-adjust: exact !important;
        }
        
        .stTabs [data-baseweb="tab-list"] { display: none !important; }
        .stTabs [data-baseweb="tab-panel"] { 
            display: block !important; 
            page-break-inside: avoid; 
        }
        
        .tw, .synth-tbl { page-break-inside: avoid; overflow: visible !important; }
    }
    </style>""",unsafe_allow_html=True)

# ============================================================
def main():
    try: locale.setlocale(locale.LC_ALL,'fr_FR.UTF-8')
    except Exception:
        try: locale.setlocale(locale.LC_ALL,'fr_FR')
        except Exception: pass
    inject_custom_css()
    
    # ✅ CORRECTION 2 : Toujours lire la date depuis date.txt
    fichier_date=get_date_from_file()

    if "hse_affiche" not in st.session_state: st.session_state.hse_affiche=False
    if not st.session_state.hse_affiche:
        c=random.choice(CONSIGNES_HSE)
        st.markdown("""<div style="min-height:100vh;display:flex;flex-direction:column;align-items:center;justify-content:center;background:linear-gradient(135deg,#1a365d,#2d3748,#1a365d);padding:40px">
        <div style="font-size:64px;margin-bottom:20px">🦺</div>
        <h1 style="text-align:center;font-size:46px;color:#fff;font-weight:900;margin:0">HSE - CONSIGNE DE SECURITE</h1>
        <p style="text-align:center;color:rgba(255,255,255,.6);font-size:22px;margin-top:8px;letter-spacing:3px;text-transform:uppercase">Securite - Sante - Environnement</p>
        <div style="background:linear-gradient(135deg,#f6e05e,#ed8936);padding:36px 48px;border-radius:20px;font-size:32px;font-weight:700;text-align:center;margin:40px 0;color:#1a202c;max-width:800px;box-shadow:0 20px 60px rgba(0,0,0,.3)">⚠️ %s</div>
        <h2 style="text-align:center;color:#48bb78;font-size:36px;font-weight:900">Aucun travail n'est plus urgent que la securite</h2>
        <div style="margin-top:40px;width:200px;height:4px;background:rgba(255,255,255,.1);border-radius:2px;overflow:hidden"><div style="width:100%%;height:100%%;background:linear-gradient(90deg,#48bb78,#38a169);border-radius:2px;animation:ld 5.5s ease-in-out forwards"></div></div>
        <style>@keyframes ld{from{width:0}to{width:100%%}}</style></div>"""%c,unsafe_allow_html=True)
        time.sleep(6); st.session_state.hse_affiche=True; st.rerun(); st.stop()

    def ckpi(n,d,sz=100): return np.where(d==0,sz,(n/d)*100)
    def cpiv(df,f,c,p):
        return pd.pivot_table(df[f],index="Poste travail princ.",columns=c,values="Ordre",aggfunc="count",fill_value=0).reindex(p,fill_value=0)
    def get_text_col(df):
        for c in ["Désignation","Designation","Désignation OT","Texte ordre","Texte","Description","Libellé","Libelle"]:
            if c in df.columns: return c
        for c in df.columns:
            if df[c].dtype=='object' and any(kw in str(c).lower() for kw in ['sign','text','desc','libell']):
                return c
        return None
    def build_statut_pivot(df_sub, posts):
        if df_sub.empty:
            return pd.DataFrame(index=posts, columns=["CRÉÉ","LANC","CLOT","TCLO","Total"]).fillna(0).astype(int)
        piv=pd.pivot_table(df_sub, index="Poste travail princ.", columns="Statut OT", values="Ordre", aggfunc="count", fill_value=0)
        for s in ["CRÉÉ","LANC","CLOT","TCLO"]:
            if s not in piv.columns: piv[s]=0
        piv["Total"]=piv[["CRÉÉ","LANC","CLOT","TCLO"]].sum(axis=1)
        return piv.reindex(posts, fill_value=0).fillna(0).astype(int)
    def html_statut_pivot(piv_df, table_class):
        cols=["Poste de travail","CRÉÉ","LANC","CLOT","TCLO","Total"]
        h='<table class="tw %s"><thead><tr>'%table_class+''.join('<th>%s</th>'%c for c in cols)+'</tr></thead><tbody>'
        for poste,row in piv_df.iterrows():
            h+='<tr><td style="font-weight:600">%s</td>'%poste
            for c in ["CRÉÉ","LANC","CLOT","TCLO"]:
                h+='<td style="text-align:center">%d</td>'%int(row.get(c,0))
            h+='<td style="text-align:center;font-weight:800">%d</td>'%int(row.get("Total",0))
            h+='</tr>'
        h+='<tr class="tr"><td>Total</td>'
        for c in ["CRÉÉ","LANC","CLOT","TCLO"]:
            h+='<td style="text-align:center">%d</td>'%int(piv_df[c].sum())
        h+='<td style="text-align:center">%d</td>'%int(piv_df["Total"].sum())
        h+='</tr></tbody></table>'
        return h
        
    def show_pie_pair(piv_df, title_prefix):
        global_counts=piv_df[["CRÉÉ","LANC","CLOT","TCLO"]].sum()
        global_counts=global_counts[global_counts>0]
        realised=global_counts.get("CLOT",0)+global_counts.get("TCLO",0)
        not_realised=global_counts.sum()-realised
        if not global_counts.empty:
            fig1=px.pie(global_counts, names=global_counts.index, values=global_counts.values,
                title="%s — Par Statut OT"%title_prefix,
                color_discrete_sequence=["#e53e3e","#d69e2e","#38a169","#3182ce"])
            fig1.update_traces(textposition='inside',textinfo='percent+value',textfont_size=11)
            fig1.update_layout(margin=dict(t=40,b=10,l=10,r=10),height=300,legend=dict(font_size=10,orientation="h",yanchor="bottom",y=-0.1))
            st.plotly_chart(fig1,use_container_width=True)
        else:
            st.markdown('<div class="es">Aucune donnee</div>',unsafe_allow_html=True)
            
        if global_counts.sum()>0:
            pie2_data=pd.DataFrame({"Statut":["Réalisés (CLOT+TCLO)","Non Réalisés"],"Nombre":[realised,not_realised]})
            fig2=px.pie(pie2_data, names="Statut", values="Nombre",
                title="%s — Réalisés vs Non Réalisés"%title_prefix,
                color="Statut", color_discrete_map={"Réalisés (CLOT+TCLO)":"#38a169","Non Réalisés":"#e53e3e"})
            fig2.update_traces(textposition='inside',textinfo='percent+value',textfont_size=11)
            fig2.update_layout(margin=dict(t=40,b=10,l=10,r=10),height=300,legend=dict(font_size=10,orientation="h",yanchor="bottom",y=-0.1))
            st.plotly_chart(fig2,use_container_width=True)
        else:
            st.markdown('<div class="es">Aucune donnee</div>',unsafe_allow_html=True)

    def show_simple_pie(piv_df, title, keep_non_carac=False):
        if not keep_non_carac and "NON CARACTERISE" in piv_df.columns:
            piv_df = piv_df.drop(columns=["NON CARACTERISE"])
        counts = piv_df.sum()
        counts = counts[counts > 0]
        if not counts.empty:
            color_map = {"CARACTERISE": "#38a169", "NON CARACTERISE": "#e53e3e"}
            colors = [color_map.get(str(c), None) for c in counts.index]
            fig = px.pie(counts, names=counts.index, values=counts.values, title=title,
                color=counts.index,
                color_discrete_map={k: v for k, v in zip(counts.index, colors)} if any(colors) else None)
            fig.update_traces(textposition='inside', textinfo='percent+value', textfont_size=11)
            fig.update_layout(margin=dict(t=40, b=10, l=10, r=10), height=280,
                              legend=dict(font_size=10, orientation="h", yanchor="bottom", y=-0.1))
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.markdown('<div class="es">Aucune donnee</div>', unsafe_allow_html=True)

    def calc_kpis(df_i, av_i, now_ts, posts):
        res={}; df=df_i.copy(); av=av_i.copy()
        res['dfp']=df
        filt_corr=(df["Nº appel pl.entret."].fillna(0)==0)&(df["Contient SOPL"]==1)
        an=cpiv(df,filt_corr,"Statut OT",posts)
        for c in ["CLOT","CRÉÉ","LANC","TCLO"]: an[c]=an.get(c,0)
        an["OT_CLOTURES"]=an["CLOT"]+an["TCLO"]
        an["TOTAL_OT"]=an[["CLOT","CRÉÉ","LANC","TCLO"]].sum(axis=1)
        an["TAUX_REALISATION_CORRECTIF/PT"]=np.where(an["TOTAL_OT"]==0,100.0,ckpi(an["OT_CLOTURES"],an["TOTAL_OT"]))
        
        pr=cpiv(df,(df["Statut OT"]=="CRÉÉ")&(df["Statut utilisateur"].str.contains("CRPR",na=False)),"ap",posts)
        for c in ["<1 mois",">3 mois","1 mois < <3 mois","Inconnu"]: pr[c]=pr.get(c,0)
        pr["Total"]=pr[["<1 mois","1 mois < <3 mois",">3 mois","Inconnu"]].sum(axis=1)
        pr["OT préparation <1 mois"]=ckpi(pr["<1 mois"],pr["Total"]); pr["OT préparation >3 mois"]=ckpi(pr[">3 mois"],pr["Total"],0); pr["OT préparation 1mois< <3mois"]=ckpi(pr["1 mois < <3 mois"],pr["Total"],0)
        
        pl=cpiv(df,(df["Statut OT"]=="LANC")&(df["Statut utilisateur"].str.contains("ATPL",case=False,na=False)),"alp",posts)
        for c in ["<1 mois",">3 mois","1 mois < <3 mois","Inconnu"]: pl[c]=pl.get(c,0)
        pl["Total"]=pl[["<1 mois","1 mois < <3 mois",">3 mois","Inconnu"]].sum(axis=1)
        pl["OT planification <1 mois"]=ckpi(pl["<1 mois"],pl["Total"]); pl["OT planification >3 mois"]=ckpi(pl[">3 mois"],pl["Total"],0); pl["OT planification 1mois< <3mois"]=ckpi(pl["1 mois < <3 mois"],pl["Total"],0)
        
        ex=cpiv(df,(df["Statut OT"]=="LANC")&(df["Contient SOPL"]==1),"aex",posts)
        for c in ["<1 mois",">3 mois","1 mois < <3 mois","Inconnu"]: ex[c]=ex.get(c,0)
        ex["Total"]=ex[["<1 mois","1 mois < <3 mois",">3 mois","Inconnu"]].sum(axis=1)
        ex["OT exécution <1 mois"]=ckpi(ex["<1 mois"],ex["Total"]); ex["OT exécution >3 mois"]=ckpi(ex[">3 mois"],ex["Total"],0); ex["OT exécution 1mois< <3mois"]=ckpi(ex["1 mois < <3 mois"],ex["Total"],0)
        
        la=pd.pivot_table(df[df["Statut OT"]=="LANC"],index="Poste travail princ.",columns="OT LANC ESTIME",values="Ordre",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
        for c in ["OUI","NON"]: la[c]=la.get(c,0)
        la["Total"]=la["OUI"]+la["NON"]; la["OT LANC ESTIME"]=ckpi(la["OUI"],la["Total"])
        
        pc=pd.pivot_table(df[df["Statut OT"]=="CRÉÉ"],index="Poste travail princ.",columns="Backlog preparation",values="Ordre",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
        for c in ["CARACTERISE","NON CARACTERISE"]: pc[c]=pc.get(c,0)
        pc["Total"]=pc["CARACTERISE"]+pc["NON CARACTERISE"]; pc["Backlog préparation caractérisé"]=ckpi(pc["CARACTERISE"],pc["Total"])
        
        plc=pd.pivot_table(df[df["Statut OT"]=="LANC"],index="Poste travail princ.",columns="Backlog planification",values="Ordre",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
        for c in ["CARACTERISE","NON CARACTERISE"]: plc[c]=plc.get(c,0)
        plc["Total"]=plc["CARACTERISE"]+plc["NON CARACTERISE"]; plc["Backlog planification caractérisé"]=ckpi(plc["CARACTERISE"],plc["Total"])
        
        for kn,cn in [("OT CONFIME","OT CONFIME"),("OT_COR_EGAL","OT_COR_EGAL")]:
            pv=pd.pivot_table(df,index="Poste travail princ.",columns=cn,values="Ordre",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
            for c in ["OUI","NON"]: pv[c]=pv.get(c,0)
            pv["Total"]=pv["OUI"]+pv["NON"]; pv[cn]=ckpi(pv["OUI"],pv["Total"]); res[kn.lower().replace(" ","_")]=pv
            
        avf=av.copy(); res['avf']=avf
        tca=pd.pivot_table(avf,index="Poste travail princ.",columns="Statut utilisateur",values="Avis",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
        for c in ["APRQ","APRV","APRV AVAU","REJT"]: tca[c]=tca.get(c,0)
        tca["Total"]=tca[["APRQ","APRV","APRV AVAU","REJT"]].sum(axis=1); tca["appel avis approuvé"]=ckpi(tca["APRV"],tca["Total"])
        
        g_num=df[(df["Statut OT"].isin(["CLOT","TCLO"]))&(df["_tw_num"]==350)].groupby("Poste travail princ.")["Ordre"].count()
        g_den=df[(df["Contient SOPL"]==1)&(df["_tw_num"]==350)].groupby("Poste travail princ.")["Ordre"].count()
        g_df=pd.DataFrame({"_n":g_num,"_d":g_den}).reindex(posts,fill_value=0)
        g_df["Performance Graissage"]=np.where(g_df["_d"]==0,100.0,(g_df["_n"]/g_df["_d"])*100)
        
        ins_types=[290,300,310]
        ins_base=(df["_tw_num"].isin(ins_types))&(df["Date de début planifiée"].notna())&(df["Date de début planifiée"]<=now_ts)
        ins_num=df[(df["Statut OT"].isin(["CLOT","TCLO"]))&ins_base].groupby("Poste travail princ.")["Ordre"].count()
        ins_den=df[(df["Contient SOPL"]==1)&ins_base].groupby("Poste travail princ.")["Ordre"].count()
        ins_df=pd.DataFrame({"_n":ins_num,"_d":ins_den}).reindex(posts,fill_value=0)
        ins_df["Performance Inspection"]=np.where(ins_df["_d"]==0,100.0,(ins_df["_n"]/ins_df["_d"])*100)
        
        sys_base=(df["_tw_num"]==360)&(df["Date de début planifiée"].notna())&(df["Date de début planifiée"]<=now_ts)
        sys_num=df[(df["Statut OT"].isin(["CLOT","TCLO"]))&sys_base].groupby("Poste travail princ.")["Ordre"].count()
        sys_den=df[(df["Contient SOPL"]==1)&sys_base].groupby("Poste travail princ.")["Ordre"].count()
        sys_df=pd.DataFrame({"_n":sys_num,"_d":sys_den}).reindex(posts,fill_value=0)
        sys_df["Performance Appels Systématiques"]=np.where(sys_df["_d"]==0,100.0,(sys_df["_n"]/sys_df["_d"])*100)
        
        fiab_s=pd.Series(100.0,index=posts); avpan_s=pd.Series(100.0,index=posts)
        res['ckdf']=pd.DataFrame({
            "TAUX_REALISATION_CORRECTIF/PT":an["TAUX_REALISATION_CORRECTIF/PT"],
            "OT préparation <1 mois":pr["OT préparation <1 mois"],"OT préparation >3 mois":pr["OT préparation >3 mois"],"OT préparation 1mois< <3mois":pr["OT préparation 1mois< <3mois"],
            "OT planification <1 mois":pl["OT planification <1 mois"],"OT planification >3 mois":pl["OT planification >3 mois"],"OT planification 1mois< <3mois":pl["OT planification 1mois< <3mois"],
            "OT exécution <1 mois":ex["OT exécution <1 mois"],"OT exécution >3 mois":ex["OT exécution >3 mois"],"OT exécution 1mois< <3mois":ex["OT exécution 1mois< <3mois"],
            "Performance Graissage":g_df["Performance Graissage"],"Performance Inspection":ins_df["Performance Inspection"],"Performance Appels Systématiques":sys_df["Performance Appels Systématiques"],
            "appel avis approuvé":tca["appel avis approuvé"],"OT LANC ESTIME":la["OT LANC ESTIME"],
            "Backlog préparation caractérisé":pc["Backlog préparation caractérisé"],"Backlog planification caractérisé":plc["Backlog planification caractérisé"],
            "OT CONFIME":res['ot_confime']["OT CONFIME"],"OT_COR_EGAL":res['ot_cor_egal']["OT_COR_EGAL"],
            "OT Fiabilité":fiab_s,"Total Avis de Panne":avpan_s
        })
        return res

    # ============================================================
    # FONCTION KS (complétée)
    # ============================================================
    def ks(v,c):
        try: val=float(v)
        except Exception: return ""
        if c in ["OT préparation <1 mois","OT planification <1 mois","OT exécution <1 mois"]:
            return "background:#c6efce;color:#006100;font-weight:600" if val>=80 else ("background:#ffeb9c;color:#9c6500;font-weight:600" if val>=75 else "background:#ffc7ce;color:#9c0006;font-weight:600")
        if c in ["OT préparation 1mois< <3mois","OT planification 1mois< <3mois","OT exécution 1mois< <3mois"]:
            return "background:#c6efce;color:#006100;font-weight:600" if val<=15 else "background:#ffc7ce;color:#9c0006;font-weight:600"
        if c in ["OT préparation >3 mois","OT planification >3 mois","OT exécution >3 mois"]:
            return "background:#c6efce;color:#006100;font-weight:600" if val<=5 else "background:#ffc7ce;color:#9c0006;font-weight:600"
        if c=="TAUX_REALISATION_CORRECTIF/PT":
            return "background:#c6efce;color:#006100;font-weight:600" if val>=85 else ("background:#ffeb9c;color:#9c6500;font-weight:600" if val>=80 else "background:#ffc7ce;color:#9c0006;font-weight:600")
        if c=="appel avis approuvé":
            return "background:#c6efce;color:#006100;font-weight:600" if val>=95 else ("background:#ffeb9c;color:#9c6500;font-weight:600" if val>=90 else "background:#ffc7ce;color:#9c0006;font-weight:600")
        if c in ["OT LANC ESTIME","Backlog préparation caractérisé","Backlog planification caractérisé","OT CONFIME","OT_COR_EGAL"]:
            return "background:#c6efce;color:#006100;font-weight:600" if val>=100 else ("background:#ffeb9c;color:#9c6500;font-weight:600" if val>=95 else "background:#ffc7ce;color:#9c0006;font-weight:600")
        if c in ["Performance Graissage","Performance Inspection","Performance Appels Systématiques"]:
            return "background:#c6efce;color:#006100;font-weight:600" if val>=95 else ("background:#ffeb9c;color:#9c6500;font-weight:600" if val>90 else "background:#ffc7ce;color:#9c0006;font-weight:600")
        if c in ["OT Fiabilité","Total Avis de Panne"]:
            return "background:#c6efce;color:#006100;font-weight:600" if val>=100 else "background:#ffc7ce;color:#9c0006;font-weight:600"
        if c in ["Score Performance","Score Qualite"]:
            return "background:#c6efce;color:#006100;font-weight:600" if val>=90 else ("background:#ffeb9c;color:#9c6500;font-weight:600" if val>=75 else "background:#ffc7ce;color:#9c0006;font-weight:600")
        return ""

    # ============================================================
    # CHARGEMENT FICHIERS
    # ============================================================
    st.sidebar.markdown("### 📁 Fichiers de donnees")
    ot_file = st.sidebar.file_uploader("Fichier OT (Excel)", type=["xlsx","xls"], key="ot_up")
    av_file = st.sidebar.file_uploader("Fichier Avis (Excel)", type=["xlsx","xls"], key="av_up")

    if not ot_file or not av_file:
        st.markdown("""<div style="min-height:60vh;display:flex;flex-direction:column;align-items:center;justify-content:center">
        <div style="font-size:80px;margin-bottom:20px">📂</div>
        <h2 style="color:#1e3a5f;font-size:32px;font-weight:800">Chargement des donnees</h2>
        <p style="color:#718096;font-size:18px;margin-top:10px">Veuillez charger les fichiers Excel OT et Avis depuis la barre laterale.</p>
        </div>""", unsafe_allow_html=True)
        st.stop()

    df, avf, apm, now_ts = prepare_data(ot_file.read(), av_file.read(), fichier_date)

    if not apm:
        st.markdown('<div class="es">Aucun poste de travail trouve (SF1/SF2).</div>', unsafe_allow_html=True)
        st.stop()

    # ============================================================
    # SIDEBAR - FILTRES
    # ============================================================
    st.sidebar.markdown("### 🔍 Filtres")
    sel_posts = st.sidebar.multiselect("Postes de travail", apm, default=apm, key="sp")
    if not sel_posts: sel_posts = apm

    # ============================================================
    # CALCUL KPIs
    # ============================================================
    df_f = df[df["Poste travail princ."].isin(sel_posts)]
    avf_f = avf[avf["Poste travail princ."].isin(sel_posts)]
    res = calc_kpis(df_f, avf_f, now_ts, sel_posts)
    ckdf = res['ckdf'].copy()

    # Scores globaux
    def calc_score(row, kpi_list):
        vals = [min(float(row[k]), CIBLE[k]) for k in kpi_list if k in row.index and pd.notna(row[k])]
        return round(sum(vals)/len(kpi_list),2) if kpi_list else 100.0

    ckdf["Score Performance"] = ckdf.apply(lambda r: calc_score(r, QK), axis=1)
    ckdf["Score Qualite"] = ckdf.apply(lambda r: calc_score(r, PK), axis=1)
    ckdf["Score Global"] = ((ckdf["Score Performance"] + ckdf["Score Qualite"]) / 2).round(2)

    sp_mean = ckdf["Score Performance"].mean()
    sq_mean = ckdf["Score Qualite"].mean()
    sg_mean = ckdf["Score Global"].mean()
    nb_anom_p = sum(1 for k in QK if k in ckdf.columns and any(ckdf[k] < CIBLE.get(k,0) for _ in [1]))
    nb_anom_q = sum(1 for k in PK if k in ckdf.columns and any(ckdf[k] < CIBLE.get(k,0) for _ in [1]))

    # ============================================================
    # HEADER
    # ============================================================
    logo_b64 = get_logo_base64()
    st.markdown("""<div class="mh">
        <img src="data:image/png;base64,%s" class="logo" onerror="this.style.display='none'">
        <h1>TABLEAU DE BORD KPI</h1>
        <div class="db">📅 %s</div>
    </div>""" % (logo_b64 if logo_b64 else "", fichier_date), unsafe_allow_html=True)

    # ============================================================
    # CARTES RESUME
    # ============================================================
    st.markdown("""<div class="cr">
        <div class="cc c1"><div class="cv">%.1f%%</div><div class="cl">Score Performance</div></div>
        <div class="cc c2"><div class="cv">%.1f%%</div><div class="cl">Score Qualite</div></div>
        <div class="cc c3"><div class="cv">%.1f%%</div><div class="cl">Score Global</div></div>
        <div class="cc c4"><div class="cv">%d</div><div class="cl">Postes selectionnes</div></div>
    </div>""" % (sp_mean, sq_mean, sg_mean, len(sel_posts)), unsafe_allow_html=True)

    # ============================================================
    # ONGLETS
    # ============================================================
    tab_synthese, tab_perf, tab_qual, tab_anom, tab_suivi, tab_actions = st.tabs([
        "📊 Synthese", "⚡ Performance", "✅ Qualite", "⚠️ Anomalies", "📈 Suivi", "📋 Plan d'actions"
    ])

    # ============================================================
    # ONGLET SYNTHESE
    # ============================================================
    with tab_synthese:
        st.markdown('<div class="stl s">Scores globaux par poste de travail</div>', unsafe_allow_html=True)
        synth_cols = ["Poste de travail"] + QK + ["Score Performance"] + PK + ["Score Qualite","Score Global"]
        synth_rows = []
        for poste in sel_posts:
            row = {"Poste de travail": poste}
            for k in ALL_KPI:
                if k in ckdf.columns and poste in ckdf.index:
                    row[k] = round(float(ckdf.loc[poste,k]),1)
                else:
                    row[k] = ""
            row["Score Performance"] = round(float(ckdf.loc[poste,"Score Performance"]),1)
            row["Score Qualite"] = round(float(ckdf.loc[poste,"Score Qualite"]),1)
            row["Score Global"] = round(float(ckdf.loc[poste,"Score Global"]),1)
            synth_rows.append(row)
        cible_row = {"Poste de travail": "Cible"}
        for k in ALL_KPI: cible_row[k] = CIBLE.get(k,"")
        cible_row["Score Performance"] = 100; cible_row["Score Qualite"] = 100; cible_row["Score Global"] = 100
        synth_rows.append(cible_row)
        total_row = {"Poste de travail": "Moyenne"}
        for k in ALL_KPI:
            if k in ckdf.columns:
                total_row[k] = round(float(ckdf[k].mean()),1)
            else:
                total_row[k] = ""
        total_row["Score Performance"] = round(sp_mean,1); total_row["Score Qualite"] = round(sq_mean,1); total_row["Score Global"] = round(sg_mean,1)
        synth_rows.append(total_row)

        h = '<table class="tw st"><thead><tr>' + ''.join('<th>%s</th>'%c for c in synth_cols) + '</tr></thead><tbody>'
        for i,r in enumerate(synth_rows):
            cls = "tr" if r["Poste de travail"] in ("Cible","Moyenne") else ""
            h += '<tr class="%s"><td class="poste-cell">%s</td>' % (cls, r["Poste de travail"])
            for c in synth_cols[1:]:
                v = r.get(c,"")
                style = ks(v,c) if cls=="" and v!="" else ""
                h += '<td style="text-align:center;%s">%s</td>' % (style, v)
            h += '</tr>'
        h += '</tbody></table>'
        st.markdown(h, unsafe_allow_html=True)

        # Graphique barres groupees
        st.markdown('<div class="stl s">Comparaison Performance / Qualite par poste</div>', unsafe_allow_html=True)
        bar_data = ckdf[["Score Performance","Score Qualite"]].reset_index()
        bar_data.columns = ["Poste","Performance","Qualite"]
        bar_data_m = bar_data.melt(id_vars="Poste", value_vars=["Performance","Qualite"], var_name="Type", value_name="Score")
        fig_bar = px.bar(bar_data_m, x="Poste", y="Score", color="Type", barmode="group",
            color_discrete_map={"Performance":"#2b6cb0","Qualite":"#276749"},
            title="Scores par poste")
        fig_bar.update_layout(margin=dict(t=40,b=80,l=10,r=10), height=400, xaxis_tickangle=-45)
        fig_bar.add_hline(y=90, line_dash="dash", line_color="#e53e3e", annotation_text="Cible 90%")
        st.plotly_chart(fig_bar, use_container_width=True)

    # ============================================================
    # ONGLET PERFORMANCE
    # ============================================================
    with tab_perf:
        st.markdown('<div class="stl p">Indicateurs de Performance</div>', unsafe_allow_html=True)
        pcols = ["Poste de travail"] + QK + ["Score Performance"]
        h = '<table class="tw pt"><thead><tr>' + ''.join('<th>%s</th>'%c for c in pcols) + '</tr></thead><tbody>'
        for poste in sel_posts:
            h += '<tr><td style="font-weight:600">%s</td>' % poste
            for k in QK:
                v = round(float(ckdf.loc[poste,k]),1) if k in ckdf.columns and poste in ckdf.index else ""
                h += '<td style="text-align:center;%s">%s</td>' % (ks(v,k), v)
            sp = round(float(ckdf.loc[poste,"Score Performance"]),1)
            h += '<td style="text-align:center;%s">%s</td>' % (ks(sp,"Score Performance"), sp)
            h += '</tr>'
        cible_p = {"Poste de travail":"Cible"}
        for k in QK: cible_p[k] = CIBLE.get(k,"")
        cible_p["Score Performance"] = 100
        h += '<tr class="tr"><td>Cible</td>'
        for k in QK: h += '<td style="text-align:center">%s</td>' % CIBLE.get(k,"")
        h += '<td style="text-align:center">100</td></tr>'
        h += '</tbody></table>'
        st.markdown(h, unsafe_allow_html=True)

        # Barres individuelles avec cible
        st.markdown('<div class="stl p">Detail par KPI de Performance</div>', unsafe_allow_html=True)
        for kpi in QK:
            target = CIBLE.get(kpi, 100)
            bh = '<div class="ca"><div class="ct">%s (Cible: %s%%)</div>' % (kpi, target)
            for poste in sel_posts:
                val = float(ckdf.loc[poste,kpi]) if kpi in ckdf.columns and poste in ckdf.index else 0
                pct = min(val, 120)
                target_pct = min(target, 120)
                color = "#38a169" if val >= target else ("#d69e2e" if val >= target*0.9 else "#e53e3e")
                bh += '<div class="car"><div class="cal">%s</div><div class="cab"><div class="caf" style="width:%.1f%%;background:%s"></div><div class="target-mark" style="left:%.1f%%"></div></div><div class="cav-out">%.1f%%</div><div class="cav-tgt">%s%%</div></div>' % (poste, pct, color, target_pct, val, target)
            bh += '</div>'
            st.markdown(bh, unsafe_allow_html=True)

        # Pie statut OT correctif
        st.markdown('<div class="stl p">Repartition des OT Correctifs par Statut</div>', unsafe_allow_html=True)
        piv_stat = build_statut_pivot(df_f[(df_f["Nº appel pl.entret."].fillna(0)==0)&(df_f["Contient SOPL"]==1)], sel_posts)
        col_p1, col_p2 = st.columns(2)
        with col_p1:
            show_pie_pair(piv_stat, "OT Correctifs")
        with col_p2:
            st.markdown(html_statut_pivot(piv_stat, "pt"), unsafe_allow_html=True)

    # ============================================================
    # ONGLET QUALITE
    # ============================================================
    with tab_qual:
        st.markdown('<div class="stl q">Indicateurs de Qualite</div>', unsafe_allow_html=True)
        qcols = ["Poste de travail"] + PK + ["Score Qualite"]
        h = '<table class="tw qt"><thead><tr>' + ''.join('<th>%s</th>'%c for c in qcols) + '</tr></thead><tbody>'
        for poste in sel_posts:
            h += '<tr><td style="font-weight:600">%s</td>' % poste
            for k in PK:
                v = round(float(ckdf.loc[poste,k]),1) if k in ckdf.columns and poste in ckdf.index else ""
                h += '<td style="text-align:center;%s">%s</td>' % (ks(v,k), v)
            sq = round(float(ckdf.loc[poste,"Score Qualite"]),1)
            h += '<td style="text-align:center;%s">%s</td>' % (ks(sq,"Score Qualite"), sq)
            h += '</tr>'
        h += '<tr class="tr"><td>Cible</td>'
        for k in PK: h += '<td style="text-align:center">%s</td>' % CIBLE.get(k,"")
        h += '<td style="text-align:center">100</td></tr>'
        h += '</tbody></table>'
        st.markdown(h, unsafe_allow_html=True)

        # Barres individuelles qualite
        st.markdown('<div class="stl q">Detail par KPI de Qualite</div>', unsafe_allow_html=True)
        for kpi in PK:
            target = CIBLE.get(kpi, 100)
            bh = '<div class="ca"><div class="ct">%s (Cible: %s%%)</div>' % (kpi, target)
            for poste in sel_posts:
                val = float(ckdf.loc[poste,kpi]) if kpi in ckdf.columns and poste in ckdf.index else 0
                pct = min(val, 120)
                target_pct = min(target, 120)
                color = "#38a169" if val >= target else ("#d69e2e" if val >= target*0.9 else "#e53e3e")
                bh += '<div class="car"><div class="cal">%s</div><div class="cab"><div class="caf" style="width:%.1f%%;background:%s"></div><div class="target-mark" style="left:%.1f%%"></div></div><div class="cav-out">%.1f%%</div><div class="cav-tgt">%s%%</div></div>' % (poste, pct, color, target_pct, val, target)
            bh += '</div>'
            st.markdown(bh, unsafe_allow_html=True)

        # Pies backlog
        st.markdown('<div class="stl q">Backlog Preparation et Planification</div>', unsafe_allow_html=True)
        col_q1, col_q2 = st.columns(2)
        with col_q1:
            piv_bp = pd.pivot_table(df_f[df_f["Statut OT"]=="CRÉÉ"], index="Poste travail princ.", columns="Backlog preparation", values="Ordre", aggfunc="count", fill_value=0).reindex(sel_posts, fill_value=0)
            for c in ["CARACTERISE","NON CARACTERISE"]: piv_bp[c]=piv_bp.get(c,0)
            show_simple_pie(piv_bp, "Backlog Preparation")
        with col_q2:
            piv_bpl = pd.pivot_table(df_f[df_f["Statut OT"]=="LANC"], index="Poste travail princ.", columns="Backlog planification", values="Ordre", aggfunc="count", fill_value=0).reindex(sel_posts, fill_value=0)
            for c in ["CARACTERISE","NON CARACTERISE"]: piv_bpl[c]=piv_bpl.get(c,0)
            show_simple_pie(piv_bpl, "Backlog Planification")

    # ============================================================
    # ONGLET ANOMALIES
    # ============================================================
    with tab_anom:
        ano_p_rows = []; ano_q_rows = []
        for poste in sel_posts:
            for k in QK:
                if k in ckdf.columns:
                    v = float(ckdf.loc[poste,k])
                    if v < CIBLE.get(k,0):
                        ano_p_rows.append({"Poste de travail":poste,"KPI":k,"Valeur":round(v,1),"Cible":CIBLE.get(k,""),"Ecart":round(v-CIBLE.get(k,0),1),"Responsable":KPI_RESP_MAP.get(k,""),"Action":ACT_MAP.get(k,"")})
            for k in PK:
                if k in ckdf.columns:
                    v = float(ckdf.loc[poste,k])
                    if v < CIBLE.get(k,0):
                        ano_q_rows.append({"Poste de travail":poste,"KPI":k,"Valeur":round(v,1),"Cible":CIBLE.get(k,""),"Ecart":round(v-CIBLE.get(k,0),1),"Responsable":KPI_RESP_MAP.get(k,""),"Action":ACT_MAP.get(k,"")})

        st.markdown('<div class="stl a">Anomalies Performance (%d)</div>' % len(ano_p_rows), unsafe_allow_html=True)
        if ano_p_rows:
            ac = ["Poste de travail","KPI","Valeur","Cible","Ecart","Responsable","Action"]
            ah = '<table class="tw at"><thead><tr>' + ''.join('<th>%s</th>'%c for c in ac) + '</tr></thead><tbody>'
            for r in ano_p_rows:
                ah += '<tr>' + ''.join('<td>%s</td>'%r.get(c,"") for c in ac) + '</tr>'
            ah += '</tbody></table>'
            st.markdown(ah, unsafe_allow_html=True)
        else:
            st.markdown('<div class="es">Aucune anomalie de performance</div>', unsafe_allow_html=True)

        st.markdown('<div class="stl a">Anomalies Qualite (%d)</div>' % len(ano_q_rows), unsafe_allow_html=True)
        if ano_q_rows:
            ac = ["Poste de travail","KPI","Valeur","Cible","Ecart","Responsable","Action"]
            ah = '<table class="tw at"><thead><tr>' + ''.join('<th>%s</th>'%c for c in ac) + '</tr></thead><tbody>'
            for r in ano_q_rows:
                ah += '<tr>' + ''.join('<td>%s</td>'%r.get(c,"") for c in ac) + '</tr>'
            ah += '</tbody></table>'
            st.markdown(ah, unsafe_allow_html=True)
        else:
            st.markdown('<div class="es">Aucune anomalie de qualite</div>', unsafe_allow_html=True)

    # ============================================================
    # ONGLET SUIVI
    # ============================================================
    with tab_suivi:
        hist_path = os.path.join("kpis","indicateurs_kpis.xlsx")
        hist_df = load_historical_kpis(hist_path)
        if hist_df.empty:
            st.markdown('<div class="es">Aucun historique disponible. Exportez les KPIs pour commencer le suivi.</div>', unsafe_allow_html=True)
        else:
            var_df = calculate_variations(hist_df)
            journal = generate_journal(var_df)
            top5, bot5 = calculate_rankings(var_df)

            st.markdown('<div class="stl s">Journal des variations significatives (|Ecart| >= 5%%)</div>', unsafe_allow_html=True)
            if not journal.empty:
                jc = ["Date precedente","Date actuelle","Poste","Type","KPI","Valeur precedente","Valeur actuelle","Ecart","Ecart %","Tendance","Sens"]
                jh = '<table class="tw st"><thead><tr>' + ''.join('<th>%s</th>'%c for c in jc) + '</tr></thead><tbody>'
                for _,r in journal.iterrows():
                    sens_color = "#006100" if r["Sens"]=="Amelioration" else ("#9c0006" if r["Sens"]=="Degradation" else "#9c6500")
                    jh += '<tr>' + ''.join('<td>%s</td>'%r.get(c,"") for c in jc[:-1])
                    jh += '<td style="color:%s;font-weight:700">%s</td>' % (sens_color, r["Sens"])
                    jh += '</tr>'
                jh += '</tbody></table>'
                st.markdown(jh, unsafe_allow_html=True)
            else:
                st.markdown('<div class="es">Aucune variation significative</div>', unsafe_allow_html=True)

            col_s1, col_s2 = st.columns(2)
            with col_s1:
                st.markdown('<div class="stl s">Top 5 Postes en progression</div>', unsafe_allow_html=True)
                if not top5.empty:
                    for i,(_,r) in enumerate(top5.iterrows()):
                        st.markdown('<div class="cgr"><div class="rk" style="color:#276749">%d</div><div class="pn">%s</div><div class="ps" style="color:#276749">%+.1f</div></div>' % (i+1, r["Poste"], r["Score variation"]), unsafe_allow_html=True)
            with col_s2:
                st.markdown('<div class="stl s">Top 5 Postes en regression</div>', unsafe_allow_html=True)
                if not bot5.empty:
                    for i,(_,r) in enumerate(bot5.iterrows()):
                        st.markdown('<div class="cgr"><div class="rk" style="color:#c53030">%d</div><div class="pn">%s</div><div class="ps" style="color:#c53030">%+.1f</div></div>' % (i+1, r["Poste"], r["Score variation"]), unsafe_allow_html=True)

    # ============================================================
    # ONGLET PLAN D'ACTIONS
    # ============================================================
    with tab_actions:
        st.markdown('<div class="stl c">Plan d\'actions correctives</div>', unsafe_allow_html=True)
        all_anomalies = ano_p_rows + ano_q_rows
        if all_anomalies:
            pa_df = pd.DataFrame(all_anomalies)
            if "pa_data" not in st.session_state:
                pa_df["Statut"] = "En cours"
                pa_df["Echeance"] = ""
                pa_df["Commentaire"] = ""
                st.session_state.pa_data = pa_df
            else:
                existing = st.session_state.pa_data
                for _,r in pa_df.iterrows():
                    match = existing[(existing["Poste de travail"]==r["Poste de travail"])&(existing["KPI"]==r["KPI"])]
                    if match.empty:
                        new_row = r.copy(); new_row["Statut"]="En cours"; new_row["Echeance"]=""; new_row["Commentaire"]=""
                        existing = pd.concat([existing, pd.DataFrame([new_row])], ignore_index=True)
                st.session_state.pa_data = existing

            edited = st.data_editor(st.session_state.pa_data, use_container_width=True, num_rows="dynamic",
                column_config={"Statut": st.column_config.SelectboxColumn(options=["En cours","Termine","Reporte","Annule"])},
                key="pa_editor")
            st.session_state.pa_data = edited
        else:
            st.markdown('<div class="es">Aucune anomalie detectee. Plan d\'actions vide.</div>', unsafe_allow_html=True)

    # ============================================================
    # BOUTONS EXPORT
    # ============================================================
    st.sidebar.markdown("---")
    st.sidebar.markdown("### 💾 Exports")

    # Export Excel
    p_rows_xl = []
    for poste in sel_posts:
        row = {"Poste de travail": poste}
        for k in QK: row[k] = round(float(ckdf.loc[poste,k]),1) if k in ckdf.columns else ""
        row["Score Performance"] = round(float(ckdf.loc[poste,"Score Performance"]),1)
        p_rows_xl.append(row)
    q_rows_xl = []
    for poste in sel_posts:
        row = {"Poste de travail": poste}
        for k in PK: row[k] = round(float(ckdf.loc[poste,k]),1) if k in ckdf.columns else ""
        row["Score Qualite"] = round(float(ckdf.loc[poste,"Score Qualite"]),1)
        q_rows_xl.append(row)
    ano_p_cols = ["Poste de travail","KPI","Valeur","Cible","Ecart","Responsable","Action"]
    ano_q_cols = ano_p_cols[:]

    if st.sidebar.button("📥 Exporter KPIs Excel"):
        save_kpis_to_excel(p_rows_xl, ["Poste de travail"]+QK+["Score Performance"],
                          q_rows_xl, ["Poste de travail"]+PK+["Score Qualite"],
                          ano_p_rows, ano_p_cols, ano_q_rows, ano_q_cols, fichier_date)
        st.sidebar.success("Exporte avec succes!")

    # Export plan d'actions
    if "pa_data" in st.session_state and not st.session_state.pa_data.empty:
        @st.cache_data
        def to_excel(df):
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Plan Actions')
            return output.getvalue()
        st.sidebar.download_button("📋 Telecharger Plan Actions", data=to_excel(st.session_state.pa_data),
            file_name="plan_actions_%s.xlsx" % fichier_date.replace("/","-"), mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # Bouton imprimer
     # Bouton imprimer
    st.sidebar.markdown("""
    <button onclick="window.print()" style="
        width: 100%;
        background: linear-gradient(135deg, #1e3a5f, #2c5282);
        color: white;
        border: none;
        border-radius: 6px;
        padding: 8px 14px;
        font-weight: 700;
        font-size: 15px;
        cursor: pointer;
        margin-top: 5px;
    ">🖨️ Imprimer / PDF</button>
    """, unsafe_allow_html=True)

    # ============================================================
    # FOOTER
    # ============================================================
    st.markdown('<div class="footer">Dashboard KPI — Date de reference : %s — Genere le %s</div>' % (fichier_date, datetime.now().strftime("%d/%m/%Y %H:%M")), unsafe_allow_html=True)

if __name__ == "__main__":
    main()
