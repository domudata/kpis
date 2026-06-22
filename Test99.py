# -*- coding: utf-8 -*-
import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import numpy as np
import io, locale, random, time, os, hashlib, json, base64
from datetime import datetime
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ============================================================
st.set_page_config(layout="wide", page_title="Dashboard KPI", initial_sidebar_state="expanded")
# ============================================================

QK = ["TAUX_REALISATION_CORRECTIF/PT","OT préparation <1 mois","OT préparation >3 mois",
      "OT préparation 1mois< <3mois","OT planification <1 mois","OT planification >3 mois",
      "OT planification 1mois< <3mois","OT exécution <1 mois","OT exécution >3 mois",
      "OT exécution 1mois< <3mois",
      "Performance Graissage","Performance Inspection","Performance Appels Systématiques"]
PK = ["Taux d'approbation des Avis","OT LANC ESTIME","Backlog préparation caractérisé",
      "Backlog planification caractérisé","OT CONFIME","OT_COR_EGAL",
      "OT Fiabilité","Total Avis de Panne"]
ALL_KPI = QK + PK

CIBLE = {"TAUX_REALISATION_CORRECTIF/PT":85,"OT préparation <1 mois":80,"OT préparation >3 mois":5,
         "OT préparation 1mois< <3mois":15,"OT planification <1 mois":80,"OT planification >3 mois":5,
         "OT planification 1mois< <3mois":15,"OT exécution <1 mois":80,"OT exécution >3 mois":5,
         "OT exécution 1mois< <3mois":15,"Taux d'approbation des Avis":95,"OT LANC ESTIME":100,
         "Backlog préparation caractérisé":100,"Backlog planification caractérisé":100,
         "OT CONFIME":100,"OT_COR_EGAL":100,
         "Performance Graissage":95,"Performance Inspection":95,"Performance Appels Systématiques":95,
         "OT Fiabilité":100,"Total Avis de Panne":100}

ACT_MAP = {"TAUX_REALISATION_CORRECTIF/PT":"Ameliorer le taux de realisation des OT.",
           "OT préparation <1 mois":"Reduire l'age de preparation des OT (< 1 mois).",
           "OT préparation >3 mois":"Traiter les OT avec preparation > 3 mois.",
           "OT planification <1 mois":"Reduire l'age de planification des OT (< 1 mois).",
           "OT planification >3 mois":"Traiter les OT avec planification > 3 mois.",
           "OT exécution <1 mois":"Reduire l'age d'execution des OT (< 1 mois).",
           "OT exécution >3 mois":"Traiter les OT avec execution > 3 mois.",
           "OT LANC ESTIME":"Estimer les couts des OT lances.",
           "Backlog préparation caractérisé":"Caracteriser le backlog de preparation.",
           "Backlog planification caractérisé":"Caracteriser le backlog de planification.",
           "OT CONFIME":"Confirmer les OT termines.",
           "OT_COR_EGAL":"Rapprocher les couts reels et budgetes.",
           "Taux d'approbation des Avis":"Creer un OT pour les avis sans ordre.",
           "OT préparation 1mois< <3mois":"Reduire les OT entre 1 et 3 mois.",
           "OT planification 1mois< <3mois":"Reduire les OT entre 1 et 3 mois.",
           "OT exécution 1mois< <3mois":"Reduire les OT entre 1 et 3 mois.",
           "Performance Graissage":"Ameliorer le taux de realisation des OT de graissage (Type 350).",
           "Performance Inspection":"Ameliorer le taux de realisation des OT d'inspection (Types 290,300,310).",
           "Performance Appels Systématiques":"Ameliorer le taux de realisation des appels systematiques (Type 360).",
           "OT Fiabilité":"Maintenir la fiabilite des OT a 100%.",
           "Total Avis de Panne":"Maintenir le suivi des avis de panne a 100%."}

KPI_RESP_MAP = {
    "TAUX_REALISATION_CORRECTIF/PT": "Chef d'atelier",
    "OT préparation <1 mois": "Préparateur BM",
    "OT préparation 1mois< <3mois": "Préparateur BM",
    "OT préparation >3 mois": "Préparateur BM",
    "OT planification <1 mois": "Planificateur BM",
    "OT planification 1mois< <3mois": "Planificateur BM",
    "OT planification >3 mois": "Planificateur BM",
    "OT exécution <1 mois": "Chef d'atelier",
    "OT exécution 1mois< <3mois": "Chef d'atelier",
    "OT exécution >3 mois": "Chef d'atelier",
    "Taux d'approbation des Avis": "Chef d'atelier",
    "OT LANC ESTIME": "Fiabilité",
    "Backlog préparation caractérisé": "Préparateur BM",
    "Backlog planification caractérisé": "Planificateur BM",
    "OT CONFIME": "Agent de saisie",
    "OT_COR_EGAL": "Agent de saisie",
    "Performance Graissage": "Chef d'atelier",
    "Performance Inspection": "Chef d'atelier",
    "Performance Appels Systématiques": "Chef d'atelier",
    "OT Fiabilité": "Fiabilité",
    "Total Avis de Panne": "Fiabilité"
}

LOWER_BETTER = ["OT préparation >3 mois","OT planification >3 mois","OT exécution >3 mois",
                "OT préparation 1mois< <3mois","OT planification 1mois< <3mois","OT exécution 1mois< <3mois"]

MP_KW = ["CRPR ATPD","CRPR ATMR","CRPR ATER","CRPR ATRS","CRPR ATMO","ATPD","ATMR","ATER","ATRS","ATMO"]
MPLAN_KW = ["ATPL ATEI","ATPL ATAL","ATPL ATER","ATPL AGAR","ATPL ATHS","ATEI","ATAL","ATAS","AGAR","ATHS"]

CONSIGNES_HSE = [
    "Port obligatoire des EPI avant toute intervention.","Port obligatoire du casque de securite.",
    "Port obligatoire des lunettes de protection.","Port obligatoire des gants adaptes au travail.",
    "Utiliser les protections auditives dans les zones bruyantes.","Verifier l'absence de tension avant toute intervention electrique.",
    "Respecter la procedure de consignation et deconsignation.","Ne jamais intervenir sur un equipement en marche.",
    "Baliser et securiser la zone de travail.","Maintenir le poste de travail propre et ordonne.",
    "Verifier l'etat des outils avant utilisation.","Utiliser uniquement du materiel homologue.",
    "Respecter les permis de travail en vigueur.","Identifier les risques avant de commencer une tache.",
    "Signaler immediatement toute situation dangereuse.","Signaler tout incident ou presque accident.",
    "Ne jamais neutraliser un dispositif de securite.","Verifier les detecteurs de gaz avant utilisation.",
    "Verifier la bonne ventilation des zones de travail.","Respecter les regles des espaces confines.",
    "Controler l'atmosphere avant d'entrer dans un espace confine.","Utiliser les points d'ancrage pour les travaux en hauteur.",
    "Verifier l'etat des echafaudages avant utilisation.","Securiser les outils lors des travaux en hauteur.",
    "Ne pas travailler seul lors des operations a risque.","Controler les elingues avant chaque levage.",
    "Respecter les limites de charge des equipements.","Verifier l'etat des appareils de levage.",
    "Maintenir les voies de circulation degagees.","Respecter la signalisation de securite.",
    "Verifier les extincteurs a proximite du chantier.","Connaitre les issues de secours les plus proches.",
    "Respecter les procedures d'arret d'urgence.","Verifier les flexibles et raccords avant mise en service.",
    "Controler les fuites avant demarrage d'un equipement.","Respecter les distances de securite.",
    "Ne jamais contourner une procedure HSE.","Porter les EPI adaptes au risque identifie.",
    "Prevenir son responsable avant toute intervention particuliere.","Analyser les risques avant chaque demarrage de chantier.",
    "Verifier la stabilite des equipements.","Utiliser les bons outils pour la bonne tache.",
    "Respecter les consignes specifiques du chantier.","Ne jamais prendre de raccourci au detriment de la securite.",
    "Arreter immediatement les travaux en cas de danger.","Proteger l'environnement lors des interventions.",
    "Collecter et trier correctement les dechets.","Eviter toute pollution accidentelle.",
    "Respecter les consignes de stockage des produits dangereux.","Lire les fiches de securite avant manipulation.",
    "Verifier les equipements avant chaque prise de poste.","S'assurer de la disponibilite des moyens de secours.",
    "Communiquer clairement avec l'equipe avant intervention.","Respecter les regles de circulation des engins.",
    "Garder une vigilance permanente sur son environnement.","Prendre le temps d'effectuer le travail en securite.",
    "La securite est l'affaire de tous.","Chaque incident peut etre evite par la prevention.",
    "Aucun travail n'est plus urgent que la securite.","Zero accident commence par un comportement sur."]

def get_logo_base64():
    for path in ["logo.png", "./logo.png", "../logo.png"]:
        if os.path.exists(path):
            try:
                with open(path, "rb") as f:
                    return base64.b64encode(f.read()).decode()
            except Exception:
                pass
    return None

def get_date_from_file():
    if os.path.exists("date.txt"):
        try:
            with open("date.txt","r",encoding="utf-8") as f: return f.read().strip()
        except Exception: pass
    return "18/06/2026"

def contient_mot(t,lm):
    t=str(t); return any(m in t for l in lm for m in l.split())
def cat_age(a):
    if pd.isna(a): return "Inconnu"
    if a<=1: return "<1 mois"
    elif a>=3: return ">3 mois"
    return "1 mois < <3 mois"

def excr(df):
    if "Poste travail princ." in df.columns:
        return df[~df["Poste travail princ."].astype(str).str.contains("cresseur",case=False,na=False)].copy()
    return df

@st.cache_data(show_spinner=False)
def prepare_data(ot_bytes, av_bytes, date_str):
    raw_ot = pd.read_excel(io.BytesIO(ot_bytes))
    raw_av = pd.read_excel(io.BytesIO(av_bytes))
    raw_ot = excr(raw_ot)
    raw_av = excr(raw_av)
    
    for c in ["Créé le","Date de début planifiée","Date de clôture","Début réel","Fin réelle"]:
        if c in raw_ot.columns: raw_ot[c]=pd.to_datetime(raw_ot[c],errors="coerce")
    for c in ["Créé le","Début souhaité","Date de la clôture"]:
        if c in raw_av.columns: raw_av[c]=pd.to_datetime(raw_av[c],errors="coerce")
        
    now_ts = pd.to_datetime(date_str, format="%d/%m/%Y", errors='coerce')
    if pd.isna(now_ts): now_ts = pd.Timestamp.now()
    
    df = raw_ot.copy()
    
    df["Backlog preparation"]=np.where(df["Statut utilisateur"].apply(lambda x:contient_mot(x,MP_KW)),"CARACTERISE","NON CARACTERISE")
    df["Backlog planification"]=np.where(df["Statut utilisateur"].apply(lambda x:contient_mot(x,MPLAN_KW)),"CARACTERISE","NON CARACTERISE")
    df["Type Carac Prep"]=df["Statut utilisateur"].apply(lambda x: next((kw.split()[0] for kw in MP_KW if kw in str(x)), "NON CARACTERISE"))
    df["Type Carac Plan"]=df["Statut utilisateur"].apply(lambda x: next((kw.split()[0] for kw in MPLAN_KW if kw in str(x)), "NON CARACTERISE"))
    
    for dc,am,ac in [('Créé le',"amp","ap"),('Date de début planifiée',"amlp","alp"),('Date de début planifiée',"amex","aex")]:
        if dc in df.columns:
            df[am]=((now_ts.year-df[dc].dt.year)*12+(now_ts.month-df[dc].dt.month)).round(2)
            df[ac]=df[am].apply(cat_age)
        else: df[am]=np.nan; df[ac]="Inconnu"
        
    df["OT CONFIME"]=np.where(df["Statut système"].str.contains("CLOT|TCLO",na=False) & df["Statut système"].str.contains("CONF",na=False),"OUI","NON")
    
    df["Contient SOPL"]=df["Statut utilisateur"].str.contains("SOPL",na=False).map({True:1,False:0})
    df["OT LANC ESTIME"]=np.where(df["Total coûts budgétés"].fillna(0)==0,"NON","OUI")
    df["OT_COR_EGAL"]=np.where((df["Total coûts budgétés"].fillna(0)-df["Total coûts réels"].fillna(0))==0,"OUI","NON")
    df["_tw_num"]=pd.to_numeric(df.get("Type de travail",pd.Series(dtype=float)),errors="coerce")
    
    if "Statut système" in df.columns: df["Statut OT"]=df["Statut système"].fillna("").astype(str).str.strip().str.split().str[0]
    
    avf = raw_av[
    (
        raw_av["Ordre"].isna() |
        (raw_av["Ordre"].astype(str).str.strip() == "")
    )
    &
    (raw_av["Type d'avis"].isin(["ZU","Z4","ZR","ZP"]))
].copy()
    
    apm = sorted(df[df["Poste travail princ."].astype(str).str.startswith(("SF1","SF2"),na=False)]["Poste travail princ."].dropna().unique().tolist())
    
    return df, avf, apm, now_ts

def save_kpis_to_excel(prows,pcols,qrows,qcols,ano_p_r,ano_p_c,ano_q_r,ano_q_c,sheet_name):
    kpis_dir="kpis"; os.makedirs(kpis_dir,exist_ok=True)
    filepath=os.path.join(kpis_dir,"indicateurs_kpis.xlsx")
    sn=str(sheet_name).replace("/","-").replace("\\","-").replace("*","").replace("?","").replace("[","").replace("]","")[:31]
    hf=Font(bold=True,color="FFFFFF",size=10); hfl=PatternFill(start_color="1E3A5F",end_color="1E3A5F",fill_type="solid")
    tf=Font(bold=True,size=12,color="1E3A5F")
    tb=Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
    try: wb=load_workbook(filepath)
    except Exception: wb=Workbook()
    if "Sheet" in wb.sheetnames: del wb["Sheet"]
    if sn in wb.sheetnames: del wb[sn]
    ws=wb.create_sheet(sn); rn=1
    def ws_sec(title,cols,rows,sr):
        ws.cell(row=sr,column=1,value=title).font=tf; sr+=1
        for j,c in enumerate(cols,1):
            cl=ws.cell(row=sr,column=j,value=c); cl.font=hf; cl.fill=hfl; cl.alignment=Alignment(horizontal='center'); cl.border=tb
        sr+=1
        for r in rows:
            for j,c in enumerate(cols,1):
                cl=ws.cell(row=sr,column=j,value=r.get(c,"")); cl.border=tb; cl.alignment=Alignment(horizontal='center')
            sr+=1
        return sr+1
    rn=ws_sec("INDICATEURS DE PERFORMANCE",pcols,prows,rn)
    if ano_p_c and ano_p_r: rn=ws_sec("ANOMALIES PERFORMANCE",ano_p_c,ano_p_r,rn)
    rn=ws_sec("INDICATEURS DE QUALITE",qcols,qrows,rn)
    if ano_q_c and ano_q_r: rn=ws_sec("ANOMALIES QUALITE",ano_q_c,ano_q_r,rn)
    try: wb.save(filepath)
    except Exception: pass

def load_historical_kpis(filepath):
    if not os.path.exists(filepath): return pd.DataFrame()
    try: wb=load_workbook(filepath,read_only=True,data_only=True)
    except Exception: return pd.DataFrame()
    records=[]; section=None; headers=None
    for sheet_name in wb.sheetnames:
        try:
            ws=wb[sheet_name]; rows_data=list(ws.iter_rows(values_only=True))
            for row in rows_data:
                cell0=str(row[0]).strip() if row[0] else ""
                if "INDICATEURS DE PERFORMANCE" in cell0.upper(): section="perf"; headers=None; continue
                elif "INDICATEURS DE QUALITE" in cell0.upper(): section="qual"; headers=None; continue
                elif "ANOMALIES" in cell0.upper(): section=None; continue
                if section and headers is None and cell0:
                    headers=[str(c).strip() if c else "" for c in row]; continue
                if section and headers and cell0 and cell0 not in ("Cible","Total general",""):
                    entry={"Date":sheet_name}
                    for j,h in enumerate(headers):
                        if j<len(row): entry[h]=row[j]
                    entry["_section"]=section; records.append(entry)
        except Exception: continue
    wb.close()
    if not records: return pd.DataFrame()
    df=pd.DataFrame(records)
    df["Date_parsed"]=pd.to_datetime(df["Date"].str.replace("-","/"),format="%d/%m/%Y",errors="coerce")
    return df.sort_values("Date_parsed").reset_index(drop=True)

def calculate_variations(hist_df):
    if hist_df.empty or "Date" not in hist_df.columns: return pd.DataFrame()
    dates=sorted(hist_df["Date"].unique())
    if len(dates)<2: return pd.DataFrame()
    perf_df=hist_df[hist_df["_section"]=="perf"].copy()
    qual_df=hist_df[hist_df["_section"]=="qual"].copy()
    variations=[]
    for i in range(1,len(dates)):
        prev_date,curr_date=dates[i-1],dates[i]
        prev_perf=perf_df[perf_df["Date"]==prev_date].set_index("Poste de travail") if "Poste de travail" in perf_df.columns else pd.DataFrame()
        curr_perf=perf_df[perf_df["Date"]==curr_date].set_index("Poste de travail") if "Poste de travail" in perf_df.columns else pd.DataFrame()
        prev_qual=qual_df[qual_df["Date"]==prev_date].set_index("Poste de travail") if "Poste de travail" in qual_df.columns else pd.DataFrame()
        curr_qual=qual_df[qual_df["Date"]==curr_date].set_index("Poste de travail") if "Poste de travail" in qual_df.columns else pd.DataFrame()
        for sec_name,prev_d,curr_d,kpi_list in [("Performance",prev_perf,curr_perf,QK+["Score Performance"]),("Qualite",prev_qual,curr_qual,PK+["Score Qualite"])]:
            for poste in set(prev_d.index)&set(curr_d.index):
                for kpi in kpi_list:
                    if kpi not in prev_d.columns or kpi not in curr_d.columns: continue
                    try: pv=float(prev_d.loc[poste,kpi])
                    except Exception: continue
                    try: cv=float(curr_d.loc[poste,kpi])
                    except Exception: continue
                    diff=cv-pv; pct=(diff/pv*100) if pv!=0 else (100 if cv!=0 else 0)
                    if abs(diff)<=0.5: trend="stabilite"
                    elif diff>0.5: trend="hausse"
                    else: trend="baisse"
                    sens = "Stable"
                    if trend != "stabilite":
                        if (trend == "hausse" and kpi not in LOWER_BETTER) or (trend == "baisse" and kpi in LOWER_BETTER):
                            sens = "Amelioration"
                        else:
                            sens = "Degradation"
                    variations.append({"Date precedente":prev_date,"Date actuelle":curr_date,"Poste":poste,
                        "Type":sec_name,"KPI":kpi,"Valeur precedente":round(pv,2),"Valeur actuelle":round(cv,2),
                        "Ecart":round(diff,2),"Ecart %":round(pct,2),"Tendance":trend, "Sens":sens})
    return pd.DataFrame(variations)

def generate_journal(var_df):
    if var_df.empty: return pd.DataFrame()
    j=var_df.copy(); j["Significatif"]=j["Ecart %"].abs()>=5
    j=j[j["Significatif"]].copy()
    return j.sort_values(["Date actuelle","Ecart %"],ascending=[True,False])

def calculate_rankings(var_df):
    if var_df.empty: return pd.DataFrame(),pd.DataFrame()
    scores={}
    for poste in var_df["Poste"].unique():
        pv=var_df[var_df["Poste"]==poste].copy()
        scores[poste]=sum((-r["Ecart %"] if r["KPI"] in LOWER_BETTER else r["Ecart %"]) for _,r in pv.iterrows())
    ranked=sorted(scores.items(),key=lambda x:x[1],reverse=True)
    return pd.DataFrame(ranked[:5],columns=["Poste","Score variation"]),pd.DataFrame(ranked[-5:][::-1],columns=["Poste","Score variation"])

def inject_custom_css():
    st.markdown("""<style>
    section[data-testid="stSidebar"]{width:250px!important}
    .main .block-container{max-width:100%!important;width:100%!important;padding-left:0.5rem!important;padding-right:0.5rem!important}
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700;800&display=swap');
    
    :root{
        --primary:#1e3a5f;
        --primary-light:#2c5282;
        --success:#10b981;
        --success-dark:#059669;
        --warning:#f59e0b;
        --warning-dark:#d97706;
        --danger:#ef4444;
        --danger-dark:#dc2626;
        --info:#3b82f6;
        --border:#e2e8f0;
        --radius:10px;
    }
    
    *{box-sizing:border-box;margin:0;padding:0}
    .stApp{background:#f8fafc;font-family:'Inter',sans-serif}
    .main .block-container{padding-top:.8rem;padding-bottom:.8rem}
    
    .mh{
        background:linear-gradient(135deg,#1e3a5f 0%,#2563eb 100%);
        padding:16px 24px;
        border-radius:12px;
        margin-bottom:12px;
        box-shadow:0 8px 24px rgba(30,58,95,0.15);
        display:flex;
        align-items:center;
        gap:16px;
    }
    .mh h1{color:#fff;font-size:42px;font-weight:800;margin:0;flex:1}
    .mh .logo{height:50px;width:auto;max-width:150px;object-fit:contain;border-radius:6px}
    .mh .db{
        background:rgba(255,255,255,0.2);
        padding:6px 16px;
        border-radius:16px;
        color:#fff;
        font-size:20px;
        font-weight:600;
        border:1px solid rgba(255,255,255,0.3);
        backdrop-filter:blur(10px);
    }
    
    .cr{display:grid;grid-template-columns:repeat(4,1fr);gap:8px;margin-bottom:10px}
    .cc{
        background:#fff;
        border-radius:12px;
        padding:18px 16px;
        box-shadow:0 4px 12px rgba(0,0,0,0.06);
        border-left:4px solid;
        transition:transform 0.2s,box-shadow 0.2s;
        text-align:center;
    }
    .cc:hover{transform:translateY(-2px);box-shadow:0 8px 20px rgba(0,0,0,0.1);}
    .cc .cv{font-size:32px;font-weight:900;line-height:1.1;}
    .cc .cl{font-size:16px;color:#1e293b;font-weight:800;text-transform:uppercase;letter-spacing:.5px;margin-top:8px;}
    .cc.c1{border-left-color:#3b82f6}.cc.c1 .cv{color:#2563eb}
    .cc.c2{border-left-color:#10b981}.cc.c2 .cv{color:#059669}
    .cc.c3{border-left-color:#8b5cf6}.cc.c3 .cv{color:#7c3aed}
    .cc.c4{border-left-color:#ef4444}.cc.c4 .cv{color:#dc2626}
    .cc.c5{border-left-color:#3b82f6}.cc.c5 .cv{color:#2563eb}
    .cc.c6{border-left-color:#06b6d4}.cc.c6 .cv{color:#0891b2}
    .cc.c7{border-left-color:#f59e0b}.cc.c7 .cv{color:#d97706}
    .cc.c8{border-left-color:#f97316}.cc.c8 .cv{color:#ea580c}
    
    .stl{font-size:16px;font-weight:800;color:var(--primary);margin:10px 0 5px 0;padding-left:12px;border-left:4px solid var(--info);}
    
    .tw{width:100%;border-collapse:collapse;font-family:'Inter',sans-serif;font-size:13px;display:block;overflow-x:auto;-webkit-overflow-scrolling:touch;margin:0}
    .tw thead th{background:var(--primary);color:#fff;font-weight:700;font-size:11px;text-transform:uppercase;letter-spacing:.3px;padding:6px 8px;border:none;white-space:nowrap;position:sticky;top:0;z-index:10}
    .tw.qt thead th{background:linear-gradient(135deg,#2563eb,#3b82f6)}
    .tw.pt thead th{background:linear-gradient(135deg,#059669,#10b981)}
    .tw.at thead th{background:linear-gradient(135deg,#dc2626,#ef4444)}
    .tw.st thead th{background:linear-gradient(135deg,#d97706,#f59e0b)}
    .tw thead th:first-child{z-index:11;left:0}
    .tw tbody td:first-child{position:sticky;left:0;background:#fff;z-index:5;border-right:1px solid var(--border);color:#1e293b !important}
    .tw tbody tr:nth-child(even) td:first-child{background:#f8fafc}
    .tw tbody tr:hover td:first-child{background:#eff6ff}
    .tw tbody td{padding:5px 8px;border-bottom:1px solid var(--border);white-space:nowrap;color:#1e293b !important}
    .tw tbody tr:nth-child(even) td{background:#f8fafc}
    .tw tbody tr:hover td{background:#eff6ff!important}
    .cb td{background:#2563eb!important;color:#fff!important;font-weight:700!important;font-size:12px!important}
    
    .plan-action-table { width:100%; border-collapse:collapse; font-family:Inter,sans-serif; font-size:12px; border:1px solid #cbd5e1; }
    .plan-action-table th { background:#1e3a5f; color:#fff; font-weight:700; padding:8px 6px; border:1px solid #1e3a5f; }
    .plan-action-table td { padding:6px 8px; border:1px solid #cbd5e1; text-align:center; vertical-align:middle;}
    .plan-action-table td:first-child { text-align:left; font-weight:800; }
    
    .stTabs [data-baseweb="tab-list"]{gap:6px;background:#e2e8f0;padding:6px;border-radius:8px;margin-bottom:8px}
    .stTabs [data-baseweb="tab"]{
        border-radius:6px;
        padding:12px 22px;
        font-weight:700;
        font-size:20px;
        line-height:1.5;
        min-height:48px;
    }
    .stTabs [data-baseweb="tab"] span,
    .stTabs [data-baseweb="tab"] > div{
        font-size:22px !important;
    }
    .stTabs [aria-selected="true"]{
        background:#fff!important;
        color:var(--primary)!important;
        box-shadow:0 3px 8px rgba(0,0,0,.1);
        font-size:21px;
    }
    .stTabs [data-baseweb="tab"] svg{width:22px;height:22px}
    
    .ca{background:#fff;border-radius:var(--radius);padding:12px;margin-top:6px;border:1px solid var(--border);box-shadow:0 1px 4px rgba(0,0,0,.02)}
    .ca .ct{font-size:14px;font-weight:700;margin-bottom:8px;padding-bottom:5px;border-bottom:1px solid var(--border)}
    .car{display:flex;align-items:center;margin-bottom:6px;font-size:12px}
    .car:last-child{margin-bottom:0}
    .car .cal{width:260px;font-weight:600;color:var(--primary);text-align:right;padding-right:8px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .car .cab{flex:1;height:26px;background:#edf2f7;border-radius:4px;overflow:visible;position:relative}
    .car .caf{height:100%;border-radius:4px;transition:width .3s}
    
    .car .target-mark{
        position:absolute;
        top:-4px;
        bottom:-4px;
        width:3px;
        background:var(--info) !important; 
        z-index:20;
        transform:translateX(-50%);
        box-shadow:0 0 6px rgba(59,130,246,.9),0 0 2px rgba(0,0,0,.4);
        border-radius:2px;
    }
    .car .cav-out{font-size:12px;font-weight:800;color:#1e293b;min-width:55px;text-align:right;padding-left:6px}
    .car .cav-tgt{font-size:10px;font-weight:700;color:#1e293b;min-width:42px;text-align:right;padding-left:4px;opacity:.7}
    .gbr{display:flex;align-items:center;padding:3px 0;font-size:12px;border-bottom:1px solid #f1f5f9}
    .gbr:last-child{border:none}
    .gbr-l{width:160px;font-weight:600;color:#1e293b;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;font-size:11px;padding-right:10px;}
    .gbr-g{display:flex;align-items:center;gap:4px;flex:1;position:relative}
    .gbr-target{position:absolute;left:90%;top:-4px;bottom:-4px;width:3px;background:var(--primary) !important;z-index:10;box-shadow:0 0 6px rgba(30,58,95,.8);border-radius:2px}
    .gbr-target-label{position:absolute;left:90%;top:-20px;transform:translateX(-50%);font-size:9px;font-weight:800;color:#fff;background:var(--primary) !important;padding:1px 5px;border-radius:3px;white-space:nowrap;z-index:11;box-shadow:0 1px 3px rgba(0,0,0,.2)}
    .gbr-w{flex:1;height:22px;background:#f1f5f9;border-radius:3px;overflow:hidden}
    .gbr-f{height:100%;border-radius:3px}
    .gbr-v{font-size:11px;font-weight:800;min-width:48px;text-align:right;color:#1e293b}
    .gbr-legend{display:flex;gap:14px;margin-bottom:10px;font-size:12px;font-weight:700;align-items:center}
    .gbr-legend span{display:flex;align-items:center;gap:5px}
    .gbr-legend i{display:inline-block;width:14px;height:14px;border-radius:2px}
    .gbr-legend .target-icon{display:inline-block;width:3px;height:14px;background:var(--primary) !important;border-radius:1px;box-shadow:0 0 3px rgba(30,58,95,.6)}
    .cg{display:grid;grid-template-columns:1fr 1fr;gap:6px}
    .cg>div{background:#fff;border-radius:var(--radius);padding:8px 10px;border:1px solid var(--border)}
    .cg .ct{font-size:13px;font-weight:700;margin-bottom:4px;padding-bottom:3px;border-bottom:1px solid var(--border)}
    .cgr{display:flex;align-items:center;padding:3px 0;font-size:12px;border-bottom:1px solid #f1f5f9}
    .cgr:last-child{border:none}
    .cgr .rk{width:18px;font-weight:800;text-align:center}
    .cgr .pn{flex:1;font-weight:600;color:#1e293b;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .cgr .ps{font-weight:800;min-width:55px;text-align:right}
    .dgrid{display:grid;grid-template-columns:1fr 1fr;gap:6px}
    .stButton>button[kind="primary"]{background:linear-gradient(135deg,var(--primary),var(--primary-light));border:none;border-radius:6px;padding:8px 14px;font-weight:700;font-size:15px;width:100%}
    ::-webkit-scrollbar{width:5px;height:5px}::-webkit-scrollbar-track{background:#f1f1f1}::-webkit-scrollbar-thumb{background:#cbd5e0;border-radius:3px}
    
    div[data-testid="stSidebar"]{
        background:linear-gradient(180deg,#1e40af 0%,#1e3a8a 50%,#1e3a5f 100%)!important;
    }
    div[data-testid="stSidebar"]*{color:rgba(255,255,255,.9)!important}
    div[data-testid="stSidebar"] .stSelectbox label,div[data-testid="stSidebar"] .stMultiSelect label,div[data-testid="stSidebar"] .stDateInput label,div[data-testid="stSidebar"] .stCheckbox label,div[data-testid="stSidebar"] .stTextInput label{color:rgba(255,255,255,.9)!important;font-weight:600;font-size:13px;text-transform:uppercase;letter-spacing:.5px}
    div[data-testid="stSidebar"] div[data-testid="stWidget"]{background:rgba(255,255,255,.1);border-radius:6px;padding:5px 10px;margin-bottom:5px;border:1px solid rgba(255,255,255,.15)}
    div[data-testid="stSidebar"] .stSelectbox>div>div,div[data-testid="stSidebar"] .stMultiSelect>div>div,div[data-testid="stSidebar"] .stDateInput>div>div,div[data-testid="stSidebar"] .stTextInput>div>div{background:rgba(255,255,255,.95)!important;border-radius:5px}
    
    .es{text-align:center;padding:14px;color:#64748b;font-size:14px}
    .synth-tbl{width:100%;border-collapse:collapse;font-family:'Inter',sans-serif;font-size:12px}
    .synth-tbl thead th{background:var(--primary);color:#fff;font-weight:700;font-size:11px;padding:5px 8px;border:none;white-space:nowrap;position:sticky;top:0}
    .synth-tbl tbody td{padding:4px 8px;border-bottom:1px solid var(--border);text-align:center;color:#1e293b !important}
    .synth-tbl tbody tr:nth-child(even) td{background:#f8fafc}
    .synth-tbl tbody tr:hover td{background:#eff6ff!important}
    .synth-tbl .poste-cell{text-align:left;font-weight:700;white-space:nowrap;min-width:140px;color:#1e293b !important}
    
    [data-testid="stHeaderActionElements"]{display:none !important;}
    [data-testid="stActionButtonContainer"]{display:none !important;}
    
    .footer {
        text-align: center;
        margin-top: 30px;
        padding: 15px;
        color: #64748b;
        font-size: 13px;
        border-top: 1px solid var(--border);
        font-weight: 600;
    }

    div[data-testid="stDataEditor"] table, 
    div[data-testid="stDataEditor"] th, 
    div[data-testid="stDataEditor"] td {
        font-size: 18px !important;
        line-height: 1.4 !important;
        white-space: normal !important;
        word-wrap: break-word !important;
    }
    div[data-testid="stDataEditor"] [data-testid="stMarkdownContainer"] {
        font-size: 18px !important;
    }
    div[data-testid="stDataEditor"] [data-testid="stTable"] {
        overflow-x: hidden !important;
        width: 100% !important;
    }

    @media(max-width:768px){
        .cr{grid-template-columns:repeat(2,1fr)}
        .mh{padding:8px 10px;gap:8px}
        .mh h1{font-size:18px}
        .mh .logo{height:35px;max-width:70px}
        .mh .db{font-size:11px;padding:2px 8px}
        .cg,.dgrid{grid-template-columns:1fr}
        .car{flex-wrap:wrap;gap:2px}
        .car .cal{width:100%;text-align:left;padding-right:0;margin-bottom:2px}
        .car .cab{flex:1 1 70%}
        .car .cav-out,.car .cav-tgt{flex:1 1 15%;min-width:40px}
        .gbr{flex-direction:column;align-items:flex-start;gap:4px}
        .gbr-l{width:100%;margin-bottom:2px}
        .gbr-g{width:100%;flex-wrap:wrap}
        .gbr-w{flex:1 1 45%}
        .gbr-v{flex:1 1 10%;min-width:40px}
        .tw{font-size:10px}
        .tw thead th,.tw tbody td{padding:3px 4px}
        .stl{font-size:13px}
        .stTabs [data-baseweb="tab"]{padding:8px 12px;font-size:15px}
        .stTabs [data-baseweb="tab"] span{font-size:16px !important}
        div[data-testid="stDataEditor"] table, div[data-testid="stDataEditor"] th, div[data-testid="stDataEditor"] td { font-size: 14px !important; }
    }
    
    @media print {
        section[data-testid="stSidebar"], 
        header[data-testid="stHeader"], 
        div[data-testid="stToolbar"], 
        div[data-testid="stHeaderActionElements"],
        footer, 
        .stDeployButton, 
        #MainMenu { display: none !important; }
        .main .block-container { 
            padding-top: 0 !important; 
            padding-left: 0 !important; 
            padding-right: 0 !important; 
            max-width: 100% !important; 
        }
        .stButton, .stDownloadButton { display: none !important; }
        * {
            -webkit-print-color-adjust: exact !important;
            print-color-adjust: exact !important;
        }
        .tw, .synth-tbl { page-break-inside: avoid; overflow: visible !important; }
    }
    </style>""",unsafe_allow_html=True)

# ============================================================
def main():
    try: locale.setlocale(locale.LC_ALL,'fr_FR.UTF-8')
    except Exception:
        try: locale.setlocale(locale.LC_ALL,'fr_FR')
        except Exception: pass
    inject_custom_css()
    fichier_date=get_date_from_file()

    if "hse_affiche" not in st.session_state: st.session_state.hse_affiche=False
    if not st.session_state.hse_affiche:
        c=random.choice(CONSIGNES_HSE)
        st.markdown("""<div style="min-height:100vh;display:flex;flex-direction:column;align-items:center;justify-content:center;background:linear-gradient(135deg,#1a365d,#2d3748,#1a365d);padding:40px">
        <div style="font-size:64px;margin-bottom:20px">🦺</div>
        <h1 style="text-align:center;font-size:46px;color:#fff;font-weight:900;margin:0">HSE - CONSIGNE DE SECURITE</h1>
        <p style="text-align:center;color:rgba(255,255,255,.6);font-size:22px;margin-top:8px;letter-spacing:3px;text-transform:uppercase">Securite - Sante - Environnement</p>
        <div style="background:linear-gradient(135deg,#f6e05e,#ed8936);padding:36px 48px;border-radius:20px;font-size:32px;font-weight:700;text-align:center;margin:40px 0;color:#1a202c;max-width:800px;box-shadow:0 20px 60px rgba(0,0,0,.3)">⚠️ %s</div>
        <h2 style="text-align:center;color:#48bb78;font-size:36px;font-weight:900">Aucun travail n'est plus urgent que la securite</h2>
        <div style="margin-top:40px;width:200px;height:4px;background:rgba(255,255,255,.1);border-radius:2px;overflow:hidden"><div style="width:100%%;height:100%%;background:linear-gradient(90deg,#48bb78,#38a169);border-radius:2px;animation:ld 5.5s ease-in-out forwards"></div></div>
        <style>@keyframes ld{from{width:0}to{width:100%%}}</style></div>"""%c,unsafe_allow_html=True)
        time.sleep(6); st.session_state.hse_affiche=True; st.rerun(); st.stop()

    def ckpi(n,d,sz=100): return np.where(d==0,sz,(n/d)*100)
    
    def cpiv(df,f,c,p):
        return pd.pivot_table(df[f],index="Poste travail princ.",columns=c,values="Ordre",aggfunc="count",fill_value=0).reindex(p,fill_value=0)
        
    def get_text_col(df):
        for c in ["Désignation","Designation","Désignation OT","Texte ordre","Texte","Description","Libellé","Libelle"]:
            if c in df.columns: return c
        for c in df.columns:
            if df[c].dtype=='object' and any(kw in str(c).lower() for kw in ['sign','text','desc','libell']):
                return c
        return None
        
    def build_statut_pivot(df_sub, posts):
        if df_sub.empty:
            return pd.DataFrame(index=posts, columns=["CRÉÉ","LANC","CLOT","TCLO","Total"]).fillna(0).astype(int)
        piv=pd.pivot_table(df_sub, index="Poste travail princ.", columns="Statut OT", values="Ordre", aggfunc="count", fill_value=0)
        for s in ["CRÉÉ","LANC","CLOT","TCLO"]:
            if s not in piv.columns: piv[s]=0
        piv["Total"]=piv[["CRÉÉ","LANC","CLOT","TCLO"]].sum(axis=1)
        return piv.reindex(posts, fill_value=0).fillna(0).astype(int)
        
    def html_statut_pivot(piv_df, table_class):
        cols=["Poste de travail","CRÉÉ","LANC","CLOT","TCLO","Total"]
        statut_colors = {
            "CRÉÉ": "background:#fef3c7;color:#92400e;font-weight:600;",
            "LANC": "background:#dbeafe;color:#1e40af;font-weight:600;",
            "CLOT": "background:#d1fae5;color:#065f46;font-weight:600;",
            "TCLO": "background:#a7f3d0;color:#064e3b;font-weight:600;",
            "Total": "background:#ede9fe;color:#5b21b6;font-weight:700;"
        }
        h='<table class="tw %s"><thead><tr>'%table_class+''.join('<th>%s</th>'%c for c in cols)+'</tr></thead><tbody>'
        for poste,row in piv_df.iterrows():
            h+='<tr><td style="font-weight:700">%s</td>'%poste
            for c in ["CRÉÉ","LANC","CLOT","TCLO"]:
                h+='<td style="text-align:center;%s">%d</td>'%(statut_colors[c], int(row.get(c,0)))
            h+='<td style="text-align:center;%s">%d</td>'%(statut_colors["Total"], int(row.get("Total",0)))
            h+='</tr>'
        h+='<tr class="tr"><td style="font-weight:800">Total</td>'
        for c in ["CRÉÉ","LANC","CLOT","TCLO"]:
            h+='<td style="text-align:center;font-weight:800;%s">%d</td>'%(statut_colors[c], int(piv_df[c].sum()))
        h+='<td style="text-align:center;font-weight:800;%s">%d</td>'%(statut_colors["Total"], int(piv_df["Total"].sum()))
        h+='</tr></tbody></table>'
        return h
        
    def show_pie_pair(piv_df, title_prefix):
        global_counts = piv_df[["CRÉÉ","LANC","CLOT","TCLO"]].sum()
        global_counts = global_counts[global_counts > 0]
        realised = global_counts.get("CLOT", 0) + global_counts.get("TCLO", 0)
        not_realised = global_counts.sum() - realised
        
        if global_counts.empty:
            st.markdown('<div class="es">Aucune donnee</div>', unsafe_allow_html=True)
            return
            
        colors = ["#8b5cf6", "#f59e0b", "#10b981", "#3b82f6"]
        fig = make_subplots(rows=1, cols=2, specs=[[{"type":"domain"},{"type":"domain"}]], 
                            subplot_titles=(f"{title_prefix} — Par Statut OT", f"{title_prefix} — Réalisés vs Non Réalisés"))
        
        fig.add_trace(go.Pie(labels=global_counts.index, values=global_counts.values, hole=0.4, 
                             textinfo='percent+label', 
                             texttemplate='%{label}<br>%{percent:.1%}<br>(%{value})', 
                             textposition='inside',
                             insidetextorientation='radial',
                             textfont=dict(size=14, color='white', family='Inter, sans-serif'),
                             marker=dict(colors=colors, line=dict(color='#FFFFFF', width=3))), 1, 1)
                             
        pie2_data = pd.Series([realised, not_realised], index=["Réalisés (CLOT+TCLO)", "Non Réalisés"])
        
        fig.add_trace(go.Pie(labels=pie2_data.index, values=pie2_data.values, hole=0.5, 
                             textinfo='percent+label', 
                             texttemplate='%{label}<br>%{percent:.1%}<br>(%{value})', 
                             textposition='inside',
                             insidetextorientation='radial',
                             textfont=dict(size=14, color='white', family='Inter, sans-serif'),
                             marker=dict(colors=["#10b981", "#8b5cf6"], line=dict(color='#FFFFFF', width=3))), 1, 2)
                             
        fig.update_layout(margin=dict(t=80, b=20, l=20, r=20), height=450, 
                          legend=dict(orientation="h", yanchor="bottom", y=-0.12, x=0.5, xanchor="center"))
                          
        st.plotly_chart(fig, use_container_width=True)

    def show_simple_pie(piv_df, title, keep_non_carac=False):
        if not keep_non_carac and "NON CARACTERISE" in piv_df.columns:
            piv_df = piv_df.drop(columns=["NON CARACTERISE"])
            
        counts = piv_df.sum()
        counts = counts[counts > 0]
        
        if counts.empty:
            st.markdown('<div class="es">Aucune donnee</div>', unsafe_allow_html=True)
            return
            
        color_map = {"CARACTERISE": "#10b981", "NON CARACTERISE": "#f97316"}
        type_palette = ['#3b82f6', '#10b981', '#f59e0b', '#8b5cf6', '#06b6d4', '#14b8a6', '#6366f1', '#0ea5e9', '#d946ef', '#a855f7']
        
        colors = []
        palette_idx = 0
        for c in counts.index:
            c_str = str(c)
            if c_str in color_map:
                colors.append(color_map[c_str])
            else:
                colors.append(type_palette[palette_idx % len(type_palette)])
                palette_idx += 1
        
        total_sum = counts.sum()
        pull_list = [0.05 if (v/total_sum)*100 < 10 else 0 for v in counts.values]
        
        fig = go.Figure(
            go.Pie(
                labels=counts.index,
                values=counts.values,
                hole=0.4,
                sort=False,
                textinfo="percent",
                textposition="outside",
                pull=pull_list,
                marker=dict(
                    colors=colors,
                    line=dict(color="white", width=2)
                )
            )
        )

        fig.update_traces(
            hovertemplate="<b>%{label}</b><br>Nombre : %{value}<br>Pourcentage : %{percent}<extra></extra>",
            textfont=dict(size=13, family='Inter, sans-serif')
        )

        fig.update_layout(
            title=dict(text=title, x=0.5, xanchor='center', font=dict(size=16)), 
            height=500, 
            showlegend=True,
            legend=dict(orientation="h", yanchor="bottom", y=-0.15, x=0.5, xanchor="center"),
            margin=dict(t=80, b=80, l=40, r=40)
        )
                          
        st.plotly_chart(fig, use_container_width=True)

    # ======================================================================
    # NOUVEAU FILTRE D'ANOMALIE : basé sur CIBLE et LOWER_BETTER
    # ======================================================================
    def is_anomaly(kpi_name, value):
        """
        Nouveau filtre : un KPI est en anomalie s'il n'atteint pas sa cible.
        - Pour les KPI LOWER_BETTER : anomalie si valeur > cible
        - Pour les autres KPI : anomalie si valeur < cible
        """
        target = CIBLE.get(kpi_name, 100)
        try:
            v = float(value)
        except (ValueError, TypeError):
            return False
        if kpi_name in LOWER_BETTER:
            return v > target
        else:
            return v < target

    def build_anomalies_table(ckdf, kpi_list, posts):
        """
        Construit la matrice d'anomalies pour une liste de KPIs et postes donnés.
        Retourne un DataFrame avec 1=anomalie, 0=OK.
        """
        anom_matrix = pd.DataFrame(index=posts, columns=kpi_list, data=0)
        for kpi in kpi_list:
            if kpi not in ckdf.columns:
                continue
            for poste in posts:
                if poste in ckdf.index:
                    val = ckdf.loc[poste, kpi]
                    if is_anomaly(kpi, val):
                        anom_matrix.loc[poste, kpi] = 1
        return anom_matrix.astype(int)

    def html_anomalies_heatmap(anom_matrix, kpi_list, section_label, table_class):
        """
        Génère un tableau HTML heatmap des anomalies par KPI et Poste.
        1 = cellule rouge foncé (anomalie), 0 = cellule verte clair (OK)
        Avec totaux par ligne et par colonne.
        """
        h = '<table class="tw %s"><thead><tr>' % table_class
        h += '<th>Poste de travail</th>'
        for kpi in kpi_list:
            short = kpi.replace("OT préparation ", "Prép ").replace("OT planification ", "Plan ").replace("OT exécution ", "Exéc ")
            short = short.replace("1mois< <3mois", "1-3m").replace("Performance ", "Perf ")
            h += '<th style="font-size:10px;writing-mode:vertical-rl;text-orientation:mixed;max-width:30px;padding:8px 3px;">%s</th>' % short
        h += '<th>Nb Anomalies</th></tr></thead><tbody>'
        
        total_par_kpi = {kpi: 0 for kpi in kpi_list}
        total_anomalies_global = 0
        
        for poste in anom_matrix.index:
            h += '<tr><td style="font-weight:700;white-space:nowrap;font-size:11px;">%s</td>' % poste
            nb_ano_poste = 0
            for kpi in kpi_list:
                val = int(anom_matrix.loc[poste, kpi])
                total_par_kpi[kpi] += val
                if val == 1:
                    h += '<td style="text-align:center;background:#dc2626;color:#fff;font-weight:800;font-size:14px;">&#10005;</td>'
                    nb_ano_poste += 1
                    total_anomalies_global += 1
                else:
                    h += '<td style="text-align:center;background:#d1fae5;color:#065f46;font-size:12px;">&#10003;</td>'
            # Style du nombre d'anomalies par poste
            if nb_ano_poste == 0:
                style = "background:#d1fae5;color:#065f46;font-weight:800;"
            elif nb_ano_poste <= 2:
                style = "background:#fef3c7;color:#92400e;font-weight:800;"
            else:
                style = "background:#dc2626;color:#fff;font-weight:800;"
            h += '<td style="text-align:center;%s">%d</td>' % (style, nb_ano_poste)
            h += '</tr>'
        
        # Ligne totaux par KPI
        h += '<tr style="background:#1e3a5f;"><td style="font-weight:800;color:#fff;">Nb Anomalies</td>'
        for kpi in kpi_list:
            v = total_par_kpi[kpi]
            if v == 0:
                h += '<td style="text-align:center;color:#d1fae5;font-weight:800;">%d</td>' % v
            else:
                h += '<td style="text-align:center;color:#fbbf24;font-weight:800;">%d</td>' % v
        h += '<td style="text-align:center;color:#fbbf24;font-weight:900;font-size:15px;">%d</td>' % total_anomalies_global
        h += '</tr></tbody></table>'
        return h, total_par_kpi, total_anomalies_global

    def html_anomalies_detail(anom_matrix, ckdf, kpi_list, posts):
        """
        Génère un tableau détaillé des anomalies avec valeur réelle, cible, écart et responsable.
        Une ligne par anomalie détectée.
        """
        rows = []
        for poste in posts:
            for kpi in kpi_list:
                if kpi not in ckdf.columns:
                    continue
                if int(anom_matrix.loc[poste, kpi]) == 1:
                    val = float(ckdf.loc[poste, kpi])
                    cible = CIBLE.get(kpi, 100)
                    if kpi in LOWER_BETTER:
                        ecart = val - cible
                    else:
                        ecart = cible - val
                    resp = KPI_RESP_MAP.get(kpi, "")
                    action = ACT_MAP.get(kpi, "")
                    rows.append({
                        "Poste de travail": poste,
                        "KPI": kpi,
                        "Valeur (%)": round(val, 1),
                        "Cible (%)": cible,
                        "Ecart (pts)": round(ecart, 1),
                        "Responsable": resp,
                        "Action corrective": action
                    })
        return rows

    # ======================================================================

    def calc_kpis(df_i, av_i, now_ts, posts):
        res={}; df=df_i.copy(); av=av_i.copy()
        res['dfp']=df
        filt_corr=(df["Nº appel pl.entret."].fillna(0)==0)&(df["Contient SOPL"]==1)
        an=cpiv(df,filt_corr,"Statut OT",posts)
        for c in ["CLOT","CRÉÉ","LANC","TCLO"]: an[c]=an.get(c,0)
        an["OT_CLOTURES"]=an["CLOT"]+an["TCLO"]
        an["TOTAL_OT"]=an[["CLOT","CRÉÉ","LANC","TCLO"]].sum(axis=1)
        an["TAUX_REALISATION_CORRECTIF/PT"]=np.where(an["TOTAL_OT"]==0,100.0,ckpi(an["OT_CLOTURES"],an["TOTAL_OT"]))
        
        pr=cpiv(df,(df["Statut OT"]=="CRÉÉ")&(df["Statut utilisateur"].str.contains("CRPR",na=False)),"ap",posts)
        for c in ["<1 mois",">3 mois","1 mois < <3 mois","Inconnu"]: pr[c]=pr.get(c,0)
        pr["Total"]=pr[["<1 mois","1 mois < <3 mois",">3 mois","Inconnu"]].sum(axis=1)
        pr["OT préparation <1 mois"]=ckpi(pr["<1 mois"],pr["Total"]); pr["OT préparation >3 mois"]=ckpi(pr[">3 mois"],pr["Total"],0); pr["OT préparation 1mois< <3mois"]=ckpi(pr["1 mois < <3 mois"],pr["Total"],0)
        
        pl=cpiv(df,(df["Statut OT"]=="LANC")&(df["Statut utilisateur"].str.contains("ATPL",case=False,na=False)),"alp",posts)
        for c in ["<1 mois",">3 mois","1 mois < <3 mois","Inconnu"]: pl[c]=pl.get(c,0)
        pl["Total"]=pl[["<1 mois","1 mois < <3 mois",">3 mois","Inconnu"]].sum(axis=1)
        pl["OT planification <1 mois"]=ckpi(pl["<1 mois"],pl["Total"]); pl["OT planification >3 mois"]=ckpi(pl[">3 mois"],pl["Total"],0); pl["OT planification 1mois< <3mois"]=ckpi(pl["1 mois < <3 mois"],pl["Total"],0)
        
        ex=cpiv(df,(df["Statut OT"]=="LANC")&(df["Contient SOPL"]==1),"aex",posts)
        for c in ["<1 mois",">3 mois","1 mois < <3 mois","Inconnu"]: ex[c]=ex.get(c,0)
        ex["Total"]=ex[["<1 mois","1 mois < <3 mois",">3 mois","Inconnu"]].sum(axis=1)
        ex["OT exécution <1 mois"]=ckpi(ex["<1 mois"],ex["Total"]); ex["OT exécution >3 mois"]=ckpi(ex[">3 mois"],ex["Total"],0); ex["OT exécution 1mois< <3mois"]=ckpi(ex["1 mois < <3 mois"],ex["Total"],0)
        
        la=pd.pivot_table(df[df["Statut OT"]=="LANC"],index="Poste travail princ.",columns="OT LANC ESTIME",values="Ordre",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
        for c in ["OUI","NON"]: la[c]=la.get(c,0)
        la["Total"]=la["OUI"]+la["NON"]; la["OT LANC ESTIME"]=ckpi(la["OUI"],la["Total"])
        
        pc=pd.pivot_table(df[df["Statut OT"]=="CRÉÉ"],index="Poste travail princ.",columns="Backlog preparation",values="Ordre",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
        for c in ["CARACTERISE","NON CARACTERISE"]: pc[c]=pc.get(c,0)
        pc["Total"]=pc["CARACTERISE"]+pc["NON CARACTERISE"]; pc["Backlog préparation caractérisé"]=ckpi(pc["CARACTERISE"],pc["Total"])
        
        plc=pd.pivot_table(df[(df["Statut OT"]=="LANC")&(df["Contient SOPL"]==0)],index="Poste travail princ.",columns="Backlog planification",values="Ordre",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
        for c in ["CARACTERISE","NON CARACTERISE"]: plc[c]=plc.get(c,0)
        plc["Total"]=plc["CARACTERISE"]+plc["NON CARACTERISE"]; plc["Backlog planification caractérisé"]=ckpi(plc["CARACTERISE"],plc["Total"])
        
        for kn,cn in [("OT CONFIME","OT CONFIME"),("OT_COR_EGAL","OT_COR_EGAL")]:
            pv=pd.pivot_table(df[df["Statut OT"].isin(["CLOT","TCLO"])],index="Poste travail princ.",columns="OT_COR_EGAL",values="Ordre",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
            for c in ["OUI","NON"]: pv[c]=pv.get(c,0)
            pv["Total"]=pv["OUI"]+pv["NON"]; pv[cn]=ckpi(pv["OUI"],pv["Total"]); res[kn.lower().replace(" ","_")]=pv
            
        avf=av.copy(); res['avf']=avf
        tca=pd.pivot_table(avf,index="Poste travail princ.",columns="Statut utilisateur",values="Avis",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
        for c in ["APRQ","APRV","APRV AVAU","REJT"]: tca[c]=tca.get(c,0)
        tca["Total"]=tca[["APRQ","APRV","APRV AVAU","REJT"]].sum(axis=1); tca["Taux d'approbation des Avis"] = ckpi(tca["APRV"] ,tca["Total"])
        
        g_num=df[(df["Statut OT"].isin(["CLOT","TCLO"]))&(df["_tw_num"]==350)].groupby("Poste travail princ.")["Ordre"].count()
        g_den=df[(df["Contient SOPL"]==1)&(df["_tw_num"]==350)].groupby("Poste travail princ.")["Ordre"].count()
        g_df=pd.DataFrame({"_n":g_num,"_d":g_den}).reindex(posts,fill_value=0)
        g_df["Performance Graissage"]=np.where(g_df["_d"]==0,100.0,(g_df["_n"]/g_df["_d"])*100)
        
        ins_types=[290,300,310]
        ins_base=(df["_tw_num"].isin(ins_types))&(df["Date de début planifiée"].notna())&(df["Date de début planifiée"]<=now_ts)
        ins_num=df[(df["Statut OT"].isin(["CLOT","TCLO"]))&ins_base].groupby("Poste travail princ.")["Ordre"].count()
        ins_den=df[(df["Contient SOPL"]==1)&ins_base].groupby("Poste travail princ.")["Ordre"].count()
        ins_df=pd.DataFrame({"_n":ins_num,"_d":ins_den}).reindex(posts,fill_value=0)
        ins_df["Performance Inspection"]=np.where(ins_df["_d"]==0,100.0,(ins_df["_n"]/ins_df["_d"])*100)
        
        sys_base=(df["_tw_num"]==360)&(df["Date de début planifiée"].notna())&(df["Date de début planifiée"]<=now_ts)
        sys_num=df[(df["Statut OT"].isin(["CLOT","TCLO"]))&sys_base].groupby("Poste travail princ.")["Ordre"].count()
        sys_den=df[(df["Contient SOPL"]==1)&sys_base].groupby("Poste travail princ.")["Ordre"].count()
        sys_df=pd.DataFrame({"_n":sys_num,"_d":sys_den}).reindex(posts,fill_value=0)
        sys_df["Performance Appels Systématiques"]=np.where(sys_df["_d"]==0,100.0,(sys_df["_n"]/sys_df["_d"])*100)
        
        fiab_s=pd.Series(100.0,index=posts); avpan_s=pd.Series(100.0,index=posts)
        res['ckdf']=pd.DataFrame({
            "TAUX_REALISATION_CORRECTIF/PT":an["TAUX_REALISATION_CORRECTIF/PT"],
            "OT préparation <1 mois":pr["OT préparation <1 mois"],"OT préparation >3 mois":pr["OT préparation >3 mois"],"OT préparation 1mois< <3mois":pr["OT préparation 1mois< <3mois"],
            "OT planification <1 mois":pl["OT planification <1 mois"],"OT planification >3 mois":pl["OT planification >3 mois"],"OT planification 1mois< <3mois":pl["OT planification 1mois< <3mois"],
            "OT exécution <1 mois":ex["OT exécution <1 mois"],"OT exécution >3 mois":ex["OT exécution >3 mois"],"OT exécution 1mois< <3mois":ex["OT exécution 1mois< <3mois"],
            "Performance Graissage":g_df["Performance Graissage"],"Performance Inspection":ins_df["Performance Inspection"],"Performance Appels Systématiques":sys_df["Performance Appels Systématiques"],
            "Taux d'approbation des Avis":tca["Taux d'approbation des Avis"],"OT LANC ESTIME":la["OT LANC ESTIME"],
            "Backlog préparation caractérisé":pc["Backlog préparation caractérisé"],"Backlog planification caractérisé":plc["Backlog planification caractérisé"],
            "OT CONFIME":res['ot_confime']["OT CONFIME"],"OT_COR_EGAL":res['ot_cor_egal']["OT_COR_EGAL"],
            "OT Fiabilité":fiab_s,"Total Avis de Panne":avpan_s
        })
        return res

    def get_bar_color(kpi, val):
        try: v = float(val)
        except: return "#cbd5e0"
        if kpi in ["OT préparation <1 mois","OT planification <1 mois","OT exécution <1 mois"]:
            if v>=80: return "#38a169"
            elif v>=75: return "#f59e0b"
            else: return "#e53e3e"
        if kpi in ["OT préparation 1mois< <3mois","OT planification 1mois< <3mois","OT exécution 1mois< <3mois"]:
            return "#38a169" if v<=15 else "#e53e3e"
        if kpi in ["OT préparation >3 mois","OT planification >3 mois","OT exécution >3 mois"]:
            return "#38a169" if v<=5 else "#e53e3e"
        if kpi=="TAUX_REALISATION_CORRECTIF/PT":
            if v>=85: return "#38a169"
            elif v>=80: return "#f59e0b"
            else: return "#e53e3e"
        if kpi=="Taux d'approbation des Avis":
            if v>=95: return "#38a169"
            elif v>=90: return "#f59e0b"
            else: return "#e53e3e"
        if kpi in ["OT LANC ESTIME","Backlog préparation caractérisé","Backlog planification caractérisé","OT CONFIME","OT_COR_EGAL"]:
            if v>=100: return "#38a169"
            elif v>=95: return "#f59e0b"
            else: return "#e53e3e"
        if kpi in ["Performance Graissage","Performance Inspection","Performance Appels Systématiques"]:
            if v>=95: return "#38a169"
            elif v>90: return "#f59e0b"
            else: return "#e53e3e"
        if kpi in ["OT Fiabilité","Total Avis de Panne"]:
            return "#38a169" if v>=100 else "#f59e0b"
        if v>=90: return "#38a169"
        elif v>=80: return "#f59e0b"
        else: return "#e53e3e"

    def ks(v,c):
        try: val=float(v)
        except Exception: return ""
        if c in ["OT préparation <1 mois","OT planification <1 mois","OT exécution <1 mois"]:
            if val>=80: return "✅"
            elif val>=75: return "⚠️"
            else: return "❌"
        if c in ["OT préparation 1mois< <3mois","OT planification 1mois< <3mois","OT exécution 1mois< <3mois"]:
            return "✅" if val<=15 else "❌"
        if c in ["OT préparation >3 mois","OT planification >3 mois","OT exécution >3 mois"]:
            return "✅" if val<=5 else "❌"
        if c=="TAUX_REALISATION_CORRECTIF/PT":
            if val>=85: return "✅"
            elif val>=80: return "⚠️"
            else: return "❌"
        if c=="Taux d'approbation des Avis":
            if val>=95: return "✅"
            elif val>=90: return "⚠️"
            else: return "❌"
        if c in ["OT LANC ESTIME","Backlog préparation caractérisé","Backlog planification caractérisé","OT CONFIME","OT_COR_EGAL"]:
            if val>=100: return "✅"
            elif val>=95: return "⚠️"
            else: return "❌"
        if c in ["Performance Graissage","Performance Inspection","Performance Appels Systématiques"]:
            if val>=95: return "✅"
            elif val>90: return "⚠️"
            else: return "❌"
        if c in ["OT Fiabilité","Total Avis de Panne"]:
            return "✅" if val>=100 else "⚠️"
        if val>=90: return "✅"
        elif val>=80: return "⚠️"
        else: return "❌"

    def html_kpi_table(ckdf, kpi_list, posts, table_class, show_score=False):
        cols = ["Poste de travail"] + kpi_list
        if show_score: cols.append("Score")
        h = '<table class="tw %s"><thead><tr>' % table_class
        for c in cols:
            short = c.replace("OT préparation ", "Prép ").replace("OT planification ", "Plan ").replace("OT exécution ", "Exéc ")
            short = short.replace("1mois< <3mois", "1-3m").replace("Performance ", "Perf ")
            h += '<th style="font-size:10px;">%s</th>' % short
        h += '</tr></thead><tbody>'
        for poste in posts:
            h += '<tr><td style="font-weight:700;white-space:nowrap;font-size:11px;">%s</td>' % poste
            score_vals = []
            for kpi in kpi_list:
                if kpi in ckdf.columns and poste in ckdf.index:
                    val = ckdf.loc[poste, kpi]
                    color = get_bar_color(kpi, val)
                    status = ks(val, kpi)
                    h += '<td style="text-align:center;color:%s;font-weight:700;font-size:12px;">%s %.1f%%</td>' % (color, status, val)
                    if kpi in CIBLE:
                        target = CIBLE[kpi]
                        if kpi in LOWER_BETTER:
                            score_vals.append(max(0, (target - val) / target * 100) if target > 0 else 100)
                        else:
                            score_vals.append(min(100, val / target * 100) if target > 0 else 100)
                    else:
                        score_vals.append(min(100, val))
                else:
                    h += '<td style="text-align:center;color:#94a3b8;">-</td>'
                    score_vals.append(0)
            if show_score:
                sc = np.mean(score_vals) if score_vals else 0
                sc_color = "#38a169" if sc >= 90 else "#f59e0b" if sc >= 75 else "#e53e3e"
                h += '<td style="text-align:center;color:%s;font-weight:900;font-size:13px;">%.1f%%</td>' % (sc_color, sc)
            h += '</tr>'
        # Ligne Cible
        h += '<tr class="cb"><td style="font-weight:800;">Cible</td>'
        for kpi in kpi_list:
            h += '<td style="text-align:center;">%d%%</td>' % CIBLE.get(kpi, 100)
        if show_score:
            h += '<td style="text-align:center;">100%%</td>'
        h += '</tr></tbody></table>'
        return h

    # ===== SIDEBAR =====
    with st.sidebar:
        logo_b64 = get_logo_base64()
        if logo_b64:
            st.markdown('<img src="data:image/png;base64,%s" style="width:100%%;border-radius:8px;margin-bottom:10px;">' % logo_b64, unsafe_allow_html=True)
        
        st.markdown('<div style="text-align:center;color:#fff;font-size:13px;font-weight:700;padding:8px 0;border-bottom:1px solid rgba(255,255,255,.2);margin-bottom:10px;">📅 %s</div>' % fichier_date, unsafe_allow_html=True)
        
        ot_file = st.file_uploader("Fichier OT (ot.xlsx)", type=["xlsx"], key="ot_up")
        av_file = st.file_uploader("Fichier Avis (avis.xlsx)", type=["xlsx"], key="av_up")
        
        st.markdown("---")
        sel_posts = st.multiselect("Filtrer par poste", [], key="sp")
        sel_resp = st.multiselect("Filtrer par responsable", sorted(set(KPI_RESP_MAP.values())), key="sr")
        
        st.markdown("---")
        show_plan = st.checkbox("Afficher Plan d'action", value=True, key="splan")
        show_anomalies_detail = st.checkbox("Afficher Detail Anomalies", value=True, key="sanodet")
        
        st.markdown("---")
        st.markdown('<div style="font-size:11px;color:rgba(255,255,255,.5);text-align:center;">Dashboard KPI v2.0<br>Filtre anomalies par cible</div>', unsafe_allow_html=True)

    # ===== MAIN =====
    if not ot_file or not av_file:
        st.markdown("""<div style="display:flex;flex-direction:column;align-items:center;justify-content:center;min-height:60vh;">
        <div style="font-size:80px;margin-bottom:20px;">📁</div>
        <h2 style="color:#1e3a5f;font-weight:800;">Chargement des donnees</h2>
        <p style="color:#64748b;font-size:16px;margin-top:8px;">Veuillez charger les fichiers <b>ot.xlsx</b> et <b>avis.xlsx</b> depuis la barre laterale.</p>
        </div>""", unsafe_allow_html=True)
        st.stop()

    with st.spinner("Traitement des donnees..."):
        df, avf, apm, now_ts = prepare_data(ot_file.read(), av_file.read(), fichier_date)

    if not apm:
        st.markdown('<div class="es">Aucun poste SF1/SF2 trouve dans les donnees.</div>', unsafe_allow_html=True)
        st.stop()

    posts = sel_posts if sel_posts else apm
    
    # Filtrer par responsable si sélectionné
    if sel_resp:
        kpis_for_resp = [k for k, v in KPI_RESP_MAP.items() if v in sel_resp]
    else:
        kpis_for_resp = None

    with st.spinner("Calcul des KPIs..."):
        res = calc_kpis(df, avf, now_ts, apm)
    ckdf = res['ckdf']

    # ===== HEADER =====
    logo_html = ''
    logo_b64 = get_logo_base64()
    if logo_b64:
        logo_html = '<img class="logo" src="data:image/png;base64,%s" alt="Logo">' % logo_b64
    st.markdown('<div class="mh">%s<h1>DASHBOARD KPI MAINTENANCE</h1><div class="db">📅 %s</div></div>' % (logo_html, fichier_date), unsafe_allow_html=True)

    # ===== CARTES RESUME =====
    perf_score = ckdf.loc[posts, QK].apply(lambda col: np.where(col >= np.array([CIBLE.get(k,100) for k in QK]), 100, col / np.array([CIBLE.get(k,100) for k in QK]) * 100)).mean().mean() if not ckdf.loc[posts, QK].empty else 0
    qual_score = ckdf.loc[posts, PK].apply(lambda col: np.where(col >= np.array([CIBLE.get(k,100) for k in PK]), 100, col / np.array([CIBLE.get(k,100) for k in PK]) * 100)).mean().mean() if not ckdf.loc[posts, PK].empty else 0
    
    # NOUVEAU FILTRE : compter les anomalies avec is_anomaly
    anom_perf_matrix = build_anomalies_table(ckdf, QK, posts)
    anom_qual_matrix = build_anomalies_table(ckdf, PK, posts)
    nb_ano_perf = int(anom_perf_matrix.sum().sum())
    nb_ano_qual = int(anom_qual_matrix.sum().sum())
    nb_ano_total = nb_ano_perf + nb_ano_qual
    nb_kpi_total = len(QK) * len(posts) + len(PK) * len(posts)
    pct_conformite = round((1 - nb_ano_total / max(nb_kpi_total, 1)) * 100, 1)

    st.markdown('<div class="cr">'
        '<div class="cc c1"><div class="cv">%.1f%%</div><div class="cl">Score Performance</div></div>'
        '<div class="cc c2"><div class="cv">%.1f%%</div><div class="cl">Score Qualite</div></div>'
        '<div class="cc c3"><div class="cv">%d</div><div class="cl">Postes</div></div>'
        '<div class="cc c4"><div class="cv">%d</div><div class="cl">Anomalies Total</div></div>'
        '<div class="cc c5"><div class="cv">%d</div><div class="cl">Anom. Performance</div></div>'
        '<div class="cc c6"><div class="cv">%d</div><div class="cl">Anom. Qualite</div></div>'
        '<div class="cc c7"><div class="cv">%d</div><div class="cl">KPIs Suivis</div></div>'
        '<div class="cc c8"><div class="cv">%.1f%%</div><div class="cl">Conformite</div></div>'
        '</div>' % (perf_score, qual_score, len(posts), nb_ano_total, nb_ano_perf, nb_ano_qual, len(ALL_KPI), pct_conformite), unsafe_allow_html=True)

    # ===== ONGLETS =====
    tab1, tab2, tab3, tab4, tab5, tab6, tab7 = st.tabs([
        "📊 Performance", "🎯 Qualite", "🔴 Anomalies", 
        "📋 Plan d'action", "📈 Tendances", "🔍 Synthese", "🗂️ Detail"
    ])

    # ===== TAB 1 : PERFORMANCE =====
    with tab1:
        st.markdown('<div class="stl">Indicateurs de Performance par Poste</div>', unsafe_allow_html=True)
        display_qk = QK if not kpis_for_resp else [k for k in QK if k in kpis_for_resp]
        st.markdown(html_kpi_table(ckdf, display_qk, posts, "pt", show_score=True), unsafe_allow_html=True)
        
        st.markdown('<div class="stl">Distribution des Statuts OT</div>', unsafe_allow_html=True)
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("**OT Correctifs (sans appel)**")
            corr_piv = build_statut_pivot(df[(df["Nº appel pl.entret."].fillna(0)==0)&(df["Contient SOPL"]==1)], posts)
            st.markdown(html_statut_pivot(corr_piv, "pt"), unsafe_allow_html=True)
        with col2:
            show_pie_pair(corr_piv, "Correctifs")
        
        st.markdown('<div class="stl">Age de Preparation (OT CRÉÉ avec CRPR)</div>', unsafe_allow_html=True)
        c1, c2 = st.columns(2)
        with c1:
            pr_piv = pd.pivot_table(df[(df["Statut OT"]=="CRÉÉ")&(df["Statut utilisateur"].str.contains("CRPR",na=False))],
                index="Poste travail princ.", columns="ap", values="Ordre", aggfunc="count", fill_value=0).reindex(posts, fill_value=0)
            for c in ["<1 mois",">3 mois","1 mois < <3 mois","Inconnu"]:
                if c not in pr_piv.columns: pr_piv[c]=0
            st.markdown(html_statut_pivot(pr_piv, "pt"), unsafe_allow_html=True)
        with c2:
            show_simple_pie(pr_piv[["<1 mois","1 mois < <3 mois",">3 mois"]], "Répartition Age Préparation")
        
        st.markdown('<div class="stl">Age de Planification (OT LANC avec ATPL)</div>', unsafe_allow_html=True)
        c1, c2 = st.columns(2)
        with c1:
            pl_piv = pd.pivot_table(df[(df["Statut OT"]=="LANC")&(df["Statut utilisateur"].str.contains("ATPL",case=False,na=False))],
                index="Poste travail princ.", columns="alp", values="Ordre", aggfunc="count", fill_value=0).reindex(posts, fill_value=0)
            for c in ["<1 mois",">3 mois","1 mois < <3 mois","Inconnu"]:
                if c not in pl_piv.columns: pl_piv[c]=0
            st.markdown(html_statut_pivot(pl_piv, "pt"), unsafe_allow_html=True)
        with c2:
            show_simple_pie(pl_piv[["<1 mois","1 mois < <3 mois",">3 mois"]], "Répartition Age Planification")
        
        st.markdown('<div class="stl">Age d\'Execution (OT LANC avec SOPL)</div>', unsafe_allow_html=True)
        c1, c2 = st.columns(2)
        with c1:
            ex_piv = pd.pivot_table(df[(df["Statut OT"]=="LANC")&(df["Contient SOPL"]==1)],
                index="Poste travail princ.", columns="aex", values="Ordre", aggfunc="count", fill_value=0).reindex(posts, fill_value=0)
            for c in ["<1 mois",">3 mois","1 mois < <3 mois","Inconnu"]:
                if c not in ex_piv.columns: ex_piv[c]=0
            st.markdown(html_statut_pivot(ex_piv, "pt"), unsafe_allow_html=True)
        with c2:
            show_simple_pie(ex_piv[["<1 mois","1 mois < <3 mois",">3 mois"]], "Répartition Age Exécution")

    # ===== TAB 2 : QUALITE =====
    with tab2:
        st.markdown('<div class="stl">Indicateurs de Qualite par Poste</div>', unsafe_allow_html=True)
        display_pk = PK if not kpis_for_resp else [k for k in PK if k in kpis_for_resp]
        st.markdown(html_kpi_table(ckdf, display_pk, posts, "qt", show_score=True), unsafe_allow_html=True)
        
        st.markdown('<div class="stl">Backlog Preparation (OT CRÉÉ)</div>', unsafe_allow_html=True)
        c1, c2 = st.columns(2)
        with c1:
            bp_piv = pd.pivot_table(df[df["Statut OT"]=="CRÉÉ"], index="Poste travail princ.", columns="Backlog preparation", values="Ordre", aggfunc="count", fill_value=0).reindex(posts, fill_value=0)
            for c in ["CARACTERISE","NON CARACTERISE"]:
                if c not in bp_piv.columns: bp_piv[c]=0
            st.markdown(html_statut_pivot(bp_piv, "qt"), unsafe_allow_html=True)
        with c2:
            show_simple_pie(bp_piv, "Backlog Préparation")
        
        st.markdown('<div class="stl">Backlog Planification (OT LANC sans SOPL)</div>', unsafe_allow_html=True)
        c1, c2 = st.columns(2)
        with c1:
            bpl_piv = pd.pivot_table(df[(df["Statut OT"]=="LANC")&(df["Contient SOPL"]==0)], index="Poste travail princ.", columns="Backlog planification", values="Ordre", aggfunc="count", fill_value=0).reindex(posts, fill_value=0)
            for c in ["CARACTERISE","NON CARACTERISE"]:
                if c not in bpl_piv.columns: bpl_piv[c]=0
            st.markdown(html_statut_pivot(bpl_piv, "qt"), unsafe_allow_html=True)
        with c2:
            show_simple_pie(bpl_piv, "Backlog Planification")
        
        st.markdown('<div class="stl">Avis sans Ordre (Non approuves)</div>', unsafe_allow_html=True)
        if not avf.empty:
            av_piv = pd.pivot_table(avf, index="Poste travail princ.", columns="Type d'avis", values="Avis", aggfunc="count", fill_value=0).reindex(posts, fill_value=0)
            st.markdown(html_statut_pivot(av_piv, "qt"), unsafe_allow_html=True)
        else:
            st.markdown('<div class="es">Aucun avis sans ordre</div>', unsafe_allow_html=True)

    # ===== TAB 3 : ANOMALIES (NOUVEAU FILTRE) =====
    with tab3:
        st.markdown('<div class="stl">Nombre d\'anomalies par KPI et Poste (à traiter pour atteindre 100%)</div>', unsafe_allow_html=True)
        st.markdown('<p style="font-size:12px;color:#64748b;margin-bottom:10px;padding-left:12px;">🔴 = Anomalie (valeur < cible pour KPIs normaux, valeur > cible pour KPIs inverses) &nbsp;|&nbsp; 🟢 = Conforme &nbsp;|&nbsp; Filtre basé sur les seuils du référentiel CIBLE</p>', unsafe_allow_html=True)
        
        # --- Performance ---
        st.markdown('<div class="stl" style="border-left-color:#059669;">Anomalies Performance (%d sur %d combinaisons)</div>' % (nb_ano_perf, len(display_qk)*len(posts)), unsafe_allow_html=True)
        if display_qk:
            html_perf, tp_perf, _ = html_anomalies_heatmap(anom_perf_matrix[display_qk], display_qk, "Performance", "at")
            st.markdown(html_perf, unsafe_allow_html=True)
            
            # Top KPIs avec le plus d'anomalies
            st.markdown('<div class="stl" style="border-left-color:#f59e0b;font-size:14px;">Classement des KPIs Performance par nombre d\'anomalies</div>', unsafe_allow_html=True)
            sorted_pkpi = sorted(tp_perf.items(), key=lambda x: x[1], reverse=True)
            h_rank = '<table class="tw st"><thead><tr><th>Rang</th><th>KPI</th><th>Nb Anomalies</th><th>Nb Postes</th><th>Taux Anomalie</th><th>Cible</th></tr></thead><tbody>'
            for i, (kpi, nb) in enumerate(sorted_pkpi, 1):
                cible = CIBLE.get(kpi, 100)
                taux = round(nb / len(posts) * 100, 1)
                sens = "↓ Plus bas = mieux" if kpi in LOWER_BETTER else "↑ Plus haut = mieux"
                bg = "background:#fee2e2;color:#991b1b;" if nb > 0 else "background:#d1fae5;color:#065f46;"
                h_rank += '<tr><td style="text-align:center;font-weight:800;">%d</td><td style="font-weight:700;">%s</td><td style="text-align:center;%sfont-weight:800;">%d</td><td style="text-align:center;">%d</td><td style="text-align:center;%sfont-weight:700;">%s%%</td><td style="text-align:center;">%d%% (%s)</td></tr>' % (i, kpi, bg, nb, len(posts), bg, taux, cible, sens)
            h_rank += '</tbody></table>'
            st.markdown(h_rank, unsafe_allow_html=True)
        else:
            st.markdown('<div class="es">Aucun KPI performance à afficher avec ce filtre</div>', unsafe_allow_html=True)
        
        st.markdown('<hr style="margin:20px 0;border-color:#e2e8f0;">', unsafe_allow_html=True)
        
        # --- Qualite ---
        st.markdown('<div class="stl" style="border-left-color:#2563eb;">Anomalies Qualite (%d sur %d combinaisons)</div>' % (nb_ano_qual, len(display_pk)*len(posts)), unsafe_allow_html=True)
        if display_pk:
            html_qual, tp_qual, _ = html_anomalies_heatmap(anom_qual_matrix[display_pk], display_pk, "Qualite", "at")
            st.markdown(html_qual, unsafe_allow_html=True)
            
            st.markdown('<div class="stl" style="border-left-color:#f59e0b;font-size:14px;">Classement des KPIs Qualite par nombre d\'anomalies</div>', unsafe_allow_html=True)
            sorted_qkpi = sorted(tp_qual.items(), key=lambda x: x[1], reverse=True)
            h_rank = '<table class="tw st"><thead><tr><th>Rang</th><th>KPI</th><th>Nb Anomalies</th><th>Nb Postes</th><th>Taux Anomalie</th><th>Cible</th></tr></thead><tbody>'
            for i, (kpi, nb) in enumerate(sorted_qkpi, 1):
                cible = CIBLE.get(kpi, 100)
                taux = round(nb / len(posts) * 100, 1)
                sens = "↓ Plus bas = mieux" if kpi in LOWER_BETTER else "↑ Plus haut = mieux"
                bg = "background:#fee2e2;color:#991b1b;" if nb > 0 else "background:#d1fae5;color:#065f46;"
                h_rank += '<tr><td style="text-align:center;font-weight:800;">%d</td><td style="font-weight:700;">%s</td><td style="text-align:center;%sfont-weight:800;">%d</td><td style="text-align:center;">%d</td><td style="text-align:center;%sfont-weight:700;">%s%%</td><td style="text-align:center;">%d%% (%s)</td></tr>' % (i, kpi, bg, nb, len(posts), bg, taux, cible, sens)
            h_rank += '</tbody></table>'
            st.markdown(h_rank, unsafe_allow_html=True)
        else:
            st.markdown('<div class="es">Aucun KPI qualite à afficher avec ce filtre</div>', unsafe_allow_html=True)
        
        st.markdown('<hr style="margin:20px 0;border-color:#e2e8f0;">', unsafe_allow_html=True)
        
        # --- Anomalies par Poste (classement) ---
        st.markdown('<div class="stl" style="border-left-color:#dc2626;">Classement des Postes par nombre d\'anomalies (tous KPIs confondus)</div>', unsafe_allow_html=True)
        poste_anomalies = []
        for poste in posts:
            nb_p = int(anom_perf_matrix.loc[poste].sum()) if poste in anom_perf_matrix.index else 0
            nb_q = int(anom_qual_matrix.loc[poste].sum()) if poste in anom_qual_matrix.index else 0
            poste_anomalies.append({"Poste": poste, "Anom. Perf.": nb_p, "Anom. Qual.": nb_q, "Total": nb_p + nb_q})
        poste_anomalies.sort(key=lambda x: x["Total"], reverse=True)
        
        h_pr = '<table class="tw at"><thead><tr><th>Rang</th><th>Poste de travail</th><th>Anom. Performance</th><th>Anom. Qualite</th><th>Total Anomalies</th><th>Statut</th></tr></thead><tbody>'
        for i, pa in enumerate(poste_anomalies, 1):
            if pa["Total"] == 0:
                statut = '<span style="background:#d1fae5;color:#065f46;padding:3px 10px;border-radius:12px;font-weight:700;font-size:11px;">CONFORME</span>'
            elif pa["Total"] <= 3:
                statut = '<span style="background:#fef3c7;color:#92400e;padding:3px 10px;border-radius:12px;font-weight:700;font-size:11px;">A AMELIORER</span>'
            else:
                statut = '<span style="background:#fee2e2;color:#991b1b;padding:3px 10px;border-radius:12px;font-weight:700;font-size:11px;">CRITIQUE</span>'
            h_pr += '<tr><td style="text-align:center;font-weight:800;">%d</td><td style="font-weight:700;">%s</td><td style="text-align:center;">%d</td><td style="text-align:center;">%d</td><td style="text-align:center;font-weight:900;color:#dc2626;">%d</td><td style="text-align:center;">%s</td></tr>' % (i, pa["Poste"], pa["Anom. Perf."], pa["Anom. Qual."], pa["Total"], statut)
        h_pr += '</tbody></table>'
        st.markdown(h_pr, unsafe_allow_html=True)

        # --- Détail des anomalies (optionnel) ---
        if show_anomalies_detail:
            st.markdown('<hr style="margin:20px 0;border-color:#e2e8f0;">', unsafe_allow_html=True)
            st.markdown('<div class="stl" style="border-left-color:#7c3aed;">Detail des Anomalies (liste exhaustive avec actions correctives)</div>', unsafe_allow_html=True)
            
            all_anom_rows = html_anomalies_detail(anom_perf_matrix, ckdf, display_qk, posts)
            all_anom_rows += html_anomalies_detail(anom_qual_matrix, ckdf, display_pk, posts)
            all_anom_rows.sort(key=lambda x: x["Ecart (pts)"], reverse=True)
            
            if all_anom_rows:
                h_det = '<table class="tw at"><thead><tr><th>Poste</th><th>KPI</th><th>Valeur</th><th>Cible</th><th>Ecart</th><th>Responsable</th><th>Action corrective</th></tr></thead><tbody>'
                for r in all_anom_rows:
                    ecart_bg = "background:#fee2e2;color:#991b1b;" if r["Ecart (pts)"] > 20 else "background:#fef3c7;color:#92400e;" if r["Ecart (pts)"] > 5 else "background:#dbeafe;color:#1e40af;"
                    h_det += '<tr><td style="font-weight:700;">%s</td><td style="font-weight:600;font-size:11px;">%s</td><td style="text-align:center;font-weight:700;">%.1f%%</td><td style="text-align:center;">%d%%</td><td style="text-align:center;%sfont-weight:800;">%.1f pts</td><td style="text-align:center;font-size:11px;">%s</td><td style="text-align:left;font-size:11px;max-width:300px;">%s</td></tr>' % (
                        r["Poste de travail"], r["KPI"], r["Valeur (%)"], r["Cible (%)"], 
                        ecart_bg, r["Ecart (pts)"], r["Responsable"], r["Action corrective"])
                h_det += '</tbody></table>'
                st.markdown(h_det, unsafe_allow_html=True)
            else:
                st.markdown('<div class="es" style="background:#d1fae5;color:#065f46;border-radius:8px;padding:20px;font-weight:700;">✅ Aucune anomalie detectee - Tous les KPIs atteignent leurs cibles !</div>', unsafe_allow_html=True)

    # ===== TAB 4 : PLAN D'ACTION =====
    with tab4:
        if show_plan:
            st.markdown('<div class="stl">Plan d\'Action - Anomalies a traiter pour atteindre 100%</div>', unsafe_allow_html=True)
            st.markdown('<p style="font-size:12px;color:#64748b;margin-bottom:10px;padding-left:12px;">Ce plan est genere automatiquement a partir du nouveau filtre d\'anomalies (valeur vs cible).</p>', unsafe_allow_html=True)
            
            plan_rows = html_anomalies_detail(anom_perf_matrix, ckdf, display_qk, posts)
            plan_rows += html_anomalies_detail(anom_qual_matrix, ckdf, display_pk, posts)
            plan_rows.sort(key=lambda x: (x["Poste de travail"], x["Ecart (pts)"]), reverse=False)
            
            if plan_rows:
                h_plan = '<table class="plan-action-table"><thead><tr><th>Poste de travail</th><th>KPI en anomalie</th><th>Valeur act.</th><th>Cible</th><th>Ecart</th><th>Responsable</th><th>Action corrective</th><th>Priorite</th></tr></thead><tbody>'
                for i, r in enumerate(plan_rows, 1):
                    if r["Ecart (pts)"] > 30:
                        prio = '<span style="background:#dc2626;color:#fff;padding:2px 8px;border-radius:4px;font-weight:800;font-size:11px;">P1</span>'
                    elif r["Ecart (pts)"] > 10:
                        prio = '<span style="background:#f59e0b;color:#fff;padding:2px 8px;border-radius:4px;font-weight:800;font-size:11px;">P2</span>'
                    else:
                        prio = '<span style="background:#3b82f6;color:#fff;padding:2px 8px;border-radius:4px;font-weight:800;font-size:11px;">P3</span>'
                    h_plan += '<tr><td style="font-weight:800;">%s</td><td style="font-weight:600;font-size:11px;">%s</td><td>%.1f%%</td><td>%d%%</td><td style="font-weight:800;color:#dc2626;">%.1f</td><td style="font-size:11px;">%s</td><td style="text-align:left;font-size:11px;max-width:280px;">%s</td><td>%s</td></tr>' % (
                        r["Poste de travail"], r["KPI"], r["Valeur (%)"], r["Cible (%)"],
                        r["Ecart (pts)"], r["Responsable"], r["Action corrective"], prio)
                h_plan += '</tbody></table>'
                st.markdown(h_plan, unsafe_allow_html=True)
                
                st.markdown('<div style="display:flex;gap:20px;margin-top:10px;font-size:12px;font-weight:700;">'
                    '<span><span style="display:inline-block;width:14px;height:14px;background:#dc2626;border-radius:3px;vertical-align:middle;margin-right:5px;"></span>P1 : Ecart > 30 pts</span>'
                    '<span><span style="display:inline-block;width:14px;height:14px;background:#f59e0b;border-radius:3px;vertical-align:middle;margin-right:5px;"></span>P2 : Ecart > 10 pts</span>'
                    '<span><span style="display:inline-block;width:14px;height:14px;background:#3b82f6;border-radius:3px;vertical-align:middle;margin-right:5px;"></span>P3 : Ecart <= 10 pts</span>'
                    '</div>', unsafe_allow_html=True)
            else:
                st.markdown('<div class="es" style="background:#d1fae5;color:#065f46;border-radius:8px;padding:20px;font-weight:700;">✅ Aucune action requise - Tous les indicateurs sont conformes !</div>', unsafe_allow_html=True)
        else:
            st.markdown('<div class="es">Activer "Afficher Plan d\'action" dans la barre laterale.</div>', unsafe_allow_html=True)

    # ===== TAB 5 : TENDANCES =====
    with tab5:
        st.markdown('<div class="stl">Tendances et Variations</div>', unsafe_allow_html=True)
        hist_path = os.path.join("kpis", "indicateurs_kpis.xlsx")
        hist_df = load_historical_kpis(hist_path)
        if hist_df.empty:
            st.markdown('<div class="es">Aucune donnee historique trouvee. Les tendances apparaitront apres au moins 2 sauvegardes.</div>', unsafe_allow_html=True)
        else:
            var_df = calculate_variations(hist_df)
            if var_df.empty:
                st.markdown('<div class="es">Pas assez de donnees pour calculer les variations.</div>', unsafe_allow_html=True)
            else:
                journal = generate_journal(var_df)
                if journal.empty:
                    st.markdown('<div class="es">Aucune variation significative detectee (ecart < 5%%).</div>', unsafe_allow_html=True)
                else:
                    st.markdown('<p style="font-size:12px;color:#64748b;margin-bottom:8px;padding-left:12px;">Variations significatives (|ecart| >= 5%%) entre les periodes</p>', unsafe_allow_html=True)
                    h_j = '<table class="tw st"><thead><tr><th>Poste</th><th>Type</th><th>KPI</th><th>Avant</th><th>Apres</th><th>Ecart</th><th>Ecart %%</th><th>Sens</th></tr></thead><tbody>'
                    for _, r in journal.iterrows():
                        sens_color = "#059669" if r["Sens"]=="Amelioration" else "#dc2626" if r["Sens"]=="Degradation" else "#64748b"
                        h_j += '<tr><td style="font-weight:700;">%s</td><td>%s</td><td style="font-size:11px;">%s</td><td>%.1f%%</td><td>%.1f%%</td><td style="font-weight:700;">%+.1f</td><td style="font-weight:800;color:%s;">%+.1f%%</td><td style="font-weight:800;color:%s;">%s</td></tr>' % (
                            r["Poste"], r["Type"], r["KPI"], r["Valeur precedente"], r["Valeur actuelle"],
                            r["Ecart"], sens_color, r["Ecart %"], sens_color, r["Sens"])
                    h_j += '</tbody></table>'
                    st.markdown(h_j, unsafe_allow_html=True)
                
                top5, bot5 = calculate_rankings(var_df)
                if not top5.empty and not bot5.empty:
                    c1, c2 = st.columns(2)
                    with c1:
                        st.markdown('<div class="stl" style="border-left-color:#059669;font-size:14px;">Top 5 Amelioration</div>', unsafe_allow_html=True)
                        h_t = '<table class="tw pt"><thead><tr><th>Rang</th><th>Poste</th><th>Score</th></tr></thead><tbody>'
                        for i, (_, r) in enumerate(top5.iterrows(), 1):
                            h_t += '<tr><td style="text-align:center;font-weight:800;">%d</td><td style="font-weight:700;">%s</td><td style="text-align:center;color:#059669;font-weight:800;">%+.1f</td></tr>' % (i, r["Poste"], r["Score variation"])
                        h_t += '</tbody></table>'
                        st.markdown(h_t, unsafe_allow_html=True)
                    with c2:
                        st.markdown('<div class="stl" style="border-left-color:#dc2626;font-size:14px;">Top 5 Degradation</div>', unsafe_allow_html=True)
                        h_b = '<table class="tw at"><thead><tr><th>Rang</th><th>Poste</th><th>Score</th></tr></thead><tbody>'
                        for i, (_, r) in enumerate(bot5.iterrows(), 1):
                            h_b += '<tr><td style="text-align:center;font-weight:800;">%d</td><td style="font-weight:700;">%s</td><td style="text-align:center;color:#dc2626;font-weight:800;">%+.1f</td></tr>' % (i, r["Poste"], r["Score variation"])
                        h_b += '</tbody></table>'
                        st.markdown(h_b, unsafe_allow_html=True)

    # ===== TAB 6 : SYNTHESE =====
    with tab6:
        st.markdown('<div class="stl">Synthese Globale</div>', unsafe_allow_html=True)
        c1, c2 = st.columns(2)
        with c1:
            st.markdown('<div class="ca"><div class="ct">Performance par Poste</div>', unsafe_allow_html=True)
            for poste in posts:
                vals = []
                for kpi in QK:
                    if kpi in ckdf.columns and poste in ckdf.index:
                        v = ckdf.loc[poste, kpi]
                        target = CIBLE.get(kpi, 100)
                        if kpi in LOWER_BETTER:
                            vals.append(max(0, min(100, (target - v) / target * 100)) if target > 0 else 100)
                        else:
                            vals.append(min(100, v / target * 100) if target > 0 else 100)
                sc = np.mean(vals) if vals else 0
                color = "#059669" if sc >= 90 else "#f59e0b" if sc >= 75 else "#dc2626"
                st.markdown('<div class="gbr"><div class="gbr-l">%s</div><div class="gbr-g"><div class="gbr-w"><div class="gbr-f" style="width:%.1f%%;background:%s;"></div></div><div class="gbr-v" style="color:%s;">%.1f%%</div></div></div>' % (poste, sc, color, color, sc), unsafe_allow_html=True)
            st.markdown('</div>', unsafe_allow_html=True)
        with c2:
            st.markdown('<div class="ca"><div class="ct">Qualite par Poste</div>', unsafe_allow_html=True)
            for poste in posts:
                vals = []
                for kpi in PK:
                    if kpi in ckdf.columns and poste in ckdf.index:
                        v = ckdf.loc[poste, kpi]
                        target = CIBLE.get(kpi, 100)
                        if kpi in LOWER_BETTER:
                            vals.append(max(0, min(100, (target - v) / target * 100)) if target > 0 else 100)
                        else:
                            vals.append(min(100, v / target * 100) if target > 0 else 100)
                sc = np.mean(vals) if vals else 0
                color = "#059669" if sc >= 90 else "#f59e0b" if sc >= 75 else "#dc2626"
                st.markdown('<div class="gbr"><div class="gbr-l">%s</div><div class="gbr-g"><div class="gbr-w"><div class="gbr-f" style="width:%.1f%%;background:%s;"></div></div><div class="gbr-v" style="color:%s;">%.1f%%</div></div></div>' % (poste, sc, color, color, sc), unsafe_allow_html=True)
            st.markdown('</div>', unsafe_allow_html=True)
        
        # Graphique radar
        if len(posts) <= 8:
            fig = go.Figure()
            colors_radar = ['#3b82f6','#10b981','#f59e0b','#ef4444','#8b5cf6','#06b6d4','#f97316','#14b8a6']
            for idx, poste in enumerate(posts):
                vals = []
                for kpi in QK[:6]:
                    if kpi in ckdf.columns and poste in ckdf.index:
                        vals.append(float(ckdf.loc[poste, kpi]))
                    else:
                        vals.append(0)
                fig.add_trace(go.Scatterpolar(r=vals, theta=[k.replace("OT ","").replace("préparation ","Prép ").replace("planification ","Plan ").replace("exécution ","Exéc ") for k in QK[:6]], fill='toself', name=poste, line_color=colors_radar[idx % len(colors_radar)]))
            fig.update_layout(polar=dict(radialaxis=dict(visible=True, range=[0, 100])), showlegend=True, height=500, legend=dict(orientation="h", y=-0.1))
            st.plotly_chart(fig, use_container_width=True)

    # ===== TAB 7 : DETAIL =====
    with tab7:
        st.markdown('<div class="stl">Detail Complet par Poste</div>', unsafe_allow_html=True)
        for poste in posts:
            with st.expander("📋 %s" % poste, expanded=False):
                st.markdown('<table class="tw qt"><thead><tr><th>KPI</th><th>Valeur</th><th>Cible</th><th>Ecart</th><th>Statut</th><th>Responsable</th></tr></thead><tbody>', unsafe_allow_html=True)
                for kpi in ALL_KPI:
                    if kpi not in ckdf.columns or poste not in ckdf.index:
                        continue
                    val = float(ckdf.loc[poste, kpi])
                    cible = CIBLE.get(kpi, 100)
                    if kpi in LOWER_BETTER:
                        ecart = val - cible
                    else:
                        ecart = cible - val
                    is_ok = not is_anomaly(kpi, val)
                    resp = KPI_RESP_MAP.get(kpi, "")
                    status = "✅ OK" if is_ok else "❌ Anomalie"
                    bg = "background:#d1fae5;" if is_ok else "background:#fee2e2;"
                    st.markdown('<tr><td style="font-weight:600;font-size:11px;">%s</td><td style="text-align:center;font-weight:700;">%.1f%%</td><td style="text-align:center;">%d%%</td><td style="text-align:center;font-weight:800;%s">%+.1f</td><td style="text-align:center;%sfont-weight:700;">%s</td><td style="text-align:center;font-size:11px;">%s</td></tr>' % (kpi, val, cible, bg, ecart, bg, status, resp), unsafe_allow_html=True)
                st.markdown('</tbody></table>', unsafe_allow_html=True)

    # ===== SAUVEGARDE EXCEL =====
    prows = []
    for poste in posts:
        row = {"Poste de travail": poste}
        for kpi in QK:
            if kpi in ckdf.columns and poste in ckdf.index:
                row[kpi] = round(float(ckdf.loc[poste, kpi]), 1)
            else:
                row[kpi] = 0
        score_vals = []
        for kpi in QK:
            if kpi in row:
                target = CIBLE.get(kpi, 100)
                if kpi in LOWER_BETTER:
                    score_vals.append(max(0, min(100, (target - row[kpi]) / target * 100)) if target > 0 else 100)
                else:
                    score_vals.append(min(100, row[kpi] / target * 100) if target > 0 else 100)
        row["Score Performance"] = round(np.mean(score_vals), 1) if score_vals else 0
        prows.append(row)
    
    qrows = []
    for poste in posts:
        row = {"Poste de travail": poste}
        for kpi in PK:
            if kpi in ckdf.columns and poste in ckdf.index:
                row[kpi] = round(float(ckdf.loc[poste, kpi]), 1)
            else:
                row[kpi] = 0
        score_vals = []
        for kpi in PK:
            if kpi in row:
                target = CIBLE.get(kpi, 100)
                if kpi in LOWER_BETTER:
                    score_vals.append(max(0, min(100, (target - row[kpi]) / target * 100)) if target > 0 else 100)
                else:
                    score_vals.append(min(100, row[kpi] / target * 100) if target > 0 else 100)
        row["Score Qualite"] = round(np.mean(score_vals), 1) if score_vals else 0
        qrows.append(row)
    
    # Nouveau filtre pour les anomalies Excel
    ano_p_rows = html_anomalies_detail(anom_perf_matrix, ckdf, QK, posts)
    ano_q_rows = html_anomalies_detail(anom_qual_matrix, ckdf, PK, posts)
    
    pcols = ["Poste de travail"] + QK + ["Score Performance"]
    qcols = ["Poste de travail"] + PK + ["Score Qualite"]
    ano_cols = ["Poste de travail", "KPI", "Valeur (%)", "Cible (%)", "Ecart (pts)", "Responsable", "Action corrective"]
    
    save_kpis_to_excel(prows, pcols, qrows, qcols, ano_p_rows, ano_cols, ano_q_rows, ano_cols, fichier_date)
    
    st.markdown('<div class="footer">Dashboard KPI Maintenance — Filtre anomalies par cible — Donnees du %s</div>' % fichier_date, unsafe_allow_html=True)

if __name__ == "__main__":
    main()
