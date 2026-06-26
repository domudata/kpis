# -*- coding: utf-8 -*-
import streamlit as st

import streamlit.components.v1 as components
import pandas as pd
import numpy as np
import io, locale, random, time, os, hashlib, json, base64
from datetime import datetime
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openai import OpenAI

client = OpenAI(api_key=st.secrets["OPENAI_API_KEY"])
# ============================================================
st.set_page_config(layout="wide", page_title="Dashboard KPI", initial_sidebar_state="expanded")
# ============================================================

QK = ["TAUX_REALISATION_CORRECTIF/PT","OT préparation <1 mois","OT préparation >3 mois",
      "OT préparation 1mois< <3mois","OT planification <1 mois","OT planification >3 mois",
      "OT planification 1mois< <3mois","OT exécution <1 mois","OT exécution >3 mois",
      "OT exécution 1mois< <3mois",
      "Performance Graissage","Performance Inspection","Performance Appels Systématiques"]
PK = ["Taux d'approbation des Avis","OT LANC ESTIME","Backlog préparation caractérisé",
      "Backlog planification caractérisé","OT CONFIME","OT_COR_EGAL",
      "OT Fiabilité","Total Avis de Panne"]
ALL_KPI = QK + PK

CIBLE = {"TAUX_REALISATION_CORRECTIF/PT":85,"OT préparation <1 mois":80,"OT préparation >3 mois":5,
         "OT préparation 1mois< <3mois":15,"OT planification <1 mois":80,"OT planification >3 mois":5,
         "OT planification 1mois< <3mois":15,"OT exécution <1 mois":80,"OT exécution >3 mois":5,
         "OT exécution 1mois< <3mois":15,"Taux d'approbation des Avis":95,"OT LANC ESTIME":100,
         "Backlog préparation caractérisé":100,"Backlog planification caractérisé":100,
         "OT CONFIME":100,"OT_COR_EGAL":100,
         "Performance Graissage":95,"Performance Inspection":95,"Performance Appels Systématiques":95,
         "OT Fiabilité":100,"Total Avis de Panne":100}

ACT_MAP = {"TAUX_REALISATION_CORRECTIF/PT":"Ameliorer le taux de realisation des OT.",
           "OT préparation <1 mois":"Reduire l'age de preparation des OT (< 1 mois).",
           "OT préparation >3 mois":"Traiter les OT avec preparation > 3 mois.",
           "OT planification <1 mois":"Reduire l'age de planification des OT (< 1 mois).",
           "OT planification >3 mois":"Traiter les OT avec planification > 3 mois.",
           "OT exécution <1 mois":"Reduire l'age d'execution des OT (< 1 mois).",
           "OT exécution >3 mois":"Traiter les OT avec execution > 3 mois.",
           "OT LANC ESTIME":"Estimer les couts des OT lances.",
           "Backlog préparation caractérisé":"Caracteriser le backlog de preparation.",
           "Backlog planification caractérisé":"Caracteriser le backlog de planification.",
           "OT CONFIME":"Confirmer les OT termines.",
           "OT_COR_EGAL":"Rapprocher les couts reels et budgetes.",
           "Taux d'approbation des Avis":"Creer un OT pour les avis sans ordre.",
           "OT préparation 1mois< <3mois":"Reduire les OT entre 1 et 3 mois.",
           "OT planification 1mois< <3mois":"Reduire les OT entre 1 et 3 mois.",
           "OT exécution 1mois< <3mois":"Reduire les OT entre 1 et 3 mois.",
           "Performance Graissage":"Ameliorer le taux de realisation des OT de graissage (Type 350).",
           "Performance Inspection":"Ameliorer le taux de realisation des OT d'inspection (Types 290,300,310).",
           "Performance Appels Systématiques":"Ameliorer le taux de realisation des appels systematiques (Type 360).",
           "OT Fiabilité":"Maintenir la fiabilite des OT a 100%.",
           "Total Avis de Panne":"Maintenir le suivi des avis de panne a 100%."}

KPI_RESP_MAP = {
    "TAUX_REALISATION_CORRECTIF/PT": "Chef d'atelier",
    "OT préparation <1 mois": "Préparateur BM",
    "OT préparation 1mois< <3mois": "Préparateur BM",
    "OT préparation >3 mois": "Préparateur BM",
    "OT planification <1 mois": "Planificateur BM",
    "OT planification 1mois< <3mois": "Planificateur BM",
    "OT planification >3 mois": "Planificateur BM",
    "OT exécution <1 mois": "Chef d'atelier",
    "OT exécution 1mois< <3mois": "Chef d'atelier",
    "OT exécution >3 mois": "Chef d'atelier",
    "Taux d'approbation des Avis": "Chef d'atelier",
    "OT LANC ESTIME": "Fiabilité",
    "Backlog préparation caractérisé": "Préparateur BM",
    "Backlog planification caractérisé": "Planificateur BM",
    "OT CONFIME": "Agent de saisie",
    "OT_COR_EGAL": "Agent de saisie",
    "Performance Graissage": "Chef d'atelier",
    "Performance Inspection": "Chef d'atelier",
    "Performance Appels Systématiques": "Chef d'atelier",
    "OT Fiabilité": "Fiabilité",
    "Total Avis de Panne": "Fiabilité"
}

LOWER_BETTER = ["OT préparation >3 mois","OT planification >3 mois","OT exécution >3 mois",
                "OT préparation 1mois< <3mois","OT planification 1mois< <3mois","OT exécution 1mois< <3mois"]

MP_KW = ["CRPR ATPD","CRPR ATMR","CRPR ATER","CRPR ATRS","CRPR ATMO","ATPD","ATMR","ATER","ATRS","ATMO"]
MPLAN_KW = ["ATPL ATEI","ATPL ATAL","ATPL ATER","ATPL AGAR","ATPL ATHS","ATEI","ATAL","ATAS","AGAR","ATHS"]

CONSIGNES_HSE = [
    "Port obligatoire des EPI avant toute intervention.","Port obligatoire du casque de securite.",
    "Port obligatoire des lunettes de protection.","Port obligatoire des gants adaptes au travail.",
    "Utiliser les protections auditives dans les zones bruyantes.","Verifier l'absence de tension avant toute intervention electrique.",
    "Respecter la procedure de consignation et deconsignation.","Ne jamais intervenir sur un equipement en marche.",
    "Baliser et securiser la zone de travail.","Maintenir le poste de travail propre et ordonne.",
    "Verifier l'etat des outils avant utilisation.","Utiliser uniquement du materiel homologue.",
    "Respecter les permis de travail en vigueur.","Identifier les risques avant de commencer une tache.",
    "Signaler immediatement toute situation dangereuse.","Signaler tout incident ou presque accident.",
    "Ne jamais neutraliser un dispositif de securite.","Verifier les detecteurs de gaz avant utilisation.",
    "Verifier la bonne ventilation des zones de travail.","Respecter les regles des espaces confines.",
    "Controler l'atmosphere avant d'entrer dans un espace confine.","Utiliser les points d'ancrage pour les travaux en hauteur.",
    "Verifier l'etat des echafaudages avant utilisation.","Securiser les outils lors des travaux en hauteur.",
    "Ne pas travailler seul lors des operations a risque.","Controler les elingues avant chaque levage.",
    "Respecter les limites de charge des equipements.","Verifier l'etat des appareils de levage.",
    "Maintenir les voies de circulation degagees.","Respecter la signalisation de securite.",
    "Verifier les extincteurs a proximite du chantier.","Connaitre les issues de secours les plus proches.",
    "Respecter les procedures d'arret d'urgence.","Verifier les flexibles et raccords avant mise en service.",
    "Controler les fuites avant demarrage d'un equipement.","Respecter les distances de securite.",
    "Ne jamais contourner une procedure HSE.","Porter les EPI adaptes au risque identifie.",
    "Prevenir son responsable avant toute intervention particuliere.","Analyser les risques avant chaque demarrage de chantier.",
    "Verifier la stabilite des equipements.","Utiliser les bons outils pour la bonne tache.",
    "Respecter les consignes specifiques du chantier.","Ne jamais prendre de raccourci au detriment de la securite.",
    "Arreter immediatement les travaux en cas de danger.","Proteger l'environnement lors des interventions.",
    "Collecter et trier correctement les dechets.","Eviter toute pollution accidentelle.",
    "Respecter les consignes de stockage des produits dangereux.","Lire les fiches de securite avant manipulation.",
    "Verifier les equipements avant chaque prise de poste.","S'assurer de la disponibilite des moyens de secours.",
    "Communiquer clairement avec l'equipe avant intervention.","Respecter les regles de circulation des engins.",
    "Garder une vigilance permanente sur son environnement.","Prendre le temps d'effectuer le travail en securite.",
    "La securite est l'affaire de tous.","Chaque incident peut etre evite par la prevention.",
    "Aucun travail n'est plus urgent que la securite.","Zero accident commence par un comportement sur."]

def get_logo_base64():
    for path in ["logo.png", "./logo.png", "../logo.png"]:
        if os.path.exists(path):
            try:
                with open(path, "rb") as f:
                    return base64.b64encode(f.read()).decode()
            except Exception:
                pass
    return None

def get_date_from_file():
    if os.path.exists("date.txt"):
        try:
            with open("date.txt","r",encoding="utf-8") as f: return f.read().strip()
        except Exception: pass
    return pd.Timestamp.today().strftime("%d/%m/%Y")

def contient_mot(t,lm):
    t=str(t); return any(m in t for l in lm for m in l.split())
def cat_age(a):
    if pd.isna(a): return "Inconnu"
    if a<=1: return "<1 mois"
    elif a>=3: return ">3 mois"
    return "1 mois < <3 mois"

def excr(df):
    if "Poste travail princ." in df.columns:
        return df[~df["Poste travail princ."].astype(str).str.contains("cresseur",case=False,na=False)].copy()
    return df

@st.cache_data(show_spinner=False)
def read_excel_safe(bytes_data):
    """Lit un fichier Excel en détectant automatiquement le vrai format."""
    bio = io.BytesIO(bytes_data)
    
    # Détection du format via les magic bytes
    header = bytes_data[:8]
    
    if header[:4] in (b'PK\x03\x04', b'PK\x05\x06'):
        # Format ZIP → .xlsx / .xlsm
        for engine in ['openpyxl', 'calamine']:
            try:
                return pd.read_excel(bio, engine=engine)
            except Exception:
                bio.seek(0)
                continue
    
    if header == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1':
        # Format OLE2 → .xls (ancien format binaire)
        for engine in ['xlrd', 'calamine']:
            try:
                return pd.read_excel(bio, engine=engine)
            except Exception:
                bio.seek(0)
                continue
    
    # Dernier recours : essai de tous les moteurs
    for engine in ['openpyxl', 'xlrd', 'calamine']:
        try:
            bio.seek(0)
            return pd.read_excel(bio, engine=engine)
        except Exception:
            continue
    
    raise ValueError(
        "Format de fichier non reconnu. Le fichier n'est ni un .xlsx ni un .xls valide.\n"
        "Vérifiez que le fichier n'est pas corrompu ou protégé par mot de passe."
    )


@st.cache_data(show_spinner=False)
def prepare_data(ot_bytes, av_bytes, date_str):
    raw_ot = read_excel_safe(ot_bytes)       # ← remplacé
    raw_av = read_excel_safe(av_bytes)       # ← remplacé
    raw_ot = excr(raw_ot)
    raw_av = excr(raw_av)
    
    # ... le reste de la fonction reste IDENTIQUE ...
    for c in ["Créé le","Date de début planifiée","Date de clôture","Début réel","Fin réelle"]:
        if c in raw_ot.columns: raw_ot[c]=pd.to_datetime(raw_ot[c],errors="coerce")
    for c in ["Créé le","Début souhaité","Date de la clôture"]:
        if c in raw_av.columns: raw_av[c]=pd.to_datetime(raw_av[c],errors="coerce")
        
   
    now_ts = pd.Timestamp.today()
    df = raw_ot.copy()
    
    df["Backlog preparation"]=np.where(df["Statut utilisateur"].apply(lambda x:contient_mot(x,MP_KW)),"CARACTERISE","NON CARACTERISE")
    df["Backlog planification"]=np.where(df["Statut utilisateur"].apply(lambda x:contient_mot(x,MPLAN_KW)),"CARACTERISE","NON CARACTERISE")
    df["Type Carac Prep"]=df["Statut utilisateur"].apply(lambda x: next((kw.split()[0] for kw in MP_KW if kw in str(x)), "NON CARACTERISE"))
    df["Type Carac Plan"]=df["Statut utilisateur"].apply(lambda x: next((kw.split()[0] for kw in MPLAN_KW if kw in str(x)), "NON CARACTERISE"))
    
    for dc,am,ac in [('Créé le',"amp","ap"),('Date de début planifiée',"amlp","alp"),('Date de début planifiée',"amex","aex")]:
        if dc in df.columns:
            df[am]=((now_ts.year-df[dc].dt.year)*12+(now_ts.month-df[dc].dt.month)).round(2)
            df[ac]=df[am].apply(cat_age)
        else: df[am]=np.nan; df[ac]="Inconnu"
        
    df["OT CONFIME"]=np.where(df["Statut système"].str.contains("CLOT|TCLO",na=False) & df["Statut système"].str.contains("CONF",na=False),"OUI","NON")
    
    df["Contient SOPL"]=df["Statut utilisateur"].str.contains("SOPL",na=False).map({True:1,False:0})
    df["OT LANC ESTIME"]=np.where(df["Total coûts budgétés"].fillna(0)==0,"NON","OUI")
    df["OT_COR_EGAL"]=np.where((df["Total coûts budgétés"].fillna(0)-df["Total coûts réels"].fillna(0))==0,"OUI","NON")
    df["_tw_num"]=pd.to_numeric(df.get("Type de travail",pd.Series(dtype=float)),errors="coerce")
    
    if "Statut système" in df.columns: df["Statut OT"]=df["Statut système"].fillna("").astype(str).str.strip().str.split().str[0]
    
    avf = raw_av[
        (
            raw_av["Ordre"].isna() |
            (raw_av["Ordre"].astype(str).str.strip() == "")
        )
        &
        (raw_av["Type d'avis"].isin(["ZU","Z4","ZR","ZP"]))
    ].copy()
    
    apm = sorted(df[df["Poste travail princ."].astype(str).str.startswith(("SF1","SF2"),na=False)]["Poste travail princ."].dropna().unique().tolist())
    
    return df, avf, apm, now_ts
def save_kpis_to_excel(prows,pcols,qrows,qcols,ano_p_r,ano_p_c,ano_q_r,ano_q_c,sheet_name):
    kpis_dir="kpis"; os.makedirs(kpis_dir,exist_ok=True)
    filepath=os.path.join(kpis_dir,"indicateurs_kpis.xlsx")
    sn=str(sheet_name).replace("/","-").replace("\\","-").replace("*","").replace("?","").replace("[","").replace("]","")[:31]
    hf=Font(bold=True,color="FFFFFF",size=10); hfl=PatternFill(start_color="1E3A5F",end_color="1E3A5F",fill_type="solid")
    tf=Font(bold=True,size=12,color="1E3A5F")
    tb=Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
    try: wb=load_workbook(filepath)
    except Exception: wb=Workbook()
    if "Sheet" in wb.sheetnames: del wb["Sheet"]
    if sn in wb.sheetnames: del wb[sn]
    ws=wb.create_sheet(sn); rn=1
    def ws_sec(title,cols,rows,sr):
        ws.cell(row=sr,column=1,value=title).font=tf; sr+=1
        for j,c in enumerate(cols,1):
            cl=ws.cell(row=sr,column=j,value=c); cl.font=hf; cl.fill=hfl; cl.alignment=Alignment(horizontal='center'); cl.border=tb
        sr+=1
        for r in rows:
            for j,c in enumerate(cols,1):
                cl=ws.cell(row=sr,column=j,value=r.get(c,"")); cl.border=tb; cl.alignment=Alignment(horizontal='center')
            sr+=1
        return sr+1
    rn=ws_sec("INDICATEURS DE PERFORMANCE",pcols,prows,rn)
    if ano_p_c and ano_p_r: rn=ws_sec("ANOMALIES PERFORMANCE",ano_p_c,ano_p_r,rn)
    rn=ws_sec("INDICATEURS DE QUALITE",qcols,qrows,rn)
    if ano_q_c and ano_q_r: rn=ws_sec("ANOMALIES QUALITE",ano_q_c,ano_q_r,rn)
    try: wb.save(filepath)
    except Exception: pass

def load_historical_kpis(filepath):
    if not os.path.exists(filepath): return pd.DataFrame()
    try: wb=load_workbook(filepath,read_only=True,data_only=True)
    except Exception: return pd.DataFrame()
    records=[]; section=None; headers=None
    for sheet_name in wb.sheetnames:
        try:
            ws=wb[sheet_name]; rows_data=list(ws.iter_rows(values_only=True))
            for row in rows_data:
                cell0=str(row[0]).strip() if row[0] else ""
                if "INDICATEURS DE PERFORMANCE" in cell0.upper(): section="perf"; headers=None; continue
                elif "INDICATEURS DE QUALITE" in cell0.upper(): section="qual"; headers=None; continue
                elif "ANOMALIES" in cell0.upper(): section=None; continue
                if section and headers is None and cell0:
                    headers=[str(c).strip() if c else "" for c in row]; continue
                if section and headers and cell0 and cell0 not in ("Cible","Total general",""):
                    entry={"Date":sheet_name}
                    for j,h in enumerate(headers):
                        if j<len(row): entry[h]=row[j]
                    entry["_section"]=section; records.append(entry)
        except Exception: continue
    wb.close()
    if not records: return pd.DataFrame()
    df=pd.DataFrame(records)
    df["Date_parsed"]=pd.to_datetime(df["Date"].str.replace("-","/"),format="%d/%m/%Y",errors="coerce")
    return df.sort_values("Date_parsed").reset_index(drop=True)

def calculate_variations(hist_df):
    if hist_df.empty or "Date" not in hist_df.columns: return pd.DataFrame()
    dates=sorted(hist_df["Date"].unique())
    if len(dates)<2: return pd.DataFrame()
    perf_df=hist_df[hist_df["_section"]=="perf"].copy()
    qual_df=hist_df[hist_df["_section"]=="qual"].copy()
    variations=[]
    for i in range(1,len(dates)):
        prev_date,curr_date=dates[i-1],dates[i]
        prev_perf=perf_df[perf_df["Date"]==prev_date].set_index("Poste de travail") if "Poste de travail" in perf_df.columns else pd.DataFrame()
        curr_perf=perf_df[perf_df["Date"]==curr_date].set_index("Poste de travail") if "Poste de travail" in perf_df.columns else pd.DataFrame()
        prev_qual=qual_df[qual_df["Date"]==prev_date].set_index("Poste de travail") if "Poste de travail" in qual_df.columns else pd.DataFrame()
        curr_qual=qual_df[qual_df["Date"]==curr_date].set_index("Poste de travail") if "Poste de travail" in qual_df.columns else pd.DataFrame()
        for sec_name,prev_d,curr_d,kpi_list in [("Performance",prev_perf,curr_perf,QK+["Score Performance"]),("Qualite",prev_qual,curr_qual,PK+["Score Qualite"])]:
            for poste in set(prev_d.index)&set(curr_d.index):
                for kpi in kpi_list:
                    if kpi not in prev_d.columns or kpi not in curr_d.columns: continue
                    try: pv=float(prev_d.loc[poste,kpi])
                    except Exception: continue
                    try: cv=float(curr_d.loc[poste,kpi])
                    except Exception: continue
                    diff=cv-pv; pct=(diff/pv*100) if pv!=0 else (100 if cv!=0 else 0)
                    if abs(diff)<=0.5: trend="stabilite"
                    elif diff>0.5: trend="hausse"
                    else: trend="baisse"
                    sens = "Stable"
                    if trend != "stabilite":
                        if (trend == "hausse" and kpi not in LOWER_BETTER) or (trend == "baisse" and kpi in LOWER_BETTER):
                            sens = "Amelioration"
                        else:
                            sens = "Degradation"
                    variations.append({"Date precedente":prev_date,"Date actuelle":curr_date,"Poste":poste,
                        "Type":sec_name,"KPI":kpi,"Valeur precedente":round(pv,2),"Valeur actuelle":round(cv,2),
                        "Ecart":round(diff,2),"Ecart %":round(pct,2),"Tendance":trend, "Sens":sens})
    return pd.DataFrame(variations)

def generate_journal(var_df):
    if var_df.empty: return pd.DataFrame()
    j=var_df.copy(); j["Significatif"]=j["Ecart %"].abs()>=5
    j=j[j["Significatif"]].copy()
    return j.sort_values(["Date actuelle","Ecart %"],ascending=[True,False])

def calculate_rankings(var_df):
    if var_df.empty: return pd.DataFrame(),pd.DataFrame()
    scores={}
    for poste in var_df["Poste"].unique():
        pv=var_df[var_df["Poste"]==poste].copy()
        scores[poste]=sum((-r["Ecart %"] if r["KPI"] in LOWER_BETTER else r["Ecart %"]) for _,r in pv.iterrows())
    ranked=sorted(scores.items(),key=lambda x:x[1],reverse=True)
    return pd.DataFrame(ranked[:5],columns=["Poste","Score variation"]),pd.DataFrame(ranked[-5:][::-1],columns=["Poste","Score variation"])

def inject_custom_css():
    st.markdown("""<style>
    section[data-testid="stSidebar"]{width:250px!important}
    .main .block-container{max-width:100%!important;width:100%!important;padding-left:0.5rem!important;padding-right:0.5rem!important}
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700;800&display=swap');
    
    :root{
        --primary:#1e3a5f;
        --primary-light:#2c5282;
        --success:#10b981;
        --success-dark:#059669;
        --warning:#f59e0b;
        --warning-dark:#d97706;
        --danger:#ef4444;
        --danger-dark:#dc2626;
        --info:#3b82f6;
        --border:#e2e8f0;
        --radius:10px;
    }
    
    *{box-sizing:border-box;margin:0;padding:0}
    .stApp{background:#f8fafc;font-family:'Inter',sans-serif}
    .main .block-container{padding-top:.8rem;padding-bottom:.8rem}
    
    .mh{
        background:linear-gradient(135deg,#1e3a5f 0%,#2563eb 100%);
        padding:16px 24px;
        border-radius:12px;
        margin-bottom:12px;
        box-shadow:0 8px 24px rgba(30,58,95,0.15);
        display:flex;
        align-items:center;
        gap:16px;
    }
    .mh h1{color:#fff;font-size:42px;font-weight:800;margin:0;flex:1}
    .mh .logo{height:50px;width:auto;max-width:150px;object-fit:contain;border-radius:6px}
    .mh .db{
        background:rgba(255,255,255,0.2);
        padding:6px 16px;
        border-radius:16px;
        color:#fff;
        font-size:20px;
        font-weight:600;
        border:1px solid rgba(255,255,255,0.3);
        backdrop-filter:blur(10px);
    }
    
    .cr{display:grid;grid-template-columns:repeat(4,1fr);gap:8px;margin-bottom:10px}
    .cc{
        background:#fff;
        border-radius:12px;
        padding:18px 16px;
        box-shadow:0 4px 12px rgba(0,0,0,0.06);
        border-left:4px solid;
        transition:transform 0.2s,box-shadow 0.2s;
        text-align:center;
    }
    .cc:hover{transform:translateY(-2px);box-shadow:0 8px 20px rgba(0,0,0,0.1);}
    .cc .cv{font-size:32px;font-weight:900;line-height:1.1;}
    .cc .cl{font-size:16px;color:#1e293b;font-weight:800;text-transform:uppercase;letter-spacing:.5px;margin-top:8px;}
    .cc.c1{border-left-color:#3b82f6}.cc.c1 .cv{color:#2563eb}
    .cc.c2{border-left-color:#10b981}.cc.c2 .cv{color:#059669}
    .cc.c3{border-left-color:#8b5cf6}.cc.c3 .cv{color:#7c3aed}
    .cc.c4{border-left-color:#ef4444}.cc.c4 .cv{color:#dc2626}
    .cc.c5{border-left-color:#3b82f6}.cc.c5 .cv{color:#2563eb}
    .cc.c6{border-left-color:#06b6d4}.cc.c6 .cv{color:#0891b2}
    .cc.c7{border-left-color:#f59e0b}.cc.c7 .cv{color:#d97706}
    .cc.c8{border-left-color:#f97316}.cc.c8 .cv{color:#ea580c}

    .cc .cv-var{font-size:13px;font-weight:800;margin-top:6px;line-height:1.2;min-height:16px;letter-spacing:.3px}
    .cc .cv-var.positive{color:#10b981}
    .cc .cv-var.negative{color:#ef4444}
    .cc .cv-var.neutral{color:#eab308}
    
    .stl{font-size:16px;font-weight:800;color:var(--primary);margin:10px 0 5px 0;padding-left:12px;border-left:4px solid var(--info);}
    
    .tw{width:100%;border-collapse:collapse;font-family:'Inter',sans-serif;font-size:13px;display:block;overflow-x:auto;-webkit-overflow-scrolling:touch;margin:0}
    .tw thead th{background:var(--primary);color:#fff;font-weight:700;font-size:11px;text-transform:uppercase;letter-spacing:.3px;padding:6px 8px;border:none;white-space:nowrap;position:sticky;top:0;z-index:10}
    .tw.qt thead th{background:linear-gradient(135deg,#2563eb,#3b82f6)}
    .tw.pt thead th{background:linear-gradient(135deg,#059669,#10b981)}
    .tw.at thead th{background:linear-gradient(135deg,#dc2626,#ef4444)}
    .tw.st thead th{background:linear-gradient(135deg,#d97706,#f59e0b)}
    .tw thead th:first-child{z-index:11;left:0}
    .tw tbody td:first-child{position:sticky;left:0;background:#fff;z-index:5;border-right:1px solid var(--border);color:#1e293b !important}
    .tw tbody tr:nth-child(even) td:first-child{background:#f8fafc}
    .tw tbody tr:hover td:first-child{background:#eff6ff}
    .tw tbody td{padding:5px 8px;border-bottom:1px solid var(--border);white-space:nowrap;color:#1e293b !important}
    .tw tbody tr:nth-child(even) td{background:#f8fafc}
    .tw tbody tr:hover td{background:#eff6ff!important}
    .cb td{background:#2563eb!important;color:#fff!important;font-weight:700!important;font-size:12px!important}
    
    /* Styles pour le tableau Plan d'action avec bordures bien visibles */
    .plan-action-table { width:100%; border-collapse:collapse; font-family:Inter,sans-serif; font-size:12px; border:1px solid #cbd5e1; }
    .plan-action-table th { background:#1e3a5f; color:#fff; font-weight:700; padding:8px 6px; border:1px solid #1e3a5f; }
    .plan-action-table td { padding:6px 8px; border:1px solid #cbd5e1; text-align:center; vertical-align:middle;}
    .plan-action-table td:first-child { text-align:left; font-weight:800; }
    
    .stTabs [data-baseweb="tab-list"]{gap:6px;background:#e2e8f0;padding:6px;border-radius:8px;margin-bottom:8px}
    .stTabs [data-baseweb="tab"]{
        border-radius:6px;
        padding:12px 22px;
        font-weight:700;
        font-size:20px;
        line-height:1.5;
        min-height:48px;
    }
    .stTabs [data-baseweb="tab"] span,
    .stTabs [data-baseweb="tab"] > div{
        font-size:22px !important;
    }
    .stTabs [aria-selected="true"]{
        background:#fff!important;
        color:var(--primary)!important;
        box-shadow:0 3px 8px rgba(0,0,0,.1);
        font-size:21px;
    }
    .stTabs [data-baseweb="tab"] svg{width:22px;height:22px}
    
    .ca{background:#fff;border-radius:var(--radius);padding:12px;margin-top:6px;border:1px solid var(--border);box-shadow:0 1px 4px rgba(0,0,0,.02)}
    .ca .ct{font-size:14px;font-weight:700;margin-bottom:8px;padding-bottom:5px;border-bottom:1px solid var(--border)}
    .car{display:flex;align-items:center;margin-bottom:6px;font-size:12px}
    .car:last-child{margin-bottom:0}
    .car .cal{width:260px;font-weight:600;color:var(--primary);text-align:right;padding-right:8px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .car .cab{flex:1;height:26px;background:#edf2f7;border-radius:4px;overflow:visible;position:relative}
    .car .caf{height:100%;border-radius:4px;transition:width .3s}
    
    .car .target-mark{
        position:absolute;
        top:-4px;
        bottom:-4px;
        width:3px;
        background:var(--info) !important; 
        z-index:20;
        transform:translateX(-50%);
        box-shadow:0 0 6px rgba(59,130,246,.9),0 0 2px rgba(0,0,0,.4);
        border-radius:2px;
    }
    .car .cav-out{font-size:12px;font-weight:800;color:#1e293b;min-width:55px;text-align:right;padding-left:6px}
    .car .cav-tgt{font-size:10px;font-weight:700;color:#1e293b;min-width:42px;text-align:right;padding-left:4px;opacity:.7}
    .gbr{display:flex;align-items:center;padding:3px 0;font-size:12px;border-bottom:1px solid #f1f5f9}
    .gbr:last-child{border:none}
    .gbr-l{width:160px;font-weight:600;color:#1e293b;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;font-size:11px;padding-right:10px;}
    .gbr-g{display:flex;align-items:center;gap:4px;flex:1;position:relative}
    .gbr-target{position:absolute;left:90%;top:-4px;bottom:-4px;width:3px;background:var(--primary) !important;z-index:10;box-shadow:0 0 6px rgba(30,58,95,.8);border-radius:2px}
    .gbr-target-label{position:absolute;left:90%;top:-20px;transform:translateX(-50%);font-size:9px;font-weight:800;color:#fff;background:var(--primary) !important;padding:1px 5px;border-radius:3px;white-space:nowrap;z-index:11;box-shadow:0 1px 3px rgba(0,0,0,.2)}
    .gbr-w{flex:1;height:22px;background:#f1f5f9;border-radius:3px;overflow:hidden}
    .gbr-f{height:100%;border-radius:3px}
    .gbr-v{font-size:11px;font-weight:800;min-width:48px;text-align:right;color:#1e293b}
    .gbr-legend{display:flex;gap:14px;margin-bottom:10px;font-size:12px;font-weight:700;align-items:center}
    .gbr-legend span{display:flex;align-items:center;gap:5px}
    .gbr-legend i{display:inline-block;width:14px;height:14px;border-radius:2px}
    .gbr-legend .target-icon{display:inline-block;width:3px;height:14px;background:var(--primary) !important;border-radius:1px;box-shadow:0 0 3px rgba(30,58,95,.6)}
    .cg{display:grid;grid-template-columns:1fr 1fr;gap:6px}
    .cg>div{background:#fff;border-radius:var(--radius);padding:8px 10px;border:1px solid var(--border)}
    .cg .ct{font-size:13px;font-weight:700;margin-bottom:4px;padding-bottom:3px;border-bottom:1px solid var(--border)}
    .cgr{display:flex;align-items:center;padding:3px 0;font-size:12px;border-bottom:1px solid #f1f5f9}
    .cgr:last-child{border:none}
    .cgr .rk{width:18px;font-weight:800;text-align:center}
    .cgr .pn{flex:1;font-weight:600;color:#1e293b;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .cgr .ps{font-weight:800;min-width:55px;text-align:right}
    .dgrid{display:grid;grid-template-columns:1fr 1fr;gap:6px}
    .stButton>button[kind="primary"]{background:linear-gradient(135deg,var(--primary),var(--primary-light));border:none;border-radius:6px;padding:8px 14px;font-weight:700;font-size:15px;width:100%}
    ::-webkit-scrollbar{width:5px;height:5px}::-webkit-scrollbar-track{background:#f1f1f1}::-webkit-scrollbar-thumb{background:#cbd5e0;border-radius:3px}
    
    div[data-testid="stSidebar"]{
        background:linear-gradient(180deg,#1e40af 0%,#1e3a8a 50%,#1e3a5f 100%)!important;
    }
    div[data-testid="stSidebar"]*{color:rgba(255,255,255,.9)!important}
    div[data-testid="stSidebar"] .stSelectbox label,div[data-testid="stSidebar"] .stMultiSelect label,div[data-testid="stSidebar"] .stDateInput label,div[data-testid="stSidebar"] .stCheckbox label,div[data-testid="stSidebar"] .stTextInput label{color:rgba(255,255,255,.9)!important;font-weight:600;font-size:13px;text-transform:uppercase;letter-spacing:.5px}
    div[data-testid="stSidebar"] div[data-testid="stWidget"]{background:rgba(255,255,255,.1);border-radius:6px;padding:5px 10px;margin-bottom:5px;border:1px solid rgba(255,255,255,.15)}
    div[data-testid="stSidebar"] .stSelectbox>div>div,div[data-testid="stSidebar"] .stMultiSelect>div>div,div[data-testid="stSidebar"] .stDateInput>div>div,div[data-testid="stSidebar"] .stTextInput>div>div{background:rgba(255,255,255,.95)!important;border-radius:5px}
    
    .es{text-align:center;padding:14px;color:#64748b;font-size:14px}
    .synth-tbl{width:100%;border-collapse:collapse;font-family:'Inter',sans-serif;font-size:12px}
    .synth-tbl thead th{background:var(--primary);color:#fff;font-weight:700;font-size:11px;padding:5px 8px;border:none;white-space:nowrap;position:sticky;top:0}
    .synth-tbl tbody td{padding:4px 8px;border-bottom:1px solid var(--border);text-align:center;color:#1e293b !important}
    .synth-tbl tbody tr:nth-child(even) td{background:#f8fafc}
    .synth-tbl tbody tr:hover td{background:#eff6ff!important}
    .synth-tbl .poste-cell{text-align:left;font-weight:700;white-space:nowrap;min-width:140px;color:#1e293b !important}
    
    [data-testid="stHeaderActionElements"]{display:none !important;}
    [data-testid="stActionButtonContainer"]{display:none !important;}
    
    .footer {
        text-align: center;
        margin-top: 30px;
        padding: 15px;
        color: #64748b;
        font-size: 13px;
        border-top: 1px solid var(--border);
        font-weight: 600;
    }

    div[data-testid="stDataEditor"] table, 
    div[data-testid="stDataEditor"] th, 
    div[data-testid="stDataEditor"] td {
        font-size: 18px !important;
        line-height: 1.4 !important;
        white-space: normal !important;
        word-wrap: break-word !important;
    }
    div[data-testid="stDataEditor"] [data-testid="stMarkdownContainer"] {
        font-size: 18px !important;
    }
    div[data-testid="stDataEditor"] [data-testid="stTable"] {
        overflow-x: hidden !important;
        width: 100% !important;
    }
    
        /* Styles pour les cellules de sparklines */
    .tw td svg { display: block; margin: 0 auto; }
    .spark-cell { text-align: center; vertical-align: middle; padding: 8px 5px !important; }
    @media(max-width:768px){
        .cr{grid-template-columns:repeat(2,1fr)}
        .mh{padding:8px 10px;gap:8px}
        .mh h1{font-size:18px}
        .mh .logo{height:35px;max-width:70px}
        .mh .db{font-size:11px;padding:2px 8px}
        .cg,.dgrid{grid-template-columns:1fr}
        .car{flex-wrap:wrap;gap:2px}
        .car .cal{width:100%;text-align:left;padding-right:0;margin-bottom:2px}
        .car .cab{flex:1 1 70%}
        .car .cav-out,.car .cav-tgt{flex:1 1 15%;min-width:40px}
        .gbr{flex-direction:column;align-items:flex-start;gap:4px}
        .gbr-l{width:100%;margin-bottom:2px}
        .gbr-g{width:100%;flex-wrap:wrap}
        .gbr-w{flex:1 1 45%}
        .gbr-v{flex:1 1 10%;min-width:40px}
        .tw{font-size:10px}
        .tw thead th,.tw tbody td{padding:3px 4px}
        .stl{font-size:13px}
        .stTabs [data-baseweb="tab"]{padding:8px 12px;font-size:15px}
        .stTabs [data-baseweb="tab"] span{font-size:16px !important}
        div[data-testid="stDataEditor"] table, div[data-testid="stDataEditor"] th, div[data-testid="stDataEditor"] td { font-size: 14px !important; }
    }
    
    @media print {
        section[data-testid="stSidebar"], 
        header[data-testid="stHeader"], 
        div[data-testid="stToolbar"], 
        div[data-testid="stHeaderActionElements"],
        footer, 
        .stDeployButton, 
        #MainMenu { display: none !important; }
        .main .block-container { 
            padding-top: 0 !important; 
            padding-left: 0 !important; 
            padding-right: 0 !important; 
            max-width: 100% !important; 
        }
        .stButton, .stDownloadButton { display: none !important; }
        * {
            -webkit-print-color-adjust: exact !important;
            print-color-adjust: exact !important;
        }
        .tw, .synth-tbl { page-break-inside: avoid; overflow: visible !important; }
    }
    </style>""",unsafe_allow_html=True)

# ============================================================
def get_previous_card_values(hist_df):
    """Récupère les valeurs précédentes des 8 KPI du header depuis hist_df.
    La 'valeur précédente' = avant-dernière date enregistrée dans le fichier historique.
    """
    prev = {
        "OT Analysés": None,
        "Score Performance Global": None,
        "Score Qualité Global": None,
        "Anomalies Totales": None,
        "Performance SF1": None,
        "Qualité SF1": None,
        "Performance SF2": None,
        "Qualité SF2": None,
    }
    if hist_df is None or hist_df.empty or "Date_parsed" not in hist_df.columns:
        return prev

    dates_parsed = sorted(hist_df["Date_parsed"].dropna().unique())
    if len(dates_parsed) < 2:
        return prev  # Aucune période précédente → tout restera None → gris 0.0 %

    prev_date = dates_parsed[-2]
    prev_data = hist_df[hist_df["Date_parsed"] == prev_date]
    prev_perf = prev_data[prev_data["_section"] == "perf"]
    prev_qual = prev_data[prev_data["_section"] == "qual"]

    # Score Performance Global (ligne "Total general" de la section Performance)
    if not prev_perf.empty and "Score Performance" in prev_perf.columns:
        tg = prev_perf[prev_perf["Poste de travail"].astype(str) == "Total general"]
        if not tg.empty:
            try:
                prev["Score Performance Global"] = float(tg.iloc[0]["Score Performance"])
            except Exception:
                pass

    # Score Qualité Global (ligne "Total general" de la section Qualité)
    if not prev_qual.empty and "Score Qualite" in prev_qual.columns:
        tg = prev_qual[prev_qual["Poste de travail"].astype(str) == "Total general"]
        if not tg.empty:
            try:
                prev["Score Qualité Global"] = float(tg.iloc[0]["Score Qualite"])
            except Exception:
                pass

    # Moyennes SF1 / SF2 pour Performance et Qualité
    for section_df, score_col, key_prefix in [
        (prev_perf, "Score Performance", "Performance"),
        (prev_qual, "Score Qualite", "Qualité"),
    ]:
        if score_col not in section_df.columns:
            continue
        sf1_vals, sf2_vals = [], []
        for _, row in section_df.iterrows():
            poste = str(row.get("Poste de travail", ""))
            if poste in ("Total general", "CIBLE", "", "nan", "None"):
                continue
            try:
                v = float(row[score_col])
            except Exception:
                continue
            if poste.startswith("SF1"):
                sf1_vals.append(v)
            elif poste.startswith("SF2"):
                sf2_vals.append(v)
        if sf1_vals:
            prev[f"{key_prefix} SF1"] = sum(sf1_vals) / len(sf1_vals)
        if sf2_vals:
            prev[f"{key_prefix} SF2"] = sum(sf2_vals) / len(sf2_vals)

    # OT Analysés et Anomalies Totales ne sont pas stockés dans hist_df
    # → restent None → affichage "➜ 0.0 %"
    return prev


def format_card_variation(current, previous):
    """Génère le HTML de la variation à afficher sous la valeur d'une carte KPI."""
    # Pas de valeur précédente → gris 0.0 %
    if previous is None:
        return '<div class="cv-var neutral">➜ 0.0 %</div>'
    try:
        current = float(current)
        previous = float(previous)
    except (ValueError, TypeError):
        return '<div class="cv-var neutral">➜ 0.0 %</div>'
    if previous == 0:
        return '<div class="cv-var neutral">➜ 0.0 %</div>'

    pct = ((current - previous) / previous) * 100
    if pct > 0.05:
        return '<div class="cv-var positive">▲ +%.1f %%</div>' % pct
    elif pct < -0.05:
        return '<div class="cv-var negative">▼ −%.1f %%</div>' % abs(pct)
    else:
        return '<div class="cv-var neutral">➜ 0.0 %</div>'
def main():
    try: locale.setlocale(locale.LC_ALL,'fr_FR.UTF-8')
    except Exception:
        try: locale.setlocale(locale.LC_ALL,'fr_FR')
        except Exception: pass
    inject_custom_css()
    fichier_date=get_date_from_file()

    if "hse_affiche" not in st.session_state: st.session_state.hse_affiche=False
    if not st.session_state.hse_affiche:
        c=random.choice(CONSIGNES_HSE)
        st.markdown("""<div style="min-height:100vh;display:flex;flex-direction:column;align-items:center;justify-content:center;background:linear-gradient(135deg,#1a365d,#2d3748,#1a365d);padding:40px">
        <div style="font-size:64px;margin-bottom:20px">🦺</div>
        <h1 style="text-align:center;font-size:46px;color:#fff;font-weight:900;margin:0">HSE - CONSIGNE DE SECURITE</h1>
        <p style="text-align:center;color:rgba(255,255,255,.6);font-size:22px;margin-top:8px;letter-spacing:3px;text-transform:uppercase">Securite - Sante - Environnement</p>
        <div style="background:linear-gradient(135deg,#f6e05e,#ed8936);padding:36px 48px;border-radius:20px;font-size:32px;font-weight:700;text-align:center;margin:40px 0;color:#1a202c;max-width:800px;box-shadow:0 20px 60px rgba(0,0,0,.3)">⚠️ %s</div>
        <h2 style="text-align:center;color:#48bb78;font-size:36px;font-weight:900">Aucun travail n'est plus urgent que la securite</h2>
        <div style="margin-top:40px;width:200px;height:4px;background:rgba(255,255,255,.1);border-radius:2px;overflow:hidden"><div style="width:100%%;height:100%%;background:linear-gradient(90deg,#48bb78,#38a169);border-radius:2px;animation:ld 5.5s ease-in-out forwards"></div></div>
        <style>@keyframes ld{from{width:0}to{width:100%%}}</style></div>"""%c,unsafe_allow_html=True)
        time.sleep(6); st.session_state.hse_affiche=True; st.rerun(); st.stop()

    def ckpi(n,d,sz=100): return np.where(d==0,sz,(n/d)*100)
    
    def cpiv(df,f,c,p):
        return pd.pivot_table(df[f],index="Poste travail princ.",columns=c,values="Ordre",aggfunc="count",fill_value=0).reindex(p,fill_value=0)
        
    def get_text_col(df):
        for c in ["Désignation","Designation","Désignation OT","Texte ordre","Texte","Description","Libellé","Libelle"]:
            if c in df.columns: return c
        for c in df.columns:
            if df[c].dtype=='object' and any(kw in str(c).lower() for kw in ['sign','text','desc','libell']):
                return c
        return None
        
    def build_statut_pivot(df_sub, posts):
        if df_sub.empty:
            return pd.DataFrame(index=posts, columns=["CRÉÉ","LANC","CLOT","TCLO","Total"]).fillna(0).astype(int)
        piv=pd.pivot_table(df_sub, index="Poste travail princ.", columns="Statut OT", values="Ordre", aggfunc="count", fill_value=0)
        for s in ["CRÉÉ","LANC","CLOT","TCLO"]:
            if s not in piv.columns: piv[s]=0
        piv["Total"]=piv[["CRÉÉ","LANC","CLOT","TCLO"]].sum(axis=1)
        return piv.reindex(posts, fill_value=0).fillna(0).astype(int)
        
    def html_statut_pivot(piv_df, table_class):
        cols=["Poste de travail","CRÉÉ","LANC","CLOT","TCLO","Total"]
        statut_colors = {
            "CRÉÉ": "background:#fef3c7;color:#92400e;font-weight:600;",
            "LANC": "background:#dbeafe;color:#1e40af;font-weight:600;",
            "CLOT": "background:#d1fae5;color:#065f46;font-weight:600;",
            "TCLO": "background:#a7f3d0;color:#064e3b;font-weight:600;",
            "Total": "background:#ede9fe;color:#5b21b6;font-weight:700;"
        }
        h='<table class="tw %s"><thead><tr>'%table_class+''.join('<th>%s</th>'%c for c in cols)+'</tr></thead><tbody>'
        for poste,row in piv_df.iterrows():
            h+='<tr><td style="font-weight:700">%s</td>'%poste
            for c in ["CRÉÉ","LANC","CLOT","TCLO"]:
                h+='<td style="text-align:center;%s">%d</td>'%(statut_colors[c], int(row.get(c,0)))
            h+='<td style="text-align:center;%s">%d</td>'%(statut_colors["Total"], int(row.get("Total",0)))
            h+='</tr>'
        h+='<tr class="tr"><td style="font-weight:800">Total</td>'
        for c in ["CRÉÉ","LANC","CLOT","TCLO"]:
            h+='<td style="text-align:center;font-weight:800;%s">%d</td>'%(statut_colors[c], int(piv_df[c].sum()))
        h+='<td style="text-align:center;font-weight:800;%s">%d</td>'%(statut_colors["Total"], int(piv_df["Total"].sum()))
        h+='</tr></tbody></table>'
        return h
        
    def show_pie_pair(piv_df, title_prefix):
        global_counts = piv_df[["CRÉÉ","LANC","CLOT","TCLO"]].sum()
        global_counts = global_counts[global_counts > 0]
        realised = global_counts.get("CLOT", 0) + global_counts.get("TCLO", 0)
        not_realised = global_counts.sum() - realised
        
        if global_counts.empty:
            st.markdown('<div class="es">Aucune donnee</div>', unsafe_allow_html=True)
            return
            
        colors = ["#8b5cf6", "#f59e0b", "#10b981", "#3b82f6"]
        fig = make_subplots(rows=1, cols=2, specs=[[{"type":"domain"},{"type":"domain"}]], 
                            subplot_titles=(f"{title_prefix} — Par Statut OT", f"{title_prefix} — Réalisés vs Non Réalisés"))
        
        fig.add_trace(go.Pie(labels=global_counts.index, values=global_counts.values, hole=0.4, 
                             textinfo='percent+label', 
                             texttemplate='%{label}<br>%{percent:.1%}<br>(%{value})', 
                             textposition='inside',
                             insidetextorientation='radial',
                             textfont=dict(size=14, color='white', family='Inter, sans-serif'),
                             marker=dict(colors=colors, line=dict(color='#FFFFFF', width=3))), 1, 1)
                             
        pie2_data = pd.Series([realised, not_realised], index=["Réalisés (CLOT+TCLO)", "Non Réalisés"])
        
        fig.add_trace(go.Pie(labels=pie2_data.index, values=pie2_data.values, hole=0.5, 
                             textinfo='percent+label', 
                             texttemplate='%{label}<br>%{percent:.1%}<br>(%{value})', 
                             textposition='inside',
                             insidetextorientation='radial',
                             textfont=dict(size=14, color='white', family='Inter, sans-serif'),
                             marker=dict(colors=["#10b981", "#8b5cf6"], line=dict(color='#FFFFFF', width=3))), 1, 2)
                             
        fig.update_layout(margin=dict(t=80, b=20, l=20, r=20), height=450, 
                          legend=dict(orientation="h", yanchor="bottom", y=-0.12, x=0.5, xanchor="center"))
                          
        st.plotly_chart(fig, use_container_width=True)

    def show_simple_pie(piv_df, title, keep_non_carac=False):
        if not keep_non_carac and "NON CARACTERISE" in piv_df.columns:
            piv_df = piv_df.drop(columns=["NON CARACTERISE"])
            
        counts = piv_df.sum()
        counts = counts[counts > 0]
        
        if counts.empty:
            st.markdown('<div class="es">Aucune donnee</div>', unsafe_allow_html=True)
            return
            
        color_map = {"CARACTERISE": "#10b981", "NON CARACTERISE": "#f97316"}
        type_palette = ['#3b82f6', '#10b981', '#f59e0b', '#8b5cf6', '#06b6d4', '#14b8a6', '#6366f1', '#0ea5e9', '#d946ef', '#a855f7']
        
        colors = []
        palette_idx = 0
        for c in counts.index:
            c_str = str(c)
            if c_str in color_map:
                colors.append(color_map[c_str])
            else:
                colors.append(type_palette[palette_idx % len(type_palette)])
                palette_idx += 1
        
        total_sum = counts.sum()
        pull_list = [0.05 if (v/total_sum)*100 < 10 else 0 for v in counts.values]
        
        fig = go.Figure(
            go.Pie(
                labels=counts.index,
                values=counts.values,
                hole=0.4,
                sort=False,
                textinfo="percent",
                textposition="outside",
                pull=pull_list,
                marker=dict(
                    colors=colors,
                    line=dict(color="white", width=2)
                )
            )
        )

        fig.update_traces(
            hovertemplate="<b>%{label}</b><br>Nombre : %{value}<br>Pourcentage : %{percent}<extra></extra>",
            textfont=dict(size=13, family='Inter, sans-serif')
        )

        fig.update_layout(
            title=dict(text=title, x=0.5, xanchor='center', font=dict(size=16)), 
            height=500, 
            showlegend=True,
            legend=dict(orientation="h", yanchor="bottom", y=-0.15, x=0.5, xanchor="center"),
            margin=dict(t=80, b=80, l=40, r=40)
        )
                          
        st.plotly_chart(fig, use_container_width=True)

    def calc_kpis(df_i, av_i, now_ts, posts):
        res={}; df=df_i.copy(); av=av_i.copy()
        res['dfp']=df
        filt_corr=(df["Nº appel pl.entret."].fillna(0)==0)&(df["Contient SOPL"]==1)
        an=cpiv(df,filt_corr,"Statut OT",posts)
        for c in ["CLOT","CRÉÉ","LANC","TCLO"]: an[c]=an.get(c,0)
        an["OT_CLOTURES"]=an["CLOT"]+an["TCLO"]
        an["TOTAL_OT"]=an[["CLOT","CRÉÉ","LANC","TCLO"]].sum(axis=1)
        an["TAUX_REALISATION_CORRECTIF/PT"]=np.where(an["TOTAL_OT"]==0,100.0,ckpi(an["OT_CLOTURES"],an["TOTAL_OT"]))
        
        pr = cpiv(
    df,
    (df["Statut OT"]=="CRÉÉ") &
    (df["Statut utilisateur"].str.contains(r"\bCRPR\b", case=False, na=False)),
    "ap",
    posts
)
        for c in ["<1 mois",">3 mois","1 mois < <3 mois","Inconnu"]: pr[c]=pr.get(c,0)
        pr["Total"]=pr[["<1 mois","1 mois < <3 mois",">3 mois","Inconnu"]].sum(axis=1)
        pr["OT préparation <1 mois"]=ckpi(pr["<1 mois"],pr["Total"]); pr["OT préparation >3 mois"]=ckpi(pr[">3 mois"],pr["Total"],0); pr["OT préparation 1mois< <3mois"]=ckpi(pr["1 mois < <3 mois"],pr["Total"],0)
        
        pl=cpiv(df,(df["Statut OT"]=="LANC")&(df["Statut utilisateur"].str.contains("ATPL",case=False,na=False)),"alp",posts)
        for c in ["<1 mois",">3 mois","1 mois < <3 mois","Inconnu"]: pl[c]=pl.get(c,0)
        pl["Total"]=pl[["<1 mois","1 mois < <3 mois",">3 mois","Inconnu"]].sum(axis=1)
        pl["OT planification <1 mois"]=ckpi(pl["<1 mois"],pl["Total"]); pl["OT planification >3 mois"]=ckpi(pl[">3 mois"],pl["Total"],0); pl["OT planification 1mois< <3mois"]=ckpi(pl["1 mois < <3 mois"],pl["Total"],0)
        
        ex=cpiv(df,(df["Statut OT"]=="LANC")&(df["Contient SOPL"]==1),"aex",posts)
        for c in ["<1 mois",">3 mois","1 mois < <3 mois","Inconnu"]: ex[c]=ex.get(c,0)
        ex["Total"]=ex[["<1 mois","1 mois < <3 mois",">3 mois","Inconnu"]].sum(axis=1)
        ex["OT exécution <1 mois"]=ckpi(ex["<1 mois"],ex["Total"]); ex["OT exécution >3 mois"]=ckpi(ex[">3 mois"],ex["Total"],0); ex["OT exécution 1mois< <3mois"]=ckpi(ex["1 mois < <3 mois"],ex["Total"],0)
        
        la=pd.pivot_table(df[df["Statut OT"]=="LANC"],index="Poste travail princ.",columns="OT LANC ESTIME",values="Ordre",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
        for c in ["OUI","NON"]: la[c]=la.get(c,0)
        la["Total"]=la["OUI"]+la["NON"]; la["OT LANC ESTIME"]=ckpi(la["OUI"],la["Total"])
        
        pc=pd.pivot_table(df[df["Statut OT"]=="CRÉÉ"],index="Poste travail princ.",columns="Backlog preparation",values="Ordre",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
        for c in ["CARACTERISE","NON CARACTERISE"]: pc[c]=pc.get(c,0)
        pc["Total"]=pc["CARACTERISE"]+pc["NON CARACTERISE"]; pc["Backlog préparation caractérisé"]=ckpi(pc["CARACTERISE"],pc["Total"])
        
        plc=pd.pivot_table(df[(df["Statut OT"]=="LANC")&(df["Contient SOPL"]==0)],index="Poste travail princ.",columns="Backlog planification",values="Ordre",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
        for c in ["CARACTERISE","NON CARACTERISE"]: plc[c]=plc.get(c,0)
        plc["Total"]=plc["CARACTERISE"]+plc["NON CARACTERISE"]; plc["Backlog planification caractérisé"]=ckpi(plc["CARACTERISE"],plc["Total"])
        
        for kn,cn in [("OT CONFIME","OT CONFIME"),("OT_COR_EGAL","OT_COR_EGAL")]:
            pv=pd.pivot_table(df[df["Statut OT"].isin(["CLOT","TCLO"])],index="Poste travail princ.",columns="OT_COR_EGAL",values="Ordre",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
            for c in ["OUI","NON"]: pv[c]=pv.get(c,0)
            pv["Total"]=pv["OUI"]+pv["NON"]; pv[cn]=ckpi(pv["OUI"],pv["Total"]); res[kn.lower().replace(" ","_")]=pv
            
        avf=av.copy(); res['avf']=avf
        tca=pd.pivot_table(avf,index="Poste travail princ.",columns="Statut utilisateur",values="Avis",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
        for c in ["APRQ","APRV","APRV AVAU","REJT"]: tca[c]=tca.get(c,0)
        tca["Total"]=tca[["APRQ","APRV","APRV AVAU","REJT"]].sum(axis=1); tca["Taux d'approbation des Avis"] = ckpi(tca["APRV"] ,tca["Total"])
        
        g_num=df[(df["Statut OT"].isin(["CLOT","TCLO"]))&(df["_tw_num"]==350)].groupby("Poste travail princ.")["Ordre"].count()
        g_den=df[(df["Contient SOPL"]==1)&(df["_tw_num"]==350)].groupby("Poste travail princ.")["Ordre"].count()
        g_df=pd.DataFrame({"_n":g_num,"_d":g_den}).reindex(posts,fill_value=0)
        g_df["Performance Graissage"]=np.where(g_df["_d"]==0,100.0,(g_df["_n"]/g_df["_d"])*100)
        
        ins_types=[290,300,310]
        ins_base=(df["_tw_num"].isin(ins_types))&(df["Date de début planifiée"].notna())&(df["Date de début planifiée"]<=now_ts)
        ins_num=df[(df["Statut OT"].isin(["CLOT","TCLO"]))&ins_base].groupby("Poste travail princ.")["Ordre"].count()
        ins_den=df[(df["Contient SOPL"]==1)&ins_base].groupby("Poste travail princ.")["Ordre"].count()
        ins_df=pd.DataFrame({"_n":ins_num,"_d":ins_den}).reindex(posts,fill_value=0)
        ins_df["Performance Inspection"]=np.where(ins_df["_d"]==0,100.0,(ins_df["_n"]/ins_df["_d"])*100)
        
        sys_base=(df["_tw_num"]==360)&(df["Date de début planifiée"].notna())&(df["Date de début planifiée"]<=now_ts)
        sys_num=df[(df["Statut OT"].isin(["CLOT","TCLO"]))&sys_base].groupby("Poste travail princ.")["Ordre"].count()
        sys_den=df[(df["Contient SOPL"]==1)&sys_base].groupby("Poste travail princ.")["Ordre"].count()
        sys_df=pd.DataFrame({"_n":sys_num,"_d":sys_den}).reindex(posts,fill_value=0)
        sys_df["Performance Appels Systématiques"]=np.where(sys_df["_d"]==0,100.0,(sys_df["_n"]/sys_df["_d"])*100)
        
        fiab_s=pd.Series(100.0,index=posts); avpan_s=pd.Series(100.0,index=posts)
        res['ckdf']=pd.DataFrame({
            "TAUX_REALISATION_CORRECTIF/PT":an["TAUX_REALISATION_CORRECTIF/PT"],
            "OT préparation <1 mois":pr["OT préparation <1 mois"],"OT préparation >3 mois":pr["OT préparation >3 mois"],"OT préparation 1mois< <3mois":pr["OT préparation 1mois< <3mois"],
            "OT planification <1 mois":pl["OT planification <1 mois"],"OT planification >3 mois":pl["OT planification >3 mois"],"OT planification 1mois< <3mois":pl["OT planification 1mois< <3mois"],
            "OT exécution <1 mois":ex["OT exécution <1 mois"],"OT exécution >3 mois":ex["OT exécution >3 mois"],"OT exécution 1mois< <3mois":ex["OT exécution 1mois< <3mois"],
            "Performance Graissage":g_df["Performance Graissage"],"Performance Inspection":ins_df["Performance Inspection"],"Performance Appels Systématiques":sys_df["Performance Appels Systématiques"],
            "Taux d'approbation des Avis":tca["Taux d'approbation des Avis"],"OT LANC ESTIME":la["OT LANC ESTIME"],
            "Backlog préparation caractérisé":pc["Backlog préparation caractérisé"],"Backlog planification caractérisé":plc["Backlog planification caractérisé"],
            "OT CONFIME":res['ot_confime']["OT CONFIME"],"OT_COR_EGAL":res['ot_cor_egal']["OT_COR_EGAL"],
            "OT Fiabilité":fiab_s,"Total Avis de Panne":avpan_s
        })
        return res

    def get_bar_color(kpi, val):
        try: v = float(val)
        except: return "#cbd5e0"
        if kpi in ["OT préparation <1 mois","OT planification <1 mois","OT exécution <1 mois"]:
            if v>=80: return "#38a169"
            elif v>=75: return "#f59e0b"
            else: return "#e53e3e"
        if kpi in ["OT préparation 1mois< <3mois","OT planification 1mois< <3mois","OT exécution 1mois< <3mois"]:
            return "#38a169" if v<=15 else "#e53e3e"
        if kpi in ["OT préparation >3 mois","OT planification >3 mois","OT exécution >3 mois"]:
            return "#38a169" if v<=5 else "#e53e3e"
        if kpi=="TAUX_REALISATION_CORRECTIF/PT":
            if v>=85: return "#38a169"
            elif v>=80: return "#f59e0b"
            else: return "#e53e3e"
        if kpi=="Taux d'approbation des Avis":
            if v>=95: return "#38a169"
            elif v>=90: return "#f59e0b"
            else: return "#e53e3e"
        if kpi in ["OT LANC ESTIME","Backlog préparation caractérisé","Backlog planification caractérisé","OT CONFIME","OT_COR_EGAL"]:
            if v>=100: return "#38a169"
            elif v>=95: return "#f59e0b"
            else: return "#e53e3e"
        if kpi in ["Performance Graissage","Performance Inspection","Performance Appels Systématiques"]:
            if v>=95: return "#38a169"
            elif v>90: return "#f59e0b"
            else: return "#e53e3e"
        if kpi in ["OT Fiabilité","Total Avis de Panne"]:
            return "#38a169" if v>=100 else "#f59e0b"
        if v>=90: return "#38a169"
        elif v>=80: return "#f59e0b"
        else: return "#e53e3e"

    def ks(v,c):
        try: val=float(v)
        except Exception: return ""
        if c in ["OT préparation <1 mois","OT planification <1 mois","OT exécution <1 mois"]:
            return "background:#c6efce;color:#006100;font-weight:600" if val>=80 else ("background:#ffeb9c;color:#9c6500;font-weight:600" if val>=75 else "background:#ffc7ce;color:#9c0006;font-weight:600")
        if c in ["OT préparation 1mois< <3mois","OT planification 1mois< <3mois","OT exécution 1mois< <3mois"]:
            return "background:#c6efce;color:#006100;font-weight:600" if val<=15 else "background:#ffc7ce;color:#9c0006;font-weight:600"
        if c in ["OT préparation >3 mois","OT planification >3 mois","OT exécution >3 mois"]:
            return "background:#c6efce;color:#006100;font-weight:600" if val<=5 else "background:#ffc7ce;color:#9c0006;font-weight:600"
        if c=="TAUX_REALISATION_CORRECTIF/PT":
            return "background:#c6efce;color:#006100;font-weight:600" if val>=85 else ("background:#ffeb9c;color:#9c6500;font-weight:600" if val>=80 else "background:#ffc7ce;color:#9c0006;font-weight:600")
        if c=="Taux d'approbation des Avis":
            return "background:#c6efce;color:#006100;font-weight:600" if val>=95 else ("background:#ffeb9c;color:#9c6500;font-weight:600" if val>=90 else "background:#ffc7ce;color:#9c0006;font-weight:600")
        if c in ["OT LANC ESTIME","Backlog préparation caractérisé","Backlog planification caractérisé","OT CONFIME","OT_COR_EGAL"]:
            return "background:#c6efce;color:#006100;font-weight:600" if val>=100 else ("background:#ffeb9c;color:#9c6500;font-weight:600" if val>=95 else "background:#ffc7ce;color:#9c0006;font-weight:600")
        if c in ["Performance Graissage","Performance Inspection","Performance Appels Systématiques"]:
            return "background:#c6efce;color:#006100;font-weight:600" if val>=95 else ("background:#ffeb9c;color:#9c6500;font-weight:600" if val>90 else "background:#ffc7ce;color:#9c0006;font-weight:600")
        if c in ["OT Fiabilité","Total Avis de Panne"]:
            return "background:#c6efce;color:#006100;font-weight:600" if val>=100 else "background:#ffeb9c;color:#9c6500;font-weight:600"
        return ""
        
    def cs(v):
        try: val=float(str(v).replace(' %','').strip())
        except Exception: return ""
        return "background:#c6efce;color:#006100;font-weight:700" if val>=90 else ("background:#ffeb9c;color:#9c6500;font-weight:700" if val>=80 else "background:#ffc7ce;color:#9c0006;font-weight:700")
        
    def kas(v):
        try: val=int(v)
        except Exception: return ""
        if val==0: return "background:#c6efce;color:#006100;font-weight:600"
        if val<=3: return "background:#ffeb9c;color:#9c6500;font-weight:600"
        if val<=10: return "background:#ffc7ce;color:#9c0006;font-weight:600"
        return "background:#ff9999;color:#7f1d1d;font-weight:800"
        
    def gscore(k,a,t):
        if pd.isna(a) or pd.isna(t): return 0
        if k in ["OT préparation <1 mois","OT planification <1 mois","OT exécution <1 mois"]: return 1 if a>=75 else 0
        if k in ["OT préparation 1mois< <3mois","OT planification 1mois< <3mois","OT exécution 1mois< <3mois"]: return 1 if a<=15 else 0
        if k in ["OT préparation >3 mois","OT planification >3 mois","OT exécution >3 mois"]: return 1 if a<=5 else 0
        if k=="TAUX_REALISATION_CORRECTIF/PT": return 1 if a>=80 else 0
        if k=="Taux d'approbation des Avis": return 1 if a>=90 else 0
        if k in ["OT LANC ESTIME","Backlog préparation caractérisé","Backlog planification caractérisé","OT CONFIME","OT_COR_EGAL"]: return 1 if a>=95 else 0
        if k in ["Performance Graissage","Performance Inspection","Performance Appels Systématiques"]: return 1 if a>=95 else 0
        if k in ["OT Fiabilité","Total Avis de Panne"]: return 1 if a>=100 else 0
        return 0
        
    def is_lb(k): return k in LOWER_BETTER

    def html_table(rows,cols,tc,sc_col=None):
        h='<table class="tw %s"><thead><tr>'%tc+''.join('<th>%s</th>'%c for c in cols)+'</tr></thead><tbody>'
        for r in rows:
            is_cible = r.get("_t")=="cible"
            is_total = r.get("_t")=="total"
            rc="cb" if is_cible else ""
            h+='<tr class="%s">'%rc
            for c in cols:
                v=r.get(c,"")
                if is_cible:
                    h+='<td style="background:#1e3a5f;color:#FFFFFF;font-weight:bold;text-align:center;">%s</td>'%v
                elif is_total:
                    s=cs(v) if sc_col and c in sc_col else ks(v,c)
                    total_style = "font-weight:800;font-size:12px;text-align:center;"
                    if s:
                        clean_s = s.replace("font-weight:600","").replace("font-weight:700","")
                        total_style += clean_s
                    h+='<td style="%s">%s</td>'%(total_style,v)
                else:
                    s=cs(v) if sc_col and c in sc_col else ks(v,c)
                    h+='<td style="%s">%s</td>'%(s or "",v)
            h+='</tr>'
        return h+'</tbody></table>'
        
    def html_anomaly_table(rows,cols,tc):
        h='<table class="tw %s"><thead><tr>'%tc+''.join('<th>%s</th>'%c for c in cols)+'</tr></thead><tbody>'
        for r in rows:
            rc="tr" if r.get("Poste de travail")=="Total" else ""
            h+='<tr class="%s">'%rc
            for c in cols:
                v=r.get(c,"")
                if c=="Poste de travail": h+='<td style="font-weight:700">%s</td>'%v
                elif c=="Total Anomalies": h+='<td style="text-align:center;font-weight:800">%s</td>'%v
                else:
                    s=kas(v)
                    h+='<td style="%s;text-align:center">%s</td>'%(s or "",v)
            h+='</tr>'
        return h+'</tbody></table>'

    def html_actions_table(kpi_list,actuals,targets,act_map):
        h='<table class="tw at"><thead><tr><th>KPI</th><th>Valeur Actuelle</th><th>Cible</th><th>Ecart</th><th>Statut</th><th>Action Recommandee</th></tr></thead><tbody>'
        for k in kpi_list:
            av=actuals.get(k,0); tv=targets.get(k,100); diff=av-tv
            met=av<=tv if is_lb(k) else av>=tv
            status="ATTEINT" if met else "NON ATTEINT"
            st_s="background:#c6efce;color:#006100;font-weight:700" if met else "background:#ffc7ce;color:#9c0006;font-weight:700"
            ec_clr="#059669" if met else "#dc2626"
            action="Objectif atteint" if met else act_map.get(k,"")
            h+='<tr><td style="font-weight:600">%s</td><td>%.1f%%</td><td>%.0f%%</td><td style="color:%s;font-weight:700">%+.1f%%</td><td style="%s">%s</td><td style="color:#4a5568">%s</td></tr>'%(k,av,tv,ec_clr,diff,st_s,status,action)
        return h+'</tbody></table>'
        
    def html_plan_actions_table(rows, title, accent_color, anomaly_dfs):
        """Tableau HTML professionnel avec bordures visibles et liens de téléchargement rapides CSV"""
        if not rows:
            return '<div class="ca" style="margin-bottom:10px;"><div class="ct" style="color:%s;border-bottom:2px solid %s;">%s</div><div class="es" style="padding:20px;">✅ Aucune action requise — Tous les KPIs sont conformes !</div></div>' % (accent_color, accent_color, title)

        rows_sorted = sorted(rows, key=lambda x: (x["poste"], -abs(x["ecart"])))

        from itertools import groupby
        grouped = [(k, list(g)) for k, g in groupby(rows_sorted, key=lambda x: x["poste"])]

        h = '<div class="ca" style="margin-bottom:12px;border:1px solid #e2e8f0;border-radius:10px;overflow:hidden;">'
        h += '<div style="background:linear-gradient(135deg,%s,%s);padding:10px 14px;color:#fff;font-size:15px;font-weight:800;display:flex;justify-content:space-between;align-items:center;">' % (accent_color, accent_color)
        h += '<span>%s</span><span style="background:rgba(255,255,255,0.2);padding:3px 12px;border-radius:14px;font-size:13px;">%d action(s)</span></div>' % (title, len(rows))

        h += '<table class="plan-action-table">'
        h += '<thead><tr>'
        headers = ["Poste de travail","KPI","Nécessite Action","Écart","Nb Anomalies","Responsable","Action Recommandée","Délai"]
        for hdr in headers:
            h += '<th>%s</th>' % hdr
        h += '</tr></thead><tbody>'

        row_idx = 0
        for poste, group_rows in grouped:
            rowspan = len(group_rows)
            first = True
            for r in group_rows:
                bg = "#ffffff" if row_idx % 2 == 0 else "#f8fafc"
                h += '<tr style="background:%s;">' % bg

                if first:
                    poste_bg = "#eff6ff" if accent_color == "#3b82f6" else "#f0fdf4"
                    h += '<td rowspan="%d" style="color:%s;background:%s;border-right:3px solid %s;">%s</td>' % (
                        rowspan, accent_color, poste_bg, accent_color, poste)
                    first = False

                h += '<td style="text-align:left;font-weight:600;color:#2d3748;">%s</td>' % r["kpi"]

                if r["needs_action"]:
                    h += '<td><span style="background:#e53e3e;color:#fff;padding:2px 10px;border-radius:12px;font-size:10px;font-weight:700;">OUI</span></td>'
                else:
                    h += '<td><span style="background:#38a169;color:#fff;padding:2px 10px;border-radius:12px;font-size:10px;font-weight:700;">NON</span></td>'

                ecart = r["ecart"]
                lower = r["kpi"] in LOWER_BETTER
                is_bad = (ecart < 0 and not lower) or (ecart > 0 and lower)
                ec_clr = "#dc2626" if is_bad else "#059669"
                h += '<td style="font-weight:800;color:%s;">%+.1f%%</td>' % (ec_clr, ecart)

                nb = r["nb_anom"]
                if nb == 0:
                    h += '<td style="font-weight:800;color:#065f46;">0</td>'
                else:
                    if nb <= 3:  nb_clr = "#92400e"
                    elif nb <= 10: nb_clr = "#991b1b"
                    else:          nb_clr = "#7f1d1d"
                    
                    link_html = str(nb)
                    try:
                        kpi_name = r["kpi"]
                        poste_name = r["poste"]
                        if kpi_name in anomaly_dfs:
                            df_anom = anomaly_dfs[kpi_name]
                            if "Poste travail princ." in df_anom.columns:
                                df_poste = df_anom[df_anom["Poste travail princ."] == poste_name]
                            else:
                                df_poste = df_anom
                                
                            if not df_poste.empty:
                                # Génération CSV très rapide pour le téléchargement
                                csv_data = df_poste.to_csv(index=False, sep=';')
                                b64 = base64.b64encode(csv_data.encode('utf-8')).decode()
                                safe_filename = f"{poste_name}_{kpi_name}".replace("/", "-").replace("\\", "-").replace(" ", "_").replace("<","").replace(">","")[:50]
                                href = f'<a href="data:text/csv;charset=utf-8;base64,{b64}" download="{safe_filename}.csv" style="color:%s;text-decoration:underline;cursor:pointer;font-weight:800;">%d</a>' % (nb_clr, nb)
                                link_html = href
                    except Exception:
                        pass
                        
                    h += '<td>%s</td>' % link_html

                h += '<td style="font-weight:600;color:#4a5568;">%s</td>' % r["responsable"]
                h += '<td style="text-align:left;color:#4a5568;">%s</td>' % r["action"]
                h += '<td style="color:#a0aec0;">—</td>'

                h += '</tr>'
                row_idx += 1

        h += '</tbody></table></div>'
        return h
        
    def html_classement(scores,accent):
        sp=sorted(scores.items(),key=lambda x:x[1],reverse=True)
        met_p=[(p,s) for p,s in sp if s>=80]; not_p=[(p,s) for p,s in sp if s<80]
        t5=met_p[:5]; b5=not_p[-5:] if len(not_p)>5 else not_p
        h='<div class="cg"><div><div class="ct" style="color:#10b981">Top 5 — Objectif Atteint</div>'
        if t5:
            for i,(p,s) in enumerate(t5): h+='<div class="cgr"><span class="rk" style="color:%s">%s</span><span class="pn">%s</span><span class="ps" style="%s">%.2f%%</span></div>'%(accent,i+1,p,cs("%.2f"%s),s)
        else: h+='<div style="padding:6px;font-size:12px;color:#64748b">Aucun poste</div>'
        h+='</div><div><div class="ct" style="color:#f97316">Bottom 5 — Non Atteint</div>'
        if b5:
            for i,(p,s) in enumerate(reversed(b5)): h+='<div class="cgr"><span class="rk" style="color:#f97316">%s</span><span class="pn">%s</span><span class="ps" style="%s">%.2f%%</span></div>'%(len(b5)-i,p,cs("%.2f"%s),s)
        else: h+='<div style="padding:6px;font-size:12px;color:#10b981">Tous atteints</div>'
        h+='</div></div>'; return h
        
    def html_kpi_bars(kpi_list,actuals,targets,title,color_ok,color_fail):
        h='<div class="ca"><div class="ct" style="color:%s">%s</div>'%(color_ok,title)
        h+='<div class="gbr-legend"><span><span style="display:inline-block;width:3px;height:14px;background:#3b82f6;border-radius:1px;box-shadow:0 0 3px rgba(59,130,246,.6);margin-right:5px;vertical-align:middle;"></span> Cible</span></div>'
        for k in kpi_list:
            av=actuals.get(k,0)
            tv=targets.get(k,100)
            met=av<=tv if is_lb(k) else av>=tv
            bw=min(max(av,0),100)
            bg=get_bar_color(k, av)
            tv_pos=min(max(tv,0),100)
            h+=('<div class="car"><div class="cal">%s</div><div class="cab"><div class="caf" style="width:%s%%;background:%s"></div><div class="target-mark" style="position:absolute;top:-5px;bottom:-5px;width:4px;background:#3b82f6;z-index:20;left:%s%%;transform:translateX(-50%%);box-shadow:0 0 6px rgba(59,130,246,1);border-radius:2px;"></div></div><div class="cav-out">%.1f%%</div><div class="cav-tgt">/%.0f%%</div></div>')%(k,bw,bg,tv_pos,av,tv)
        return h+'</div>'
        
    def html_grouped_bars(posts,pscores,qscores,title):
        h='<div class="ca"><div class="ct" style="color:#1e3a5f">%s</div>'%title
        h+='<div style="display:flex;align-items:center;margin-bottom:8px;padding-bottom:5px;border-bottom:1px solid #e2e8f0;">'
        h+='<div class="gbr-l"></div>'
        h+='<div class="gbr-g">'
        h+='<div style="flex:1;text-align:center;font-weight:800;color:#2563eb;font-size:14px;">Performance</div>'
        h+='<div style="min-width:48px;"></div>'
        h+='<div style="flex:1;text-align:center;font-weight:800;color:#059669;font-size:14px;">Qualite</div>'
        h+='<div style="min-width:48px;"></div>'
        h+='</div></div>'
        
        sorted_posts = sorted(posts,key=lambda x:(pscores.get(x,0)+qscores.get(x,0))/2,reverse=True)
        for p in sorted_posts:
            pv,qv=pscores.get(p,0),qscores.get(p,0)
            p_color = get_bar_color(None, pv)
            q_color = get_bar_color(None, qv)
            h+='<div class="gbr"><div class="gbr-l">%s</div><div class="gbr-g"><div class="gbr-w"><div class="gbr-f" style="width:%s%%;background:%s"></div></div><div class="gbr-v">%.1f%%</div><div class="gbr-w"><div class="gbr-f" style="width:%s%%;background:%s"></div></div><div class="gbr-v">%.1f%%</div></div></div>'%(p,min(max(pv,0),100),p_color,pv,min(max(qv,0),100),q_color,qv)
        return h+'</div>'
        
    def html_synthese_table(synth_data,kpi_list,posts):
        h='<table class="synth-tbl"><thead><tr><th style="min-width:160px;text-align:left">Poste de travail</th>'
        for kpi in kpi_list: h+='<th>%s</th>'%kpi
        h+='</tr></thead><tbody>'
        for poste in posts:
            h+='<tr><td class="poste-cell">%s</td>'%poste
            for kpi in kpi_list:
                info=synth_data.get(poste,{}).get(kpi,{})
                diff=info.get("diff","—")
                if diff != "—":
                    try:
                        d = float(diff)
                        if d > 0: clr="#d1fae5"
                        elif d < 0: clr="#fee2e2"
                        else: clr=""
                        h+='<td style="background:%s;text-align:center;font-weight:700">%s</td>'%(clr,diff)
                    except:
                        h+='<td style="text-align:center">—</td>'
                else:
                    h+='<td style="text-align:center">—</td>'
            h+='</tr>'
        h+='</tbody></table>'
        return h
        
    def export_btn(df,filename):
        buf=io.BytesIO(); df.to_excel(buf,index=False,engine='openpyxl'); buf.seek(0)
        st.download_button("📥 Exporter Excel",data=buf,file_name=filename,mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        
    def html_generic_pivot(piv_df, table_class, title):
        piv_df = piv_df.copy()
        piv_df["Total"] = piv_df.sum(axis=1)
        cols = ["Poste de travail"] + [str(c) for c in piv_df.columns]
        h='<div class="ca"><div class="ct" style="color:#1e3a5f">%s</div>'%title
        h+='<table class="tw %s"><thead><tr>'%table_class+''.join('<th>%s</th>'%c for c in cols)+'</tr></thead><tbody>'
        
        def get_style(col_name):
            if col_name == "CARACTERISE": return "background:#d1fae5;color:#065f46;font-weight:600;"
            if col_name == "NON CARACTERISE": return "background:#fee2e2;color:#991b1b;font-weight:600;"
            if col_name == "Total": return "background:#ede9fe;color:#5b21b6;font-weight:700;"
            return "background:#f8fafc;color:#1e293b;font-weight:600;"
            
        for poste,row in piv_df.iterrows():
            h+='<tr><td style="font-weight:700">%s</td>'%poste
            for c in piv_df.columns:
                h+='<td style="text-align:center;%s">%d</td>'%(get_style(c), int(row.get(c,0)))
            h+='</tr>'
        h+='<tr class="tr"><td style="font-weight:800">Total</td>'
        for c in piv_df.columns:
            h+='<td style="text-align:center;font-weight:800;%s">%d</td>'%(get_style(c), int(piv_df[c].sum()))
        h+='</tr></tbody></table></div>'
        return h

    # ===================== LOAD CACHED DATA =====================
    ot_bytes, av_bytes = None, None
    if os.path.exists("ot.xlsx") and os.path.exists("avis.xlsx"):
        with open("ot.xlsx", "rb") as f: ot_bytes = f.read()
        with open("avis.xlsx", "rb") as f: av_bytes = f.read()

    if ot_bytes and av_bytes:
        df_full, av_full, apm, now_ts = prepare_data(ot_bytes, av_bytes, fichier_date)
    else:
        df_full, av_full, apm, now_ts = pd.DataFrame(), pd.DataFrame(), [], pd.Timestamp.now()

    # ===================== SIDEBAR =====================
    with st.sidebar:
        logo_b64 = get_logo_base64()
        if logo_b64:
            st.markdown('<div style="display:flex;justify-content:center;padding:10px 0 15px 0;border-bottom:1px solid rgba(255,255,255,0.1);margin-bottom:10px;"><img src="data:image/png;base64,%s" style="max-width:100%%;height:auto;max-height:200px;object-fit:contain;border-radius:4px;"></div>'%logo_b64,unsafe_allow_html=True)
        else:
            st.markdown("""<div style="padding:10px 0 4px 0"><div style="font-size:22px;margin-bottom:2px">⚙️</div><div style="font-size:14px;font-weight:800;color:white">Filtres & Parametres</div><div style="font-size:11px;color:rgba(255,255,255,.5);text-transform:uppercase;letter-spacing:1px">Configuration</div></div>""",unsafe_allow_html=True)
        
        if st.button("🔄 Rafraîchir le cache", use_container_width=True):
            st.cache_data.clear()
            st.rerun()

        
            
        show_filters=st.checkbox("✅ Afficher les filtres",value=True,key="show_filters")
        if show_filters:
            unf=st.toggle("📁 Charger nouveaux fichiers",value=False,key="tf")
            ot_f=av_f=None
            if unf:
                pwd = st.text_input("Mot de passe administrateur", type="password")
                if pwd == "779900":
                    ot_f = st.file_uploader("Fichier OT", type=["xlsx"], key="uot")
                    av_f = st.file_uploader("Fichier AVIS", type=["xlsx"], key="uav")
                    new_date = st.text_input("Entrez la date (JJ/MM/AAAA)", value=fichier_date)
                    if st.button("💾 Sauvegarder et Appliquer"):
                        try:
                            datetime.strptime(new_date, "%d/%m/%Y")
                            if ot_f is not None:
                                with open("ot.xlsx", "wb") as f: f.write(ot_f.getbuffer())
                            if av_f is not None:
                                with open("avis.xlsx", "wb") as f: f.write(av_f.getbuffer())
                            with open("date.txt", "w", encoding="utf-8") as f: f.write(new_date)
                            st.success("Fichiers et date mis à jour avec succès !")
                            time.sleep(2)
                            st.cache_data.clear()
                            st.rerun()
                        except ValueError:
                            st.error("Format de date invalide. Veuillez utiliser JJ/MM/AAAA.")
                elif pwd != "":
                    st.error("Mot de passe incorrect.")
            else:
                st.markdown("""<div style="background:rgba(255,255,255,.1);padding:6px 10px;border-radius:6px;border:1px solid rgba(255,255,255,.15)"><div style="font-size:11px;color:rgba(255,255,255,.5);text-transform:uppercase;letter-spacing:1px">Donnees</div><div style="font-size:14px;color:white;font-weight:600;margin-top:2px">📅 %s</div></div>"""%fichier_date,unsafe_allow_html=True)
            st.markdown("---"); st.markdown("**🎯 Postes**")
            sp=st.multiselect("Poste",["All"]+apm,["All"],key="sp")
            st.markdown("**🏭 Atelier**")
            sa=st.multiselect("Atelier",["All","Sulfurique (PS)","Phosphorique (PP)","Centrale (CU)","Engrais (TSP/REX)","Feed (MCP/DCP)"],["All"],key="sa")
            st.markdown("**🏢 Division**")
            sd=st.multiselect("Division",["All","SF1","SF2"],["All"],key="sd")
            st.markdown("---"); st.markdown("**📅 Periode**")
            dr=st.date_input("Date debut planifiee",value=(datetime(2025,1,1).date(),datetime.today().date()),format="DD/MM/YYYY",key="dr")
        else:
            unf=False; ot_f=av_f=None; sp=["All"]; sa=["All"]; sd=["All"]
            dr=(datetime(2025,1,1).date(),datetime.today().date())

    # ===================== APPLY FAST FILTERS & CALCULATIONS =====================
    if not df_full.empty:
        try:
            if unf and ot_f is not None and av_f is not None:
                df_full, av_full, apm, now_ts = prepare_data(ot_f.getvalue(), av_f.getvalue(), fichier_date)
                
            if "All" in sp or not sp: sp=apm
            if "All" in sa or not sa: sa=["All"]
            if "All" in sd or not sd: sd=["All"]
            sdt=pd.to_datetime(dr[0]) if len(dr)==2 else pd.to_datetime(datetime(2025,1,1))
            edt=pd.to_datetime(dr[1]) if len(dr)==2 else pd.to_datetime(datetime.today())

            def mf(poste):
                p=str(poste).upper()
                if "All" not in sa:
                    m=False
                    if "Sulfurique (PS)" in sa and "PS" in p: m=True
                    if "Phosphorique (PP)" in sa and "PP" in p: m=True
                    if "Engrais (TSP/REX)" in sa and ("TSP" in p or "REX" in p): m=True
                    if "Feed (MCP/DCP)" in sa and ("MCP" in p or "DCP" in p): m=True
                    if "Centrale (CU)" in sa and "CU" in p: m=True
                    if not m: return False
                if "All" not in sd:
                    m=False
                    if "SF1" in sd and "SF1" in p: m=True
                    if "SF2" in sd and "SF2" in p: m=True
                    if not m: return False
                return True

            vp=[p for p in apm if mf(p) and p in sp]
            
            df = df_full[(df_full["Poste travail princ."].isin(vp)) & (df_full["Date de début planifiée"].between(sdt,edt))].copy()
            avdf = av_full[av_full["Poste travail princ."].isin(vp)].copy()
            if "Créé le" in avdf.columns:
                avdf = avdf[avdf["Créé le"].between(sdt,edt)]
                
            df_dash = df_full[df_full["Poste travail princ."].isin(vp)].copy()
            avdf_dash = av_full[av_full["Poste travail princ."].isin(vp)].copy()
            
            res = calc_kpis(df, avdf, now_ts, vp)
            res_d = calc_kpis(df_dash, avdf_dash, now_ts, vp)

            ckdf=res['ckdf']; dfp=res['dfp']; avf=res['avf']; ckdf_d=res_d['ckdf']
            pa={k:round(ckdf[k].mean(),2) for k in QK}; qa={k:round(ckdf[k].mean(),2) for k in PK}
            pscores={}; qscores={}
            for poste in ckdf.index:
                r=ckdf.loc[poste]
                pscores[poste]=(sum(gscore(k,r[k],CIBLE[k]) for k in QK if k in r.index)/len(QK)*100) if QK else 0
                qscores[poste]=(sum(gscore(k,r[k],CIBLE[k]) for k in PK if k in r.index)/len(PK)*100) if PK else 0

            sf1_posts = [p for p in vp if str(p).startswith("SF1")]
            sf2_posts = [p for p in vp if str(p).startswith("SF2")]
            sf1_p_score = np.mean([pscores[p] for p in sf1_posts]) if sf1_posts else 0
            sf1_q_score = np.mean([qscores[p] for p in sf1_posts]) if sf1_posts else 0
            sf2_p_score = np.mean([pscores[p] for p in sf2_posts]) if sf2_posts else 0
            sf2_q_score = np.mean([qscores[p] for p in sf2_posts]) if sf2_posts else 0

            # ANOMALIES
            ano_map = {}
            ano_map["TAUX_REALISATION_CORRECTIF/PT"] = dfp[(dfp["Nº appel pl.entret."].fillna(0)==0)&(dfp["Contient SOPL"]==1)&(~dfp["Statut OT"].isin(["CLOT","TCLO"]))].groupby("Poste travail princ.")["Ordre"].count()
            prep_filt = (dfp["Statut OT"]=="CRÉÉ")&(dfp["Statut utilisateur"].str.contains("CRPR",na=False))
            ano_map["OT préparation <1 mois"] = dfp[prep_filt & (dfp["ap"]!="<1 mois")].groupby("Poste travail princ.")["Ordre"].count()
            ano_map["OT préparation >3 mois"] = dfp[prep_filt & (dfp["ap"]==">3 mois")].groupby("Poste travail princ.")["Ordre"].count()
            ano_map["OT préparation 1mois< <3mois"] = dfp[prep_filt & (dfp["ap"]=="1 mois < <3 mois")].groupby("Poste travail princ.")["Ordre"].count()
            plan_filt = (dfp["Statut OT"]=="LANC")&(dfp["Statut utilisateur"].str.contains("ATPL",case=False,na=False))
            ano_map["OT planification <1 mois"] = dfp[plan_filt & (dfp["alp"]!="<1 mois")].groupby("Poste travail princ.")["Ordre"].count()
            ano_map["OT planification >3 mois"] = dfp[plan_filt & (dfp["alp"]==">3 mois")].groupby("Poste travail princ.")["Ordre"].count()
            ano_map["OT planification 1mois< <3mois"] = dfp[plan_filt & (dfp["alp"]=="1 mois < <3 mois")].groupby("Poste travail princ.")["Ordre"].count()
            exec_filt = (dfp["Statut OT"]=="LANC")&(dfp["Contient SOPL"]==1)
            ano_map["OT exécution <1 mois"] = dfp[exec_filt & (dfp["aex"]!="<1 mois")].groupby("Poste travail princ.")["Ordre"].count()
            ano_map["OT exécution >3 mois"] = dfp[exec_filt & (dfp["aex"]==">3 mois")].groupby("Poste travail princ.")["Ordre"].count()
            ano_map["OT exécution 1mois< <3mois"] = dfp[exec_filt & (dfp["aex"]=="1 mois < <3 mois")].groupby("Poste travail princ.")["Ordre"].count()
            perf_filt = (dfp["Contient SOPL"]==1)&(~dfp["Statut OT"].isin(["CLOT","TCLO"]))
            ano_map["Performance Graissage"] = dfp[perf_filt & (dfp["_tw_num"]==350)].groupby("Poste travail princ.")["Ordre"].count()
            ano_map["Performance Inspection"] = dfp[perf_filt & (dfp["_tw_num"].isin([290,300,310]))&(dfp["Date de début planifiée"]<=now_ts)].groupby("Poste travail princ.")["Ordre"].count()
            ano_map["Performance Appels Systématiques"] = dfp[perf_filt & (dfp["_tw_num"]==360)&(dfp["Date de début planifiée"]<=now_ts)].groupby("Poste travail princ.")["Ordre"].count()
            avf_tot = avf.groupby("Poste travail princ.")["Avis"].count()
            avf_aprv = avf[avf["Statut utilisateur"].isin(["APRV","APRV AVAU"])].groupby("Poste travail princ.")["Avis"].count()
            ano_map["Taux d'approbation des Avis"] = avf_tot.sub(avf_aprv, fill_value=0)
            ano_map["OT LANC ESTIME"] = dfp[(dfp["Statut OT"]=="LANC")&(dfp["OT LANC ESTIME"]=="NON")].groupby("Poste travail princ.")["Ordre"].count()
            ano_map["Backlog préparation caractérisé"] = dfp[(dfp["Statut OT"]=="CRÉÉ")&(dfp["Backlog preparation"]=="NON CARACTERISE")].groupby("Poste travail princ.")["Ordre"].count()
            ano_map["Backlog planification caractérisé"] = dfp[(dfp["Statut OT"]=="LANC")&(dfp["Backlog planification"]=="NON CARACTERISE")].groupby("Poste travail princ.")["Ordre"].count()
            ano_map["OT CONFIME"] = dfp[dfp["OT CONFIME"]=="NON"].groupby("Poste travail princ.")["Ordre"].count()
            ano_map["OT_COR_EGAL"] = dfp[dfp["OT_COR_EGAL"]=="NON"].groupby("Poste travail princ.")["Ordre"].count()
            
            ano_p_rows = []
            for poste in vp:
                r = {"Poste de travail": poste}
                total = 0
                for kpi in QK:
                    cnt = int(ano_map.get(kpi, pd.Series()).get(poste, 0))
                    r[kpi] = cnt
                    total += cnt
                r["Total Anomalies"] = total
                ano_p_rows.append(r)
            tot_row_p = {"Poste de travail": "Total"}
            tot_tot = 0
            for kpi in QK:
                s = sum([r[kpi] for r in ano_p_rows])
                tot_row_p[kpi] = s
                tot_tot += s
            tot_row_p["Total Anomalies"] = tot_tot
            ano_p_rows.append(tot_row_p)

            ano_q_rows = []
            for poste in vp:
                r = {"Poste de travail": poste}
                total = 0
                for kpi in PK:
                    if kpi in ["OT Fiabilité", "Total Avis de Panne"]:
                        cnt = 0
                    else:
                        cnt = int(ano_map.get(kpi, pd.Series()).get(poste, 0))
                    r[kpi] = cnt
                    total += cnt
                r["Total Anomalies"] = total
                ano_q_rows.append(r)
            tot_row_q = {"Poste de travail": "Total"}
            tot_tot = 0
            for kpi in PK:
                s = sum([r[kpi] for r in ano_q_rows])
                tot_row_q[kpi] = s
                tot_tot += s
            tot_row_q["Total Anomalies"] = tot_tot
            ano_q_rows.append(tot_row_q)

            # ==========================================
            # TABLE ROWS AVEC NOUVELLE METHODE TOTAL
            # ==========================================
            pcols=["Poste de travail"]+QK+["Score Performance"]
            qcols=["Poste de travail"]+PK+["Score Qualite"]
            prows=[]; qrows=[]
            
            for poste in ckdf.index:
                r=ckdf.loc[poste]; 
                prw={"Poste de travail":poste}
                for k in QK: prw[k]="%.1f"%r[k] if k in r.index else "0.0"
                prw["Score Performance"]="%.2f"%pscores.get(poste,0); 
                prows.append(prw)
                
                qrw={"Poste de travail":poste}
                for k in PK: qrw[k]="%.1f"%r[k] if k in r.index else "0.0"
                qrw["Score Qualite"]="%.2f"%qscores.get(poste,0); 
                qrows.append(qrw)
            
            cible_p={"Poste de travail":"CIBLE"}
            for k in QK: cible_p[k]="%.0f"%CIBLE.get(k,100)
            cible_p["Score Performance"]="%.0f"%100; 
            prows.append({"_t":"cible",**cible_p})
            
            cible_q={"Poste de travail":"CIBLE"}
            for k in PK: cible_q[k]="%.0f"%CIBLE.get(k,100)
            cible_q["Score Qualite"]="%.0f"%100; 
            qrows.append({"_t":"cible",**cible_q})
            
            # Ligne TOTAL GENERAL (Méthode : 1 si Vert/Jaune, 0 si Rouge -> Moyenne)
            tot_p={"Poste de travail":"Total general"}
            for k in QK:
                conform_count = 0
                total_count = 0
                for rw in prows:
                    if k in rw and rw.get("_t") not in ("cible", "total"):
                        try:
                            conform_count += gscore(k,100-float(rw[k]) if k in ["OT préparation >3 mois","OT préparation 1mois< <3mois","OT planification >3 mois","OT planification 1mois< <3mois","OT exécution >3 mois","OT exécution 1mois< <3mois"] else float(rw[k]),CIBLE.get(k,100))
                            total_count += 1
                        except:
                            pass
                tot_p[k] = "%.1f" % ((conform_count / total_count) * 100) if total_count > 0 else "0.0"
            tot_p["Score Performance"]="%.2f"%(sum(pscores.values())/len(pscores)) if pscores else "0.00"
            prows.append({"_t":"total",**tot_p})
            
            tot_q={"Poste de travail":"Total general"}
            for k in PK:
                conform_count = 0
                total_count = 0
                for rw in qrows:
                    if k in rw and rw.get("_t") not in ("cible", "total"):
                        try:
                            val = float(rw[k])
                            target = CIBLE.get(k, 100)
                            conform_count += gscore(k, val, target)
                            total_count += 1
                        except:
                            pass
                tot_q[k] = "%.1f" % ((conform_count / total_count) * 100) if total_count > 0 else "0.0"
            tot_q["Score Qualite"]="%.2f"%(sum(qscores.values())/len(qscores)) if qscores else "0.00"
            qrows.append({"_t":"total",**tot_q})

            # BACKLOG PIVOTS
            prep_backlog_df = dfp[dfp["Statut OT"]=="CRÉÉ"].copy()
            plan_backlog_df = dfp[dfp["Statut OT"]=="LANC"].copy()
            
            piv_carac_prep_stat = pd.pivot_table(prep_backlog_df, index="Poste travail princ.", columns="Backlog preparation", values="Ordre", aggfunc="count", fill_value=0).reindex(vp, fill_value=0)
            prep_carac_df = prep_backlog_df[prep_backlog_df["Backlog preparation"]=="CARACTERISE"]
            piv_carac_prep_type = pd.pivot_table(prep_carac_df, index="Poste travail princ.", columns="Type Carac Prep", values="Ordre", aggfunc="count", fill_value=0).reindex(vp, fill_value=0)
            
            piv_carac_plan_stat = pd.pivot_table(plan_backlog_df, index="Poste travail princ.", columns="Backlog planification", values="Ordre", aggfunc="count", fill_value=0).reindex(vp, fill_value=0)
            plan_carac_df = plan_backlog_df[plan_backlog_df["Backlog planification"]=="CARACTERISE"]
            piv_carac_plan_type = pd.pivot_table(plan_carac_df, index="Poste travail princ.", columns="Type Carac Plan", values="Ordre", aggfunc="count", fill_value=0).reindex(vp, fill_value=0)

            text_col=get_text_col(dfp)
            oms_df_sub=dfp[dfp[text_col].astype(str).str.contains("OMS",case=False,na=False)] if text_col else pd.DataFrame()
            thm_df_sub=dfp[dfp[text_col].astype(str).str.contains("THERMO|THERMOGRAPH",case=False,na=False)] if text_col else pd.DataFrame()
            piv_oms=build_statut_pivot(oms_df_sub,vp)
            piv_thm=build_statut_pivot(thm_df_sub,vp)
            piv_all=build_statut_pivot(dfp,vp)

            ano_p_cols = ["Poste de travail"] + QK + ["Total Anomalies"]
            ano_q_cols = ["Poste de travail"] + PK + ["Total Anomalies"]
            save_kpis_to_excel(prows,pcols,qrows,qcols,ano_p_rows,ano_p_cols,ano_q_rows,ano_q_cols,fichier_date)

            # ANOMALIES DETAILED EXPORT (Pour les liens de téléchargement du Plan d'action)
            anomaly_dfs = {}
            anomaly_dfs["TAUX_REALISATION_CORRECTIF/PT"] = dfp[(dfp["Nº appel pl.entret."].fillna(0)==0)&(dfp["Contient SOPL"]==1)&(~dfp["Statut OT"].isin(["CLOT","TCLO"]))].copy()
            anomaly_dfs["OT préparation <1 mois"] = dfp[prep_filt & (dfp["ap"]!="<1 mois")].copy()
            anomaly_dfs["OT préparation >3 mois"] = dfp[prep_filt & (dfp["ap"]==">3 mois")].copy()
            anomaly_dfs["OT préparation 1mois< <3mois"] = dfp[prep_filt & (dfp["ap"]=="1 mois < <3 mois")].copy()
            anomaly_dfs["OT planification <1 mois"] = dfp[plan_filt & (dfp["alp"]!="<1 mois")].copy()
            anomaly_dfs["OT planification >3 mois"] = dfp[plan_filt & (dfp["alp"]==">3 mois")].copy()
            anomaly_dfs["OT planification 1mois< <3mois"] = dfp[plan_filt & (dfp["alp"]=="1 mois < <3 mois")].copy()
            anomaly_dfs["OT exécution <1 mois"] = dfp[exec_filt & (dfp["aex"]!="<1 mois")].copy()
            anomaly_dfs["OT exécution >3 mois"] = dfp[exec_filt & (dfp["aex"]==">3 mois")].copy()
            anomaly_dfs["OT exécution 1mois< <3mois"] = dfp[exec_filt & (dfp["aex"]=="1 mois < <3 mois")].copy()
            anomaly_dfs["Performance Graissage"] = dfp[perf_filt & (dfp["_tw_num"]==350)].copy()
            anomaly_dfs["Performance Inspection"] = dfp[perf_filt & (dfp["_tw_num"].isin([290,300,310]))&(dfp["Date de début planifiée"]<=now_ts)].copy()
            anomaly_dfs["Performance Appels Systématiques"] = dfp[perf_filt & (dfp["_tw_num"]==360)&(dfp["Date de début planifiée"]<=now_ts)].copy()
            anomaly_dfs["Taux d'approbation des Avis"] = avf[~avf["Statut utilisateur"].isin(["APRV","APRV AVAU"])].copy()
            anomaly_dfs["OT LANC ESTIME"] = dfp[(dfp["Statut OT"]=="LANC")&(dfp["OT LANC ESTIME"]=="NON")].copy()
            anomaly_dfs["Backlog préparation caractérisé"] = dfp[(dfp["Statut OT"]=="CRÉÉ")&(dfp["Backlog preparation"]=="NON CARACTERISE")].copy()
            anomaly_dfs["Backlog planification caractérisé"] = dfp[(dfp["Statut OT"]=="LANC")&(dfp["Backlog planification"]=="NON CARACTERISE")].copy()
            anomaly_dfs["OT CONFIME"] = dfp[dfp["OT CONFIME"]=="NON"].copy()
            anomaly_dfs["OT_COR_EGAL"] = dfp[dfp["OT_COR_EGAL"]=="NON"].copy()
            
            hist_filepath=os.path.join("kpis","indicateurs_kpis.xlsx")
            hist_df=load_historical_kpis(hist_filepath)
            var_df=calculate_variations(hist_df)
            journal_df=generate_journal(var_df)
            top5_df,bot5_df=calculate_rankings(var_df)

            # SYNTHESE DATA
            synth_perf={}; synth_qual={}
            if not var_df.empty and "Date precedente" in var_df.columns:
                for poste in vp:
                    synth_perf[poste]={}; synth_qual[poste]={}
                    pv=var_df[var_df["Poste"]==poste]
                    for kpi in QK:
                        kpi_v=pv[pv["KPI"]==kpi]
                        if not kpi_v.empty:
                            last=kpi_v.iloc[-1]
                            synth_perf[poste][kpi]={"diff":"%+.1f"%last["Ecart"]}
                        else:
                            synth_perf[poste][kpi]={"diff":"—"}
                    for kpi in PK:
                        kpi_v=pv[pv["KPI"]==kpi]
                        if not kpi_v.empty:
                            last=kpi_v.iloc[-1]
                            synth_qual[poste][kpi]={"diff":"%+.1f"%last["Ecart"]}
                        else:
                            synth_qual[poste][kpi]={"diff":"—"}


            # ==========================================
            # CONSTRUCTION DONNÉES PLAN D'ACTIONS (1 ligne par Poste + KPI)
            # ==========================================
            plan_actions_rows = []
            for poste in vp:
                if poste not in ckdf.index:
                    continue
                poste_data = ckdf.loc[poste]

                for kpi in ALL_KPI:
                    actual = float(poste_data.get(kpi, 100))
                    target = CIBLE.get(kpi, 100)
                    lower  = is_lb(kpi)

                    if lower:
                        needs_action = actual > target
                    else:
                        needs_action = actual < target
                    ecart = actual - target

                    nb_anom = int(ano_map.get(kpi, pd.Series()).get(poste, 0))

                    if needs_action or nb_anom > 0:
                        plan_actions_rows.append({
                            "poste": poste,
                            "kpi": kpi,
                            "needs_action": needs_action,
                            "ecart": ecart,
                            "nb_anom": nb_anom,
                            "responsable": KPI_RESP_MAP.get(kpi, "Non assigné"),
                            "action": ACT_MAP.get(kpi, ""),
                            "delai": ""
                        })

            sf1_rows = [r for r in plan_actions_rows if str(r["poste"]).startswith("SF1")]
            sf2_rows = [r for r in plan_actions_rows if str(r["poste"]).startswith("SF2")]


            # RENDER
            avg_p_score=sum(pa.values())/len(pa) if pa else 0
            avg_q_score=sum(qa.values())/len(qa) if qa else 0
            total_ano_p=sum([r["Total Anomalies"] for r in ano_p_rows if r.get("Poste de travail")!="Total"])
            total_ano_q=sum([r["Total Anomalies"] for r in ano_q_rows if r.get("Poste de travail")!="Total"])
            total_ot=len(df)
            logo_b64 = get_logo_base64()
            if logo_b64:
               st.markdown(f'<div class="mh"><img src="data:image/png;base64,{logo_b64}" class="logo" alt="Logo"><h1>Tableau de Bord KPIs Performance & Qualite</h1><span class="db">📅 {fichier_date}</span></div>', unsafe_allow_html=True)
            else:
                st.markdown('<div class="mh"><h1>Tableau de Bord KPIs Performance & Qualite</h1><span class="db">📅 18/06/2026</span></div>',unsafe_allow_html=True)
            
            # === Variations automatiques depuis hist_df ===
            prev_values = get_previous_card_values(hist_df)

            var_ot = format_card_variation(total_ot,                  prev_values.get("OT Analysés"))
            var_sp = format_card_variation(avg_p_score,               prev_values.get("Score Performance Global"))
            var_sq = format_card_variation(avg_q_score,               prev_values.get("Score Qualité Global"))
            var_at = format_card_variation(total_ano_p + total_ano_q, prev_values.get("Anomalies Totales"))
            var_p1 = format_card_variation(sf1_p_score,               prev_values.get("Performance SF1"))
            var_q1 = format_card_variation(sf1_q_score,               prev_values.get("Qualité SF1"))
            var_p2 = format_card_variation(sf2_p_score,               prev_values.get("Performance SF2"))
            var_q2 = format_card_variation(sf2_q_score,               prev_values.get("Qualité SF2"))

            st.markdown(
                '<div class="cr">'
                '<div class="cc c1"><div class="cv">%d</div>%s<div class="cl">OT Analyses</div></div>'
                '<div class="cc c2"><div class="cv">%.1f%%</div>%s<div class="cl">Score Performance Global</div></div>'
                '<div class="cc c3"><div class="cv">%.1f%%</div>%s<div class="cl">Score Qualite Global</div></div>'
                '<div class="cc c4"><div class="cv">%d</div>%s<div class="cl">Anomalies Totales</div></div>'
                '</div>' % (total_ot, var_ot,
                            avg_p_score, var_sp,
                            avg_q_score, var_sq,
                            total_ano_p + total_ano_q, var_at),
                unsafe_allow_html=True
            )

            st.markdown(
                '<div class="cr">'
                '<div class="cc c5"><div class="cv">%.1f%%</div>%s<div class="cl">Performance SF1</div></div>'
                '<div class="cc c6"><div class="cv">%.1f%%</div>%s<div class="cl">Qualite SF1</div></div>'
                '<div class="cc c7"><div class="cv">%.1f%%</div>%s<div class="cl">Performance SF2</div></div>'
                '<div class="cc c8"><div class="cv">%.1f%%</div>%s<div class="cl">Qualite SF2</div></div>'
                '</div>' % (sf1_p_score, var_p1,
                            sf1_q_score, var_q1,
                            sf2_p_score, var_p2,
                            sf2_q_score, var_q2),
                unsafe_allow_html=True
            )

            tabs=st.tabs(["🏠 Tableau de Bord","📈 Performance","✅ Qualite","📂 Backlog","📋 Suivi & Evolution","🎯 Plan d'action"])

      
            
            with tabs[0]:
                st.markdown('<div class="stl p">Scores globaux par poste</div>',unsafe_allow_html=True)
                st.markdown(html_grouped_bars(vp,pscores,qscores,"Comparaison Performance / Qualite par poste"),unsafe_allow_html=True)
                col1,col2=st.columns(2)
                with col1:
                    st.markdown('<div class="stl p">Indicateurs de Performance</div>',unsafe_allow_html=True)
                    st.markdown(html_kpi_bars(QK,pa,CIBLE,"Taux moyens — Performance","#10b981","#8b5cf6"),unsafe_allow_html=True)
                with col2:
                    st.markdown('<div class="stl q">Indicateurs de Qualite</div>',unsafe_allow_html=True)
                    st.markdown(html_kpi_bars(PK,qa,CIBLE,"Taux moyens — Qualite","#3b82f6","#8b5cf6"),unsafe_allow_html=True)
                st.markdown('<div class="stl c">Classement Performance</div>',unsafe_allow_html=True)
                st.markdown(html_classement(pscores,"#10b981"),unsafe_allow_html=True)
                st.markdown('<div class="stl c">Classement Qualite</div>',unsafe_allow_html=True)
                st.markdown(html_classement(qscores,"#3b82f6"),unsafe_allow_html=True)

            with tabs[1]:
                st.markdown('<div class="stl p">Detail des indicateurs de Performance</div>',unsafe_allow_html=True)
                st.markdown(html_table(prows,pcols,"pt",["Score Performance"]),unsafe_allow_html=True)
                st.markdown('<div class="stl a">Nombre d\'anomalies par KPI et Poste (à traiter pour atteindre 100%)</div>',unsafe_allow_html=True)
                st.markdown(html_anomaly_table(ano_p_rows,ano_p_cols,"at"),unsafe_allow_html=True)
                st.markdown('<div class="stl a">Actions recommandees — Performance</div>',unsafe_allow_html=True)
                st.markdown(html_actions_table(QK,pa,CIBLE,ACT_MAP),unsafe_allow_html=True)

            with tabs[2]:
                st.markdown('<div class="stl q">Detail des indicateurs de Qualite</div>',unsafe_allow_html=True)
                st.markdown(html_table(qrows,qcols,"qt",["Score Qualite"]),unsafe_allow_html=True)
                st.markdown('<div class="stl a">Nombre d\'anomalies par KPI et Poste (à traiter pour atteindre 100%)</div>',unsafe_allow_html=True)
                st.markdown(html_anomaly_table(ano_q_rows,ano_q_cols,"at"),unsafe_allow_html=True)
                st.markdown('<div class="stl a">Actions recommandees — Qualite</div>',unsafe_allow_html=True)
                st.markdown(html_actions_table(PK,qa,CIBLE,ACT_MAP),unsafe_allow_html=True)

            with tabs[3]:
                st.markdown('<div class="stl c">Caractérisation Backlog Préparation</div>',unsafe_allow_html=True)
                c1, c2 = st.columns([0.5, 0.5], vertical_alignment="center")
                with c1:
                    st.markdown(html_generic_pivot(piv_carac_prep_stat, "omt", "Synthèse Caractérisé / Non Caractérisé"),unsafe_allow_html=True)
                with c2:
                    show_simple_pie(piv_carac_prep_stat, "Répartition Globale Caractérisé / Non Caractérisé", keep_non_carac=True)
                    show_simple_pie(piv_carac_prep_type, "Répartition par Type de Caractérisation", keep_non_carac=False)

                st.markdown('<div class="stl c">Caractérisation Backlog Planification</div>',unsafe_allow_html=True)
                c5, c6 = st.columns([0.5, 0.5], vertical_alignment="center")
                with c5:
                    st.markdown(html_generic_pivot(piv_carac_plan_stat, "omt", "Synthèse Caractérisé / Non Caractérisé"),unsafe_allow_html=True)
                with c6:
                    show_simple_pie(piv_carac_plan_stat, "Répartition Globale Caractérisé / Non Caractérisé", keep_non_carac=True)
                    show_simple_pie(piv_carac_plan_type, "Répartition par Type de Caractérisation", keep_non_carac=False)

                st.markdown('<div class="stl p">Statuts OT par Poste de Travail</div>',unsafe_allow_html=True)
                
                st.markdown('<div class="stl s">OT OMS par Poste et Statut OT</div>',unsafe_allow_html=True)
                c_oms1, c_oms2 = st.columns([0.5, 0.5], vertical_alignment="center")
                with c_oms1: st.markdown(html_statut_pivot(piv_oms,"omt"),unsafe_allow_html=True)
                with c_oms2: show_pie_pair(piv_oms,"OT OMS")
                
                st.markdown('<div class="stl s">OT Thermographie par Poste et Statut OT</div>',unsafe_allow_html=True)
                c_thm1, c_thm2 = st.columns([0.5, 0.5], vertical_alignment="center")
                with c_thm1: st.markdown(html_statut_pivot(piv_thm,"tht"),unsafe_allow_html=True)
                with c_thm2: show_pie_pair(piv_thm,"OT Thermographie")
                
                st.markdown('<div class="stl s">Tous les OT par Poste et Statut OT</div>',unsafe_allow_html=True)
                c_all1, c_all2 = st.columns([0.5, 0.5], vertical_alignment="center")
                with c_all1: st.markdown(html_statut_pivot(piv_all,"pt"),unsafe_allow_html=True)
                with c_all2: show_pie_pair(piv_all,"Tous les OT")

            with tabs[4]:
                  
                min_date = var_df["Date precedente"].min() if not var_df.empty else "?"
                max_date = var_df["Date actuelle"].max() if not var_df.empty else "?"

                # --- 1. Bouton Masquer/Afficher les tableaux de synthèse ---
                if "show_synth" not in st.session_state:
                    st.session_state.show_synth = False

                btn_label = "▼ Masquer les détails" if st.session_state.show_synth else "▶ Voir plus de détails"
                if st.button(btn_label, key="btn_synth"):
                    st.session_state.show_synth = not st.session_state.show_synth
                    st.rerun()

                if st.session_state.show_synth:
                    st.markdown(f'<div class="stl s">Synthèse d\'évolution Performance entre {min_date} et {max_date}</div>',unsafe_allow_html=True)
                    if synth_perf and any(any(v.get("diff","—")!="—" for v in d.values()) for d in synth_perf.values()):
                        st.markdown(html_synthese_table(synth_perf,QK,vp),unsafe_allow_html=True)
                    else:
                        st.markdown('<div class="es">Pas assez de donnees historiques pour calculer la synthese Performance. Au moins 2 periodes sont necessaires.</div>',unsafe_allow_html=True)
                        
                    st.markdown(f'<div class="stl s">Synthèse d\'évolution Qualité entre {min_date} et {max_date}</div>',unsafe_allow_html=True)
                    if synth_qual and any(any(v.get("diff","—")!="—" for v in d.values()) for d in synth_qual.values()):
                        st.markdown(html_synthese_table(synth_qual,PK,vp),unsafe_allow_html=True)
                    else:
                        st.markdown('<div class="es">Pas assez de donnees historiques pour calculer la synthese Qualite. Au moins 2 periodes sont necessaires.</div>',unsafe_allow_html=True)
                
                st.markdown("---")

                # --- 2, 3, 4 & 5. Nouveau tableau de suivi par poste avec Sparklines ---
                st.markdown('<div class="stl s">Suivi Sparklines par Poste de Travail</div>',unsafe_allow_html=True)

                def get_spark_color(v):
                    if pd.isna(v): return "#cbd5e0"
                    if v >= 90: return "#10b981" # Vert
                    elif v >= 80: return "#f59e0b" # Jaune
                    else: return "#ef4444" # Rouge

                def get_sparkline_html(scores):
                    n = len(scores)
                    if n == 0: return ""
                    W, H = 130, 35
                    pad = 5
                    def get_xy(i, v):
                        x = pad + (i / (n - 1) * (W - 2 * pad)) if n > 1 else W / 2
                        v_disp = max(0, min(100, v if pd.notna(v) else 0))
                        y = H - pad - (v_disp / 100 * (H - 2 * pad))
                        return x, y

                    svg = f'<svg width="{W}" height="{H}" viewBox="0 0 {W} {H}" xmlns="http://www.w3.org/2000/svg">'
                    # Lignes (segments)
                    for i in range(n - 1):
                        x1, y1 = get_xy(i, scores[i])
                        x2, y2 = get_xy(i + 1, scores[i+1])
                        col = get_spark_color(scores[i+1])
                        svg += f'<line x1="{x1:.1f}" y1="{y1:.1f}" x2="{x2:.1f}" y2="{y2:.1f}" stroke="{col}" stroke-width="2.5" />'
                    # Points
                    for i, v in enumerate(scores):
                        x, y = get_xy(i, v)
                        col = get_spark_color(v)
                        svg += f'<circle cx="{x:.1f}" cy="{y:.1f}" r="3.5" fill="{col}" />'
                    svg += '</svg>'
                    return svg

                def get_comparison_html(scores):
                    if len(scores) == 0:
                        return '<span style="color:#94a3b8">N/A</span>'
                    if len(scores) == 1:
                        return '<span style="color:#94a3b8">Première mesure disponible</span>'
                    
                    prev = scores[-2]
                    curr = scores[-1]
                    if prev == 0:
                        return '<span style="color:#94a3b8">➜ Stable</span>'
                    
                    pct = ((curr - prev) / prev) * 100
                    if pct > 0.05:
                        return f'<span style="color:#10b981;font-weight:600">▲ +{pct:.1f} % — Amélioration</span>'
                    elif pct < -0.05:
                        return f'<span style="color:#ef4444;font-weight:600">▼ {pct:.1f} % — Dégradation</span>'
                    else:
                        return '<span style="color:#94a3b8;font-weight:600">➜ Stable</span>'

                # Préparation des données depuis hist_df
                if not hist_df.empty and "Poste de travail" in hist_df.columns:
                    # Filtrer pour ne garder que les postes actuellement sélectionnés et valides
                    valid_postes = [p for p in vp if p in hist_df["Poste de travail"].unique()]
                    valid_postes = sorted(valid_postes)

                    perf_df_h = hist_df[(hist_df["_section"]=="perf") & (hist_df["Poste de travail"].isin(valid_postes))]
                    qual_df_h = hist_df[(hist_df["_section"]=="qual") & (hist_df["Poste de travail"].isin(valid_postes))]

                    # Construction du tableau HTML
                    h = '<table class="tw st"><thead><tr>'
                    h += '<th>Poste de travail</th><th>Sparkline Performance</th><th>Comparaison Performance</th>'
                    h += '<th>Sparkline Qualité</th><th>Comparaison Qualité</th>'
                    h += '</tr></thead><tbody>'

                    for poste in valid_postes:
                        p_data = perf_df_h[perf_df_h["Poste de travail"]==poste].sort_values("Date_parsed")
                        q_data = qual_df_h[qual_df_h["Poste de travail"]==poste].sort_values("Date_parsed")
                        
                        p_scores = p_data["Score Performance"].astype(float).tolist() if "Score Performance" in p_data.columns else []
                        q_scores = q_data["Score Qualite"].astype(float).tolist() if "Score Qualite" in q_data.columns else []

                        h += f'<tr><td style="font-weight:700">{poste}</td>'
                        h += f'<td class="spark-cell">{get_sparkline_html(p_scores)}</td>'
                        h += f'<td class="spark-cell">{get_comparison_html(p_scores)}</td>'
                        h += f'<td class="spark-cell">{get_sparkline_html(q_scores)}</td>'
                        h += f'<td class="spark-cell">{get_comparison_html(q_scores)}</td>'
                        h += '</tr>'

                    h += '</tbody></table>'
                    st.markdown(h, unsafe_allow_html=True)
                else:
                    st.markdown('<div class="es">Pas assez de données historiques pour générer les sparklines.</div>', unsafe_allow_html=True)

                st.markdown("---")
                st.markdown('<div class="stl s">Journal des variations significatives</div>',unsafe_allow_html=True)
                if not journal_df.empty:
                    st.dataframe(journal_df[["Date precedente","Date actuelle","Poste","Type","KPI","Valeur precedente","Valeur actuelle","Ecart %","Sens"]].reset_index(drop=True),use_container_width=True,height=400)
                else:
                    st.markdown('<div class="es">Aucune variation significative detectee (ecart >= 5%% entre deux periodes)</div>',unsafe_allow_html=True)

                if not top5_df.empty:
                    c1,c2=st.columns(2)
                    with c1:
                        st.markdown('<div class="stl p">Top 5 Postes — Amelioration</div>',unsafe_allow_html=True)
                        st.dataframe(top5_df,use_container_width=True)
                    with c2:
                        st.markdown('<div class="stl a">Bottom 5 Postes — Degradation</div>',unsafe_allow_html=True)
                        st.dataframe(bot5_df,use_container_width=True)

            with tabs[5]:
                st.markdown('<div class="stl a">📋 Plan d\'action</div>', unsafe_allow_html=True)

                # ── Bouton de téléchargement PDF ──
                col_pdf, col_metrics = st.columns([1, 3])
                with col_pdf:
                    if st.button("📥 Télécharger en PDF", use_container_width=True):
                        components.html("<script>window.print();</script>", height=0, width=0)

                with col_metrics:
                    mc1, mc2, mc3 = st.columns(3)
                    with mc1: st.metric("🔔 Total Actions Requises", len(plan_actions_rows))
                    with mc2: st.metric("🏭 Actions SF1", len(sf1_rows))
                    with mc3: st.metric("🏭 Actions SF2", len(sf2_rows))

                st.write("")

                # ── Tableaux HTML professionnels (SF1 / SF2 séparés) ──
                st.markdown(html_plan_actions_table(sf1_rows, "SF1 — Plan d'Actions", "#3b82f6", anomaly_dfs), unsafe_allow_html=True)
                st.markdown(html_plan_actions_table(sf2_rows, "SF2 — Plan d'Actions", "#10b981", anomaly_dfs), unsafe_allow_html=True)

                if not plan_actions_rows:
                    st.markdown('<div class="es">🎉 Aucune anomalie détectée. Tous les KPIs sont aux normes !</div>', unsafe_allow_html=True)

        except Exception as e:
            st.error("Erreur lors du chargement des donnees : %s"%str(e))
            st.markdown('<div class="es">Veuillez verifier que les fichiers ot.xlsx et avis.xlsx sont presents dans le repertoire.</div>',unsafe_allow_html=True)
    else:
        st.markdown('<div class="es">📁 Veuillez charger les fichiers OT et AVIS via le panneau de filtres.</div>',unsafe_allow_html=True)

    st.markdown('<div class="footer">Bureau Méthodes Maroc Chimie – © 2026 Tous droits réservés</div>', unsafe_allow_html=True)

if __name__ == "__main__":
    main()
