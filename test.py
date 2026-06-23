# -*- coding: utf-8 -*-
import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import numpy as np
import io, locale, random, time, os, hashlib, json, base64
from datetime import datetime
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ============================================================
st.set_page_config(layout="wide", page_title="Dashboard KPI", initial_sidebar_state="expanded")
# ============================================================

QK = ["TAUX_REALISATION_CORRECTIF/PT","OT préparation <1 mois","OT préparation >3 mois",
      "OT préparation 1mois< <3mois","OT planification <1 mois","OT planification >3 mois",
      "OT planification 1mois< <3mois","OT exécution <1 mois","OT exécution >3 mois",
      "OT exécution 1mois< <3mois",
      "Performance Graissage","Performance Inspection","Performance Appels Systématiques"]
PK = ["Taux d'approbation des Avis","OT LANC ESTIME","Backlog préparation caractérisé",
      "Backlog planification caractérisé","OT CONFIME","OT_COR_EGAL",
      "OT Fiabilité","Total Avis de Panne"]
ALL_KPI = QK + PK

CIBLE = {"TAUX_REALISATION_CORRECTIF/PT":85,"OT préparation <1 mois":80,"OT préparation >3 mois":5,
         "OT préparation 1mois< <3mois":15,"OT planification <1 mois":80,"OT planification >3 mois":5,
         "OT planification 1mois< <3mois":15,"OT exécution <1 mois":80,"OT exécution >3 mois":5,
         "OT exécution 1mois< <3mois":15,"Taux d'approbation des Avis":95,"OT LANC ESTIME":100,
         "Backlog préparation caractérisé":100,"Backlog planification caractérisé":100,
         "OT CONFIME":100,"OT_COR_EGAL":100,
         "Performance Graissage":95,"Performance Inspection":95,"Performance Appels Systématiques":95,
         "OT Fiabilité":100,"Total Avis de Panne":100}

ACT_MAP = {"TAUX_REALISATION_CORRECTIF/PT":"Ameliorer le taux de realisation des OT.",
           "OT préparation <1 mois":"Reduire l'age de preparation des OT (< 1 mois).",
           "OT préparation >3 mois":"Traiter les OT avec preparation > 3 mois.",
           "OT planification <1 mois":"Reduire l'age de planification des OT (< 1 mois).",
           "OT planification >3 mois":"Traiter les OT avec planification > 3 mois.",
           "OT exécution <1 mois":"Reduire l'age d'execution des OT (< 1 mois).",
           "OT exécution >3 mois":"Traiter les OT avec execution > 3 mois.",
           "OT LANC ESTIME":"Estimer les couts des OT lances.",
           "Backlog préparation caractérisé":"Caracteriser le backlog de preparation.",
           "Backlog planification caractérisé":"Caracteriser le backlog de planification.",
           "OT CONFIME":"Confirmer les OT termines.",
           "OT_COR_EGAL":"Rapprocher les couts reels et budgetes.",
           "Taux d'approbation des Avis":"Creer un OT pour les avis sans ordre.",
           "OT préparation 1mois< <3mois":"Reduire les OT entre 1 et 3 mois.",
           "OT planification 1mois< <3mois":"Reduire les OT entre 1 et 3 mois.",
           "OT exécution 1mois< <3mois":"Reduire les OT entre 1 et 3 mois.",
           "Performance Graissage":"Ameliorer le taux de realisation des OT de graissage (Type 350).",
           "Performance Inspection":"Ameliorer le taux de realisation des OT d'inspection (Types 290,300,310).",
           "Performance Appels Systématiques":"Ameliorer le taux de realisation des appels systematiques (Type 360).",
           "OT Fiabilité":"Maintenir la fiabilite des OT a 100%.",
           "Total Avis de Panne":"Maintenir le suivi des avis de panne a 100%."}

KPI_RESP_MAP = {
    "TAUX_REALISATION_CORRECTIF/PT": "Chef d'atelier",
    "OT préparation <1 mois": "Préparateur BM",
    "OT préparation 1mois< <3mois": "Préparateur BM",
    "OT préparation >3 mois": "Préparateur BM",
    "OT planification <1 mois": "Planificateur BM",
    "OT planification 1mois< <3mois": "Planificateur BM",
    "OT planification >3 mois": "Planificateur BM",
    "OT exécution <1 mois": "Chef d'atelier",
    "OT exécution 1mois< <3mois": "Chef d'atelier",
    "OT exécution >3 mois": "Chef d'atelier",
    "Taux d'approbation des Avis": "Chef d'atelier",
    "OT LANC ESTIME": "Fiabilité",
    "Backlog préparation caractérisé": "Préparateur BM",
    "Backlog planification caractérisé": "Planificateur BM",
    "OT CONFIME": "Agent de saisie",
    "OT_COR_EGAL": "Agent de saisie",
    "Performance Graissage": "Chef d'atelier",
    "Performance Inspection": "Chef d'atelier",
    "Performance Appels Systématiques": "Chef d'atelier",
    "OT Fiabilité": "Fiabilité",
    "Total Avis de Panne": "Fiabilité"
}

LOWER_BETTER = ["OT préparation >3 mois","OT planification >3 mois","OT exécution >3 mois",
                "OT préparation 1mois< <3mois","OT planification 1mois< <3mois","OT exécution 1mois< <3mois"]

MP_KW = ["CRPR ATPD","CRPR ATMR","CRPR ATER","CRPR ATRS","CRPR ATMO","ATPD","ATMR","ATER","ATRS","ATMO"]
MPLAN_KW = ["ATPL ATEI","ATPL ATAL","ATPL ATER","ATPL AGAR","ATPL ATHS","ATEI","ATAL","ATAS","AGAR","ATHS"]

CONSIGNES_HSE = [
    "Port obligatoire des EPI avant toute intervention.","Port obligatoire du casque de securite.",
    "Port obligatoire des lunettes de protection.","Port obligatoire des gants adaptes au travail.",
    "Utiliser les protections auditives dans les zones bruyantes.","Verifier l'absence de tension avant toute intervention electrique.",
    "Respecter la procedure de consignation et deconsignation.","Ne jamais intervenir sur un equipement en marche.",
    "Baliser et securiser la zone de travail.","Maintenir le poste de travail propre et ordonne.",
    "Verifier l'etat des outils avant utilisation.","Utiliser uniquement du materiel homologue.",
    "Respecter les permis de travail en vigueur.","Identifier les risques avant de commencer une tache.",
    "Signaler immediatement toute situation dangereuse.","Signaler tout incident ou presque accident.",
    "Ne jamais neutraliser un dispositif de securite.","Verifier les detecteurs de gaz avant utilisation.",
    "Verifier la bonne ventilation des zones de travail.","Respecter les regles des espaces confines.",
    "Controler l'atmosphere avant d'entrer dans un espace confine.","Utiliser les points d'ancrage pour les travaux en hauteur.",
    "Verifier l'etat des echafaudages avant utilisation.","Securiser les outils lors des travaux en hauteur.",
    "Ne pas travailler seul lors des operations a risque.","Controler les elingues avant chaque levage.",
    "Respecter les limites de charge des equipements.","Verifier l'etat des appareils de levage.",
    "Maintenir les voies de circulation degagees.","Respecter la signalisation de securite.",
    "Verifier les extincteurs a proximite du chantier.","Connaitre les issues de secours les plus proches.",
    "Respecter les procedures d'arret d'urgence.","Verifier les flexibles et raccords avant mise en service.",
    "Controler les fuites avant demarrage d'un equipement.","Respecter les distances de securite.",
    "Ne jamais contourner une procedure HSE.","Porter les EPI adaptes au risque identifie.",
    "Prevenir son responsable avant toute intervention particuliere.","Analyser les risques avant chaque demarrage de chantier.",
    "Verifier la stabilite des equipements.","Utiliser les bons outils pour la bonne tache.",
    "Respecter les consignes specifiques du chantier.","Ne jamais prendre de raccourci au detriment de la securite.",
    "Arreter immediatement les travaux en cas de danger.","Proteger l'environnement lors des interventions.",
    "Collecter et trier correctement les dechets.","Eviter toute pollution accidentelle.",
    "Respecter les consignes de stockage des produits dangereux.","Lire les fiches de securite avant manipulation.",
    "Verifier les equipements avant chaque prise de poste.","S'assurer de la disponibilite des moyens de secours.",
    "Communiquer clairement avec l'equipe avant intervention.","Respecter les regles de circulation des engins.",
    "Garder une vigilance permanente sur son environnement.","Prendre le temps d'effectuer le travail en securite.",
    "La securite est l'affaire de tous.","Chaque incident peut etre evite par la prevention.",
    "Aucun travail n'est plus urgent que la securite.","Zero accident commence par un comportement sur."]

def get_logo_base64():
    for path in ["logo.png", "./logo.png", "../logo.png"]:
        if os.path.exists(path):
            try:
                with open(path, "rb") as f:
                    return base64.b64encode(f.read()).decode()
            except Exception:
                pass
    return None

def get_date_from_file():
    if os.path.exists("date.txt"):
        try:
            with open("date.txt","r",encoding="utf-8") as f: return f.read().strip()
        except Exception: pass
    return pd.Timestamp.today().strftime("%d/%m/%Y")

def contient_mot(t,lm):
    t=str(t); return any(m in t for l in lm for m in l.split())
def cat_age(a):
    if pd.isna(a): return "Inconnu"
    if a<=1: return "<1 mois"
    elif a>=3: return ">3 mois"
    return "1 mois < <3 mois"

def excr(df):
    if "Poste travail princ." in df.columns:
        return df[~df["Poste travail princ."].astype(str).str.contains("cresseur",case=False,na=False)].copy()
    return df

@st.cache_data(show_spinner=False)
def read_excel_safe(bytes_data):
    """Lit un fichier Excel en détectant automatiquement le vrai format."""
    bio = io.BytesIO(bytes_data)
    header = bytes_data[:8]
    if header[:4] in (b'PK\x03\x04', b'PK\x05\x06'):
        for engine in ['openpyxl', 'calamine']:
            try:
                return pd.read_excel(bio, engine=engine)
            except Exception:
                bio.seek(0)
                continue
    if header == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1':
        for engine in ['xlrd', 'calamine']:
            try:
                return pd.read_excel(bio, engine=engine)
            except Exception:
                bio.seek(0)
                continue
    for engine in ['openpyxl', 'xlrd', 'calamine']:
        try:
            bio.seek(0)
            return pd.read_excel(bio, engine=engine)
        except Exception:
            continue
    raise ValueError(
        "Format de fichier non reconnu. Le fichier n'est ni un .xlsx ni un .xls valide.\n"
        "Vérifiez que le fichier n'est pas corrompu ou protégé par mot de passe."
    )


@st.cache_data(show_spinner=False)
def prepare_data(ot_bytes, av_bytes, date_str):
    raw_ot = read_excel_safe(ot_bytes)
    raw_av = read_excel_safe(av_bytes)
    raw_ot = excr(raw_ot)
    raw_av = excr(raw_av)
    
    for c in ["Créé le","Date de début planifiée","Date de clôture","Début réel","Fin réelle"]:
        if c in raw_ot.columns: raw_ot[c]=pd.to_datetime(raw_ot[c],errors="coerce")
    for c in ["Créé le","Début souhaité","Date de la clôture"]:
        if c in raw_av.columns: raw_av[c]=pd.to_datetime(raw_av[c],errors="coerce")
        
    now_ts = pd.Timestamp.today()
    df = raw_ot.copy()
    
    df["Backlog preparation"]=np.where(df["Statut utilisateur"].apply(lambda x:contient_mot(x,MP_KW)),"CARACTERISE","NON CARACTERISE")
    df["Backlog planification"]=np.where(df["Statut utilisateur"].apply(lambda x:contient_mot(x,MPLAN_KW)),"CARACTERISE","NON CARACTERISE")
    df["Type Carac Prep"]=df["Statut utilisateur"].apply(lambda x: next((kw.split()[0] for kw in MP_KW if kw in str(x)), "NON CARACTERISE"))
    df["Type Carac Plan"]=df["Statut utilisateur"].apply(lambda x: next((kw.split()[0] for kw in MPLAN_KW if kw in str(x)), "NON CARACTERISE"))
    
    for dc,am,ac in [('Créé le',"amp","ap"),('Date de début planifiée',"amlp","alp"),('Date de début planifiée',"amex","aex")]:
        if dc in df.columns:
            df[am]=((now_ts.year-df[dc].dt.year)*12+(now_ts.month-df[dc].dt.month)).round(2)
            df[ac]=df[am].apply(cat_age)
        else: df[am]=np.nan; df[ac]="Inconnu"
        
    df["OT CONFIME"]=np.where(df["Statut système"].str.contains("CLOT|TCLO",na=False) & df["Statut système"].str.contains("CONF",na=False),"OUI","NON")
    
    df["Contient SOPL"]=df["Statut utilisateur"].str.contains("SOPL",na=False).map({True:1,False:0})
    df["OT LANC ESTIME"]=np.where(df["Total coûts budgétés"].fillna(0)==0,"NON","OUI")
    df["OT_COR_EGAL"]=np.where((df["Total coûts budgétés"].fillna(0)-df["Total coûts réels"].fillna(0))==0,"OUI","NON")
    df["_tw_num"]=pd.to_numeric(df.get("Type de travail",pd.Series(dtype=float)),errors="coerce")
    
    if "Statut système" in df.columns: df["Statut OT"]=df["Statut système"].fillna("").astype(str).str.strip().str.split().str[0]
    
    avf = raw_av[
        (
            raw_av["Ordre"].isna() |
            (raw_av["Ordre"].astype(str).str.strip() == "")
        )
        &
        (raw_av["Type d'avis"].isin(["ZU","Z4","ZR","ZP"]))
    ].copy()
    
    apm = sorted(df[df["Poste travail princ."].astype(str).str.startswith(("SF1","SF2"),na=False)]["Poste travail princ."].dropna().unique().tolist())
    
    return df, avf, apm, now_ts


def save_kpis_to_excel(prows,pcols,qrows,qcols,ano_p_r,ano_p_c,ano_q_r,ano_q_c,sheet_name):
    kpis_dir="kpis"; os.makedirs(kpis_dir,exist_ok=True)
    filepath=os.path.join(kpis_dir,"indicateurs_kpis.xlsx")
    sn=str(sheet_name).replace("/","-").replace("\\","-").replace("*","").replace("?","").replace("[","").replace("]","")[:31]
    hf=Font(bold=True,color="FFFFFF",size=10); hfl=PatternFill(start_color="1E3A5F",end_color="1E3A5F",fill_type="solid")
    tf=Font(bold=True,size=12,color="1E3A5F")
    tb=Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
    try: wb=load_workbook(filepath)
    except Exception: wb=Workbook()
    if "Sheet" in wb.sheetnames: del wb["Sheet"]
    if sn in wb.sheetnames: del wb[sn]
    ws=wb.create_sheet(sn); rn=1
    def ws_sec(title,cols,rows,sr):
        ws.cell(row=sr,column=1,value=title).font=tf; sr+=1
        for j,c in enumerate(cols,1):
            cl=ws.cell(row=sr,column=j,value=c); cl.font=hf; cl.fill=hfl; cl.alignment=Alignment(horizontal='center'); cl.border=tb
        sr+=1
        for r in rows:
            for j,c in enumerate(cols,1):
                cl=ws.cell(row=sr,column=j,value=r.get(c,"")); cl.border=tb; cl.alignment=Alignment(horizontal='center')
            sr+=1
        return sr+1
    rn=ws_sec("INDICATEURS DE PERFORMANCE",pcols,prows,rn)
    if ano_p_c and ano_p_r: rn=ws_sec("ANOMALIES PERFORMANCE (filtre PK)",ano_p_c,ano_p_r,rn)
    rn=ws_sec("INDICATEURS DE QUALITE",qcols,qrows,rn)
    if ano_q_c and ano_q_r: rn=ws_sec("ANOMALIES QUALITE (filtre QK)",ano_q_c,ano_q_r,rn)
    try: wb.save(filepath)
    except Exception: pass

def load_historical_kpis(filepath):
    if not os.path.exists(filepath): return pd.DataFrame()
    try: wb=load_workbook(filepath,read_only=True,data_only=True)
    except Exception: return pd.DataFrame()
    records=[]; section=None; headers=None
    for sheet_name in wb.sheetnames:
        try:
            ws=wb[sheet_name]; rows_data=list(ws.iter_rows(values_only=True))
            for row in rows_data:
                cell0=str(row[0]).strip() if row[0] else ""
                if "INDICATEURS DE PERFORMANCE" in cell0.upper(): section="perf"; headers=None; continue
                elif "INDICATEURS DE QUALITE" in cell0.upper(): section="qual"; headers=None; continue
                elif "ANOMALIES" in cell0.upper(): section=None; continue
                if section and headers is None and cell0:
                    headers=[str(c).strip() if c else "" for c in row]; continue
                if section and headers and cell0 and cell0 not in ("Cible","Total general",""):
                    entry={"Date":sheet_name}
                    for j,h in enumerate(headers):
                        if j<len(row): entry[h]=row[j]
                    entry["_section"]=section; records.append(entry)
        except Exception: continue
    wb.close()
    if not records: return pd.DataFrame()
    df=pd.DataFrame(records)
    df["Date_parsed"]=pd.to_datetime(df["Date"].str.replace("-","/"),format="%d/%m/%Y",errors="coerce")
    return df.sort_values("Date_parsed").reset_index(drop=True)

def calculate_variations(hist_df):
    if hist_df.empty or "Date" not in hist_df.columns: return pd.DataFrame()
    dates=sorted(hist_df["Date"].unique())
    if len(dates)<2: return pd.DataFrame()
    perf_df=hist_df[hist_df["_section"]=="perf"].copy()
    qual_df=hist_df[hist_df["_section"]=="qual"].copy()
    variations=[]
    for i in range(1,len(dates)):
        prev_date,curr_date=dates[i-1],dates[i]
        prev_perf=perf_df[perf_df["Date"]==prev_date].set_index("Poste de travail") if "Poste de travail" in perf_df.columns else pd.DataFrame()
        curr_perf=perf_df[perf_df["Date"]==curr_date].set_index("Poste de travail") if "Poste de travail" in perf_df.columns else pd.DataFrame()
        prev_qual=qual_df[qual_df["Date"]==prev_date].set_index("Poste de travail") if "Poste de travail" in qual_df.columns else pd.DataFrame()
        curr_qual=qual_df[qual_df["Date"]==curr_date].set_index("Poste de travail") if "Poste de travail" in qual_df.columns else pd.DataFrame()
        for sec_name,prev_d,curr_d,kpi_list in [("Performance",prev_perf,curr_perf,PK+["Score Performance"]),("Qualite",prev_qual,curr_qual,QK+["Score Qualite"])]:
            for poste in set(prev_d.index)&set(curr_d.index):
                for kpi in kpi_list:
                    if kpi not in prev_d.columns or kpi not in curr_d.columns: continue
                    try: pv=float(prev_d.loc[poste,kpi])
                    except Exception: continue
                    try: cv=float(curr_d.loc[poste,kpi])
                    except Exception: continue
                    diff=cv-pv; pct=(diff/pv*100) if pv!=0 else (100 if cv!=0 else 0)
                    if abs(diff)<=0.5: trend="stabilite"
                    elif diff>0.5: trend="hausse"
                    else: trend="baisse"
                    sens = "Stable"
                    if trend != "stabilite":
                        if (trend == "hausse" and kpi not in LOWER_BETTER) or (trend == "baisse" and kpi in LOWER_BETTER):
                            sens = "Amelioration"
                        else:
                            sens = "Degradation"
                    variations.append({"Date precedente":prev_date,"Date actuelle":curr_date,"Poste":poste,
                        "Type":sec_name,"KPI":kpi,"Valeur precedente":round(pv,2),"Valeur actuelle":round(cv,2),
                        "Ecart":round(diff,2),"Ecart %":round(pct,2),"Tendance":trend, "Sens":sens})
    return pd.DataFrame(variations)

def generate_journal(var_df):
    if var_df.empty: return pd.DataFrame()
    j=var_df.copy(); j["Significatif"]=j["Ecart %"].abs()>=5
    j=j[j["Significatif"]].copy()
    return j.sort_values(["Date actuelle","Ecart %"],ascending=[True,False])

def calculate_rankings(var_df):
    if var_df.empty: return pd.DataFrame(),pd.DataFrame()
    scores={}
    for poste in var_df["Poste"].unique():
        pv=var_df[var_df["Poste"]==poste].copy()
        scores[poste]=sum((-r["Ecart %"] if r["KPI"] in LOWER_BETTER else r["Ecart %"]) for _,r in pv.iterrows())
    ranked=sorted(scores.items(),key=lambda x:x[1],reverse=True)
    return pd.DataFrame(ranked[:5],columns=["Poste","Score variation"]),pd.DataFrame(ranked[-5:][::-1],columns=["Poste","Score variation"])


# ============================================================
# FONCTION MODIFIEE : Calcul des anomalies avec filtre par type de KPI
# ============================================================
def calculer_anomalies_par_kpi_poste(ckdf, posts, kpi_list, label_section):
    """
    Calcule les anomalies par KPI et Poste pour atteindre 100%.
    
    Parameters:
    -----------
    ckdf : DataFrame - Contient les valeurs des KPI par poste (index = postes)
    posts : list - Liste des postes de travail
    kpi_list : list - Liste des KPI à analyser :
        - PK pour anomalies de Performance (filtre Detail des indicateurs de Performance)
        - QK pour anomalies de Qualité (filtre Detail des indicateurs de Qualité)
    label_section : str - "Performance" ou "Qualité"
    
    Returns:
    --------
    tuple : (rows_list, cols_list, dataframe)
    """
    rows = []
    cols = ["Poste de travail", "KPI", "Valeur actuelle", "Cible", "Ecart à traiter", "Action corrective", "Responsable"]
    
    for kpi in kpi_list:
        cible = CIBLE.get(kpi, 100)
        for poste in posts:
            val = ckdf.get(kpi, pd.Series(dtype=float)).get(poste, 100)
            if pd.notna(val):
                val = float(val)
                ecart = 0
                is_anomalie = False
                
                if kpi in LOWER_BETTER:
                    if val > cible:
                        ecart = val - cible
                        is_anomalie = True
                else:
                    if val < cible:
                        ecart = cible - val
                        is_anomalie = True
                
                if is_anomalie:
                    rows.append({
                        "Poste de travail": poste,
                        "KPI": kpi,
                        "Valeur actuelle": round(val, 2),
                        "Cible": cible,
                        "Ecart à traiter": round(ecart, 2),
                        "Action corrective": ACT_MAP.get(kpi, ""),
                        "Responsable": KPI_RESP_MAP.get(kpi, "")
                    })
    
    return rows, cols, pd.DataFrame(rows)


def inject_custom_css():
    st.markdown("""<style>
    section[data-testid="stSidebar"]{width:250px!important}
    .main .block-container{max-width:100%!important;width:100%!important;padding-left:0.5rem!important;padding-right:0.5rem!important}
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700;800&display=swap');
    
    :root{
        --primary:#1e3a5f;
        --primary-light:#2c5282;
        --success:#10b981;
        --success-dark:#059669;
        --warning:#f59e0b;
        --warning-dark:#d97706;
        --danger:#ef4444;
        --danger-dark:#dc2626;
        --info:#3b82f6;
        --border:#e2e8f0;
        --radius:10px;
    }
    
    *{box-sizing:border-box;margin:0;padding:0}
    .stApp{background:#f8fafc;font-family:'Inter',sans-serif}
    .main .block-container{padding-top:.8rem;padding-bottom:.8rem}
    
    .mh{
        background:linear-gradient(135deg,#1e3a5f 0%,#2563eb 100%);
        padding:16px 24px;
        border-radius:12px;
        margin-bottom:12px;
        box-shadow:0 8px 24px rgba(30,58,95,0.15);
        display:flex;
        align-items:center;
        gap:16px;
    }
    .mh h1{color:#fff;font-size:42px;font-weight:800;margin:0;flex:1}
    .mh .logo{height:50px;width:auto;max-width:150px;object-fit:contain;border-radius:6px}
    .mh .db{
        background:rgba(255,255,255,0.2);
        padding:6px 16px;
        border-radius:16px;
        color:#fff;
        font-size:20px;
        font-weight:600;
        border:1px solid rgba(255,255,255,0.3);
        backdrop-filter:blur(10px);
    }
    
    .cr{display:grid;grid-template-columns:repeat(4,1fr);gap:8px;margin-bottom:10px}
    .cc{
        background:#fff;
        border-radius:12px;
        padding:18px 16px;
        box-shadow:0 4px 12px rgba(0,0,0,0.06);
        border-left:4px solid;
        transition:transform 0.2s,box-shadow 0.2s;
        text-align:center;
    }
    .cc:hover{transform:translateY(-2px);box-shadow:0 8px 20px rgba(0,0,0,0.1);}
    .cc .cv{font-size:32px;font-weight:900;line-height:1.1;}
    .cc .cl{font-size:16px;color:#1e293b;font-weight:800;text-transform:uppercase;letter-spacing:.5px;margin-top:8px;}
    .cc.c1{border-left-color:#3b82f6}.cc.c1 .cv{color:#2563eb}
    .cc.c2{border-left-color:#10b981}.cc.c2 .cv{color:#059669}
    .cc.c3{border-left-color:#8b5cf6}.cc.c3 .cv{color:#7c3aed}
    .cc.c4{border-left-color:#ef4444}.cc.c4 .cv{color:#dc2626}
    .cc.c5{border-left-color:#3b82f6}.cc.c5 .cv{color:#2563eb}
    .cc.c6{border-left-color:#06b6d4}.cc.c6 .cv{color:#0891b2}
    .cc.c7{border-left-color:#f59e0b}.cc.c7 .cv{color:#d97706}
    .cc.c8{border-left-color:#f97316}.cc.c8 .cv{color:#ea580c}
    
    .stl{font-size:16px;font-weight:800;color:var(--primary);margin:10px 0 5px 0;padding-left:12px;border-left:4px solid var(--info);}
    
    .tw{width:100%;border-collapse:collapse;font-family:'Inter',sans-serif;font-size:13px;display:block;overflow-x:auto;-webkit-overflow-scrolling:touch;margin:0}
    .tw thead th{background:var(--primary);color:#fff;font-weight:700;font-size:11px;text-transform:uppercase;letter-spacing:.3px;padding:6px 8px;border:none;white-space:nowrap;position:sticky;top:0;z-index:10}
    .tw.qt thead th{background:linear-gradient(135deg,#2563eb,#3b82f6)}
    .tw.pt thead th{background:linear-gradient(135deg,#059669,#10b981)}
    .tw.at thead th{background:linear-gradient(135deg,#dc2626,#ef4444)}
    .tw.st thead th{background:linear-gradient(135deg,#d97706,#f59e0b)}
    .tw thead th:first-child{z-index:11;left:0}
    .tw tbody td:first-child{position:sticky;left:0;background:#fff;z-index:5;border-right:1px solid var(--border);color:#1e293b !important}
    .tw tbody tr:nth-child(even) td:first-child{background:#f8fafc}
    .tw tbody tr:hover td:first-child{background:#eff6ff}
    .tw tbody td{padding:5px 8px;border-bottom:1px solid var(--border);white-space:nowrap;color:#1e293b !important}
    .tw tbody tr:nth-child(even) td{background:#f8fafc}
    .tw tbody tr:hover td{background:#eff6ff!important}
    .cb td{background:#2563eb!important;color:#fff!important;font-weight:700!important;font-size:12px!important}
    
    .plan-action-table { width:100%; border-collapse:collapse; font-family:Inter,sans-serif; font-size:12px; border:1px solid #cbd5e1; }
    .plan-action-table th { background:#1e3a5f; color:#fff; font-weight:700; padding:8px 6px; border:1px solid #1e3a5f; }
    .plan-action-table td { padding:6px 8px; border:1px solid #cbd5e1; text-align:center; vertical-align:middle;}
    .plan-action-table td:first-child { text-align:left; font-weight:800; }
    
    .stTabs [data-baseweb="tab-list"]{gap:6px;background:#e2e8f0;padding:6px;border-radius:8px;margin-bottom:8px}
    .stTabs [data-baseweb="tab"]{
        border-radius:6px;
        padding:12px 22px;
        font-weight:700;
        font-size:20px;
        line-height:1.5;
        min-height:48px;
    }
    .stTabs [data-baseweb="tab"] span,
    .stTabs [data-baseweb="tab"] > div{
        font-size:22px !important;
    }
    .stTabs [aria-selected="true"]{
        background:#fff!important;
        color:var(--primary)!important;
        box-shadow:0 3px 8px rgba(0,0,0,.1);
        font-size:21px;
    }
    .stTabs [data-baseweb="tab"] svg{width:22px;height:22px}
    
    .ca{background:#fff;border-radius:var(--radius);padding:12px;margin-top:6px;border:1px solid var(--border);box-shadow:0 1px 4px rgba(0,0,0,.02)}
    .ca .ct{font-size:14px;font-weight:700;margin-bottom:8px;padding-bottom:5px;border-bottom:1px solid var(--border)}
    .car{display:flex;align-items:center;margin-bottom:6px;font-size:12px}
    .car:last-child{margin-bottom:0}
    .car .cal{width:260px;font-weight:600;color:var(--primary);text-align:right;padding-right:8px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .car .cab{flex:1;height:26px;background:#edf2f7;border-radius:4px;overflow:visible;position:relative}
    .car .caf{height:100%;border-radius:4px;transition:width .3s}
    
    .car .target-mark{
        position:absolute;
        top:-4px;
        bottom:-4px;
        width:3px;
        background:var(--info) !important; 
        z-index:20;
        transform:translateX(-50%);
        box-shadow:0 0 6px rgba(59,130,246,.9),0 0 2px rgba(0,0,0,.4);
        border-radius:2px;
    }
    .car .cav-out{font-size:12px;font-weight:800;color:#1e293b;min-width:55px;text-align:right;padding-left:6px}
    .car .cav-tgt{font-size:10px;font-weight:700;color:#1e293b;min-width:42px;text-align:right;padding-left:4px;opacity:.7}
    .gbr{display:flex;align-items:center;padding:3px 0;font-size:12px;border-bottom:1px solid #f1f5f9}
    .gbr:last-child{border:none}
    .gbr-l{width:160px;font-weight:600;color:#1e293b;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;font-size:11px;padding-right:10px;}
    .gbr-g{display:flex;align-items:center;gap:4px;flex:1;position:relative}
    .gbr-target{position:absolute;left:90%;top:-4px;bottom:-4px;width:3px;background:var(--primary) !important;z-index:10;box-shadow:0 0 6px rgba(30,58,95,.8);border-radius:2px}
    .gbr-target-label{position:absolute;left:90%;top:-20px;transform:translateX(-50%);font-size:9px;font-weight:800;color:#fff;background:var(--primary) !important;padding:1px 5px;border-radius:3px;white-space:nowrap;z-index:11;box-shadow:0 1px 3px rgba(0,0,0,.2)}
    .gbr-w{flex:1;height:22px;background:#f1f5f9;border-radius:3px;overflow:hidden}
    .gbr-f{height:100%;border-radius:3px}
    .gbr-v{font-size:11px;font-weight:800;min-width:48px;text-align:right;color:#1e293b}
    .gbr-legend{display:flex;gap:14px;margin-bottom:10px;font-size:12px;font-weight:700;align-items:center}
    .gbr-legend span{display:flex;align-items:center;gap:5px}
    .gbr-legend i{display:inline-block;width:14px;height:14px;border-radius:2px}
    .gbr-legend .target-icon{display:inline-block;width:3px;height:14px;background:var(--primary) !important;border-radius:1px;box-shadow:0 0 3px rgba(30,58,95,.6)}
    .cg{display:grid;grid-template-columns:1fr 1fr;gap:6px}
    .cg>div{background:#fff;border-radius:var(--radius);padding:8px 10px;border:1px solid var(--border)}
    .cg .ct{font-size:13px;font-weight:700;margin-bottom:4px;padding-bottom:3px;border-bottom:1px solid var(--border)}
    .cgr{display:flex;align-items:center;padding:3px 0;font-size:12px;border-bottom:1px solid #f1f5f9}
    .cgr:last-child{border:none}
    .cgr .rk{width:18px;font-weight:800;text-align:center}
    .cgr .pn{flex:1;font-weight:600;color:#1e293b;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .cgr .ps{font-weight:800;min-width:55px;text-align:right}
    .dgrid{display:grid;grid-template-columns:1fr 1fr;gap:6px}
    .stButton>button[kind="primary"]{background:linear-gradient(135deg,var(--primary),var(--primary-light));border:none;border-radius:6px;padding:8px 14px;font-weight:700;font-size:15px;width:100%}
    ::-webkit-scrollbar{width:5px;height:5px}::-webkit-scrollbar-track{background:#f1f1f1}::-webkit-scrollbar-thumb{background:#cbd5e0;border-radius:3px}
    
    div[data-testid="stSidebar"]{
        background:linear-gradient(180deg,#1e40af 0%,#1e3a8a 50%,#1e3a5f 100%)!important;
    }
    div[data-testid="stSidebar"]*{color:rgba(255,255,255,.9)!important}
    div[data-testid="stSidebar"] .stSelectbox label,div[data-testid="stSidebar"] .stMultiSelect label,div[data-testid="stSidebar"] .stDateInput label,div[data-testid="stSidebar"] .stCheckbox label,div[data-testid="stSidebar"] .stTextInput label{color:rgba(255,255,255,.9)!important;font-weight:600;font-size:13px;text-transform:uppercase;letter-spacing:.5px}
    div[data-testid="stSidebar"] div[data-testid="stWidget"]{background:rgba(255,255,255,.1);border-radius:6px;padding:5px 10px;margin-bottom:5px;border:1px solid rgba(255,255,255,.15)}
    div[data-testid="stSidebar"] .stSelectbox>div>div,div[data-testid="stSidebar"] .stMultiSelect>div>div,div[data-testid="stSidebar"] .stDateInput>div>div,div[data-testid="stSidebar"] .stTextInput>div>div{background:rgba(255,255,255,.95)!important;border-radius:5px}
    
    .es{text-align:center;padding:14px;color:#64748b;font-size:14px}
    .synth-tbl{width:100%;border-collapse:collapse;font-family:'Inter',sans-serif;font-size:12px}
    .synth-tbl thead th{background:var(--primary);color:#fff;font-weight:700;font-size:11px;padding:5px 8px;border:none;white-space:nowrap;position:sticky;top:0}
    .synth-tbl tbody td{padding:4px 8px;border-bottom:1px solid var(--border);text-align:center;color:#1e293b !important}
    .synth-tbl tbody tr:nth-child(even) td{background:#f8fafc}
    .synth-tbl tbody tr:hover td{background:#eff6ff!important}
    .synth-tbl .poste-cell{text-align:left;font-weight:700;white-space:nowrap;min-width:140px;color:#1e293b !important}
    
    [data-testid="stHeaderActionElements"]{display:none !important;}
    [data-testid="stActionButtonContainer"]{display:none !important;}
    
    .footer {
        text-align: center;
        margin-top: 30px;
        padding: 15px;
        color: #64748b;
        font-size: 13px;
        border-top: 1px solid var(--border);
        font-weight: 600;
    }

    div[data-testid="stDataEditor"] table, 
    div[data-testid="stDataEditor"] th, 
    div[data-testid="stDataEditor"] td {
        font-size: 18px !important;
        line-height: 1.4 !important;
        white-space: normal !important;
        word-wrap: break-word !important;
    }
    div[data-testid="stDataEditor"] [data-testid="stMarkdownContainer"] {
        font-size: 18px !important;
    }
    div[data-testid="stDataEditor"] [data-testid="stTable"] {
        overflow-x: hidden !important;
        width: 100% !important;
    }

    @media(max-width:768px){
        .cr{grid-template-columns:repeat(2,1fr)}
        .mh{padding:8px 10px;gap:8px}
        .mh h1{font-size:18px}
        .mh .logo{height:35px;max-width:70px}
        .mh .db{font-size:11px;padding:2px 8px}
        .cg,.dgrid{grid-template-columns:1fr}
        .car{flex-wrap:wrap;gap:2px}
        .car .cal{width:100%;text-align:left;padding-right:0;margin-bottom:2px}
        .car .cab{flex:1 1 70%}
        .car .cav-out,.car .cav-tgt{flex:1 1 15%;min-width:40px}
        .gbr{flex-direction:column;align-items:flex-start;gap:4px}
        .gbr-l{width:100%;margin-bottom:2px}
        .gbr-g{width:100%;flex-wrap:wrap}
        .gbr-w{flex:1 1 45%}
        .gbr-v{flex:1 1 10%;min-width:40px}
        .tw{font-size:10px}
        .tw thead th,.tw tbody td{padding:3px 4px}
        .stl{font-size:13px}
        .stTabs [data-baseweb="tab"]{padding:8px 12px;font-size:15px}
        .stTabs [data-baseweb="tab"] span{font-size:16px !important}
        div[data-testid="stDataEditor"] table, div[data-testid="stDataEditor"] th, div[data-testid="stDataEditor"] td { font-size: 14px !important; }
    }
    
    @media print {
        section[data-testid="stSidebar"], 
        header[data-testid="stHeader"], 
        div[data-testid="stToolbar"], 
        div[data-testid="stHeaderActionElements"],
        footer, 
        .stDeployButton, 
        #MainMenu { display: none !important; }
        .main .block-container { 
            padding-top: 0 !important; 
            padding-left: 0 !important; 
            padding-right: 0 !important; 
            max-width: 100% !important; 
        }
        .stButton, .stDownloadButton { display: none !important; }
        * {
            -webkit-print-color-adjust: exact !important;
            print-color-adjust: exact !important;
        }
        .tw, .synth-tbl { page-break-inside: avoid; overflow: visible !important; }
    }
    </style>""",unsafe_allow_html=True)

# ============================================================
def main():
    try: locale.setlocale(locale.LC_ALL,'fr_FR.UTF-8')
    except Exception:
        try: locale.setlocale(locale.LC_ALL,'fr_FR')
        except Exception: pass
    inject_custom_css()
    fichier_date=get_date_from_file()

    if "hse_affiche" not in st.session_state: st.session_state.hse_affiche=False
    if not st.session_state.hse_affiche:
        c=random.choice(CONSIGNES_HSE)
        st.markdown("""<div style="min-height:100vh;display:flex;flex-direction:column;align-items:center;justify-content:center;background:linear-gradient(135deg,#1a365d,#2d3748,#1a365d);padding:40px">
        <div style="font-size:64px;margin-bottom:20px">🦺</div>
        <h1 style="text-align:center;font-size:46px;color:#fff;font-weight:900;margin:0">HSE - CONSIGNE DE SECURITE</h1>
        <p style="text-align:center;color:rgba(255,255,255,.6);font-size:22px;margin-top:8px;letter-spacing:3px;text-transform:uppercase">Securite - Sante - Environnement</p>
        <div style="background:linear-gradient(135deg,#f6e05e,#ed8936);padding:36px 48px;border-radius:20px;font-size:32px;font-weight:700;text-align:center;margin:40px 0;color:#1a202c;max-width:800px;box-shadow:0 20px 60px rgba(0,0,0,.3)">⚠️ %s</div>
        <h2 style="text-align:center;color:#48bb78;font-size:36px;font-weight:900">Aucun travail n'est plus urgent que la securite</h2>
        <div style="margin-top:40px;width:200px;height:4px;background:rgba(255,255,255,.1);border-radius:2px;overflow:hidden"><div style="width:100%%;height:100%%;background:linear-gradient(90deg,#48bb78,#38a169);border-radius:2px;animation:ld 5.5s ease-in-out forwards"></div></div>
        <style>@keyframes ld{from{width:0}to{width:100%%}}</style></div>"""%c,unsafe_allow_html=True)
        time.sleep(6); st.session_state.hse_affiche=True; st.rerun(); st.stop()

    def ckpi(n,d,sz=100): return np.where(d==0,sz,(n/d)*100)
    
    def cpiv(df,f,c,p):
        return pd.pivot_table(df[f],index="Poste travail princ.",columns=c,values="Ordre",aggfunc="count",fill_value=0).reindex(p,fill_value=0)
        
    def get_text_col(df):
        for c in ["Désignation","Designation","Désignation OT","Texte ordre","Texte","Description","Libellé","Libelle"]:
            if c in df.columns: return c
        for c in df.columns:
            if df[c].dtype=='object' and any(kw in str(c).lower() for kw in ['sign','text','desc','libell']):
                return c
        return None
        
    def build_statut_pivot(df_sub, posts):
        if df_sub.empty:
            return pd.DataFrame(index=posts, columns=["CRÉÉ","LANC","CLOT","TCLO","Total"]).fillna(0).astype(int)
        piv=pd.pivot_table(df_sub, index="Poste travail princ.", columns="Statut OT", values="Ordre", aggfunc="count", fill_value=0)
        for s in ["CRÉÉ","LANC","CLOT","TCLO"]:
            if s not in piv.columns: piv[s]=0
        piv["Total"]=piv[["CRÉÉ","LANC","CLOT","TCLO"]].sum(axis=1)
        return piv.reindex(posts, fill_value=0).fillna(0).astype(int)
        
    def html_statut_pivot(piv_df, table_class):
        cols=["Poste de travail","CRÉÉ","LANC","CLOT","TCLO","Total"]
        statut_colors = {
            "CRÉÉ": "background:#fef3c7;color:#92400e;font-weight:600;",
            "LANC": "background:#dbeafe;color:#1e40af;font-weight:600;",
            "CLOT": "background:#d1fae5;color:#065f46;font-weight:600;",
            "TCLO": "background:#a7f3d0;color:#064e3b;font-weight:600;",
            "Total": "background:#ede9fe;color:#5b21b6;font-weight:700;"
        }
        h='<table class="tw %s"><thead><tr>'%table_class+''.join('<th>%s</th>'%c for c in cols)+'</tr></thead><tbody>'
        for poste,row in piv_df.iterrows():
            h+='<tr><td style="font-weight:700">%s</td>'%poste
            for c in ["CRÉÉ","LANC","CLOT","TCLO"]:
                h+='<td style="text-align:center;%s">%d</td>'%(statut_colors[c], int(row.get(c,0)))
            h+='<td style="text-align:center;%s">%d</td>'%(statut_colors["Total"], int(row.get("Total",0)))
            h+='</tr>'
        h+='<tr class="tr"><td style="font-weight:800">Total</td>'
        for c in ["CRÉÉ","LANC","CLOT","TCLO"]:
            h+='<td style="text-align:center;font-weight:800;%s">%d</td>'%(statut_colors[c], int(piv_df[c].sum()))
        h+='<td style="text-align:center;font-weight:800;%s">%d</td>'%(statut_colors["Total"], int(piv_df["Total"].sum()))
        h+='</tr></tbody></table>'
        return h
        
    def show_pie_pair(piv_df, title_prefix):
        global_counts = piv_df[["CRÉÉ","LANC","CLOT","TCLO"]].sum()
        global_counts = global_counts[global_counts > 0]
        realised = global_counts.get("CLOT", 0) + global_counts.get("TCLO", 0)
        not_realised = global_counts.sum() - realised
        
        if global_counts.empty:
            st.markdown('<div class="es">Aucune donnee</div>', unsafe_allow_html=True)
            return
            
        colors = ["#8b5cf6", "#f59e0b", "#10b981", "#3b82f6"]
        fig = make_subplots(rows=1, cols=2, specs=[[{"type":"domain"},{"type":"domain"}]], 
                            subplot_titles=(f"{title_prefix} — Par Statut OT", f"{title_prefix} — Réalisés vs Non Réalisés"))
        
        fig.add_trace(go.Pie(labels=global_counts.index, values=global_counts.values, hole=0.4, 
                             textinfo='percent+label', 
                             texttemplate='%{label}<br>%{percent:.1%}<br>(%{value})', 
                             textposition='inside',
                             insidetextorientation='radial',
                             textfont=dict(size=14, color='white', family='Inter, sans-serif'),
                             marker=dict(colors=colors, line=dict(color='#FFFFFF', width=3))), 1, 1)
                             
        pie2_data = pd.Series([realised, not_realised], index=["Réalisés (CLOT+TCLO)", "Non Réalisés"])
        
        fig.add_trace(go.Pie(labels=pie2_data.index, values=pie2_data.values, hole=0.5, 
                             textinfo='percent+label', 
                             texttemplate='%{label}<br>%{percent:.1%}<br>(%{value})', 
                             textposition='inside',
                             insidetextorientation='radial',
                             textfont=dict(size=14, color='white', family='Inter, sans-serif'),
                             marker=dict(colors=["#10b981", "#8b5cf6"], line=dict(color='#FFFFFF', width=3))), 1, 2)
                             
        fig.update_layout(margin=dict(t=80, b=20, l=20, r=20), height=450, 
                          legend=dict(orientation="h", yanchor="bottom", y=-0.12, x=0.5, xanchor="center"))
                          
        st.plotly_chart(fig, use_container_width=True)

    def show_simple_pie(piv_df, title, keep_non_carac=False):
        if not keep_non_carac and "NON CARACTERISE" in piv_df.columns:
            piv_df = piv_df.drop(columns=["NON CARACTERISE"])
            
        counts = piv_df.sum()
        counts = counts[counts > 0]
        
        if counts.empty:
            st.markdown('<div class="es">Aucune donnee</div>', unsafe_allow_html=True)
            return
            
        color_map = {"CARACTERISE": "#10b981", "NON CARACTERISE": "#f97316"}
        type_palette = ['#3b82f6', '#10b981', '#f59e0b', '#8b5cf6', '#06b6d4', '#14b8a6', '#6366f1', '#0ea5e9', '#d946ef', '#a855f7']
        
        colors = []
        palette_idx = 0
        for c in counts.index:
            c_str = str(c)
            if c_str in color_map:
                colors.append(color_map[c_str])
            else:
                colors.append(type_palette[palette_idx % len(type_palette)])
                palette_idx += 1
        
        total_sum = counts.sum()
        pull_list = [0.05 if (v/total_sum)*100 < 10 else 0 for v in counts.values]
        
        fig = go.Figure(
            go.Pie(
                labels=counts.index,
                values=counts.values,
                hole=0.4,
                sort=False,
                textinfo="percent",
                textposition="outside",
                pull=pull_list,
                marker=dict(
                    colors=colors,
                    line=dict(color="white", width=2)
                )
            )
        )

        fig.update_traces(
            hovertemplate="<b>%{label}</b><br>Nombre : %{value}<br>Pourcentage : %{percent}<extra></extra>",
            textfont=dict(size=13, family='Inter, sans-serif')
        )

        fig.update_layout(
            title=dict(text=title, x=0.5, xanchor='center', font=dict(size=16)), 
            height=500, 
            showlegend=True,
            legend=dict(orientation="h", yanchor="bottom", y=-0.15, x=0.5, xanchor="center"),
            margin=dict(t=80, b=80, l=40, r=40)
        )
                          
        st.plotly_chart(fig, use_container_width=True)

    def calc_kpis(df_i, av_i, now_ts, posts):
        res={}; df=df_i.copy(); av=av_i.copy()
        res['dfp']=df
        filt_corr=(df["Nº appel pl.entret."].fillna(0)==0)&(df["Contient SOPL"]==1)
        an=cpiv(df,filt_corr,"Statut OT",posts)
        for c in ["CLOT","CRÉÉ","LANC","TCLO"]: an[c]=an.get(c,0)
        an["OT_CLOTURES"]=an["CLOT"]+an["TCLO"]
        an["TOTAL_OT"]=an[["CLOT","CRÉÉ","LANC","TCLO"]].sum(axis=1)
        an["TAUX_REALISATION_CORRECTIF/PT"]=np.where(an["TOTAL_OT"]==0,100.0,ckpi(an["OT_CLOTURES"],an["TOTAL_OT"]))
        
        pr = cpiv(
            df,
            (df["Statut OT"]=="CRÉÉ") &
            (df["Statut utilisateur"].str.contains(r"\bCRPR\b", case=False, na=False)),
            "ap",
            posts
        )
        for c in ["<1 mois",">3 mois","1 mois < <3 mois","Inconnu"]: pr[c]=pr.get(c,0)
        pr["Total"]=pr[["<1 mois","1 mois < <3 mois",">3 mois","Inconnu"]].sum(axis=1)
        pr["OT préparation <1 mois"]=ckpi(pr["<1 mois"],pr["Total"]); pr["OT préparation >3 mois"]=ckpi(pr[">3 mois"],pr["Total"],0); pr["OT préparation 1mois< <3mois"]=ckpi(pr["1 mois < <3 mois"],pr["Total"],0)
        
        pl=cpiv(df,(df["Statut OT"]=="LANC")&(df["Statut utilisateur"].str.contains("ATPL",case=False,na=False)),"alp",posts)
        for c in ["<1 mois",">3 mois","1 mois < <3 mois","Inconnu"]: pl[c]=pl.get(c,0)
        pl["Total"]=pl[["<1 mois","1 mois < <3 mois",">3 mois","Inconnu"]].sum(axis=1)
        pl["OT planification <1 mois"]=ckpi(pl["<1 mois"],pl["Total"]); pl["OT planification >3 mois"]=ckpi(pl[">3 mois"],pl["Total"],0); pl["OT planification 1mois< <3mois"]=ckpi(pl["1 mois < <3 mois"],pl["Total"],0)
        
        ex=cpiv(df,(df["Statut OT"]=="LANC")&(df["Contient SOPL"]==1),"aex",posts)
        for c in ["<1 mois",">3 mois","1 mois < <3 mois","Inconnu"]: ex[c]=ex.get(c,0)
        ex["Total"]=ex[["<1 mois","1 mois < <3 mois",">3 mois","Inconnu"]].sum(axis=1)
        ex["OT exécution <1 mois"]=ckpi(ex["<1 mois"],ex["Total"]); ex["OT exécution >3 mois"]=ckpi(ex[">3 mois"],ex["Total"],0); ex["OT exécution 1mois< <3mois"]=ckpi(ex["1 mois < <3 mois"],ex["Total"],0)
        
        la=pd.pivot_table(df[df["Statut OT"]=="LANC"],index="Poste travail princ.",columns="OT LANC ESTIME",values="Ordre",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
        for c in ["OUI","NON"]: la[c]=la.get(c,0)
        la["Total"]=la["OUI"]+la["NON"]; la["OT LANC ESTIME"]=ckpi(la["OUI"],la["Total"])
        
        pc=pd.pivot_table(df[df["Statut OT"]=="CRÉÉ"],index="Poste travail princ.",columns="Backlog preparation",values="Ordre",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
        for c in ["CARACTERISE","NON CARACTERISE"]: pc[c]=pc.get(c,0)
        pc["Total"]=pc["CARACTERISE"]+pc["NON CARACTERISE"]; pc["Backlog préparation caractérisé"]=ckpi(pc["CARACTERISE"],pc["Total"])
        
        plc=pd.pivot_table(df[(df["Statut OT"]=="LANC")&(df["Contient SOPL"]==0)],index="Poste travail princ.",columns="Backlog planification",values="Ordre",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
        for c in ["CARACTERISE","NON CARACTERISE"]: plc[c]=plc.get(c,0)
        plc["Total"]=plc["CARACTERISE"]+plc["NON CARACTERISE"]; plc["Backlog planification caractérisé"]=ckpi(plc["CARACTERISE"],plc["Total"])
        
        for kn,cn in [("OT CONFIME","OT CONFIME"),("OT_COR_EGAL","OT_COR_EGAL")]:
            pv=pd.pivot_table(df[df["Statut OT"].isin(["CLOT","TCLO"])],index="Poste travail princ.",columns="OT_COR_EGAL",values="Ordre",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
            for c in ["OUI","NON"]: pv[c]=pv.get(c,0)
            pv["Total"]=pv["OUI"]+pv["NON"]; pv[cn]=ckpi(pv["OUI"],pv["Total"]); res[kn.lower().replace(" ","_")]=pv
            
        avf=av.copy(); res['avf']=avf
        tca=pd.pivot_table(avf,index="Poste travail princ.",columns="Statut utilisateur",values="Avis",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
        for c in ["APRQ","APRV","APRV AVAU","REJT"]: tca[c]=tca.get(c,0)
        tca["Total"]=tca[["APRQ","APRV","APRV AVAU","REJT"]].sum(axis=1); tca["Taux d'approbation des Avis"] = ckpi(tca["APRV"] ,tca["Total"])
        
        g_num=df[(df["Statut OT"].isin(["CLOT","TCLO"]))&(df["_tw_num"]==350)].groupby("Poste travail princ.")["Ordre"].count()
        g_den=df[(df["Contient SOPL"]==1)&(df["_tw_num"]==350)].groupby("Poste travail princ.")["Ordre"].count()
        g_df=pd.DataFrame({"_n":g_num,"_d":g_den}).reindex(posts,fill_value=0)
        g_df["Performance Graissage"]=np.where(g_df["_d"]==0,100.0,(g_df["_n"]/g_df["_d"])*100)
        
        ins_types=[290,300,310]
        ins_base=(df["_tw_num"].isin(ins_types))&(df["Date de début planifiée"].notna())&(df["Date de début planifiée"]<=now_ts)
        ins_num=df[(df["Statut OT"].isin(["CLOT","TCLO"]))&ins_base].groupby("Poste travail princ.")["Ordre"].count()
        ins_den=df[(df["Contient SOPL"]==1)&ins_base].groupby("Poste travail princ.")["Ordre"].count()
        ins_df=pd.DataFrame({"_n":ins_num,"_d":ins_den}).reindex(posts,fill_value=0)
        ins_df["Performance Inspection"]=np.where(ins_df["_d"]==0,100.0,(ins_df["_n"]/ins_df["_d"])*100)
        
        sys_base=(df["_tw_num"]==360)&(df["Date de début planifiée"].notna())&(df["Date de début planifiée"]<=now_ts)
        sys_num=df[(df["Statut OT"].isin(["CLOT","TCLO"]))&sys_base].groupby("Poste travail princ.")["Ordre"].count()
        sys_den=df[(df["Contient SOPL"]==1)&sys_base].groupby("Poste travail princ.")["Ordre"].count()
        sys_df=pd.DataFrame({"_n":sys_num,"_d":sys_den}).reindex(posts,fill_value=0)
        sys_df["Performance Appels Systématiques"]=np.where(sys_df["_d"]==0,100.0,(sys_df["_n"]/sys_df["_d"])*100)
        
        fiab_s=pd.Series(100.0,index=posts); avpan_s=pd.Series(100.0,index=posts)
        res['ckdf']=pd.DataFrame({
            "TAUX_REALISATION_CORRECTIF/PT":an["TAUX_REALISATION_CORRECTIF/PT"],
            "OT préparation <1 mois":pr["OT préparation <1 mois"],"OT préparation >3 mois":pr["OT préparation >3 mois"],"OT préparation 1mois< <3mois":pr["OT préparation 1mois< <3mois"],
            "OT planification <1 mois":pl["OT planification <1 mois"],"OT planification >3 mois":pl["OT planification >3 mois"],"OT planification 1mois< <3mois":pl["OT planification 1mois< <3mois"],
            "OT exécution <1 mois":ex["OT exécution <1 mois"],"OT exécution >3 mois":ex["OT exécution >3 mois"],"OT exécution 1mois< <3mois":ex["OT exécution 1mois< <3mois"],
            "Performance Graissage":g_df["Performance Graissage"],"Performance Inspection":ins_df["Performance Inspection"],"Performance Appels Systématiques":sys_df["Performance Appels Systématiques"],
            "Taux d'approbation des Avis":tca["Taux d'approbation des Avis"],
            "OT LANC ESTIME":la["OT LANC ESTIME"],
            "Backlog préparation caractérisé":pc["Backlog préparation caractérisé"],
            "Backlog planification caractérisé":plc["Backlog planification caractérisé"],
            "OT CONFIME":res.get("ot_confime",pd.Series(100.0,index=posts))["OT CONFIME"] if "ot_confime" in res else pd.Series(100.0,index=posts),
            "OT_COR_EGAL":res.get("ot_cor_egal",pd.Series(100.0,index=posts))["OT_COR_EGAL"] if "ot_cor_egal" in res else pd.Series(100.0,index=posts),
            "OT Fiabilité":fiab_s,
            "Total Avis de Panne":avpan_s
        },index=posts)
        
        res['an']=an; res['pr']=pr; res['pl']=pl; res['ex']=ex
        res['la']=la; res['pc']=pc; res['plc']=plc; res['tca']=tca
        res['g_df']=g_df; res['ins_df']=ins_df; res['sys_df']=sys_df
        res['posts']=posts; res['now_ts']=now_ts
        
        # ============================================================
        # CALCUL DES ANOMALIES AVEC FILTRES PAR TYPE
        # ============================================================
        
        # ANOMALIES PERFORMANCE : filtre sur PK (Detail des indicateurs de Performance)
        ano_p_rows, ano_p_cols, ano_p_df = calculer_anomalies_par_kpi_poste(
            ckdf=res['ckdf'],
            posts=posts,
            kpi_list=PK,           # <-- FILTRE : uniquement les KPI de Performance
            label_section="Performance"
        )
        res['ano_p_rows'] = ano_p_rows
        res['ano_p_cols'] = ano_p_cols
        res['ano_p_df'] = ano_p_df
        
        # ANOMALIES QUALITE : filtre sur QK (Detail des indicateurs de Qualité)
        ano_q_rows, ano_q_cols, ano_q_df = calculer_anomalies_par_kpi_poste(
            ckdf=res['ckdf'],
            posts=posts,
            kpi_list=QK,           # <-- FILTRE : uniquement les KPI de Qualité
            label_section="Qualité"
        )
        res['ano_q_rows'] = ano_q_rows
        res['ano_q_cols'] = ano_q_cols
        res['ano_q_df'] = ano_q_df
        
        # ============================================================
        # CALCUL DES SCORES
        # ============================================================
        perf_kpis = PK
        qual_kpis = QK
        
        res['ckdf']['Score Performance'] = res['ckdf'][perf_kpis].mean(axis=1).round(2)
        res['ckdf']['Score Qualité'] = res['ckdf'][qual_kpis].mean(axis=1).round(2)
        res['ckdf']['Score Global'] = res['ckdf'][ALL_KPI].mean(axis=1).round(2)
        
        return res

    # ============================================================
    # SIDEBAR
    # ============================================================
    with st.sidebar:
        logo_b64 = get_logo_base64()
        if logo_b64:
            st.markdown(f'<img src="data:image/png;base64,{logo_b64}" class="logo" style="width:100%;border-radius:8px;margin-bottom:10px">', unsafe_allow_html=True)
        
        st.markdown("### 📊 Fichiers de données")
        f_ot = st.file_uploader("Fichier OT", type=["xlsx","xls"], key="ot_up")
        f_av = st.file_uploader("Fichier Avis", type=["xlsx","xls"], key="av_up")
        
        st.markdown("---")
        st.markdown("### 🔍 Filtres")
        sel_postes = st.multiselect("Postes de travail", [], key="poste_filter")
        sel_kpi_perf = st.multiselect("KPI Performance", PK, key="kpi_perf_filter")
        sel_kpi_qual = st.multiselect("KPI Qualité", QK, key="kpi_qual_filter")
        
        st.markdown("---")
        if st.button("🔄 Réinitialiser", use_container_width=True):
            st.rerun()

    # ============================================================
    # MAIN CONTENT
    # ============================================================
    if not f_ot or not f_av:
        st.markdown("""
        <div style="display:flex;align-items:center;justify-content:center;min-height:60vh;flex-direction:column">
            <div style="font-size:80px;margin-bottom:20px">📁</div>
            <h2 style="color:#1e3a5f;font-size:28px;font-weight:800">Chargement des données</h2>
            <p style="color:#64748b;font-size:16px;margin-top:10px">Veuillez charger les fichiers OT et Avis dans la barre latérale</p>
        </div>
        """, unsafe_allow_html=True)
        st.stop()

    with st.spinner("Traitement des données..."):
        df, avf, posts, now_ts = prepare_data(f_ot.read(), f_av.read(), fichier_date)
    
    if not posts:
        st.markdown('<div class="es">Aucun poste de travail trouvé dans les données.</div>', unsafe_allow_html=True)
        st.stop()
    
    # Appliquer filtre postes
    if sel_postes:
        posts = [p for p in posts if p in sel_postes]
    
    if not posts:
        st.markdown('<div class="es">Aucun poste sélectionné.</div>', unsafe_allow_html=True)
        st.stop()
    
    res = calc_kpis(df, avf, now_ts, posts)
    ckdf = res['ckdf']
    
    # ============================================================
    # HEADER
    # ============================================================
    header_html = f'<div class="mh"><h1>Dashboard KPI Maintenance</h1><div class="db">📅 {fichier_date}</div></div>'
    st.markdown(header_html, unsafe_allow_html=True)
    
    # ============================================================
    # CARTES RESUME
    # ============================================================
    sp = ckdf['Score Performance'].mean()
    sq = ckdf['Score Qualité'].mean()
    sg = ckdf['Score Global'].mean()
    nb_ano_p = len(res['ano_p_df'])
    nb_ano_q = len(res['ano_q_df'])
    nb_ano_total = nb_ano_p + nb_ano_q
    
    cards_html = f'''
    <div class="cr">
        <div class="cc c1"><div class="cv">{sp:.1f}%</div><div class="cl">Score Performance</div></div>
        <div class="cc c2"><div class="cv">{sq:.1f}%</div><div class="cl">Score Qualité</div></div>
        <div class="cc c3"><div class="cv">{sg:.1f}%</div><div class="cl">Score Global</div></div>
        <div class="cc c4"><div class="cv">{nb_ano_total}</div><div class="cl">Anomalies Totales</div></div>
    </div>
    '''
    st.markdown(cards_html, unsafe_allow_html=True)
    
    # ============================================================
    # ONGLETS
    # ============================================================
    tab1, tab2, tab3, tab4, tab5 = st.tabs(["📊 Synthèse", "📈 Performance", "🔍 Qualité", "⚠️ Anomalies", "📋 Plan d'Action"])
    
    # ============================================================
    # TAB 1 : SYNTHESE
    # ============================================================
    with tab1:
        st.markdown('<div class="stl">Résumé Global par Poste de travail</div>', unsafe_allow_html=True)
        
        synth_cols = ["Poste de travail", "Score Performance", "Score Qualité", "Score Global"]
        synth_html = '<table class="synth-tbl"><thead><tr>' + ''.join(f'<th>{c}</th>' for c in synth_cols) + '</tr></thead><tbody>'
        for poste, row in ckdf.iterrows():
            sp_val = row['Score Performance']
            sq_val = row['Score Qualité']
            sg_val = row['Score Global']
            
            def color_score(v):
                if v >= 90: return "color:#059669;font-weight:700"
                elif v >= 75: return "color:#d97706;font-weight:700"
                else: return "color:#dc2626;font-weight:700"
            
            synth_html += f'<tr><td class="poste-cell">{poste}</td>'
            synth_html += f'<td style="{color_score(sp_val)}">{sp_val:.1f}%</td>'
            synth_html += f'<td style="{color_score(sq_val)}">{sq_val:.1f}%</td>'
            synth_html += f'<td style="{color_score(sg_val)}">{sg_val:.1f}%</td></tr>'
        
        synth_html += '</tbody></table>'
        st.markdown(synth_html, unsafe_allow_html=True)
    
    # ============================================================
    # TAB 2 : PERFORMANCE (avec filtre PK)
    # ============================================================
    with tab2:
        st.markdown('<div class="stl">Detail des indicateurs de Performance</div>', unsafe_allow_html=True)
        
        # Appliquer filtre KPI Performance si sélectionné
        display_perf_kpis = sel_kpi_perf if sel_kpi_perf else PK
        
        perf_df = ckdf[["Poste de travail"] if "Poste de travail" in ckdf.columns else [] + display_perf_kpis].copy()
        if "Poste de travail" not in perf_df.columns:
            perf_df.index.name = "Poste de travail"
            perf_df = perf_df.reset_index()
        
        # Ajouter cibles et scores
        perf_html = '<table class="tw pt"><thead><tr><th>Poste de travail</th>'
        for kpi in display_perf_kpis:
            perf_html += f'<th>{kpi}<br><small>(Cible: {CIBLE.get(kpi,100)}%)</small></th>'
        perf_html += '<th>Score Performance</th></tr></thead><tbody>'
        
        for poste, row in ckdf.iterrows():
            perf_html += f'<tr><td style="font-weight:700">{poste}</td>'
            for kpi in display_perf_kpis:
                val = row.get(kpi, 100)
                cible = CIBLE.get(kpi, 100)
                if kpi in LOWER_BETTER:
                    color = "color:#059669;font-weight:700" if val <= cible else "color:#dc2626;font-weight:700"
                else:
                    color = "color:#059669;font-weight:700" if val >= cible else "color:#dc2626;font-weight:700"
                perf_html += f'<td style="text-align:center;{color}">{val:.1f}%</td>'
            sp_val = row['Score Performance']
            color_sp = "color:#059669;font-weight:800" if sp_val >= 90 else ("color:#d97706;font-weight:800" if sp_val >= 75 else "color:#dc2626;font-weight:800")
            perf_html += f'<td style="text-align:center;{color_sp}">{sp_val:.1f}%</td></tr>'
        
        perf_html += '</tbody></table>'
        st.markdown(perf_html, unsafe_allow_html=True)
    
    # ============================================================
    # TAB 3 : QUALITE (avec filtre QK)
    # ============================================================
    with tab3:
        st.markdown('<div class="stl">Detail des indicateurs de Qualité</div>', unsafe_allow_html=True)
        
        # Appliquer filtre KPI Qualité si sélectionné
        display_qual_kpis = sel_kpi_qual if sel_kpi_qual else QK
        
        qual_html = '<table class="tw qt"><thead><tr><th>Poste de travail</th>'
        for kpi in display_qual_kpis:
            qual_html += f'<th>{kpi}<br><small>(Cible: {CIBLE.get(kpi,100)}%)</small></th>'
        qual_html += '<th>Score Qualité</th></tr></thead><tbody>'
        
        for poste, row in ckdf.iterrows():
            qual_html += f'<tr><td style="font-weight:700">{poste}</td>'
            for kpi in display_qual_kpis:
                val = row.get(kpi, 100)
                cible = CIBLE.get(kpi, 100)
                if kpi in LOWER_BETTER:
                    color = "color:#059669;font-weight:700" if val <= cible else "color:#dc2626;font-weight:700"
                else:
                    color = "color:#059669;font-weight:700" if val >= cible else "color:#dc2626;font-weight:700"
                qual_html += f'<td style="text-align:center;{color}">{val:.1f}%</td>'
            sq_val = row['Score Qualité']
            color_sq = "color:#059669;font-weight:800" if sq_val >= 90 else ("color:#d97706;font-weight:800" if sq_val >= 75 else "color:#dc2626;font-weight:800")
            qual_html += f'<td style="text-align:center;{color_sq}">{sq_val:.1f}%</td></tr>'
        
        qual_html += '</tbody></table>'
        st.markdown(qual_html, unsafe_allow_html=True)
    
    # ============================================================
    # TAB 4 : ANOMALIES (AVEC FILTRES PAR TYPE DE KPI)
    # ============================================================
    with tab4:
        st.markdown('<div class="stl">Nombre d\'anomalies par KPI et Poste (à traiter pour atteindre 100%)</div>', unsafe_allow_html=True)
        
        # Sous-onglets pour séparer Performance et Qualité
        sub_tab_p, sub_tab_q, sub_tab_all = st.tabs(["⚠️ Anomalies Performance (filtre PK)", "⚠️ Anomalies Qualité (filtre QK)", "📊 Résumé Anomalies"])
        
        # --------------------------------------------------------
        # SOUS-TAB : Anomalies Performance (filtre PK)
        # --------------------------------------------------------
        with sub_tab_p:
            ano_p_df = res['ano_p_df']
            nb_ano_p = len(ano_p_df)
            
            st.markdown(f'<div style="background:#fef2f2;border:1px solid #fecaca;border-radius:8px;padding:12px 16px;margin-bottom:12px;display:flex;align-items:center;gap:10px">'
                       f'<span style="font-size:24px">🔴</span>'
                       f'<div><span style="font-weight:800;color:#dc2626;font-size:18px">{nb_ano_p}</span> '
                       f'<span style="color:#64748b;font-size:14px">anomalie(s) de Performance détectée(s) — Filtre appliqué : Detail des indicateurs de Performance (PK)</span></div></div>', 
                       unsafe_allow_html=True)
            
            if nb_ano_p > 0:
                # Pivot : Nombre d'anomalies par KPI et Poste
                ano_pivot_p = ano_p_df.pivot_table(
                    index="Poste de travail", 
                    columns="KPI", 
                    values="Ecart à traiter",
                    aggfunc="count", 
                    fill_value=0
                )
                
                ano_html_p = '<table class="tw at"><thead><tr><th>Poste de travail</th>'
                for kpi in ano_pivot_p.columns:
                    ano_html_p += f'<th>{kpi}</th>'
                ano_html_p += '<th style="background:#7f1d1d">Total Anomalies</th></tr></thead><tbody>'
                
                for poste, row in ano_pivot_p.iterrows():
                    total = row.sum()
                    ano_html_p += f'<tr><td style="font-weight:700">{poste}</td>'
                    for kpi in ano_pivot_p.columns:
                        val = int(row[kpi])
                        bg = "background:#fecaca;color:#991b1b;font-weight:700" if val > 0 else "background:#f8fafc;color:#64748b"
                        ano_html_p += f'<td style="text-align:center;{bg}">{val}</td>'
                    ano_html_p += f'<td style="text-align:center;background:#7f1d1d;color:#fff;font-weight:800">{int(total)}</td></tr>'
                
                # Ligne total
                ano_html_p += '<tr style="background:#dc2626!important"><td style="font-weight:800;color:#fff">Total</td>'
                for kpi in ano_pivot_p.columns:
                    ano_html_p += f'<td style="text-align:center;font-weight:800;color:#fff">{int(ano_pivot_p[kpi].sum())}</td>'
                ano_html_p += f'<td style="text-align:center;font-weight:900;color:#fff;font-size:14px">{int(ano_pivot_p.sum().sum())}</td></tr>'
                
                ano_html_p += '</tbody></table>'
                st.markdown(ano_html_p, unsafe_allow_html=True)
                
                # Détail des anomalies
                with st.expander("📋 Détail complet des anomalies Performance", expanded=False):
                    detail_cols = ["Poste de travail", "KPI", "Valeur actuelle", "Cible", "Ecart à traiter", "Responsable", "Action corrective"]
                    detail_html_p = '<table class="tw st"><thead><tr>' + ''.join(f'<th>{c}</th>' for c in detail_cols) + '</tr></thead><tbody>'
                    for _, r in ano_p_df.iterrows():
                        detail_html_p += '<tr>'
                        for c in detail_cols:
                            detail_html_p += f'<td>{r.get(c,"")}</td>'
                        detail_html_p += '</tr>'
                    detail_html_p += '</tbody></table>'
                    st.markdown(detail_html_p, unsafe_allow_html=True)
            else:
                st.markdown('<div class="es" style="color:#059669">✅ Aucune anomalie de Performance — Tous les KPI Performance sont à leur cible !</div>', unsafe_allow_html=True)
        
        # --------------------------------------------------------
        # SOUS-TAB : Anomalies Qualité (filtre QK)
        # --------------------------------------------------------
        with sub_tab_q:
            ano_q_df = res['ano_q_df']
            nb_ano_q = len(ano_q_df)
            
            st.markdown(f'<div style="background:#fef2f2;border:1px solid #fecaca;border-radius:8px;padding:12px 16px;margin-bottom:12px;display:flex;align-items:center;gap:10px">'
                       f'<span style="font-size:24px">🔴</span>'
                       f'<div><span style="font-weight:800;color:#dc2626;font-size:18px">{nb_ano_q}</span> '
                       f'<span style="color:#64748b;font-size:14px">anomalie(s) de Qualité détectée(s) — Filtre appliqué : Detail des indicateurs de Qualité (QK)</span></div></div>', 
                       unsafe_allow_html=True)
            
            if nb_ano_q > 0:
                # Pivot : Nombre d'anomalies par KPI et Poste
                ano_pivot_q = ano_q_df.pivot_table(
                    index="Poste de travail", 
                    columns="KPI", 
                    values="Ecart à traiter",
                    aggfunc="count", 
                    fill_value=0
                )
                
                ano_html_q = '<table class="tw at"><thead><tr><th>Poste de travail</th>'
                for kpi in ano_pivot_q.columns:
                    ano_html_q += f'<th>{kpi}</th>'
                ano_html_q += '<th style="background:#7f1d1d">Total Anomalies</th></tr></thead><tbody>'
                
                for poste, row in ano_pivot_q.iterrows():
                    total = row.sum()
                    ano_html_q += f'<tr><td style="font-weight:700">{poste}</td>'
                    for kpi in ano_pivot_q.columns:
                        val = int(row[kpi])
                        bg = "background:#fecaca;color:#991b1b;font-weight:700" if val > 0 else "background:#f8fafc;color:#64748b"
                        ano_html_q += f'<td style="text-align:center;{bg}">{val}</td>'
                    ano_html_q += f'<td style="text-align:center;background:#7f1d1d;color:#fff;font-weight:800">{int(total)}</td></tr>'
                
                # Ligne total
                ano_html_q += '<tr style="background:#dc2626!important"><td style="font-weight:800;color:#fff">Total</td>'
                for kpi in ano_pivot_q.columns:
                    ano_html_q += f'<td style="text-align:center;font-weight:800;color:#fff">{int(ano_pivot_q[kpi].sum())}</td>'
                ano_html_q += f'<td style="text-align:center;font-weight:900;color:#fff;font-size:14px">{int(ano_pivot_q.sum().sum())}</td></tr>'
                
                ano_html_q += '</tbody></table>'
                st.markdown(ano_html_q, unsafe_allow_html=True)
                
                # Détail des anomalies
                with st.expander("📋 Détail complet des anomalies Qualité", expanded=False):
                    detail_cols = ["Poste de travail", "KPI", "Valeur actuelle", "Cible", "Ecart à traiter", "Responsable", "Action corrective"]
                    detail_html_q = '<table class="tw st"><thead><tr>' + ''.join(f'<th>{c}</th>' for c in detail_cols) + '</tr></thead><tbody>'
                    for _, r in ano_q_df.iterrows():
                        detail_html_q += '<tr>'
                        for c in detail_cols:
                            detail_html_q += f'<td>{r.get(c,"")}</td>'
                        detail_html_q += '</tr>'
                    detail_html_q += '</tbody></table>'
                    st.markdown(detail_html_q, unsafe_allow_html=True)
            else:
                st.markdown('<div class="es" style="color:#059669">✅ Aucune anomalie de Qualité — Tous les KPI Qualité sont à leur cible !</div>', unsafe_allow_html=True)
        
        # --------------------------------------------------------
        # SOUS-TAB : Résumé Anomalies
        # --------------------------------------------------------
        with sub_tab_all:
            st.markdown('<div class="stl">Répartition des anomalies par type</div>', unsafe_allow_html=True)
            
            # Graphique camembert
            if nb_ano_total > 0:
                fig_ano = go.Figure()
                fig_ano.add_trace(go.Pie(
                    labels=["Anomalies Performance (PK)", "Anomalies Qualité (QK)"],
                    values=[nb_ano_p, nb_ano_q],
                    hole=0.4,
                    marker=dict(colors=["#f59e0b", "#ef4444"], line=dict(color="white", width=3)),
                    textinfo="percent+label+value",
                    texttemplate="%{label}<br>%{value} (%{percent})",
                    textfont=dict(size=14, family="Inter, sans-serif"),
                    textposition="inside"
                ))
                fig_ano.update_layout(
                    title=dict(text="Répartition des Anomalies par Catégorie", x=0.5, xanchor="center", font=dict(size=16)),
                    height=400,
                    margin=dict(t=80, b=40, l=40, r=40),
                    legend=dict(orientation="h", yanchor="bottom", y=-0.1, x=0.5, xanchor="center")
                )
                st.plotly_chart(fig_ano, use_container_width=True)
                
                # Graphique barres par poste
                all_anomalies = pd.concat([
                    res['ano_p_df'].assign(Type="Performance"),
                    res['ano_q_df'].assign(Type="Qualité")
                ], ignore_index=True)
                
                ano_by_poste = all_anomalies.groupby(["Poste de travail", "Type"]).size().unstack(fill_value=0)
                if "Performance" not in ano_by_poste.columns: ano_by_poste["Performance"] = 0
                if "Qualité" not in ano_by_poste.columns: ano_by_poste["Qualité"] = 0
                ano_by_poste["Total"] = ano_by_poste.sum(axis=1)
                ano_by_poste = ano_by_poste.sort_values("Total", ascending=True)
                
                fig_bar = go.Figure()
                fig_bar.add_trace(go.Bar(
                    y=ano_by_poste.index,
                    x=ano_by_poste["Performance"],
                    name="Performance (PK)",
                    orientation="h",
                    marker=dict(color="#f59e0b"),
                    text=ano_by_poste["Performance"],
                    textposition="outside"
                ))
                fig_bar.add_trace(go.Bar(
                    y=ano_by_poste.index,
                    x=ano_by_poste["Qualité"],
                    name="Qualité (QK)",
                    orientation="h",
                    marker=dict(color="#ef4444"),
                    text=ano_by_poste["Qualité"],
                    textposition="outside"
                ))
                fig_bar.update_layout(
                    barmode="stack",
                    title=dict(text="Nombre d'anomalies par Poste et Type", x=0.5, xanchor="center", font=dict(size=16)),
                    height=max(400, len(ano_by_poste) * 35 + 100),
                    margin=dict(t=80, b=40, l=200, r=40),
                    xaxis_title="Nombre d'anomalies",
                    legend=dict(orientation="h", yanchor="bottom", y=-0.2, x=0.5, xanchor="center")
                )
                st.plotly_chart(fig_bar, use_container_width=True)
            else:
                st.markdown('<div class="es" style="color:#059669;font-size:18px">🎉 Aucune anomalie détectée — Tous les indicateurs sont conformes !</div>', unsafe_allow_html=True)
    
    # ============================================================
    # TAB 5 : PLAN D'ACTION
    # ============================================================
    with tab5:
        st.markdown('<div class="stl">Plan d\'Action — Anomalies à traiter pour atteindre 100%</div>', unsafe_allow_html=True)
        
        all_anomalies = pd.concat([
            res['ano_p_df'].assign(Type="Performance"),
            res['ano_q_df'].assign(Type="Qualité")
        ], ignore_index=True)
        
        if len(all_anomalies) > 0:
            # Grouper par Responsable et KPI
            plan = all_anomalies.groupby(["Responsable", "KPI", "Action corrective", "Type"]).agg({
                "Poste de travail": lambda x: ", ".join(sorted(x.unique())),
                "Ecart à traiter": "sum",
                "Valeur actuelle": "mean"
            }).reset_index()
            plan.columns = ["Responsable", "KPI", "Action", "Type", "Postes concernés", "Ecart total", "Valeur moy."]
            plan = plan.sort_values(["Responsable", "Type", "Ecart total"], ascending=[True, True, False])
            
            plan_html = '<table class="plan-action-table"><thead><tr>'
            plan_html += '<th>Responsable</th><th>Type</th><th>KPI</th><th>Action</th><th>Postes concernés</th><th>Ecart total</th><th>Valeur moy.</th>'
            plan_html += '</tr></thead><tbody>'
            
            for _, r in plan.iterrows():
                type_color = "background:#fef3c7;color:#92400e" if r["Type"] == "Performance" else "background:#dbeafe;color:#1e40af"
                plan_html += f'<tr>'
                plan_html += f'<td style="font-weight:800">{r["Responsable"]}</td>'
                plan_html += f'<td style="{type_color};font-weight:700;border-radius:4px">{r["Type"]}</td>'
                plan_html += f'<td style="font-weight:600">{r["KPI"]}</td>'
                plan_html += f'<td style="text-align:left;font-size:11px">{r["Action"]}</td>'
                plan_html += f'<td style="font-size:10px">{r["Postes concernés"]}</td>'
                plan_html += f'<td style="font-weight:800;color:#dc2626">{r["Ecart total"]:.1f}%</td>'
                plan_html += f'<td style="font-weight:700">{r["Valeur moy."]:.1f}%</td>'
                plan_html += '</tr>'
            
            plan_html += '</tbody></table>'
            st.markdown(plan_html, unsafe_allow_html=True)
        else:
            st.markdown('<div class="es" style="color:#059669;font-size:18px">✅ Aucune action requise — Tous les indicateurs sont conformes !</div>', unsafe_allow_html=True)
    
    # ============================================================
    # EXPORT EXCEL
    # ============================================================
    st.markdown("---")
    col_exp1, col_exp2 = st.columns(2)
    
    with col_exp1:
        if st.button("📥 Exporter les KPIs en Excel", use_container_width=True, type="primary"):
            # Préparer les données pour export
            prow_data = []
            for poste, row in ckdf.iterrows():
                r = {"Poste de travail": poste}
                for kpi in PK:
                    r[kpi] = round(row.get(kpi, 100), 2)
                r["Score Performance"] = round(row['Score Performance'], 2)
                prow_data.append(r)
            
            qrow_data = []
            for poste, row in ckdf.iterrows():
                r = {"Poste de travail": poste}
                for kpi in QK:
                    r[kpi] = round(row.get(kpi, 100), 2)
                r["Score Qualité"] = round(row['Score Qualité'], 2)
                qrow_data.append(r)
            
            pcols_list = ["Poste de travail"] + PK + ["Score Performance"]
            qcols_list = ["Poste de travail"] + QK + ["Score Qualité"]
            
            save_kpis_to_excel(
                prow_data, pcols_list,
                qrow_data, qcols_list,
                res['ano_p_rows'], res['ano_p_cols'],
                res['ano_q_rows'], res['ano_q_cols'],
                fichier_date
            )
            st.success("✅ Fichier exporté : kpis/indicateurs_kpis.xlsx")
    
    with col_exp2:
        # Export anomalies CSV
        if st.button("📥 Exporter les Anomalies en CSV", use_container_width=True):
            all_anomalies = pd.concat([
                res['ano_p_df'].assign(Type="Performance"),
                res['ano_q_df'].assign(Type="Qualité")
            ], ignore_index=True)
            
            csv_data = all_anomalies.to_csv(index=False, encoding='utf-8-sig')
            st.download_button(
                label="⬇️ Télécharger anomalies.csv",
                data=csv_data,
                file_name=f"anomalies_{fichier_date.replace('/','_')}.csv",
                mime="text/csv",
                use_container_width=True
            )
    
    # ============================================================
    # FOOTER
    # ============================================================
    st.markdown('<div class="footer">Dashboard KPI Maintenance — Données au ' + fichier_date + ' — Filtres : Anomalies Performance (PK) & Qualité (QK)</div>', unsafe_allow_html=True)


if __name__ == "__main__":
    main()
