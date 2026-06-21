# -*- coding: utf-8 -*-
import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import numpy as np
import io, locale, random, time, os, hashlib, json, base64
from datetime import datetime
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ============================================================
st.set_page_config(layout="wide", page_title="Dashboard KPI", initial_sidebar_state="expanded")
# ============================================================

QK = ["TAUX_REALISATION_CORRECTIF/PT","OT préparation <1 mois","OT préparation >3 mois",
      "OT préparation 1mois< <3mois","OT planification <1 mois","OT planification >3 mois",
      "OT planification 1mois< <3mois","OT exécution <1 mois","OT exécution >3 mois",
      "OT exécution 1mois< <3mois",
      "Performance Graissage","Performance Inspection","Performance Appels Systématiques"]
PK = ["appel avis approuvé","OT LANC ESTIME","Backlog préparation caractérisé",
      "Backlog planification caractérisé","OT CONFIME","OT_COR_EGAL",
      "OT Fiabilité","Total Avis de Panne"]
ALL_KPI = QK + PK

CIBLE = {"TAUX_REALISATION_CORRECTIF/PT":95,"OT préparation <1 mois":80,"OT préparation >3 mois":5,
         "OT préparation 1mois< <3mois":15,"OT planification <1 mois":80,"OT planification >3 mois":5,
         "OT planification 1mois< <3mois":15,"OT exécution <1 mois":80,"OT exécution >3 mois":5,
         "OT exécution 1mois< <3mois":15,"appel avis approuvé":95,"OT LANC ESTIME":100,
         "Backlog préparation caractérisé":100,"Backlog planification caractérisé":100,
         "OT CONFIME":100,"OT_COR_EGAL":95,
         "Performance Graissage":95,"Performance Inspection":95,"Performance Appels Systématiques":85,
         "OT Fiabilité":100,"Total Avis de Panne":100}

ACT_MAP = {"TAUX_REALISATION_CORRECTIF/PT":"Ameliorer le taux de realisation des OT.",
           "OT préparation <1 mois":"Reduire l'age de preparation des OT (< 1 mois).",
           "OT préparation >3 mois":"Traiter les OT avec preparation > 3 mois.",
           "OT planification <1 mois":"Reduire l'age de planification des OT (< 1 mois).",
           "OT planification >3 mois":"Traiter les OT avec planification > 3 mois.",
           "OT exécution <1 mois":"Reduire l'age d'execution des OT (< 1 mois).",
           "OT exécution >3 mois":"Traiter les OT avec execution > 3 mois.",
           "OT LANC ESTIME":"Estimer les couts des OT lances.",
           "Backlog préparation caractérisé":"Caracteriser le backlog de preparation.",
           "Backlog planification caractérisé":"Caracteriser le backlog de planification.",
           "OT CONFIME":"Confirmer les OT termines.",
           "OT_COR_EGAL":"Rapprocher les couts reels et budgetes.",
           "appel avis approuvé":"Creer un OT pour les avis sans ordre.",
           "OT préparation 1mois< <3mois":"Reduire les OT entre 1 et 3 mois.",
           "OT planification 1mois< <3mois":"Reduire les OT entre 1 et 3 mois.",
           "OT exécution 1mois< <3mois":"Reduire les OT entre 1 et 3 mois.",
           "Performance Graissage":"Ameliorer le taux de realisation des OT de graissage (Type 350).",
           "Performance Inspection":"Ameliorer le taux de realisation des OT d'inspection (Types 290,300,310).",
           "Performance Appels Systématiques":"Ameliorer le taux de realisation des appels systematiques (Type 360).",
           "OT Fiabilité":"Maintenir la fiabilite des OT a 100%.",
           "Total Avis de Panne":"Maintenir le suivi des avis de panne a 100%."}

KPI_RESP_MAP = {
    "TAUX_REALISATION_CORRECTIF/PT": "Chef d'atelier",
    "OT préparation <1 mois": "Préparateur BM",
    "OT préparation 1mois< <3mois": "Préparateur BM",
    "OT préparation >3 mois": "Préparateur BM",
    "OT planification <1 mois": "Planificateur BM",
    "OT planification 1mois< <3mois": "Planificateur BM",
    "OT planification >3 mois": "Planificateur BM",
    "OT exécution <1 mois": "Chef d'atelier",
    "OT exécution 1mois< <3mois": "Chef d'atelier",
    "OT exécution >3 mois": "Chef d'atelier",
    "appel avis approuvé": "Chef d'atelier",
    "OT LANC ESTIME": "Fiabilité",
    "Backlog préparation caractérisé": "Préparateur BM",
    "Backlog planification caractérisé": "Planificateur BM",
    "OT CONFIME": "Agent de saisie",
    "OT_COR_EGAL": "Agent de saisie",
    "Performance Graissage": "Chef d'atelier",
    "Performance Inspection": "Chef d'atelier",
    "Performance Appels Systématiques": "Chef d'atelier",
    "OT Fiabilité": "Fiabilité",
    "Total Avis de Panne": "Fiabilité"
}

LOWER_BETTER = ["OT préparation >3 mois","OT planification >3 mois","OT exécution >3 mois",
                "OT préparation 1mois< <3mois","OT planification 1mois< <3mois","OT exécution 1mois< <3mois"]

# ============================================================
# REGLES DE COLORATION : (seuil, operateur)
# ">=" : vert si val >= seuil
# "<=" : vert si val <= seuil
# "==" : vert si val == seuil
# ============================================================
COLOR_RULES = {
    "TAUX_REALISATION_CORRECTIF/PT": (95, ">="),
    "OT préparation <1 mois": (80, ">="),
    "OT préparation >3 mois": (5, "<="),
    "OT préparation 1mois< <3mois": (15, "<="),
    "OT planification <1 mois": (80, ">="),
    "OT planification >3 mois": (5, "<="),
    "OT planification 1mois< <3mois": (15, "<="),
    "OT exécution <1 mois": (80, ">="),
    "OT exécution >3 mois": (5, "<="),
    "OT exécution 1mois< <3mois": (15, "<="),
    "Performance Graissage": (95, ">="),
    "Performance Inspection": (95, ">="),
    "Performance Appels Systématiques": (85, ">="),
    "appel avis approuvé": (95, ">="),
    "OT LANC ESTIME": (100, "=="),
    "Backlog préparation caractérisé": (100, "=="),
    "Backlog planification caractérisé": (100, "=="),
    "OT CONFIME": (100, "=="),
    "OT_COR_EGAL": (95, ">="),
    "OT Fiabilité": (100, "=="),
    "Total Avis de Panne": (100, "=="),
}

MP_KW = ["CRPR ATPD","CRPR ATMR","CRPR ATER","CRPR ATRS","CRPR ATMO","ATPD","ATMR","ATER","ATRS","ATMO"]
MPLAN_KW = ["ATPL ATEI","ATPL ATAL","ATPL ATER","ATPL AGAR","ATPL ATHS","ATEI","ATAL","ATAS","AGAR","ATHS"]

CONSIGNES_HSE = [
    "Port obligatoire des EPI avant toute intervention.","Port obligatoire du casque de securite.",
    "Port obligatoire des lunettes de protection.","Port obligatoire des gants adaptes au travail.",
    "Utiliser les protections auditives dans les zones bruyantes.","Verifier l'absence de tension avant toute intervention electrique.",
    "Respecter la procedure de consignation et deconsignation.","Ne jamais intervenir sur un equipement en marche.",
    "Baliser et securiser la zone de travail.","Maintenir le poste de travail propre et ordonne.",
    "Verifier l'etat des outils avant utilisation.","Utiliser uniquement du materiel homologue.",
    "Respecter les permis de travail en vigueur.","Identifier les risques avant de commencer une tache.",
    "Signaler immediatement toute situation dangereuse.","Signaler tout incident ou presque accident.",
    "Ne jamais neutraliser un dispositif de securite.","Verifier les detecteurs de gaz avant utilisation.",
    "Verifier la bonne ventilation des zones de travail.","Respecter les regles des espaces confines.",
    "Controler l'atmosphere avant d'entrer dans un espace confine.","Utiliser les points d'ancrage pour les travaux en hauteur.",
    "Verifier l'etat des echafaudages avant utilisation.","Securiser les outils lors des travaux en hauteur.",
    "Ne pas travailler seul lors des operations a risque.","Controler les elingues avant chaque levage.",
    "Respecter les limites de charge des equipements.","Verifier l'etat des appareils de levage.",
    "Maintenir les voies de circulation degagees.","Respecter la signalisation de securite.",
    "Verifier les extincteurs a proximite du chantier.","Connaitre les issues de secours les plus proches.",
    "Respecter les procedures d'arret d'urgence.","Verifier les flexibles et raccords avant mise en service.",
    "Controler les fuites avant demarrage d'un equipement.","Respecter les distances de securite.",
    "Ne jamais contourner une procedure HSE.","Porter les EPI adaptes au risque identifie.",
    "Prevenir son responsable avant toute intervention particuliere.","Analyser les risques avant chaque demarrage de chantier.",
    "Verifier la stabilite des equipements.","Utiliser les bons outils pour la bonne tache.",
    "Respecter les consignes specifiques du chantier.","Ne jamais prendre de raccourci au detriment de la securite.",
    "Arreter immediatement les travaux en cas de danger.","Proteger l'environnement lors des interventions.",
    "Collecter et trier correctement les dechets.","Eviter toute pollution accidentelle.",
    "Respecter les consignes de stockage des produits dangereux.","Lire les fiches de securite avant manipulation.",
    "Verifier les equipements avant chaque prise de poste.","S'assurer de la disponibilite des moyens de secours.",
    "Communiquer clairement avec l'equipe avant intervention.","Respecter les regles de circulation des engins.",
    "Garder une vigilance permanente sur son environnement.","Prendre le temps d'effectuer le travail en securite.",
    "La securite est l'affaire de tous.","Chaque incident peut etre evite par la prevention.",
    "Aucun travail n'est plus urgent que la securite.","Zero accident commence par un comportement sur."]

def get_logo_base64():
    for path in ["logo.png", "./logo.png", "../logo.png"]:
        if os.path.exists(path):
            try:
                with open(path, "rb") as f:
                    return base64.b64encode(f.read()).decode()
            except Exception:
                pass
    return None

def get_date_from_file():
    if os.path.exists("date.txt"):
        try:
            with open("date.txt","r",encoding="utf-8") as f: return f.read().strip()
        except Exception: pass
    return "18/06/2026"

def contient_mot(t,lm):
    t=str(t); return any(m in t for l in lm for m in l.split())
def cat_age(a):
    if pd.isna(a): return "Inconnu"
    if a<=1: return "<1 mois"
    elif a>=3: return ">3 mois"
    return "1 mois < <3 mois"

def excr(df):
    if "Poste travail princ." in df.columns:
        return df[~df["Poste travail princ."].astype(str).str.contains("cresseur",case=False,na=False)].copy()
    return df

# ============================================================
# FONCTIONS DE COLORATION
# ============================================================
def is_cell_good(kpi, val):
    """Retourne True si la cellule est VERTE selon la regle."""
    if kpi not in COLOR_RULES: return True
    try:
        v = float(val)
    except (ValueError, TypeError):
        return False
    seuil, op = COLOR_RULES[kpi]
    if op == ">=": return v >= seuil
    elif op == "<=": return v <= seuil
    elif op == "==": return v >= seuil - 0.01  # tolerance pour ==100
    return False

def get_cell_color_style(kpi, val):
    """Retourne le style CSS pour une cellule de valeur."""
    if is_cell_good(kpi, val):
        return "background:#10b981;color:#fff;font-weight:700;"
    else:
        return "background:#ef4444;color:#fff;font-weight:700;"

def get_total_general_style(val):
    """Style pour Total General : >=90 vert, sinon rouge."""
    try:
        v = float(val)
    except (ValueError, TypeError):
        return "background:#ef4444;color:#fff;font-weight:800;"
    if v >= 90:
        return "background:#10b981;color:#fff;font-weight:800;"
    else:
        return "background:#ef4444;color:#fff;font-weight:800;"

def get_bar_color(kpi, val):
    """Couleur pour les barres de progression."""
    try: v = float(val)
    except: return "#cbd5e0"
    if is_cell_good(kpi, v):
        return "#10b981"
    else:
        return "#ef4444"

# ============================================================
@st.cache_data(show_spinner=False)
def prepare_data(ot_bytes, av_bytes, date_str):
    raw_ot = pd.read_excel(io.BytesIO(ot_bytes))
    raw_av = pd.read_excel(io.BytesIO(av_bytes))
    raw_ot = excr(raw_ot)
    raw_av = excr(raw_av)

    for c in ["Créé le","Date de début planifiée","Date de clôture","Début réel","Fin réelle"]:
        if c in raw_ot.columns: raw_ot[c]=pd.to_datetime(raw_ot[c],errors="coerce")
    for c in ["Créé le","Début souhaité","Date de la clôture"]:
        if c in raw_av.columns: raw_av[c]=pd.to_datetime(raw_av[c],errors="coerce")

    now_ts = pd.to_datetime(date_str, format="%d/%m/%Y", errors='coerce')
    if pd.isna(now_ts): now_ts = pd.Timestamp.now()

    df = raw_ot.copy()

    df["Backlog preparation"]=np.where(df["Statut utilisateur"].apply(lambda x:contient_mot(x,MP_KW)),"CARACTERISE","NON CARACTERISE")
    df["Backlog planification"]=np.where(df["Statut utilisateur"].apply(lambda x:contient_mot(x,MPLAN_KW)),"CARACTERISE","NON CARACTERISE")
    df["Type Carac Prep"]=df["Statut utilisateur"].apply(lambda x: next((kw.split()[0] for kw in MP_KW if kw in str(x)), "NON CARACTERISE"))
    df["Type Carac Plan"]=df["Statut utilisateur"].apply(lambda x: next((kw.split()[0] for kw in MPLAN_KW if kw in str(x)), "NON CARACTERISE"))

    for dc,am,ac in [('Créé le',"amp","ap"),('Date de début planifiée',"amlp","alp"),('Date de début planifiée',"amex","aex")]:
        if dc in df.columns:
            df[am]=((now_ts.year-df[dc].dt.year)*12+(now_ts.month-df[dc].dt.month)).round(2)
            df[ac]=df[am].apply(cat_age)
        else: df[am]=np.nan; df[ac]="Inconnu"

    df["OT CONFIME"]=np.where((df["Statut système"].str.contains("CLOT",na=False)|df["Statut système"].str.contains("TCLO",na=False))&~df["Statut système"].str.contains("CONF",na=False),"NON","OUI")

    df["Contient SOPL"]=df["Statut utilisateur"].str.contains("SOPL",na=False).map({True:1,False:0})
    df["OT LANC ESTIME"]=np.where(df["Total coûts budgétés"].fillna(0)==0,"NON","OUI")
    df["OT_COR_EGAL"]=np.where((df["Total coûts budgétés"].fillna(0)-df["Total coûts réels"].fillna(0))==0,"OUI","NON")
    df["_tw_num"]=pd.to_numeric(df.get("Type de travail",pd.Series(dtype=float)),errors="coerce")

    if "Statut système" in df.columns: df["Statut OT"]=df["Statut système"].fillna("").astype(str).str.strip().str.split().str[0]

    avf = raw_av[(raw_av["Ordre"].isna())|(raw_av["Ordre"].astype(str).str.strip()=="")].copy()

    apm = sorted(df[df["Poste travail princ."].astype(str).str.startswith(("SF1","SF2"),na=False)]["Poste travail princ."].dropna().unique().tolist())

    return df, avf, apm, now_ts

def save_kpis_to_excel(prows,pcols,qrows,qcols,ano_p_r,ano_p_c,ano_q_r,ano_q_c,sheet_name):
    kpis_dir="kpis"; os.makedirs(kpis_dir,exist_ok=True)
    filepath=os.path.join(kpis_dir,"indicateurs_kpis.xlsx")
    sn=str(sheet_name).replace("/","-").replace("\\","-").replace("*","").replace("?","").replace("[","").replace("]","")[:31]
    hf=Font(bold=True,color="FFFFFF",size=10); hfl=PatternFill(start_color="1E3A5F",end_color="1E3A5F",fill_type="solid")
    tf=Font(bold=True,size=12,color="1E3A5F")
    tb=Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
    try: wb=load_workbook(filepath)
    except Exception: wb=Workbook()
    if "Sheet" in wb.sheetnames: del wb["Sheet"]
    if sn in wb.sheetnames: del wb[sn]
    ws=wb.create_sheet(sn); rn=1
    def ws_sec(title,cols,rows,sr):
        ws.cell(row=sr,column=1,value=title).font=tf; sr+=1
        for j,c in enumerate(cols,1):
            cl=ws.cell(row=sr,column=j,value=c); cl.font=hf; cl.fill=hfl; cl.alignment=Alignment(horizontal='center'); cl.border=tb
        sr+=1
        for r in rows:
            for j,c in enumerate(cols,1):
                cl=ws.cell(row=sr,column=j,value=r.get(c,"")); cl.border=tb; cl.alignment=Alignment(horizontal='center')
            sr+=1
        return sr+1
    rn=ws_sec("INDICATEURS DE PERFORMANCE",pcols,prows,rn)
    if ano_p_c and ano_p_r: rn=ws_sec("ANOMALIES PERFORMANCE",ano_p_c,ano_p_r,rn)
    rn=ws_sec("INDICATEURS DE QUALITE",qcols,qrows,rn)
    if ano_q_c and ano_q_r: rn=ws_sec("ANOMALIES QUALITE",ano_q_c,ano_q_r,rn)
    try: wb.save(filepath)
    except Exception: pass

def load_historical_kpis(filepath):
    if not os.path.exists(filepath): return pd.DataFrame()
    try: wb=load_workbook(filepath,read_only=True,data_only=True)
    except Exception: return pd.DataFrame()
    records=[]; section=None; headers=None
    for sheet_name in wb.sheetnames:
        try:
            ws=wb[sheet_name]; rows_data=list(ws.iter_rows(values_only=True))
            for row in rows_data:
                cell0=str(row[0]).strip() if row[0] else ""
                if "INDICATEURS DE PERFORMANCE" in cell0.upper(): section="perf"; headers=None; continue
                elif "INDICATEURS DE QUALITE" in cell0.upper(): section="qual"; headers=None; continue
                elif "ANOMALIES" in cell0.upper(): section=None; continue
                if section and headers is None and cell0:
                    headers=[str(c).strip() if c else "" for c in row]; continue
                if section and headers and cell0 and cell0 not in ("Cible","Total general",""):
                    entry={"Date":sheet_name}
                    for j,h in enumerate(headers):
                        if j<len(row): entry[h]=row[j]
                    entry["_section"]=section; records.append(entry)
        except Exception: continue
    wb.close()
    if not records: return pd.DataFrame()
    df=pd.DataFrame(records)
    df["Date_parsed"]=pd.to_datetime(df["Date"].str.replace("-","/"),format="%d/%m/%Y",errors="coerce")
    return df.sort_values("Date_parsed").reset_index(drop=True)

def calculate_variations(hist_df):
    if hist_df.empty or "Date" not in hist_df.columns: return pd.DataFrame()
    dates=sorted(hist_df["Date"].unique())
    if len(dates)<2: return pd.DataFrame()
    perf_df=hist_df[hist_df["_section"]=="perf"].copy()
    qual_df=hist_df[hist_df["_section"]=="qual"].copy()
    variations=[]
    for i in range(1,len(dates)):
        prev_date,curr_date=dates[i-1],dates[i]
        prev_perf=perf_df[perf_df["Date"]==prev_date].set_index("Poste de travail") if "Poste de travail" in perf_df.columns else pd.DataFrame()
        curr_perf=perf_df[perf_df["Date"]==curr_date].set_index("Poste de travail") if "Poste de travail" in perf_df.columns else pd.DataFrame()
        prev_qual=qual_df[qual_df["Date"]==prev_date].set_index("Poste de travail") if "Poste de travail" in qual_df.columns else pd.DataFrame()
        curr_qual=qual_df[qual_df["Date"]==curr_date].set_index("Poste de travail") if "Poste de travail" in qual_df.columns else pd.DataFrame()
        for sec_name,prev_d,curr_d,kpi_list in [("Performance",prev_perf,curr_perf,QK+["Score Performance"]),("Qualite",prev_qual,curr_qual,PK+["Score Qualite"])]:
            for poste in set(prev_d.index)&set(curr_d.index):
                for kpi in kpi_list:
                    if kpi not in prev_d.columns or kpi not in curr_d.columns: continue
                    try: pv=float(prev_d.loc[poste,kpi])
                    except Exception: continue
                    try: cv=float(curr_d.loc[poste,kpi])
                    except Exception: continue
                    diff=cv-pv; pct=(diff/pv*100) if pv!=0 else (100 if cv!=0 else 0)
                    if abs(diff)<=0.5: trend="stabilite"
                    elif diff>0.5: trend="hausse"
                    else: trend="baisse"
                    sens = "Stable"
                    if trend != "stabilite":
                        if (trend == "hausse" and kpi not in LOWER_BETTER) or (trend == "baisse" and kpi in LOWER_BETTER):
                            sens = "Amelioration"
                        else:
                            sens = "Degradation"
                    variations.append({"Date precedente":prev_date,"Date actuelle":curr_date,"Poste":poste,
                        "Type":sec_name,"KPI":kpi,"Valeur precedente":round(pv,2),"Valeur actuelle":round(cv,2),
                        "Ecart":round(diff,2),"Ecart %":round(pct,2),"Tendance":trend, "Sens":sens})
    return pd.DataFrame(variations)

def generate_journal(var_df):
    if var_df.empty: return pd.DataFrame()
    j=var_df.copy(); j["Significatif"]=j["Ecart %"].abs()>=5
    j=j[j["Significatif"]].copy()
    return j.sort_values(["Date actuelle","Ecart %"],ascending=[True,False])

def calculate_rankings(var_df):
    if var_df.empty: return pd.DataFrame(),pd.DataFrame()
    scores={}
    for poste in var_df["Poste"].unique():
        pv=var_df[var_df["Poste"]==poste].copy()
        scores[poste]=sum((-r["Ecart %"] if r["KPI"] in LOWER_BETTER else r["Ecart %"]) for _,r in pv.iterrows())
    ranked=sorted(scores.items(),key=lambda x:x[1],reverse=True)
    return pd.DataFrame(ranked[:5],columns=["Poste","Score variation"]),pd.DataFrame(ranked[-5:][::-1],columns=["Poste","Score variation"])

def inject_custom_css():
    st.markdown("""<style>
    section[data-testid="stSidebar"]{width:250px!important}
    .main .block-container{max-width:100%!important;width:100%!important;padding-left:0.5rem!important;padding-right:0.5rem!important}
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700;800&display=swap');

    :root{
        --primary:#1e3a5f;
        --primary-light:#2c5282;
        --success:#10b981;
        --success-dark:#059669;
        --warning:#f59e0b;
        --warning-dark:#d97706;
        --danger:#ef4444;
        --danger-dark:#dc2626;
        --info:#3b82f6;
        --border:#e2e8f0;
        --radius:10px;
    }

    *{box-sizing:border-box;margin:0;padding:0}
    .stApp{background:#f8fafc;font-family:'Inter',sans-serif}
    .main .block-container{padding-top:.8rem;padding-bottom:.8rem}

    .mh{
        background:linear-gradient(135deg,#1e3a5f 0%,#2563eb 100%);
        padding:16px 24px;
        border-radius:12px;
        margin-bottom:12px;
        box-shadow:0 8px 24px rgba(30,58,95,0.15);
        display:flex;
        align-items:center;
        gap:16px;
    }
    .mh h1{color:#fff;font-size:42px;font-weight:800;margin:0;flex:1}
    .mh .logo{height:50px;width:auto;max-width:150px;object-fit:contain;border-radius:6px}
    .mh .db{
        background:rgba(255,255,255,0.2);
        padding:6px 16px;
        border-radius:16px;
        color:#fff;
        font-size:20px;
        font-weight:600;
        border:1px solid rgba(255,255,255,0.3);
        backdrop-filter:blur(10px);
    }

    .cr{display:grid;grid-template-columns:repeat(4,1fr);gap:8px;margin-bottom:10px}
    .cc{
        background:#fff;
        border-radius:12px;
        padding:18px 16px;
        box-shadow:0 4px 12px rgba(0,0,0,0.06);
        border-left:4px solid;
        transition:transform 0.2s,box-shadow 0.2s;
        text-align:center;
    }
    .cc:hover{transform:translateY(-2px);box-shadow:0 8px 20px rgba(0,0,0,0.1);}
    .cc .cv{font-size:32px;font-weight:900;line-height:1.1;}
    .cc .cl{font-size:16px;color:#1e293b;font-weight:800;text-transform:uppercase;letter-spacing:.5px;margin-top:8px;}
    .cc.c1{border-left-color:#3b82f6}.cc.c1 .cv{color:#2563eb}
    .cc.c2{border-left-color:#10b981}.cc.c2 .cv{color:#059669}
    .cc.c3{border-left-color:#8b5cf6}.cc.c3 .cv{color:#7c3aed}
    .cc.c4{border-left-color:#ef4444}.cc.c4 .cv{color:#dc2626}
    .cc.c5{border-left-color:#3b82f6}.cc.c5 .cv{color:#2563eb}
    .cc.c6{border-left-color:#06b6d4}.cc.c6 .cv{color:#0891b2}
    .cc.c7{border-left-color:#f59e0b}.cc.c7 .cv{color:#d97706}
    .cc.c8{border-left-color:#f97316}.cc.c8 .cv{color:#ea580c}

    .stl{font-size:16px;font-weight:800;color:var(--primary);margin:10px 0 5px 0;padding-left:12px;border-left:4px solid var(--info);}

    .tw{width:100%;border-collapse:collapse;font-family:'Inter',sans-serif;font-size:13px;display:block;overflow-x:auto;-webkit-overflow-scrolling:touch;margin:0}
    .tw thead th{background:var(--primary);color:#fff;font-weight:700;font-size:11px;text-transform:uppercase;letter-spacing:.3px;padding:6px 8px;border:none;white-space:nowrap;position:sticky;top:0;z-index:10}
    .tw.qt thead th{background:linear-gradient(135deg,#2563eb,#3b82f6)}
    .tw.pt thead th{background:linear-gradient(135deg,#059669,#10b981)}
    .tw.at thead th{background:linear-gradient(135deg,#dc2626,#ef4444)}
    .tw.st thead th{background:linear-gradient(135deg,#d97706,#f59e0b)}
    .tw thead th:first-child{z-index:11;left:0}
    .tw tbody td:first-child{position:sticky;left:0;background:#fff;z-index:5;border-right:1px solid var(--border);color:#1e293b !important}
    .tw tbody tr:nth-child(even) td:first-child{background:#f8fafc}
    .tw tbody tr:hover td:first-child{background:#eff6ff}
    .tw tbody td{padding:5px 8px;border-bottom:1px solid var(--border);white-space:nowrap;color:#1e293b !important}
    .tw tbody tr:nth-child(even) td{background:#f8fafc}
    .tw tbody tr:hover td{background:#eff6ff!important}

    .plan-action-table { width:100%; border-collapse:collapse; font-family:Inter,sans-serif; font-size:12px; border:1px solid #cbd5e1; }
    .plan-action-table th { background:#1e3a5f; color:#fff; font-weight:700; padding:8px 6px; border:1px solid #1e3a5f; }
    .plan-action-table td { padding:6px 8px; border:1px solid #cbd5e1; text-align:center; vertical-align:middle;}
    .plan-action-table td:first-child { text-align:left; font-weight:800; }

    .stTabs [data-baseweb="tab-list"]{gap:6px;background:#e2e8f0;padding:6px;border-radius:8px;margin-bottom:8px}
    .stTabs [data-baseweb="tab"]{
        border-radius:6px;
        padding:12px 22px;
        font-weight:700;
        font-size:20px;
        line-height:1.5;
        min-height:48px;
    }
    .stTabs [data-baseweb="tab"] span,
    .stTabs [data-baseweb="tab"] > div{
        font-size:22px !important;
    }
    .stTabs [aria-selected="true"]{
        background:#fff!important;
        color:var(--primary)!important;
        box-shadow:0 3px 8px rgba(0,0,0,.1);
        font-size:21px;
    }
    .stTabs [data-baseweb="tab"] svg{width:22px;height:22px}

    .ca{background:#fff;border-radius:var(--radius);padding:12px;margin-top:6px;border:1px solid var(--border);box-shadow:0 1px 4px rgba(0,0,0,.02)}
    .ca .ct{font-size:14px;font-weight:700;margin-bottom:8px;padding-bottom:5px;border-bottom:1px solid var(--border)}
    .car{display:flex;align-items:center;margin-bottom:6px;font-size:12px}
    .car:last-child{margin-bottom:0}
    .car .cal{width:260px;font-weight:600;color:var(--primary);text-align:right;padding-right:8px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .car .cab{flex:1;height:26px;background:#edf2f7;border-radius:4px;overflow:visible;position:relative}
    .car .caf{height:100%;border-radius:4px;transition:width .3s}
    .car .target-mark{
        position:absolute;
        top:-4px;
        bottom:-4px;
        width:3px;
        background:var(--info) !important;
        z-index:20;
        transform:translateX(-50%);
        box-shadow:0 0 6px rgba(59,130,246,.9),0 0 2px rgba(0,0,0,.4);
        border-radius:2px;
    }
    .car .cav-out{font-size:12px;font-weight:800;color:#1e293b;min-width:55px;text-align:right;padding-left:6px}
    .car .cav-tgt{font-size:10px;font-weight:700;color:#1e293b;min-width:42px;text-align:right;padding-left:4px;opacity:.7}
    .gbr{display:flex;align-items:center;padding:3px 0;font-size:12px;border-bottom:1px solid #f1f5f9}
    .gbr:last-child{border:none}
    .gbr-l{width:160px;font-weight:600;color:#1e293b;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;font-size:11px;padding-right:10px;}
    .gbr-g{display:flex;align-items:center;gap:4px;flex:1;position:relative}
    .gbr-target{position:absolute;left:90%;top:-4px;bottom:-4px;width:3px;background:var(--primary) !important;z-index:10;box-shadow:0 0 6px rgba(30,58,95,.8);border-radius:2px}
    .gbr-target-label{position:absolute;left:90%;top:-20px;transform:translateX(-50%);font-size:9px;font-weight:800;color:#fff;background:var(--primary) !important;padding:1px 5px;border-radius:3px;white-space:nowrap;z-index:11;box-shadow:0 1px 3px rgba(0,0,0,.2)}
    .gbr-w{flex:1;height:22px;background:#f1f5f9;border-radius:3px;overflow:hidden}
    .gbr-f{height:100%;border-radius:3px}
    .gbr-v{font-size:11px;font-weight:800;min-width:48px;text-align:right;color:#1e293b}
    .gbr-legend{display:flex;gap:14px;margin-bottom:10px;font-size:12px;font-weight:700;align-items:center}
    .gbr-legend span{display:flex;align-items:center;gap:5px}
    .gbr-legend i{display:inline-block;width:14px;height:14px;border-radius:2px}
    .gbr-legend .target-icon{display:inline-block;width:3px;height:14px;background:var(--primary) !important;border-radius:1px;box-shadow:0 0 3px rgba(30,58,95,.6)}
    .cg{display:grid;grid-template-columns:1fr 1fr;gap:6px}
    .cg>div{background:#fff;border-radius:var(--radius);padding:8px 10px;border:1px solid var(--border)}
    .cg .ct{font-size:13px;font-weight:700;margin-bottom:4px;padding-bottom:3px;border-bottom:1px solid var(--border)}
    .cgr{display:flex;align-items:center;padding:3px 0;font-size:12px;border-bottom:1px solid #f1f5f9}
    .cgr:last-child{border:none}
    .cgr .rk{width:18px;font-weight:800;text-align:center}
    .cgr .pn{flex:1;font-weight:600;color:#1e293b;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .cgr .ps{font-weight:800;min-width:55px;text-align:right}
    .dgrid{display:grid;grid-template-columns:1fr 1fr;gap:6px}
    .stButton>button[kind="primary"]{background:linear-gradient(135deg,var(--primary),var(--primary-light));border:none;border-radius:6px;padding:8px 14px;font-weight:700;font-size:15px;width:100%}
    ::-webkit-scrollbar{width:5px;height:5px}::-webkit-scrollbar-track{background:#f1f1f1}::-webkit-scrollbar-thumb{background:#cbd5e0;border-radius:3px}

    div[data-testid="stSidebar"]{
        background:linear-gradient(180deg,#1e40af 0%,#1e3a8a 50%,#1e3a5f 100%)!important;
    }
    div[data-testid="stSidebar"]*{color:rgba(255,255,255,.9)!important}
    div[data-testid="stSidebar"] .stSelectbox label,div[data-testid="stSidebar"] .stMultiSelect label,div[data-testid="stSidebar"] .stDateInput label,div[data-testid="stSidebar"] .stCheckbox label,div[data-testid="stSidebar"] .stTextInput label{color:rgba(255,255,255,.9)!important;font-weight:600;font-size:13px;text-transform:uppercase;letter-spacing:.5px}
    div[data-testid="stSidebar"] div[data-testid="stWidget"]{background:rgba(255,255,255,.1);border-radius:6px;padding:5px 10px;margin-bottom:5px;border:1px solid rgba(255,255,255,.15)}
    div[data-testid="stSidebar"] .stSelectbox>div>div,div[data-testid="stSidebar"] .stMultiSelect>div>div,div[data-testid="stSidebar"] .stDateInput>div>div,div[data-testid="stSidebar"] .stTextInput>div>div{background:rgba(255,255,255,.95)!important;border-radius:5px}

    .es{text-align:center;padding:14px;color:#64748b;font-size:14px}
    .synth-tbl{width:100%;border-collapse:collapse;font-family:'Inter',sans-serif;font-size:12px}
    .synth-tbl thead th{background:var(--primary);color:#fff;font-weight:700;font-size:11px;padding:5px 8px;border:none;white-space:nowrap;position:sticky;top:0}
    .synth-tbl tbody td{padding:4px 8px;border-bottom:1px solid var(--border);text-align:center;color:#1e293b !important}
    .synth-tbl tbody tr:nth-child(even) td{background:#f8fafc}
    .synth-tbl tbody tr:hover td{background:#eff6ff!important}
    .synth-tbl .poste-cell{text-align:left;font-weight:700;white-space:nowrap;min-width:140px;color:#1e293b !important}

    [data-testid="stHeaderActionElements"]{display:none !important;}
    [data-testid="stActionButtonContainer"]{display:none !important;}

    .footer {
        text-align: center;
        margin-top: 30px;
        padding: 15px;
        color: #64748b;
        font-size: 13px;
        border-top: 1px solid var(--border);
        font-weight: 600;
    }

    div[data-testid="stDataEditor"] table,
    div[data-testid="stDataEditor"] th,
    div[data-testid="stDataEditor"] td {
        font-size: 18px !important;
        line-height: 1.4 !important;
        white-space: normal !important;
        word-wrap: break-word !important;
    }
    div[data-testid="stDataEditor"] [data-testid="stMarkdownContainer"] {
        font-size: 18px !important;
    }
    div[data-testid="stDataEditor"] [data-testid="stTable"] {
        overflow-x: hidden !important;
        width: 100% !important;
    }

    .anomaly-link{
        color:#2563eb;
        font-weight:800;
        cursor:pointer;
        text-decoration:underline;
        font-size:14px;
    }
    .anomaly-link:hover{
        color:#1d4ed8;
    }

    @media(max-width:768px){
        .cr{grid-template-columns:repeat(2,1fr)}
        .mh{padding:8px 10px;gap:8px}
        .mh h1{font-size:18px}
        .mh .logo{height:35px;max-width:70px}
        .mh .db{font-size:11px;padding:2px 8px}
        .cg,.dgrid{grid-template-columns:1fr}
        .car{flex-wrap:wrap;gap:2px}
        .car .cal{width:100%;text-align:left;padding-right:0;margin-bottom:2px}
        .car .cab{flex:1 1 70%}
        .car .cav-out,.car .cav-tgt{flex:1 1 15%;min-width:40px}
        .gbr{flex-direction:column;align-items:flex-start;gap:4px}
        .gbr-l{width:100%;margin-bottom:2px}
        .gbr-g{width:100%;flex-wrap:wrap}
        .gbr-w{flex:1 1 45%}
        .gbr-v{flex:1 1 10%;min-width:40px}
        .tw{font-size:10px}
        .tw thead th,.tw tbody td{padding:3px 4px}
        .stl{font-size:13px}
        .stTabs [data-baseweb="tab"]{padding:8px 12px;font-size:15px}
        .stTabs [data-baseweb="tab"] span{font-size:16px !important}
        div[data-testid="stDataEditor"] table, div[data-testid="stDataEditor"] th, div[data-testid="stDataEditor"] td { font-size: 14px !important; }
    }

    @media print {
        section[data-testid="stSidebar"],
        header[data-testid="stHeader"],
        div[data-testid="stToolbar"],
        div[data-testid="stHeaderActionElements"],
        footer,
        .stDeployButton,
        #MainMenu { display: none !important; }
        .main .block-container {
            padding-top: 0 !important;
            padding-left: 0 !important;
            padding-right: 0 !important;
            max-width: 100% !important;
        }
        .stButton, .stDownloadButton { display: none !important; }
        * {
            -webkit-print-color-adjust: exact !important;
            print-color-adjust: exact !important;
        }
        .tw, .synth-tbl { page-break-inside: avoid; overflow: visible !important; }
    }
    </style>""",unsafe_allow_html=True)

# ============================================================
# FONCTION PRINCIPALE DE RENDU DU TABLEAU AVEC TOTAL GENERAL
# ============================================================
def render_kpi_table_html(ckdf, kpi_list, table_class, posts, cible_dict, anomalies_raw, df_source, avf_source):
    """
    Genere le HTML du tableau KPI avec :
    - Coloration vert/rouge par cellule selon COLOR_RULES
    - Colonne Total General (ligne) : somme(cells vertes)/count * 100
    - Ligne Total General (colonne) : somme(cells vertes)/count * 100
    - Cellule coin bas-droite : global
    - Coloration Total General : >=90 vert, sinon rouge
    - Ligne Cible
    """
    if ckdf.empty or not posts:
        return '<div class="es">Aucune donnee</div>'

    n_kpi = len(kpi_list)
    n_posts = len(posts)

    # En-tetes
    headers = ["Poste de travail"] + kpi_list + ["Total Général"]

    # Pre-calculer les booleens "bon" pour chaque cellule
    good_matrix = {}  # {poste: {kpi: True/False}}
    for poste in posts:
        if poste not in ckdf.index:
            good_matrix[poste] = {k: False for k in kpi_list}
            continue
        good_matrix[poste] = {}
        for kpi in kpi_list:
            val = ckdf.loc[poste, kpi] if kpi in ckdf.columns else 0
            good_matrix[poste][kpi] = is_cell_good(kpi, val)

    # Calculer Total General par ligne (poste)
    row_totals = {}
    for poste in posts:
        count_good = sum(1 for kpi in kpi_list if good_matrix.get(poste, {}).get(kpi, False))
        row_totals[poste] = round(count_good / n_kpi * 100, 1) if n_kpi > 0 else 0

    # Calculer Total General par colonne (kpi)
    col_totals = {}
    for kpi in kpi_list:
        count_good = sum(1 for poste in posts if good_matrix.get(poste, {}).get(kpi, False))
        col_totals[kpi] = round(count_good / n_posts * 100, 1) if n_posts > 0 else 0

    # Global
    total_good = sum(1 for poste in posts for kpi in kpi_list if good_matrix.get(poste, {}).get(kpi, False))
    total_cells = n_kpi * n_posts
    global_total = round(total_good / total_cells * 100, 1) if total_cells > 0 else 0

    # Construction HTML
    h = '<table class="tw %s"><thead><tr>' % table_class
    for hdr in headers:
        h += '<th>%s</th>' % hdr
    h += '</tr></thead><tbody>'

    # Ligne Cible
    h += '<tr><td style="font-weight:800;background:#fef3c7;color:#92400e;">Cible</td>'
    for kpi in kpi_list:
        cible_val = cible_dict.get(kpi, "")
        if isinstance(cible_val, (int, float)):
            h += '<td style="text-align:center;background:#fef3c7;color:#92400e;font-weight:600;">%s%%</td>' % int(cible_val)
        else:
            h += '<td style="text-align:center;background:#fef3c7;color:#92400e;font-weight:600;">%s</td>' % cible_val
    h += '<td style="text-align:center;background:#fef3c7;color:#92400e;font-weight:700;">90%%</td>'
    h += '</tr>'

    # Lignes par poste
    for poste in posts:
        h += '<tr><td style="font-weight:700">%s</td>' % poste
        for kpi in kpi_list:
            val = ckdf.loc[poste, kpi] if poste in ckdf.index and kpi in ckdf.columns else 0
            style = get_cell_color_style(kpi, val)
            h += '<td style="text-align:center;%s">%.1f%%</td>' % (style, float(val))
        # Total General ligne
        tg = row_totals[poste]
        tg_style = get_total_general_style(tg)
        h += '<td style="text-align:center;%s">%.1f%%</td>' % (tg_style, tg)
        h += '</tr>'

    # Ligne Total General (colonne)
    h += '<tr style="background:#1e3a5f !important;">'
    h += '<td style="font-weight:800;color:#fff;background:#1e3a5f !important;">Total Général</td>'
    for kpi in kpi_list:
        ct = col_totals[kpi]
        ct_style = get_total_general_style(ct)
        h += '<td style="text-align:center;%s">%.1f%%</td>' % (ct_style, ct)
    # Coin bas-droite
    gt_style = get_total_general_style(global_total)
    h += '<td style="text-align:center;%s">%.1f%%</td>' % (gt_style, global_total)
    h += '</tr>'

    h += '</tbody></table>'
    return h


# ============================================================
# FONCTION DE RENDU DU TABLEAU DES ANOMALIES
# ============================================================
def render_anomalies_table_html(kpi_list, posts, ckdf, anomalies_raw, table_class, df_source, avf_source):
    """
    Genere le HTML du tableau des anomalies avec le NOMBRE REEL d'OT/avis a corriger.
    """
    if ckdf.empty or not posts:
        return '<div class="es">Aucune anomalie</div>'

    headers = ["Poste de travail"] + kpi_list + ["Total Anomalies"]

    # Calculer les anomalies par poste et par kpi
    ano_matrix = {}
    total_ano_by_poste = {}
    total_ano_by_kpi = {k: 0 for k in kpi_list}
    grand_total = 0

    for poste in posts:
        ano_matrix[poste] = {}
        poste_total = 0
        for kpi in kpi_list:
            # Recuperer le nombre d'anomalies brutes
            ano_count = 0
            if anomalies_raw and kpi in anomalies_raw:
                ano_s = anomalies_raw[kpi]
                if poste in ano_s.index:
                    ano_count = int(ano_s.loc[poste])
            # Si l'indicateur est bon, anomaly = 0
            if poste in ckdf.index and kpi in ckdf.columns:
                if is_cell_good(kpi, ckdf.loc[poste, kpi]):
                    ano_count = 0
            ano_matrix[poste][kpi] = ano_count
            poste_total += ano_count
            total_ano_by_kpi[kpi] += ano_count
        total_ano_by_poste[poste] = poste_total
        grand_total += poste_total

    h = '<table class="tw %s"><thead><tr>' % table_class
    for hdr in headers:
        h += '<th>%s</th>' % hdr
    h += '</tr></thead><tbody>'

    for poste in posts:
        h += '<tr><td style="font-weight:700">%s</td>' % poste
        for kpi in kpi_list:
            ano = ano_matrix[poste][kpi]
            if ano > 0:
                h += '<td style="text-align:center;background:#fef2f2;color:#dc2626;font-weight:800;">%d</td>' % ano
            else:
                h += '<td style="text-align:center;background:#f0fdf4;color:#16a34a;font-weight:600;">0</td>' % ano
        pt_total = total_ano_by_poste[poste]
        if pt_total > 0:
            h += '<td style="text-align:center;background:#fef2f2;color:#dc2626;font-weight:800;">%d</td>' % pt_total
        else:
            h += '<td style="text-align:center;background:#f0fdf4;color:#16a34a;font-weight:700;">0</td>' % pt_total
        h += '</tr>'

    # Ligne total
    h += '<tr style="background:#1e3a5f !important;">'
    h += '<td style="font-weight:800;color:#fff;background:#1e3a5f !important;">Total</td>'
    for kpi in kpi_list:
        h += '<td style="text-align:center;color:#fff;font-weight:800;background:#1e3a5f !important;">%d</td>' % total_ano_by_kpi[kpi]
    h += '<td style="text-align:center;color:#fff;font-weight:800;background:#1e3a5f !important;">%d</td>' % grand_total
    h += '</tr></tbody></table>'
    return h


# ============================================================
def main():
    try: locale.setlocale(locale.LC_ALL,'fr_FR.UTF-8')
    except Exception:
        try: locale.setlocale(locale.LC_ALL,'fr_FR')
        except Exception: pass
    inject_custom_css()
    fichier_date=get_date_from_file()

    if "hse_affiche" not in st.session_state: st.session_state.hse_affiche=False
    if not st.session_state.hse_affiche:
        c=random.choice(CONSIGNES_HSE)
        st.markdown("""<div style="min-height:100vh;display:flex;flex-direction:column;align-items:center;justify-content:center;background:linear-gradient(135deg,#1a365d,#2d3748,#1a365d);padding:40px">
        <div style="font-size:64px;margin-bottom:20px">🦺</div>
        <h1 style="text-align:center;font-size:46px;color:#fff;font-weight:900;margin:0">HSE - CONSIGNE DE SECURITE</h1>
        <p style="text-align:center;color:rgba(255,255,255,.6);font-size:22px;margin-top:8px;letter-spacing:3px;text-transform:uppercase">Securite - Sante - Environnement</p>
        <div style="background:linear-gradient(135deg,#f6e05e,#ed8936);padding:36px 48px;border-radius:20px;font-size:32px;font-weight:700;text-align:center;margin:40px 0;color:#1a202c;max-width:800px;box-shadow:0 20px 60px rgba(0,0,0,.3)">⚠️ %s</div>
        <h2 style="text-align:center;color:#48bb78;font-size:36px;font-weight:900">Aucun travail n'est plus urgent que la securite</h2>
        <div style="margin-top:40px;width:200px;height:4px;background:rgba(255,255,255,.1);border-radius:2px;overflow:hidden"><div style="width:100%%;height:100%%;background:linear-gradient(90deg,#48bb78,#38a169);border-radius:2px;animation:ld 5.5s ease-in-out forwards"></div></div>
        <style>@keyframes ld{from{width:0}to{width:100%%}}</style></div>"""%c,unsafe_allow_html=True)
        time.sleep(6); st.session_state.hse_affiche=True; st.rerun(); st.stop()

    def ckpi(n,d,sz=100): return np.where(d==0,sz,(n/d)*100)

    def cpiv(df,f,c,p):
        return pd.pivot_table(df[f],index="Poste travail princ.",columns=c,values="Ordre",aggfunc="count",fill_value=0).reindex(p,fill_value=0)

    def build_statut_pivot(df_sub, posts):
        if df_sub.empty:
            return pd.DataFrame(index=posts, columns=["CRÉÉ","LANC","CLOT","TCLO","Total"]).fillna(0).astype(int)
        piv=pd.pivot_table(df_sub, index="Poste travail princ.", columns="Statut OT", values="Ordre", aggfunc="count", fill_value=0)
        for s in ["CRÉÉ","LANC","CLOT","TCLO"]:
            if s not in piv.columns: piv[s]=0
        piv["Total"]=piv[["CRÉÉ","LANC","CLOT","TCLO"]].sum(axis=1)
        return piv.reindex(posts, fill_value=0).fillna(0).astype(int)

    def html_statut_pivot(piv_df, table_class):
        cols=["Poste de travail","CRÉÉ","LANC","CLOT","TCLO","Total"]
        statut_colors = {
            "CRÉÉ": "background:#fef3c7;color:#92400e;font-weight:600;",
            "LANC": "background:#dbeafe;color:#1e40af;font-weight:600;",
            "CLOT": "background:#d1fae5;color:#065f46;font-weight:600;",
            "TCLO": "background:#a7f3d0;color:#064e3b;font-weight:600;",
            "Total": "background:#ede9fe;color:#5b21b6;font-weight:700;"
        }
        h='<table class="tw %s"><thead><tr>'%table_class+''.join('<th>%s</th>'%c for c in cols)+'</tr></thead><tbody>'
        for poste,row in piv_df.iterrows():
            h+='<tr><td style="font-weight:700">%s</td>'%poste
            for c in ["CRÉÉ","LANC","CLOT","TCLO"]:
                h+='<td style="text-align:center;%s">%d</td>'%(statut_colors[c], int(row.get(c,0)))
            h+='<td style="text-align:center;%s">%d</td>'%(statut_colors["Total"], int(row.get("Total",0)))
            h+='</tr>'
        h+='<tr><td style="font-weight:800">Total</td>'
        for c in ["CRÉÉ","LANC","CLOT","TCLO"]:
            h+='<td style="text-align:center;font-weight:800;%s">%d</td>'%(statut_colors[c], int(piv_df[c].sum()))
        h+='<td style="text-align:center;font-weight:800;%s">%d</td>'%(statut_colors["Total"], int(piv_df["Total"].sum()))
        h+='</tr></tbody></table>'
        return h

    def show_pie_pair(piv_df, title_prefix):
        global_counts = piv_df[["CRÉÉ","LANC","CLOT","TCLO"]].sum()
        global_counts = global_counts[global_counts > 0]
        realised = global_counts.get("CLOT", 0) + global_counts.get("TCLO", 0)
        not_realised = global_counts.sum() - realised
        if global_counts.empty:
            st.markdown('<div class="es">Aucune donnee</div>', unsafe_allow_html=True)
            return
        colors = ["#8b5cf6", "#f59e0b", "#10b981", "#3b82f6"]
        fig = make_subplots(rows=1, cols=2, specs=[[{"type":"domain"},{"type":"domain"}]],
                            subplot_titles=(f"{title_prefix} — Par Statut OT", f"{title_prefix} — Réalisés vs Non Réalisés"))
        fig.add_trace(go.Pie(labels=global_counts.index, values=global_counts.values, hole=0.4,
                             textinfo='percent+label',
                             texttemplate='%{label}<br>%{percent:.1%}<br>(%{value})',
                             textposition='inside', insidetextorientation='radial',
                             textfont=dict(size=14, color='white', family='Inter, sans-serif'),
                             marker=dict(colors=colors, line=dict(color='#FFFFFF', width=3))), 1, 1)
        pie2_data = pd.Series([realised, not_realised], index=["Réalisés (CLOT+TCLO)", "Non Réalisés"])
        fig.add_trace(go.Pie(labels=pie2_data.index, values=pie2_data.values, hole=0.5,
                             textinfo='percent+label',
                             texttemplate='%{label}<br>%{percent:.1%}<br>(%{value})',
                             textposition='inside', insidetextorientation='radial',
                             textfont=dict(size=14, color='white', family='Inter, sans-serif'),
                             marker=dict(colors=["#10b981", "#8b5cf6"], line=dict(color='#FFFFFF', width=3))), 1, 2)
        fig.update_layout(margin=dict(t=80, b=20, l=20, r=20), height=450,
                          legend=dict(orientation="h", yanchor="bottom", y=-0.12, x=0.5, xanchor="center"))
        st.plotly_chart(fig, use_container_width=True)

    def show_simple_pie(piv_df, title, keep_non_carac=False):
        if not keep_non_carac and "NON CARACTERISE" in piv_df.columns:
            piv_df = piv_df.drop(columns=["NON CARACTERISE"])
        counts = piv_df.sum()
        counts = counts[counts > 0]
        if counts.empty:
            st.markdown('<div class="es">Aucune donnee</div>', unsafe_allow_html=True)
            return
        color_map = {"CARACTERISE": "#10b981", "NON CARACTERISE": "#f97316"}
        type_palette = ['#3b82f6', '#10b981', '#f59e0b', '#8b5cf6', '#06b6d4', '#14b8a6', '#6366f1', '#0ea5e9', '#d946ef', '#a855f7']
        colors = []
        palette_idx = 0
        for c in counts.index:
            c_str = str(c)
            if c_str in color_map:
                colors.append(color_map[c_str])
            else:
                colors.append(type_palette[palette_idx % len(type_palette)])
                palette_idx += 1
        total_sum = counts.sum()
        pull_list = [0.05 if (v/total_sum)*100 < 10 else 0 for v in counts.values]
        fig = go.Figure(go.Pie(labels=counts.index, values=counts.values, hole=0.4, sort=False,
                               textinfo="percent", textposition="outside", pull=pull_list,
                               marker=dict(colors=colors, line=dict(color="white", width=2))))
        fig.update_traces(hovertemplate="<b>%{label}</b><br>Nombre : %{value}<br>Pourcentage : %{percent}<extra></extra>",
                          textfont=dict(size=13, family='Inter, sans-serif'))
        fig.update_layout(title=dict(text=title, x=0.5, xanchor='center', font=dict(size=16)),
                          height=500, showlegend=True,
                          legend=dict(orientation="h", yanchor="bottom", y=-0.15, x=0.5, xanchor="center"),
                          margin=dict(t=80, b=80, l=40, r=40))
        st.plotly_chart(fig, use_container_width=True)

    # ============================================================
    # CALCUL DES KPIs AVEC ANOMALIES BRUTES
    # ============================================================
    def calc_kpis(df_i, av_i, now_ts, posts):
        res={}; df=df_i.copy(); av=av_i.copy()
        res['dfp']=df

        filt_corr=(df["Nº appel pl.entret."].fillna(0)==0)&(df["Contient SOPL"]==1)
        an=cpiv(df,filt_corr,"Statut OT",posts)
        for c in ["CLOT","CRÉÉ","LANC","TCLO"]: an[c]=an.get(c,0)
        an["OT_CLOTURES"]=an["CLOT"]+an["TCLO"]
        an["TOTAL_OT"]=an[["CLOT","CRÉÉ","LANC","TCLO"]].sum(axis=1)
        an["TAUX_REALISATION_CORRECTIF/PT"]=np.where(an["TOTAL_OT"]==0,100.0,ckpi(an["OT_CLOTURES"],an["TOTAL_OT"]))

        pr=cpiv(df,(df["Statut OT"]=="CRÉÉ")&(df["Statut utilisateur"].str.contains("CRPR",na=False)),"ap",posts)
        for c in ["<1 mois",">3 mois","1 mois < <3 mois","Inconnu"]: pr[c]=pr.get(c,0)
        pr["Total"]=pr[["<1 mois","1 mois < <3 mois",">3 mois","Inconnu"]].sum(axis=1)
        pr["OT préparation <1 mois"]=ckpi(pr["<1 mois"],pr["Total"])
        pr["OT préparation >3 mois"]=ckpi(pr[">3 mois"],pr["Total"],0)
        pr["OT préparation 1mois< <3mois"]=ckpi(pr["1 mois < <3 mois"],pr["Total"],0)

        pl=cpiv(df,(df["Statut OT"]=="LANC")&(df["Statut utilisateur"].str.contains("ATPL",case=False,na=False)),"alp",posts)
        for c in ["<1 mois",">3 mois","1 mois < <3 mois","Inconnu"]: pl[c]=pl.get(c,0)
        pl["Total"]=pl[["<1 mois","1 mois < <3 mois",">3 mois","Inconnu"]].sum(axis=1)
        pl["OT planification <1 mois"]=ckpi(pl["<1 mois"],pl["Total"])
        pl["OT planification >3 mois"]=ckpi(pl[">3 mois"],pl["Total"],0)
        pl["OT planification 1mois< <3mois"]=ckpi(pl["1 mois < <3 mois"],pl["Total"],0)

        ex=cpiv(df,(df["Statut OT"]=="LANC")&(df["Contient SOPL"]==1),"aex",posts)
        for c in ["<1 mois",">3 mois","1 mois < <3 mois","Inconnu"]: ex[c]=ex.get(c,0)
        ex["Total"]=ex[["<1 mois","1 mois < <3 mois",">3 mois","Inconnu"]].sum(axis=1)
        ex["OT exécution <1 mois"]=ckpi(ex["<1 mois"],ex["Total"])
        ex["OT exécution >3 mois"]=ckpi(ex[">3 mois"],ex["Total"],0)
        ex["OT exécution 1mois< <3mois"]=ckpi(ex["1 mois < <3 mois"],ex["Total"],0)

        la=pd.pivot_table(df[df["Statut OT"]=="LANC"],index="Poste travail princ.",columns="OT LANC ESTIME",values="Ordre",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
        for c in ["OUI","NON"]: la[c]=la.get(c,0)
        la["Total"]=la["OUI"]+la["NON"]; la["OT LANC ESTIME"]=ckpi(la["OUI"],la["Total"])

        pc=pd.pivot_table(df[df["Statut OT"]=="CRÉÉ"],index="Poste travail princ.",columns="Backlog preparation",values="Ordre",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
        for c in ["CARACTERISE","NON CARACTERISE"]: pc[c]=pc.get(c,0)
        pc["Total"]=pc["CARACTERISE"]+pc["NON CARACTERISE"]; pc["Backlog préparation caractérisé"]=ckpi(pc["CARACTERISE"],pc["Total"])

        plc=pd.pivot_table(df[df["Statut OT"]=="LANC"],index="Poste travail princ.",columns="Backlog planification",values="Ordre",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
        for c in ["CARACTERISE","NON CARACTERISE"]: plc[c]=plc.get(c,0)
        plc["Total"]=plc["CARACTERISE"]+plc["NON CARACTERISE"]; plc["Backlog planification caractérisé"]=ckpi(plc["CARACTERISE"],plc["Total"])

        for kn,cn in [("OT CONFIME","OT CONFIME"),("OT_COR_EGAL","OT_COR_EGAL")]:
            pv=pd.pivot_table(df,index="Poste travail princ.",columns=cn,values="Ordre",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
            for c in ["OUI","NON"]: pv[c]=pv.get(c,0)
            pv["Total"]=pv["OUI"]+pv["NON"]; pv[cn]=ckpi(pv["OUI"],pv["Total"])
            res[kn.lower().replace(" ","_")]=pv

        avf=av.copy(); res['avf']=avf
        tca=pd.pivot_table(avf,index="Poste travail princ.",columns="Statut utilisateur",values="Avis",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
        for c in ["APRQ","APRV","APRV AVAU","REJT"]: tca[c]=tca.get(c,0)
        tca["Total"]=tca[["APRQ","APRV","APRV AVAU","REJT"]].sum(axis=1); tca["appel avis approuvé"]=ckpi(tca["APRV"],tca["Total"])

        g_num=df[(df["Statut OT"].isin(["CLOT","TCLO"]))&(df["_tw_num"]==350)].groupby("Poste travail princ.")["Ordre"].count()
        g_den=df[(df["Contient SOPL"]==1)&(df["_tw_num"]==350)].groupby("Poste travail princ.")["Ordre"].count()
        g_df=pd.DataFrame({"_n":g_num,"_d":g_den}).reindex(posts,fill_value=0)
        g_df["Performance Graissage"]=np.where(g_df["_d"]==0,100.0,(g_df["_n"]/g_df["_d"])*100)

        ins_types=[290,300,310]
        ins_base=(df["_tw_num"].isin(ins_types))&(df["Date de début planifiée"].notna())&(df["Date de début planifiée"]<=now_ts)
        ins_num=df[(df["Statut OT"].isin(["CLOT","TCLO"]))&ins_base].groupby("Poste travail princ.")["Ordre"].count()
        ins_den=df[(df["Contient SOPL"]==1)&ins_base].groupby("Poste travail princ.")["Ordre"].count()
        ins_df=pd.DataFrame({"_n":ins_num,"_d":ins_den}).reindex(posts,fill_value=0)
        ins_df["Performance Inspection"]=np.where(ins_df["_d"]==0,100.0,(ins_df["_n"]/ins_df["_d"])*100)

        sys_base=(df["_tw_num"]==360)&(df["Date de début planifiée"].notna())&(df["Date de début planifiée"]<=now_ts)
        sys_num=df[(df["Statut OT"].isin(["CLOT","TCLO"]))&sys_base].groupby("Poste travail princ.")["Ordre"].count()
        sys_den=df[(df["Contient SOPL"]==1)&sys_base].groupby("Poste travail princ.")["Ordre"].count()
        sys_df=pd.DataFrame({"_n":sys_num,"_d":sys_den}).reindex(posts,fill_value=0)
        sys_df["Performance Appels Systématiques"]=np.where(sys_df["_d"]==0,100.0,(sys_df["_n"]/sys_df["_d"])*100)

        fiab_s=pd.Series(100.0,index=posts); avpan_s=pd.Series(100.0,index=posts)

        res['ckdf']=pd.DataFrame({
            "TAUX_REALISATION_CORRECTIF/PT":an["TAUX_REALISATION_CORRECTIF/PT"],
            "OT préparation <1 mois":pr["OT préparation <1 mois"],"OT préparation >3 mois":pr["OT préparation >3 mois"],"OT préparation 1mois< <3mois":pr["OT préparation 1mois< <3mois"],
            "OT planification <1 mois":pl["OT planification <1 mois"],"OT planification >3 mois":pl["OT planification >3 mois"],"OT planification 1mois< <3mois":pl["OT planification 1mois< <3mois"],
            "OT exécution <1 mois":ex["OT exécution <1 mois"],"OT exécution >3 mois":ex["OT exécution >3 mois"],"OT exécution 1mois< <3mois":ex["OT exécution 1mois< <3mois"],
            "Performance Graissage":g_df["Performance Graissage"],"Performance Inspection":ins_df["Performance Inspection"],"Performance Appels Systématiques":sys_df["Performance Appels Systématiques"],
            "appel avis approuvé":tca["appel avis approuvé"],"OT LANC ESTIME":la["OT LANC ESTIME"],
            "Backlog préparation caractérisé":pc["Backlog préparation caractérisé"],"Backlog planification caractérisé":plc["Backlog planification caractérisé"],
            "OT CONFIME":res['ot_confime']["OT CONFIME"],"OT_COR_EGAL":res['ot_cor_egal']["OT_COR_EGAL"],
            "OT Fiabilité":fiab_s,"Total Avis de Panne":avpan_s
        })

        # ============================================================
        # ANOMALIES BRUTES : nombre d'OT/avis a corriger pour atteindre 100%
        # ============================================================
        anomalies = {}

        # TAUX_REALISATION_CORRECTIF/PT : OT non clôturés (CRÉÉ + LANC)
        anomalies["TAUX_REALISATION_CORRECTIF/PT"] = (an["CRÉÉ"] + an["LANC"]).reindex(posts, fill_value=0).astype(int)

        # OT préparation <1 mois : OTs pas <1 mois (Total - <1 mois)
        anomalies["OT préparation <1 mois"] = (pr["Total"] - pr["<1 mois"]).reindex(posts, fill_value=0).astype(int)

        # OT préparation >3 mois : OTs >3 mois
        anomalies["OT préparation >3 mois"] = pr[">3 mois"].reindex(posts, fill_value=0).astype(int)

        # OT préparation 1mois< <3mois : OTs entre 1 et 3 mois
        anomalies["OT préparation 1mois< <3mois"] = pr["1 mois < <3 mois"].reindex(posts, fill_value=0).astype(int)

        # OT planification <1 mois
        anomalies["OT planification <1 mois"] = (pl["Total"] - pl["<1 mois"]).reindex(posts, fill_value=0).astype(int)

        # OT planification >3 mois
        anomalies["OT planification >3 mois"] = pl[">3 mois"].reindex(posts, fill_value=0).astype(int)

        # OT planification 1mois< <3mois
        anomalies["OT planification 1mois< <3mois"] = pl["1 mois < <3 mois"].reindex(posts, fill_value=0).astype(int)

        # OT exécution <1 mois
        anomalies["OT exécution <1 mois"] = (ex["Total"] - ex["<1 mois"]).reindex(posts, fill_value=0).astype(int)

        # OT exécution >3 mois
        anomalies["OT exécution >3 mois"] = ex[">3 mois"].reindex(posts, fill_value=0).astype(int)

        # OT exécution 1mois< <3mois
        anomalies["OT exécution 1mois< <3mois"] = ex["1 mois < <3 mois"].reindex(posts, fill_value=0).astype(int)

        # Performance Graissage : OTs graissage non réalisés
        anomalies["Performance Graissage"] = (g_df["_d"] - g_df["_n"]).reindex(posts, fill_value=0).astype(int)

        # Performance Inspection
        anomalies["Performance Inspection"] = (ins_df["_d"] - ins_df["_n"]).reindex(posts, fill_value=0).astype(int)

        # Performance Appels Systématiques
        anomalies["Performance Appels Systématiques"] = (sys_df["_d"] - sys_df["_n"]).reindex(posts, fill_value=0).astype(int)

        # appel avis approuvé : avis non approuvés
        anomalies["appel avis approuvé"] = (tca["Total"] - tca["APRV"]).reindex(posts, fill_value=0).astype(int)

        # OT LANC ESTIME : OTs non estimés
        anomalies["OT LANC ESTIME"] = la["NON"].reindex(posts, fill_value=0).astype(int)

        # Backlog préparation caractérisé : NON CARACTERISE
        anomalies["Backlog préparation caractérisé"] = pc["NON CARACTERISE"].reindex(posts, fill_value=0).astype(int)

        # Backlog planification caractérisé
        anomalies["Backlog planification caractérisé"] = plc["NON CARACTERISE"].reindex(posts, fill_value=0).astype(int)

        # OT CONFIME : NON confirmés
        anomalies["OT CONFIME"] = res['ot_confime']["NON"].reindex(posts, fill_value=0).astype(int)

        # OT_COR_EGAL : NON
        anomalies["OT_COR_EGAL"] = res['ot_cor_egal']["NON"].reindex(posts, fill_value=0).astype(int)

        # OT Fiabilité et Total Avis de Panne : toujours 100
        anomalies["OT Fiabilité"] = pd.Series(0, index=posts).astype(int)
        anomalies["Total Avis de Panne"] = pd.Series(0, index=posts).astype(int)

        res['anomalies'] = anomalies

        # Stocker les dataframes intermédiaires pour le plan d'action
        res['_an'] = an
        res['_pr'] = pr
        res['_pl'] = pl
        res['_ex'] = ex
        res['_la'] = la
        res['_pc'] = pc
        res['_plc'] = plc
        res['_tca'] = tca
        res['_g_df'] = g_df
        res['_ins_df'] = ins_df
        res['_sys_df'] = sys_df

        return res

    # ============================================================
    # PLAN D'ACTION : lister les OTs/avis anomalie pour un poste+kpi
    # ============================================================
    def get_anomaly_details(poste, kpi, res):
        """Retourne un DataFrame des OTs/avis en anomalie pour un poste+kpi donné."""
        df = res['dfp']
        avf = res.get('avf', pd.DataFrame())
        details = pd.DataFrame()

        if kpi == "TAUX_REALISATION_CORRECTIF/PT":
            details = df[(df["Poste travail princ."]==poste)&
                         (df["Nº appel pl.entret."].fillna(0)==0)&
                         (df["Contient SOPL"]==1)&
                         (~df["Statut OT"].isin(["CLOT","TCLO"]))][["Ordre","Statut OT","Désignation" if "Désignation" in df.columns else df.columns[0]]]

        elif kpi == "OT préparation <1 mois":
            details = df[(df["Poste travail princ."]==poste)&
                         (df["Statut OT"]=="CRÉÉ")&
                         (df["Statut utilisateur"].str.contains("CRPR",na=False))&
                         (df["ap"]>1)][["Ordre","ap","Désignation" if "Désignation" in df.columns else df.columns[0]]]

        elif kpi == "OT préparation >3 mois":
            details = df[(df["Poste travail princ."]==poste)&
                         (df["Statut OT"]=="CRÉÉ")&
                         (df["Statut utilisateur"].str.contains("CRPR",na=False))&
                         (df["ap"]>=3)][["Ordre","ap","Désignation" if "Désignation" in df.columns else df.columns[0]]]

        elif kpi == "OT préparation 1mois< <3mois":
            details = df[(df["Poste travail princ."]==poste)&
                         (df["Statut OT"]=="CRÉÉ")&
                         (df["Statut utilisateur"].str.contains("CRPR",na=False))&
                         (df["ap"]>1)&(df["ap"]<3)][["Ordre","ap","Désignation" if "Désignation" in df.columns else df.columns[0]]]

        elif kpi == "OT planification <1 mois":
            details = df[(df["Poste travail princ."]==poste)&
                         (df["Statut OT"]=="LANC")&
                         (df["Statut utilisateur"].str.contains("ATPL",case=False,na=False))&
                         (df["alp"]>1)][["Ordre","alp","Désignation" if "Désignation" in df.columns else df.columns[0]]]

        elif kpi == "OT planification >3 mois":
            details = df[(df["Poste travail princ."]==poste)&
                         (df["Statut OT"]=="LANC")&
                         (df["Statut utilisateur"].str.contains("ATPL",case=False,na=False))&
                         (df["alp"]>=3)][["Ordre","alp","Désignation" if "Désignation" in df.columns else df.columns[0]]]

        elif kpi == "OT planification 1mois< <3mois":
            details = df[(df["Poste travail princ."]==poste)&
                         (df["Statut OT"]=="LANC")&
                         (df["Statut utilisateur"].str.contains("ATPL",case=False,na=False))&
                         (df["alp"]>1)&(df["alp"]<3)][["Ordre","alp","Désignation" if "Désignation" in df.columns else df.columns[0]]]

        elif kpi == "OT exécution <1 mois":
            details = df[(df["Poste travail princ."]==poste)&
                         (df["Statut OT"]=="LANC")&
                         (df["Contient SOPL"]==1)&
                         (df["aex"]>1)][["Ordre","aex","Désignation" if "Désignation" in df.columns else df.columns[0]]]

        elif kpi == "OT exécution >3 mois":
            details = df[(df["Poste travail princ."]==poste)&
                         (df["Statut OT"]=="LANC")&
                         (df["Contient SOPL"]==1)&
                         (df["aex"]>=3)][["Ordre","aex","Désignation" if "Désignation" in df.columns else df.columns[0]]]

        elif kpi == "OT exécution 1mois< <3mois":
            details = df[(df["Poste travail princ."]==poste)&
                         (df["Statut OT"]=="LANC")&
                         (df["Contient SOPL"]==1)&
                         (df["aex"]>1)&(df["aex"]<3)][["Ordre","aex","Désignation" if "Désignation" in df.columns else df.columns[0]]]

        elif kpi == "Performance Graissage":
            details = df[(df["Poste travail princ."]==poste)&
                         (df["_tw_num"]==350)&
                         (df["Contient SOPL"]==1)&
                         (~df["Statut OT"].isin(["CLOT","TCLO"]))][["Ordre","Statut OT","Désignation" if "Désignation" in df.columns else df.columns[0]]]

        elif kpi == "Performance Inspection":
            details = df[(df["Poste travail princ."]==poste)&
                         (df["_tw_num"].isin([290,300,310]))&
                         (df["Contient SOPL"]==1)&
                         (~df["Statut OT"].isin(["CLOT","TCLO"]))][["Ordre","Statut OT","Désignation" if "Désignation" in df.columns else df.columns[0]]]

        elif kpi == "Performance Appels Systématiques":
            details = df[(df["Poste travail princ."]==poste)&
                         (df["_tw_num"]==360)&
                         (df["Contient SOPL"]==1)&
                         (~df["Statut OT"].isin(["CLOT","TCLO"]))][["Ordre","Statut OT","Désignation" if "Désignation" in df.columns else df.columns[0]]]

        elif kpi == "appel avis approuvé":
            if not avf.empty and "Poste travail princ." in avf.columns:
                details = avf[(avf["Poste travail princ."]==poste)&
                              (~avf["Statut utilisateur"].str.contains("APRV",na=False))][["Avis","Statut utilisateur","Désignation" if "Désignation" in avf.columns else avf.columns[0]]]

        elif kpi == "OT LANC ESTIME":
            details = df[(df["Poste travail princ."]==poste)&
                         (df["Statut OT"]=="LANC")&
                         (df["OT LANC ESTIME"]=="NON")][["Ordre","Statut OT","Désignation" if "Désignation" in df.columns else df.columns[0]]]

        elif kpi == "Backlog préparation caractérisé":
            details = df[(df["Poste travail princ."]==poste)&
                         (df["Statut OT"]=="CRÉÉ")&
                         (df["Backlog preparation"]=="NON CARACTERISE")][["Ordre","Statut utilisateur","Désignation" if "Désignation" in df.columns else df.columns[0]]]

        elif kpi == "Backlog planification caractérisé":
            details = df[(df["Poste travail princ."]==poste)&
                         (df["Statut OT"]=="LANC")&
                         (df["Backlog planification"]=="NON CARACTERISE")][["Ordre","Statut utilisateur","Désignation" if "Désignation" in df.columns else df.columns[0]]]

        elif kpi == "OT CONFIME":
            details = df[(df["Poste travail princ."]==poste)&
                         (df["OT CONFIME"]=="NON")][["Ordre","Statut système","Désignation" if "Désignation" in df.columns else df.columns[0]]]

        elif kpi == "OT_COR_EGAL":
            details = df[(df["Poste travail princ."]==poste)&
                         (df["OT_COR_EGAL"]=="NON")][["Ordre","Total coûts budgétés","Total coûts réels","Désignation" if "Désignation" in df.columns else df.columns[0]]]

        return details

    # ============================================================
    # BARRES DE PROGRESSION PAR POSTE
    # ============================================================
    def render_bars_by_poste(ckdf, kpi_list, posts, cible_dict, title):
        st.markdown('<div class="stl">%s</div>' % title, unsafe_allow_html=True)
        for poste in posts:
            if poste not in ckdf.index: continue
            st.markdown('<div class="ca"><div class="ct">📌 %s</div>' % poste, unsafe_allow_html=True)
            for kpi in kpi_list:
                val = float(ckdf.loc[poste, kpi]) if kpi in ckdf.columns else 0
                cible = cible_dict.get(kpi, 100)
                color = get_bar_color(kpi, val)
                pct = min(val, 120) / 120 * 100
                st.markdown('''<div class="car">
                    <div class="cal">%s</div>
                    <div class="cab">
                        <div class="caf" style="width:%.1f%%;background:%s;"></div>
                        <div class="target-mark" style="left:%.1f%%;"></div>
                    </div>
                    <div class="cav-out">%.1f%%</div>
                    <div class="cav-tgt">Cible: %s%%</div>
                </div>''' % (kpi, pct, color, min(cible, 120)/120*100, val, int(cible)), unsafe_allow_html=True)
            st.markdown('</div>', unsafe_allow_html=True)

    # ============================================================
    # BARRES GLOBALES PAR KPI
    # ============================================================
    def render_global_bars(ckdf, kpi_list, posts, cible_dict, title):
        st.markdown('<div class="stl">%s</div>' % title, unsafe_allow_html=True)
        st.markdown('<div class="gbr-legend"><span><i style="background:#10b981;"></i> Atteint</span><span><i style="background:#ef4444;"></i> Non atteint</span><span><span class="target-icon"></span> Cible</span></div>', unsafe_allow_html=True)
        for kpi in kpi_list:
            if kpi not in ckdf.columns: continue
            avg_val = ckdf[kpi].mean()
            cible = cible_dict.get(kpi, 100)
            color = get_bar_color(kpi, avg_val)
            pct = min(avg_val, 120) / 120 * 100
            st.markdown('''<div class="gbr">
                <div class="gbr-l">%s</div>
                <div class="gbr-g">
                    <div class="gbr-w"><div class="gbr-f" style="width:%.1f%%;background:%s;"></div></div>
                    <div class="gbr-target" style="left:%.1f%%;"></div>
                    <div class="gbr-target-label" style="left:%.1f%%;">%s%%</div>
                </div>
                <div class="gbr-v" style="color:%s;">%.1f%%</div>
            </div>''' % (kpi, pct, color, min(cible,120)/120*100, min(cible,120)/120*100, int(cible), color, avg_val), unsafe_allow_html=True)

    # ============================================================
    # INTERFACE PRINCIPALE
    # ============================================================
    with st.sidebar:
        logo_b64 = get_logo_base64()
        if logo_b64:
            st.markdown('<img src="data:image/png;base64,%s" style="width:100%%;border-radius:8px;margin-bottom:10px;">' % logo_b64, unsafe_allow_html=True)

        ot_file = st.file_uploader("Fichier OT (Excel)", type=["xlsx","xls"], key="ot_up")
        av_file = st.file_uploader("Fichier Avis (Excel)", type=["xlsx","xls"], key="av_up")

        st.markdown("---")
        date_input = st.text_input("Date de référence", value=fichier_date, key="date_in")
        st.markdown("---")

        if ot_file and av_file:
            try:
                df_ot, df_av, apm, now_ts = prepare_data(ot_file.getvalue(), av_file.getvalue(), date_input)
                st.success(f"Données chargées : {len(df_ot)} OTs, {len(df_av)} Avis")
                st.info(f"Date : {date_input}")
                st.info(f"{len(apm)} postes de travail détectés")
            except Exception as e:
                st.error(f"Erreur chargement : {e}")
                st.stop()
        else:
            st.warning("Veuillez charger les deux fichiers Excel.")
            st.stop()

    # Calculs
    with st.spinner("Calcul des KPIs..."):
        res = calc_kpis(df_ot, df_av, now_ts, apm)
    ckdf = res['ckdf']
    anomalies = res['anomalies']
    df_full = res['dfp']
    avf_full = res.get('avf', pd.DataFrame())

    # Header
    header_html = '<div class="mh"><h1>📊 Dashboard KPI Maintenance</h1>'
    header_html += '<div class="db">📅 %s</div>' % date_input
    header_html += '</div>'
    st.markdown(header_html, unsafe_allow_html=True)

    # Cartes résumé
    n_posts = len(apm)
    n_qk_good = sum(1 for k in QK for p in apm if p in ckdf.index and is_cell_good(k, ckdf.loc[p, k]))
    n_qk_total = len(QK) * n_posts
    score_q = round(n_qk_good / n_qk_total * 100, 1) if n_qk_total > 0 else 0

    n_pk_good = sum(1 for k in PK for p in apm if p in ckdf.index and is_cell_good(k, ckdf.loc[p, k]))
    n_pk_total = len(PK) * n_posts
    score_p = round(n_pk_good / n_pk_total * 100, 1) if n_pk_total > 0 else 0

    total_anomalies = sum(int(anomalies[k].sum()) for k in ALL_KPI if k in anomalies)
    total_good_all = n_qk_good + n_pk_good
    total_all = n_qk_total + n_pk_total
    score_global = round(total_good_all / total_all * 100, 1) if total_all > 0 else 0

    cr_html = '<div class="cr">'
    cr_html += '<div class="cc c1"><div class="cv">%d</div><div class="cl">Postes de travail</div></div>'
    cr_html += '<div class="cc c2"><div class="cv" style="color:%s;">%.1f%%</div><div class="cl">Score Performance</div></div>' % ("#059669" if score_q >= 90 else "#dc2626", score_q)
    cr_html += '<div class="cc c3"><div class="cv" style="color:%s;">%.1f%%</div><div class="cl">Score Qualité</div></div>' % ("#059669" if score_p >= 90 else "#dc2626", score_p)
    cr_html += '<div class="cc c4"><div class="cv">%d</div><div class="cl">Anomalies totales</div></div>' % total_anomalies
    cr_html += '<div class="cc c5"><div class="cv" style="color:%s;">%.1f%%</div><div class="cl">Score Global</div></div>' % ("#059669" if score_global >= 90 else "#dc2626", score_global)
    cr_html += '</div>'
    st.markdown(cr_html % n_posts, unsafe_allow_html=True)

    # ============================================================
    # ONGLETS
    # ============================================================
    tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
        "📊 Indicateurs Performance",
        "📋 Indicateurs Qualité",
        "🔍 Anomalies Performance",
        "🔍 Anomalies Qualité",
        "📝 Plan d'Action",
        "📈 Graphiques"
    ])

    # ============================================================
    # TAB 1 : PERFORMANCE
    # ============================================================
    with tab1:
        st.markdown(render_kpi_table_html(ckdf, QK, "pt", apm, CIBLE, anomalies, df_full, avf_full), unsafe_allow_html=True)
        st.markdown("---")
        render_bars_by_poste(ckdf, QK, apm, CIBLE, "Détail par Poste — Performance")
        st.markdown("---")
        render_global_bars(ckdf, QK, apm, CIBLE, "Vue Globale — Performance")

    # ============================================================
    # TAB 2 : QUALITE
    # ============================================================
    with tab2:
        st.markdown(render_kpi_table_html(ckdf, PK, "qt", apm, CIBLE, anomalies, df_full, avf_full), unsafe_allow_html=True)
        st.markdown("---")
        render_bars_by_poste(ckdf, PK, apm, CIBLE, "Détail par Poste — Qualité")
        st.markdown("---")
        render_global_bars(ckdf, PK, apm, CIBLE, "Vue Globale — Qualité")

    # ============================================================
    # TAB 3 : ANOMALIES PERFORMANCE
    # ============================================================
    with tab3:
        st.markdown(render_anomalies_table_html(QK, apm, ckdf, anomalies, "at", df_full, avf_full), unsafe_allow_html=True)
        st.markdown("---")
        st.markdown("### 🔎 Détail des anomalies par Poste et Indicateur")
        for poste in apm:
            poste_anomalies = []
            for kpi in QK:
                ano_count = 0
                if kpi in anomalies and poste in anomalies[kpi].index:
                    ano_count = int(anomalies[kpi].loc[poste])
                if poste in ckdf.index and kpi in ckdf.columns:
                    if is_cell_good(kpi, ckdf.loc[poste, kpi]):
                        ano_count = 0
                if ano_count > 0:
                    poste_anomalies.append((kpi, ano_count))
            if poste_anomalies:
                with st.expander(f"📌 {poste} — {len(poste_anomalies)} anomalie(s)"):
                    for kpi, cnt in poste_anomalies:
                        st.markdown(f"**{kpi}** : <span style='color:#dc2626;font-weight:800;'>{cnt} OT/avis à corriger</span>", unsafe_allow_html=True)
                        details = get_anomaly_details(poste, kpi, res)
                        if not details.empty:
                            st.dataframe(details, use_container_width=True, height=min(200, max(50, len(details)*35)))
                        else:
                            st.markdown('<span style="color:#64748b;">Aucun détail disponible</span>', unsafe_allow_html=True)

    # ============================================================
    # TAB 4 : ANOMALIES QUALITE
    # ============================================================
    with tab4:
        st.markdown(render_anomalies_table_html(PK, apm, ckdf, anomalies, "at", df_full, avf_full), unsafe_allow_html=True)
        st.markdown("---")
        st.markdown("### 🔎 Détail des anomalies par Poste et Indicateur")
        for poste in apm:
            poste_anomalies = []
            for kpi in PK:
                ano_count = 0
                if kpi in anomalies and poste in anomalies[kpi].index:
                    ano_count = int(anomalies[kpi].loc[poste])
                if poste in ckdf.index and kpi in ckdf.columns:
                    if is_cell_good(kpi, ckdf.loc[poste, kpi]):
                        ano_count = 0
                if ano_count > 0:
                    poste_anomalies.append((kpi, ano_count))
            if poste_anomalies:
                with st.expander(f"📌 {poste} — {len(poste_anomalies)} anomalie(s)"):
                    for kpi, cnt in poste_anomalies:
                        st.markdown(f"**{kpi}** : <span style='color:#dc2626;font-weight:800;'>{cnt} OT/avis à corriger</span>", unsafe_allow_html=True)
                        details = get_anomaly_details(poste, kpi, res)
                        if not details.empty:
                            st.dataframe(details, use_container_width=True, height=min(200, max(50, len(details)*35)))
                        else:
                            st.markdown('<span style="color:#64748b;">Aucun détail disponible</span>', unsafe_allow_html=True)

    # ============================================================
    # TAB 5 : PLAN D'ACTION
    # ============================================================
    with tab5:
        st.markdown('<div class="stl">📝 Plan d\'Action Correctif</div>', unsafe_allow_html=True)

        # Construire le plan d'action complet
        plan_rows = []
        for poste in apm:
            for kpi in ALL_KPI:
                ano_count = 0
                if kpi in anomalies and poste in anomalies[kpi].index:
                    ano_count = int(anomalies[kpi].loc[poste])
                if poste in ckdf.index and kpi in ckdf.columns:
                    if is_cell_good(kpi, ckdf.loc[poste, kpi]):
                        ano_count = 0
                if ano_count > 0:
                    val = float(ckdf.loc[poste, kpi]) if poste in ckdf.index and kpi in ckdf.columns else 0
                    plan_rows.append({
                        "Poste": poste,
                        "Indicateur": kpi,
                        "Valeur actuelle": round(val, 1),
                        "Cible": CIBLE.get(kpi, 100),
                        "Nb anomalies": ano_count,
                        "Responsable": KPI_RESP_MAP.get(kpi, ""),
                        "Action": ACT_MAP.get(kpi, "")
                    })

        if plan_rows:
            plan_df = pd.DataFrame(plan_rows)
            plan_df = plan_df.sort_values(["Poste", "Indicateur"]).reset_index(drop=True)

            # Afficher le tableau avec style
            plan_html = '<table class="plan-action-table"><thead><tr>'
            for col in plan_df.columns:
                plan_html += '<th>%s</th>' % col
            plan_html += '</tr></thead><tbody>'
            for _, row in plan_df.iterrows():
                plan_html += '<tr>'
                for col in plan_df.columns:
                    val = row[col]
                    if col == "Nb anomalies":
                        plan_html += '<td style="background:#fef2f2;color:#dc2626;font-weight:800;">%d</td>' % int(val)
                    elif col == "Valeur actuelle":
                        plan_html += '<td>%.1f%%</td>' % float(val)
                    elif col == "Cible":
                        plan_html += '<td>%s%%</td>' % int(val)
                    elif col == "Action":
                        plan_html += '<td style="text-align:left;font-size:11px;max-width:300px;">%s</td>' % val
                    else:
                        plan_html += '<td>%s</td>' % val
                plan_html += '</tr>'
            plan_html += '</tbody></table>'
            st.markdown(plan_html, unsafe_allow_html=True)

            # Détail expansible par ligne du plan d'action
            st.markdown("---")
            st.markdown("### 📋 Détail des OT/avis à corriger")
            for _, prow in plan_df.iterrows():
                poste = prow["Poste"]
                kpi = prow["Indicateur"]
                ano = int(prow["Nb anomalies"])
                with st.expander(f"📌 {poste} — {kpi} ({ano} OT/avis)"):
                    details = get_anomaly_details(poste, kpi, res)
                    if not details.empty:
                        st.dataframe(details, use_container_width=True, height=min(300, max(80, len(details)*35)))
                    else:
                        st.markdown('<span style="color:#64748b;">Aucun détail disponible pour ces anomalies</span>', unsafe_allow_html=True)
        else:
            st.markdown('<div class="es" style="background:#f0fdf4;color:#16a34a;font-size:18px;padding:30px;border-radius:12px;border:2px solid #10b981;">✅ Aucune anomalie détectée — Tous les indicateurs sont atteints !</div>', unsafe_allow_html=True)

    # ============================================================
    # TAB 6 : GRAPHIQUES
    # ============================================================
    with tab6:
        sub_tab1, sub_tab2, sub_tab3 = st.tabs(["Distribution Statuts", "Backlog", "Performance"])

        with sub_tab1:
            col_g1, col_g2 = st.columns(2)
            with col_g1:
                st.markdown("#### OT Correctifs par Poste")
                corr_df = df_full[(df_full["Nº appel pl.entret."].fillna(0)==0)&(df_full["Contient SOPL"]==1)]
                piv_corr = build_statut_pivot(corr_df, apm)
                show_pie_pair(piv_corr, "OT Correctifs")
                st.markdown(html_statut_pivot(piv_corr, "pt"), unsafe_allow_html=True)

            with col_g2:
                st.markdown("#### Tous les OT par Poste")
                piv_all = build_statut_pivot(df_full, apm)
                show_pie_pair(piv_all, "Tous les OT")
                st.markdown(html_statut_pivot(piv_all, "qt"), unsafe_allow_html=True)

        with sub_tab2:
            col_b1, col_b2 = st.columns(2)
            with col_b1:
                st.markdown("#### Backlog Préparation")
                bp = pd.pivot_table(df_full[df_full["Statut OT"]=="CRÉÉ"], index="Poste travail princ.", columns="Backlog preparation", values="Ordre", aggfunc="count", fill_value=0).reindex(apm, fill_value=0)
                for c in ["CARACTERISE","NON CARACTERISE"]:
                    if c not in bp.columns: bp[c]=0
                show_simple_pie(bp, "Backlog Préparation")

            with col_b2:
                st.markdown("#### Backlog Planification")
                blp = pd.pivot_table(df_full[df_full["Statut OT"]=="LANC"], index="Poste travail princ.", columns="Backlog planification", values="Ordre", aggfunc="count", fill_value=0).reindex(apm, fill_value=0)
                for c in ["CARACTERISE","NON CARACTERISE"]:
                    if c not in blp.columns: blp[c]=0
                show_simple_pie(blp, "Backlog Planification")

            st.markdown("---")
            st.markdown("#### Types de Caractérisation Préparation")
            tp = pd.pivot_table(df_full[(df_full["Statut OT"]=="CRÉÉ")&(df_full["Statut utilisateur"].str.contains("CRPR",na=False))],
                                index="Poste travail princ.", columns="Type Carac Prep", values="Ordre", aggfunc="count", fill_value=0).reindex(apm, fill_value=0)
            if "NON CARACTERISE" in tp.columns: tp = tp.drop(columns=["NON CARACTERISE"])
            show_simple_pie(tp, "Types Caractérisation Préparation")

            st.markdown("#### Types de Caractérisation Planification")
            tl = pd.pivot_table(df_full[(df_full["Statut OT"]=="LANC")&(df_full["Statut utilisateur"].str.contains("ATPL",case=False,na=False))],
                                index="Poste travail princ.", columns="Type Carac Plan", values="Ordre", aggfunc="count", fill_value=0).reindex(apm, fill_value=0)
            if "NON CARACTERISE" in tl.columns: tl = tl.drop(columns=["NON CARACTERISE"])
            show_simple_pie(tl, "Types Caractérisation Planification")

        with sub_tab3:
            st.markdown("#### Performance par Poste — Radar")
            for poste in apm:
                if poste not in ckdf.index: continue
                vals = [float(ckdf.loc[poste, k]) if k in ckdf.columns else 0 for k in QK]
                fig = go.Figure(go.Scatterpolar(
                    r=vals,
                    theta=QK,
                    fill='toself',
                    name=poste,
                    line_color='#2563eb'
                ))
                cibles = [CIBLE.get(k, 100) for k in QK]
                fig.add_trace(go.Scatterpolar(
                    r=cibles,
                    theta=QK,
                    fill='none',
                    name='Cible',
                    line=dict(color='#ef4444', dash='dash', width=2)
                ))
                fig.update_layout(
                    polar=dict(radialaxis=dict(visible=True, range=[0, 120])),
                    showlegend=True,
                    height=400,
                    margin=dict(t=40, b=40, l=40, r=40),
                    title=dict(text=poste, font=dict(size=14))
                )
                st.plotly_chart(fig, use_container_width=True)

    # ============================================================
    # SAUVEGARDE EXCEL
    # ============================================================
    st.markdown("---")
    col_save1, col_save2 = st.columns(2)
    with col_save1:
        if st.button("💾 Sauvegarder les KPIs en Excel", use_container_width=True):
            # Préparer les données pour Excel
            p_rows = []
            for poste in apm:
                row = {"Poste de travail": poste}
                for kpi in QK:
                    row[kpi] = round(float(ckdf.loc[poste, kpi]), 1) if poste in ckdf.index and kpi in ckdf.columns else 0
                # Total General ligne
                n_good = sum(1 for kpi in QK if is_cell_good(kpi, row[kpi]))
                row["Total Général"] = round(n_good / len(QK) * 100, 1) if QK else 0
                p_rows.append(row)
            # Ligne Total General colonne
            tg_row = {"Poste de travail": "Total Général"}
            for kpi in QK:
                n_good = sum(1 for p in apm if p in ckdf.index and is_cell_good(kpi, ckdf.loc[p, kpi]))
                tg_row[kpi] = round(n_good / len(apm) * 100, 1) if apm else 0
            total_good = sum(1 for p in apm for kpi in QK if p in ckdf.index and is_cell_good(kpi, ckdf.loc[p, kpi]))
            tg_row["Total Général"] = round(total_good / (len(QK)*len(apm)) * 100, 1) if apm and QK else 0
            p_rows.append(tg_row)
            p_cols = ["Poste de travail"] + QK + ["Total Général"]

            q_rows = []
            for poste in apm:
                row = {"Poste de travail": poste}
                for kpi in PK:
                    row[kpi] = round(float(ckdf.loc[poste, kpi]), 1) if poste in ckdf.index and kpi in ckdf.columns else 0
                n_good = sum(1 for kpi in PK if is_cell_good(kpi, row[kpi]))
                row["Total Général"] = round(n_good / len(PK) * 100, 1) if PK else 0
                q_rows.append(row)
            tg_row_q = {"Poste de travail": "Total Général"}
            for kpi in PK:
                n_good = sum(1 for p in apm if p in ckdf.index and is_cell_good(kpi, ckdf.loc[p, kpi]))
                tg_row_q[kpi] = round(n_good / len(apm) * 100, 1) if apm else 0
            total_good_q = sum(1 for p in apm for kpi in PK if p in ckdf.index and is_cell_good(kpi, ckdf.loc[p, kpi]))
            tg_row_q["Total Général"] = round(total_good_q / (len(PK)*len(apm)) * 100, 1) if apm and PK else 0
            q_rows.append(tg_row_q)
            q_cols = ["Poste de travail"] + PK + ["Total Général"]

            # Anomalies Performance
            ano_p_rows = []
            for poste in apm:
                row = {"Poste de travail": poste}
                for kpi in QK:
                    ano = 0
                    if kpi in anomalies and poste in anomalies[kpi].index:
                        ano = int(anomalies[kpi].loc[poste])
                    if poste in ckdf.index and kpi in ckdf.columns and is_cell_good(kpi, ckdf.loc[poste, kpi]):
                        ano = 0
                    row[kpi] = ano
                row["Total"] = sum(row[k] for k in QK)
                ano_p_rows.append(row)
            ano_p_cols = ["Poste de travail"] + QK + ["Total"]

            # Anomalies Qualité
            ano_q_rows = []
            for poste in apm:
                row = {"Poste de travail": poste}
                for kpi in PK:
                    ano = 0
                    if kpi in anomalies and poste in anomalies[kpi].index:
                        ano = int(anomalies[kpi].loc[poste])
                    if poste in ckdf.index and kpi in ckdf.columns and is_cell_good(kpi, ckdf.loc[poste, kpi]):
                        ano = 0
                    row[kpi] = ano
                row["Total"] = sum(row[k] for k in PK)
                ano_q_rows.append(row)
            ano_q_cols = ["Poste de travail"] + PK + ["Total"]

            save_kpis_to_excel(p_rows, p_cols, q_rows, q_cols, ano_p_rows, ano_p_cols, ano_q_rows, ano_q_cols, date_input)
            st.success("✅ Fichier sauvegardé dans kpis/indicateurs_kpis.xlsx")

    with col_save2:
        hist_path = "kpis/indicateurs_kpis.xlsx"
        if os.path.exists(hist_path):
            hist_df = load_historical_kpis(hist_path)
            if not hist_df.empty:
                var_df = calculate_variations(hist_df)
                if not var_df.empty:
                    journal_df = generate_journal(var_df)
                    if not journal_df.empty:
                        st.markdown(f"**📋 Journal des variations significatives** ({len(journal_df)} entrées)")
                        with st.expander("Voir le journal"):
                            st.dataframe(journal_df, use_container_width=True, height=300)

                        best_df, worst_df = calculate_rankings(var_df)
                        if not best_df.empty and not worst_df.empty:
                            st.markdown('<div class="cg">', unsafe_allow_html=True)
                            st.markdown('<div><div class="ct">🏆 Top 5 Améliorations</div>', unsafe_allow_html=True)
                            for i, row in best_df.iterrows():
                                st.markdown('<div class="cgr"><div class="rk">%d</div><div class="pn">%s</div><div class="ps" style="color:#059669;">%+.1f</div></div>' % (i+1, row["Poste"], row["Score variation"]), unsafe_allow_html=True)
                            st.markdown('</div>', unsafe_allow_html=True)

                            st.markdown('<div><div class="ct">⚠️ Top 5 Dégradations</div>', unsafe_allow_html=True)
                            for i, row in worst_df.iterrows():
                                st.markdown('<div class="cgr"><div class="rk">%d</div><div class="pn">%s</div><div class="ps" style="color:#dc2626;">%+.1f</div></div>' % (i+1, row["Poste"], row["Score variation"]), unsafe_allow_html=True)
                            st.markdown('</div></div>', unsafe_allow_html=True)

    # Footer
    st.markdown('<div class="footer">Dashboard KPI Maintenance — Généré le %s</div>' % datetime.now().strftime("%d/%m/%Y %H:%M"), unsafe_allow_html=True)


if __name__ == "__main__":
    main()
